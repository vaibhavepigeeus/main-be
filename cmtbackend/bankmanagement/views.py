from enum import Enum
import mimetypes
import os
from dataclasses import dataclass

import requests
from documents.views import PolicyCreateViewSet

from django.shortcuts import render
from django.shortcuts import render
import json

from django.shortcuts import render
from rest_framework.decorators import api_view, permission_classes, action
from .serializers import *
from users.serializers import *
from rest_framework.response import Response
from rest_framework import generics, viewsets

from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.generics import ListCreateAPIView, ListAPIView
from .models import *
from django.conf import settings
from django.core.mail import send_mail
from django.utils.crypto import get_random_string
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse, HttpResponse
from rest_framework.pagination import PageNumberPagination
from rest_framework_swagger import renderers
from rest_framework.decorators import api_view, renderer_classes
from django.views.decorators.csrf import csrf_exempt
from .serializers import *
from documents.serializers import *
from documents.models import *
from cmtbackend.storage_backends import create_presigned_url
from datetime import datetime, date
from rest_framework import status
import openpyxl
from django.http import HttpResponse
from .pagination import CustomPagination
from decouple import config
from django.db.models import Q
from rest_framework.views import APIView
from .utils.operations import lock_or_unlock_cash_allocation, update_cash_allocation
from .utils.dashboard_operations import *
from django.db.models import Q
import calendar
from django.db.models import Sum, FloatField, DecimalField
from users.utils import send_email
from decimal import Decimal, ROUND_HALF_UP
from django.utils import timezone
from datetime import timedelta
import logging
import traceback
import io, re
from django.db.models import QuerySet
from django.utils.dateparse import parse_datetime
from knox.auth import TokenAuthentication
from io import BytesIO
from collections import defaultdict
from django.db.models import Case, When, Sum, F, Value
from django.core.exceptions import ValidationError

logger = logging.getLogger(__name__)

from dateutil import parser
from django.db.models.functions import Cast, Coalesce
from django.db.models import DateField
from django.db import transaction

from filemanagement.views import reusable_file_upload
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

PATH = config('BANK_FILES_PATH')

logger = logging.getLogger(__name__)

def get_date(dateStr):
    format_str = "%Y/%m/%d"
    formattedDate = datetime.strptime(dateStr, format_str)
    return formattedDate


def get_accounting_month(creation_date):
    """
    Get accounting month based on given date using the following rules:
    - Dates from 3rd of current month to 2nd of next month -> current month end
    - Dates from 3rd of month -> that month's end date
    """
    # Convert string to date if needed
    if isinstance(creation_date, str):
        creation_date = datetime.strptime(creation_date, '%Y-%m-%d').date()
    # Convert datetime to date if needed
    elif isinstance(creation_date, datetime):
        creation_date = creation_date.date()
        
    # Get all accounting months
    accounting_months = AccountingMonthEnd.objects.filter(accounting_month_start_date__lte=creation_date, accounting_month_end_date__gte=creation_date).order_by('accounting_month_end_date')
    
    if not accounting_months:
        return Response(
            {
                "message": "Calendar month not defined",
                "status": "error"
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    else:
        return accounting_months.first().accounting_month_date


def upload_file(request):
    if request.method == 'POST' and request.FILES['bank_details']:
        file = request.FILES['bank_details']
        if not os.path.exists(PATH):
            os.makedirs(PATH)
        full_path = PATH + '/' + str(file)
        with open(full_path, 'wb') as actual_file:
            actual_file.write(file.read())
        uploaded_file_url = full_path
        return render(request, 'upload.html', {
            'uploaded_file_url': uploaded_file_url
        })
    return render(request, 'upload.html')


class UploadFileView(APIView):
    # CMT-25
    @staticmethod
    def get_user(request):
        user_id = request.headers.get("user-id")
        user = Users.objects.filter(id=user_id)
        return user[0]
        # token = request.META.get('HTTP_AUTHORIZATION', False)
        # if token:
        #     token = str(token).split()[1].encode("utf-8")
        #     knoxAuth = TokenAuthentication()
        #     user, auth_token = knoxAuth.authenticate_credentials(token)
        #     request.user = user
        #     return user

    def post(self, request, format=None):
        user = self.get_user(request)               # CMT-25
        if request.FILES.get('bank_details'):
            file = request.FILES['bank_details']
            file_parts = str(file).split('-')          # CMT-25
            account_no = file_parts[0]

            # 0. check file doesnt exist in bank recon table
            if BankReconciliation.objects.filter(file_name=str(file),
                                             uploaded_status='uploaded').exists():
                return Response({"message": "Duplicate file"},status=status.HTTP_400_BAD_REQUEST)

            # 1. Check file has xlsx
            from bankmanagement.utils.generalUtils import isExcelFile 
            if not isExcelFile(str(file)):
                return Response({"message": "Invalid file extention - accepted xls or xlsx"} ,status=status.HTTP_400_BAD_REQUEST)

            # 2. check in list of bank acccoun numbers
            all_bank_accounts = ['62514844', '53549925', '53242122', '56878488', 'BE0190078', 'BE0190079', 'BE0190080', '30963411988808', 'GB12LOYD30801222032537', 'GB16LOYD30801286692374', 'GB26LOYD30801211988808',
                  'GB90LOYD30801221687268', 'GB90LOYD30801223067868', 'B77LOYD30801212016893','20240021609', '13726185', '416092', '16770019', '889031312', '889031339', '889031347', '011302460502', '011248663501', '80009471527', '80009950108', '80010417022',
                  '19000073673', '19000073674', '19000073675', '7799225208', '7793085226', '7729309684', '20218316', '21000500', '29006117', '80009842677', '7729309692']

            if str(account_no).strip() not in all_bank_accounts:
                return Response({"message": "File name does not match the pattern"},status=status.HTTP_400_BAD_REQUEST)

            # 3. check file name pattern
            from bankmanagement.utils.generalUtils import extract_account_number_and_date
            data = extract_account_number_and_date(str(file))
            if not data[0]:
                if not data[1]:
                    return Response({"message": "File name does not match the pattern"},status=status.HTTP_400_BAD_REQUEST)
            
            # CMT-25
            # 4. Modify the file name to include user ID
            user_id = str(user.id)
            modified_file_name = f"{user_id}_{str(file)}"
            display_file_name = str(file)

            today = date.today()
            new_ac = get_accounting_month(today)
            if isinstance(new_ac, Response):
                return new_ac

            if not os.path.exists(PATH):
                os.makedirs(PATH)
            full_path = os.path.join(PATH, modified_file_name)
            with open(full_path, 'wb') as actual_file:
                actual_file.write(file.read())
            uploaded_file_url = full_path
            return Response({'uploaded_file_url': display_file_name}, status=status.HTTP_200_OK)
        return Response(status=status.HTTP_400_BAD_REQUEST)

    def get(self, request, format=None):
        return render(request, 'upload.html')


class BankTransactionViewSet(viewsets.ModelViewSet):
    model = BankTransaction
    serializer_class = BankTransactionSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_user(self, request):
        user_id = request.headers.get("user-id")
        user = Users.objects.filter(id=user_id)
        return user[0]

    def get_date(self, dateStr):
        format = "%d/%m/%Y"
        formattedDate = datetime.strptime(dateStr, format)
        actualDate = formattedDate.strftime("%Y-%m-%d")
        return actualDate

    def get_queryset(self):
        trans = BankTransaction.objects.filter(archived=False).order_by('Payment_Receive_Date')
        return trans

    def list(self, request):
        filter_conditions = {}
        if request.GET.get("br_id"):
            br_id = int(request.GET.get("br_id"))
            trans = BankTransaction.objects.filter(
                bank_reconciliation_id=br_id,
                archived=False
            ).order_by("id")
        else:
            bank_name = request.GET.get("bankName", None)
            accountno = request.GET.get("accountNo", None)
            transactionid = request.GET.get("transactionId", None)
            fromdate = request.GET.get("fromDate", None)
            todate = request.GET.get("toDate", None)

            assigned = request.GET.get("assigned", None)
            transaction_status = request.GET.get("allocated", None)

            page_number = int(request.GET.get("skip", 0))
            rows_per_page = int(request.GET.get("pageSize", 20))
            skip = page_number * rows_per_page
            filter_conditions = {}
            if fromdate:
                fromdate = self.get_date(fromdate)
            if todate:
                todate = self.get_date(todate)

            if bank_name:
                filter_conditions["bank_details__bank_name"] = bank_name
            if accountno:
                filter_conditions["Receiving_Bank_Account"] = accountno
            if transactionid:
                filter_conditions["Bank_Transaction_Id__contains"] = transactionid
            if fromdate and todate:
                filter_conditions["Payment_Receive_Date__range"] = [fromdate, todate]
            # if assigned==="0":
            #     filter_conditions["Assigned_User_id"]= None

            if transaction_status:
                if transaction_status == 'Allocated':
                    filter_conditions["transaction_status__in"] = ["Completed"]
                else:
                    filter_conditions["transaction_status__in"] = ["In Progress - CC", "Open", "Query - Broker", "Query - F&A"] 

            if assigned == "Unassigned":
                trans = (

                    BankTransaction.objects.filter(archived=False)
                    .exclude(Assigned_User_id__isnull=False)
                    .filter(**filter_conditions)
                    .order_by("-id")[skip: skip + rows_per_page]

                )

            elif assigned == "Unassigned":
                trans = (

                    BankTransaction.objects.filter(archived=False)
                    .exclude(Assigned_User_id__isnull=True)
                    .filter(**filter_conditions)
                    .order_by("-id")[skip: skip + rows_per_page]

                )
            else:
                trans = (

                    BankTransaction.objects.filter(archived=False)
                    .filter(**filter_conditions)
                    .order_by("-id")[skip: skip + rows_per_page]

                )

        serializer = BankTransactionSerializer(trans, many=True)
        dataa = serializer.data
        for data in dataa:
            if data['File_Name']:
                file_path = data['File_Name']

                if 'media/bank/media' in file_path:
                    file_name = file_path.split('/media/bank/media/')[-1]
                    data['File_Name'] = file_name
                else:
                    data['File_Name'] = None
        data = {
            "count": BankTransaction.objects.filter(archived=False).filter(**filter_conditions).count(),
            "data": dataa,
        }
        return Response(data)

    def retrieve(self, request, pk=None):
        if pk:
            trans = BankTransaction.objects.get(id=pk)
            serializer = BankTransactionSerializer(trans)
            dataa = serializer.data
            return Response(dataa)
    
    def get_bank_exchange_rate(self, Bank_Currency_Code, last_day_of_month):
        try:
            # First, try to get the exact match
            bank_exchange_rate_obj = BankExchangeRate.objects.get(
                currency_code=str(Bank_Currency_Code).strip(), month=last_day_of_month
            )
            return bank_exchange_rate_obj.exchange_rate, last_day_of_month
        except BankExchangeRate.DoesNotExist:
            try:
                # If no exact match, get the latest record before the last_day_of_month
                bank_exchange_rate_obj = BankExchangeRate.objects.filter(
                    currency_code=str(Bank_Currency_Code).strip(),
                    month__lt=last_day_of_month,
                    month__isnull=False,
                ).latest("month")

                parsed_date = None
                if bank_exchange_rate_obj.month:
                    parsed_date = parser.parse(bank_exchange_rate_obj.month).date()

                return bank_exchange_rate_obj.exchange_rate, parsed_date
            except BankExchangeRate.DoesNotExist:
                raise ValueError(
                    "No valid exchange rate found for the given currency code."
                )

    def create(self, request, *args, **kwargs):

        upload_date = datetime.now()
        dataa = request.data
        transactions = []
        for data in dataa:
            bank_reconciliation_id = None
            if data["Receiving_Bank_Account"] and data["Payment_Receive_Date"]:
                filter_conditions = Q()
                filter_conditions &= (Q(bank_account_no=data['Receiving_Bank_Account']))
                filter_conditions &= (Q(file_date=data['Payment_Receive_Date']))
                filter_conditions &= ~Q(uploaded_status='rejected')

                bank_recon_is_exist = BankReconciliation.objects.filter(filter_conditions)

                if not bank_recon_is_exist:
                    bank_reconciliation_id = BankReconciliationViewSet.update_bank_reconciliation(
                        Decimal(data['Receivable_Amount']), Decimal(data['Bank_Charges']),
                        "create",
                        data['bank_details'],
                        data['Receiving_Bank_Account'],
                        data['Payment_Receive_Date'])
                else:
                    bank_reconciliation_id = BankReconciliationViewSet.update_bank_reconciliation(
                        Decimal(data['Receivable_Amount']), Decimal(data['Bank_Charges']),
                        "update",
                        data['bank_details'],
                        data['Receiving_Bank_Account'],
                        data['Payment_Receive_Date'], bank_recon_is_exist[0])

            try:
                g = BankTransaction.objects.filter(archived=False, Bank_Transaction_Id__startswith="BNKTXN").latest("id")
                last_trans_id = g.Bank_Transaction_Id
                if last_trans_id:
                    ll = last_trans_id[6:]
                    trl = str(int(ll) + 1).zfill(4)
                    trans_id_gen = "BNKTXN00" + trl
            except BankTransaction.DoesNotExist:
                trans_id_gen = "BNKTXN000001"

            payment_date = datetime.strptime(data["Payment_Receive_Date"], "%Y-%m-%d")
            last_day = calendar.monthrange(payment_date.year, payment_date.month)[1]
            accounting_month = payment_date.replace(day=last_day)
            logger.info(f'Receivable amount: {data["Receivable_Amount"]} Bank currency code: {data["Bank_Currency_Code"]} Accounting month: {accounting_month.date()}')
            #calculate receivable amount in USD
            roe, roe_date = self.get_bank_exchange_rate(data["Bank_Currency_Code"],accounting_month.date())
            logger.info(f'ROE: {roe} ROE date: {roe_date}')
            receivable_amount_usd = Decimal(data["Receivable_Amount"]) / roe

            logger.info(f"Receivable amount in USD: {receivable_amount_usd}")

            try:
                if data["broker_information"]:
                    broker = BrokerInformation.objects.get(id=data["broker_information"]).broker_name
                elif data['Broker_Branch']:
                    broker = BrokerInformation.objects.get(Broker_Branch=data['Broker_Branch']).broker_name
                else:
                    broker = None
            except Exception as e:
                logger.error(f"Error getting broker information: {e}")
                broker = None

            new_ac = get_accounting_month(data["Accounting_Month"])
            if isinstance(new_ac, Response):
                return new_ac
            
            trans = BankTransaction.objects.create(
                Bank_Transaction_Id=trans_id_gen,
                Accounting_Month=new_ac,
                PT_Receving_Bank_Name=data["PT_Receving_Bank_Name"],
                Bank_Account_Name_Entity=data["Bank_Account_Name_Entity"],
                Receiving_Bank_Account=data["Receiving_Bank_Account"],
                Broker_Branch=data["Broker_Branch"],
                Broker=broker,
                Payment_Receive_Date=data["Payment_Receive_Date"],
                Payment_Reference=data["Payment_Reference"],
                Payment_Currency_Code=data["Payment_Currency_Code"],
                Bank_Currency_Code=data["Bank_Currency_Code"],
                Bank_Exchange_Rate=data["Bank_Exchange_Rate"],
                Bank_Exchange_Charges=data["Bank_Exchange_Charges"],
                Bank_Charges=data["Bank_Charges"],
                Receivable_Amount=data["Receivable_Amount"],
                Receivable_Amount_USD=receivable_amount_usd,
                TL_Fees=data["TL_Fees"],
                Currency=data["Currency"],
                Created_By=data["Created_By"],
                Analyst_Name=data["Analyst_Name"],
                Date_And_Time=upload_date,
                Uploaded_By=data["Uploaded_By"],
                Allocation_Status=data["Allocation_Status"],
                broker_information=BrokerInformation.objects.get(
                    id=data["broker_information"]
                ) if data.get("broker_information") else None,
                bank_details=BankDetails.objects.get(id=data["bank_details"]),
                bank_reconciliation_id=bank_reconciliation_id if bank_reconciliation_id else None,
                ROE=roe,
                ROE_Date=roe_date,
                auto_upload=False,
                Aging_Bucket='0 - 5',
                txn_category=data["Txn_Category"],
                transaction_status=data["transaction_status"] if data.get("transaction_status") else "Open",
                comment=data.get("comment")
            )

            trans.save()

            # Save bank txn audit
            BankTransactionAudit.objects.create(
                bank_transaction=trans,
                audit_data={
                    "field_name": "-",
                    "old_value": "-",
                    "new_value": "-",
                    "previous_edit_datetime": "-",
                    "current_edit_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "changed_by": self.get_user(request).pk,
                    "event_type": "add"
                }
            )
            
            Date_And_Time = datetime.now()
            data["Date_And_Time"] = str(Date_And_Time)
            data_items = []
            data_items.append(data)
            fields = json.dumps(data_items)
            if len(fields) > 65535:
                return Response(
                    {
                        "msg": "Length of the value for update_fields property got exceeded!"
                    }
                )
            else:
                trans.updated_fields = fields
                trans.save()
            doc_f = data["File_Name"]
            transactions.append(trans)
            print("dc", doc_f)
            if doc_f:
                print("hereee")
                trans.File_Name = data["File_Name"]
                trans.save()
                transactions.append(trans)

        serializer = BankTransactionCreateSerializer(transactions, many=True)
        m = serializer.data
        return Response(m)

    def update(self, request, *args, **kwargs):
        trans_object = self.get_object()
        data = request.data
        trans_object.Bank_Transaction_Id = data["Bank_Transaction_Id"]
        trans_object.Accounting_Month = data["Accounting_Month"]
        trans_object.PT_Receving_Bank_Name = data["PT_Receving_Bank_Name"]
        trans_object.Bank_Account_Name_Entity = data["Bank_Account_Name_Entity"]
        trans_object.Receiving_Bank_Account = data["Receiving_Bank_Account"]
        trans_object.Broker_Branch = data["Broker_Branch"]
        trans_object.Broker = data["Broker"]
        trans_object.Payment_Receive_Date = data["Payment_Receive_Date"]
        trans_object.Bank_Exchange_Rate = data["Bank_Exchange_Rate"]
        trans_object.Bank_Currency_Code = data["Bank_Currency_Code"]
        trans_object.Bank_Exchange_Charges = data["Bank_Exchange_Charges"]
        trans_object.File_Name = data["File_Name"]
        trans_object.Bank_Charges = data["Bank_Charges"]
        trans_object.Receivable_Amount = data["Receivable_Amount"]
        trans_object.TL_Fees = data["TL_Fees"]
        trans_object.Currency = data["Currency"]
        trans_object.Created_By = data["Created_By"]
        trans_object.Analyst_Name = data["Analyst_Name"]
        trans_object.Date_And_Time = data["Date_And_Time"]
        trans_object.Uploaded_By = data["Uploaded_By"]
        trans_object.Allocation_Status = data["Allocation_Status"]
        trans_object.txn_category = data["txn_category"]
        trans_object.txn_type = data["txn_type"]

        trans_object.comment = data.get("comment") if data.get("comment") else trans_object.comment
        trans_object.batch_ref = data.get("batch_ref") if data.get("batch_ref") else trans_object.batch_ref
        trans_object.transaction_status = data.get("transaction_status") if data.get(
            "transaction_status") else trans_object.transaction_status

        trans_object.broker_information = BrokerInformation.objects.get(
            id=request.data["broker_information"]
        )
        trans_object.bank_details = BankDetails.objects.get(
            id=request.data["bank_details"]
        )

        trans_object.save()

        serializer = BankTransactionCreateSerializer(trans_object)
        dataa = serializer.data
        report_tracker = CashTrackerReport.objects.filter(
            bank_txn_id=trans_object.id
        ).update(
            Accounting_Month=data["Accounting_Month"],
            PT_Receving_Bank_Account_Name=data["PT_Receving_Bank_Name"],
            Receiving_Bank_Account=data["Receiving_Bank_Account"],
            Broker=data["Broker"],
            Broker_Branch=data["Broker_Branch"],
            Payment_Receive_Date=data["Payment_Receive_Date"],
            Payment_Reference=data["Payment_Reference"],
            Payment_Currency_Code=data["Payment_Currency_Code"],
            Bank_Currency_Code=data["Bank_Currency_Code"],
            Payment_Special_Instructions=data["Payment_Reference"],
            Bank_Charges=data["Bank_Charges"],
        )

        return Response(dataa)

    def partial_update(self, request, *args, **kwargs):
        doc_object = self.get_object()
        data = request.data

        # Save bank txn audit
        for key, value in dict(data).items():
            if key not in ['id', 'updated_by']:
                # Users.objects.get(id=5).pk
                BankTransactionAudit.objects.create(
                    bank_transaction=doc_object,
                    audit_data={
                        "field_name": key,
                        "old_value": str(getattr(doc_object, key, "")),
                        "new_value": value,
                        "previous_edit_datetime": doc_object.updated_at.strftime("%Y-%m-%d %H:%M:%S") if doc_object.updated_at else "-",
                        "current_edit_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "changed_by": self.get_user(request).pk,
                        "event_type": "edit"
                    }
                )

        doc_object.Bank_Transaction_Id = data.get(
            "Bank_Transaction_Id", doc_object.Bank_Transaction_Id
        )
        doc_object.Accounting_Month = data.get(
            "Accounting_Month", doc_object.Accounting_Month
        )
        doc_object.PT_Receving_Bank_Name = data.get(
            "PT_Receving_Bank_Name", doc_object.PT_Receving_Bank_Name
        )
        doc_object.Bank_Account_Name_Entity = data.get(
            "Bank_Account_Name_Entity", doc_object.Bank_Account_Name_Entity
        )
        doc_object.Receiving_Bank_Account = data.get(
            "Receiving_Bank_Account", doc_object.Receiving_Bank_Account
        )
        doc_object.Broker_Branch = data.get("Broker_Branch", doc_object.Broker_Branch)
        doc_object.Broker = data.get("Broker", doc_object.Broker)
        doc_object.Payment_Receive_Date = data.get(
            "Payment_Receive_Date", doc_object.Payment_Receive_Date
        )
        doc_object.Payment_Reference = data.get(
            "Payment_Reference", doc_object.Payment_Reference
        )
        doc_object.Payment_Currency_Code = data.get(
            "Payment_Currency_Code", doc_object.Payment_Currency_Code
        )
        doc_object.Bank_Currency_Code = data.get(
            "Bank_Currency_Code", doc_object.Bank_Currency_Code
        )
        doc_object.Bank_Exchange_Rate = data.get(
            "Bank_Exchange_Rate", doc_object.Bank_Exchange_Rate
        )
        doc_object.Bank_Exchange_Charges = data.get(
            "Bank_Exchange_Charges", doc_object.Bank_Exchange_Charges
        )
        doc_object.Bank_Charges = data.get("Bank_Charges", doc_object.Bank_Charges)
        doc_object.Receivable_Amount = data.get(
            "Receivable_Amount", doc_object.Receivable_Amount
        )
        doc_object.TL_Fees = data.get("TL_Fees", doc_object.TL_Fees)
        doc_object.Currency = data.get("Currency", doc_object.Currency)
        doc_object.Created_By = data.get("Created_By", doc_object.Created_By)
        doc_object.Analyst_Name = data.get("Analyst_Name", doc_object.Analyst_Name)
        doc_object.Allocation_Status = data.get(
            "Allocation_Status", doc_object.Allocation_Status
        )
        doc_object.File_Name = data.get("File_Name", doc_object.File_Name)
        doc_object.Date_And_Time = data.get("Date_And_Time", doc_object.Date_And_Time)
        doc_object.Uploaded_By = data.get("Uploaded_By", doc_object.Uploaded_By)
        doc_object.updated_by = data.get("updated_by", doc_object.updated_by)
        doc_object.txn_category = data.get("txn_category", doc_object.txn_category)
        doc_object.txn_type = data.get("txn_type", doc_object.txn_type)
        doc_object.correction_type = data.get("correction_type", doc_object.correction_type)

        doc_object.updatedDateAndTime = datetime.now()
        doc_object.broker_information = (
            BrokerInformation.objects.get(id=data.get("broker_information"))
            if data.get("broker_information")
            else doc_object.broker_information
        )
        doc_object.bank_details = (
            BankDetails.objects.get(id=data.get("bank_details"))
            if data.get("bank_details")
            else doc_object.bank_details
        )
        updatedatetime = datetime.now()
        data["updatedDateAndTime"] = str(updatedatetime)
        data_items = []
        if doc_object.updated_fields:
            old_list = doc_object.updated_fields
            old_list = json.loads(old_list)
            data_items.extend(old_list)
        data_items.append(data)
        changedFields = json.dumps(data_items)
        if len(changedFields) > 65535:
            return Response(
                {"msg": "Length of the value for update_fields property got exceeded!"}
            )
        else:
            doc_object.updated_fields = changedFields

        doc_object.comment = data.get("comment") if data.get("comment") else doc_object.comment
        doc_object.batch_ref = data.get("batch_ref") if data.get("batch_ref") else doc_object.batch_ref
        doc_object.transaction_status = data.get("transaction_status") if data.get(
            "transaction_status") else doc_object.transaction_status

        doc_object.save()
        serializer = BankTransactionCreateSerializer(doc_object)
        dataa = serializer.data

        report_tracker = CashTrackerReport.objects.filter(
            bank_txn_id=doc_object.id
        ).update(
            Accounting_Month=data.get("Accounting_Month", doc_object.Accounting_Month),
            PT_Receving_Bank_Account_Name=data.get(
                "PT_Receving_Bank_Name", doc_object.PT_Receving_Bank_Name
            ),
            Receiving_Bank_Account=data.get(
                "Receiving_Bank_Account", doc_object.Receiving_Bank_Account
            ),
            Broker=data.get("Broker", doc_object.Broker),
            Broker_Branch=data.get("Broker_Branch", doc_object.Broker_Branch),
            Payment_Receive_Date=data.get(
                "Payment_Receive_Date", doc_object.Payment_Receive_Date
            ),
            Payment_Reference=data.get(
                "Payment_Reference", doc_object.Payment_Reference
            ),
            Payment_Currency_Code=data.get(
                "Payment_Currency_Code", doc_object.Payment_Currency_Code
            ),
            Bank_Currency_Code=data.get(
                "Bank_Currency_Code", doc_object.Bank_Currency_Code
            ),
            Payment_Special_Instructions=data.get(
                "Payment_Reference", doc_object.Payment_Reference
            ),
            Bank_Charges=data.get("Bank_Charges", doc_object.Bank_Charges),
        )
        return Response(dataa)

    def destroy(self, request, *args, **kwargs):
        logger.info("Starting the deletion process")
        doc_object = self.get_object()
        cash_allocation = CashAllocation.objects.filter(bank_txn=doc_object)
        if cash_allocation.exists():
            for allocation in cash_allocation:
                allocation.archived = True
                allocation.save()
                logger.debug(f"Archived cash allocation: {allocation.id}")
        doc_object.archived = True
        doc_object.save()
        logger.debug(f"Archived document object: {doc_object.id}")
        logger.info("Bank statement deleted successfully")
        return Response({"message": "bank statement deleted successfully"})


from users.models import Users


class UserAssignedForTransactionViewSet(viewsets.ModelViewSet):
    model = BankTransaction
    serializer_class = BankTransactionSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    today = date.today()

    def create(self, request, *args, **kwargs):
        data = request.data

        user_id = data["user_id_assigned"]
        users_list = []
        for user_ids in user_id:
            if Users.objects.filter(id=user_ids["id"]).exists():
                users_list.append(user_ids["id"])
            else:
                return Response(
                    {"message": f'User with id {user_ids["id"]} does not exist.'}
                )
        trans = BankTransaction.objects.filter(
            Bank_Transaction_Id=data["Bank_Transaction_Id"],
            archived=False
        ).last()
        print(trans, "ffffffffffffffffffffffffffffffffff")
        for users_id in users_list:
            trans.Assigned_Users.add(users_id)
            trans.transaction_status = "Open"
            trans.save()
        trans.assigned_date = UserAssignedForTransactionViewSet.today
        trans.save()
        serializer = BankTransactionSerializer(trans)
        dataa = serializer.data
        return Response(dataa)


class UserAssignTransactionsViewSet(viewsets.ModelViewSet):
    model = BankTransaction
    serializer_class = BankTransactionSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = BankTransaction.objects.filter(archived=False).order_by('-assigned_date')
        return objs

    def list(self, request):
        user_id = request.GET.get("user_id", None)
        if user_id:
            user_id = int(user_id)
        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        bankName = request.GET.get("bankName", None)
        transactionId = request.GET.get("transactionId", None)
        allocation_Status = request.GET.get("allocationStatus", None)
        transaction_status = request.GET.get("allocated", None)
        payment_rec_date_from = request.GET.get("payment_rec_date_from", None)
        payment_rec_date_to = request.GET.get("payment_rec_date_to", None)
        skip = page_number * rows_per_page
        filter_conditions = {}
        exclude_conditions = {}
        if bankName:
            filter_conditions["bank_details_id__bank_name__icontains"] = bankName
        if transactionId:
            filter_conditions["Bank_Transaction_Id"] = transactionId
        if user_id:
            filter_conditions["Assigned_User_id"] = user_id
        # CMT-28
        if allocation_Status and transaction_status:
            if allocation_Status:
                filter_conditions["cashallocation__allocation_status"] = allocation_Status
        else:
            if allocation_Status:
                filter_conditions["cashallocation__allocation_status"] = allocation_Status
            if transaction_status:
                if transaction_status == "Allocated":
                    filter_conditions["transaction_status"] = "Completed"
                else:
                    exclude_conditions["transaction_status"] = "Completed"

        if payment_rec_date_from and payment_rec_date_to:
            payment_rec_date_obj = datetime.strptime(payment_rec_date_from, "%d-%m-%Y")
            payment_rec_date_obj_to = datetime.strptime(payment_rec_date_to, "%d-%m-%Y")
            filter_conditions["Payment_Receive_Date__range"] = [payment_rec_date_obj, payment_rec_date_obj_to]
        trans = (
            self.get_queryset()
            .filter(**filter_conditions)
            .exclude(**exclude_conditions)
            .order_by("-assigned_date")[skip: skip + rows_per_page]
        )
        serializer = BankTransactionSerializer(trans, many=True)
        m = serializer.data
        data = {}

        data["data"] = m
        data["count"] = (
            self.get_queryset().filter(**filter_conditions).exclude(**exclude_conditions).order_by("-id").count()
        )
        return Response(data, status=status.HTTP_200_OK)

    def create(self, request):
        pass


class MultiAssignTransactionsToUserViewSet(viewsets.ModelViewSet):
    model = BankTransaction
    serializer_class = BankTransactionSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_user(self, request):
        user_id = request.headers.get("user-id")
        user = Users.objects.filter(id=user_id)
        return user[0]

    def bank_txn_audit(self, request, obj, field_name, old_value, new_value, previous_edit_datetime=None):
        BankTransactionAudit.objects.create(
            bank_transaction=obj,
            audit_data={
                "field_name": field_name,
                "old_value": old_value,
                "new_value": new_value,
                "previous_edit_datetime": previous_edit_datetime,
                "current_edit_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "changed_by": self.get_user(request).pk,
                "event_type": "edit"
            }
        )

    def create(self, request, *args, **kwargs):
        today = date.today()
        data = request.data
        user_id = data["user_login_id"]
        Bank_Transaction_Ids = data["Bank_Transaction_Ids"]
        user_id = Users.objects.get(id=user_id)
        transss = []

        for id in Bank_Transaction_Ids:
            trans = BankTransaction.objects.get(id=id["id"])

            # Save audit data for bank transaction Assigned_User field
            self.bank_txn_audit(
                request, 
                trans,
                "Assigned User", 
                trans.Assigned_User.user_name if trans.Assigned_User else "-", 
                user_id.user_name, 
                trans.updated_at.strftime("%Y-%m-%d %H:%M:%S") if trans.updated_at else "-"
            )

            trans.Assigned_User = user_id
            trans.Assigned_Users.clear()
            trans.Assigned_Users.add(user_id)
            trans.assigned_date = today

            if user_id and trans.transaction_status in ["Open", ""]:
                # Save audit data for bank transaction transaction_status field
                self.bank_txn_audit(
                    request, 
                    trans,
                    "transaction_status", 
                    trans.transaction_status, 
                    "In Progress - CC", 
                    trans.updated_at.strftime("%Y-%m-%d %H:%M:%S") if trans.updated_at else "-"
                )

                trans.transaction_status = "In Progress - CC"           # CMT-28

            trans.save()

            # ff=trans.get_assigned_users_list()
            #
            # serializer = UserSerializer(user_id)
            # data = serializer.data
            # data["assigned_users_number"] = number
            # number = number + 1
            # assigned_users_list.append(data)
            # assigned_users_list = sorted(assigned_users_list, key=lambda d: d['assigned_users_number'])
            # trans.set_assigned_users_list(assigned_users_list)
            # trans.save()
            transss.append(trans)
        serializer = BankTransactionSerializer(transss, many=True)
        email_from = settings.EMAIL_HOST_USER
        email_to = user_id.get_decrypted_email()
        recipient_list = [email_to]

        imagepath = config('MOSAIC_LOGO_IMAGE')

        subject = "New Bank Statement Transaction Assigned to You"
        body = """
        <html>
            <head>
            </head>
            <body>
                <div class="email-body">
                    <div class="email-header">
                        <img src="{imagepath}" alt="Mosaic Insurance Logo">
                    </div>
                    <p>Dear User,</p>
                    <p>A new transaction has been assigned to you. Please review it as soon as possible.</p>
                    <p>If you have any questions, email us at <a href="mailto:support@mosaicinsurance.com">support@mosaicinsurance.com</a>.</p>
                    <p>If you no longer wish to receive these email notifications, you can unsubscribe by replying to this email with "Unsubscribe" in the subject line.</p>

                    <div class="email-footer">
                        <p>Regards,<br>Mosaic Insurance</p>
                    </div>
                    <p>This message, including any attachments, may include proprietary or confidential material. Any distribution or use of this communication by anyone other than the intended recipient(s) is prohibited. If you are not the intended recipient, please notify the sender by replying to this message and then deleting it from your system.</p>
                </div>
            </body>
        </html>
        """.format(imagepath=imagepath)
        try:
            send_email(sender_email=email_from, recipient_email=recipient_list, subject=subject, body=body)
        except:
            pass
        m = serializer.data
        return Response(m)


class CashAllocationViewSet(viewsets.ModelViewSet):
    queryset = CashAllocation.objects.filter(archived=False)
    serializer_class = CashAllocationSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)    
    
    def get_user(self, request):
        user_id = request.headers.get("user-id")
        user = Users.objects.filter(id=user_id)
        return user[0]

    @staticmethod
    def get_accounting_date(date_of_month, is_date=True):
        """ This method to calculate accounting month year for current month """

        year = date_of_month.year
        month = date_of_month.month
        last_day = calendar.monthrange(year, month)[1]

        last_day_of_month = date(year, month, last_day) if is_date else datetime(year, month, last_day)

        return last_day_of_month

    @staticmethod
    def get_bank_exchange_rate(bank_transaction_obj):
        last_day_of_month = CashAllocationViewSet.get_accounting_date(bank_transaction_obj.Payment_Receive_Date)
        logger.info(f'last_day_of_month: {last_day_of_month}')
        try:
            # First, try to get the exact match
            bank_exchange_rate_obj = BankExchangeRate.objects.get(
                currency_code=str(bank_transaction_obj.Bank_Currency_Code).strip(), month=last_day_of_month
            )
            return bank_exchange_rate_obj.exchange_rate, last_day_of_month
        except BankExchangeRate.DoesNotExist:
            try:
                # If no exact match, get the latest record before the last_day_of_month
                bank_exchange_rate_obj = BankExchangeRate.objects.filter(
                    currency_code=str(bank_transaction_obj.Bank_Currency_Code).strip(),
                    month__lt=last_day_of_month,
                    month__isnull=False,
                ).latest("month")

                parsed_date = None
                if bank_exchange_rate_obj.month:
                    parsed_date = parser.parse(bank_exchange_rate_obj.month).date()

                return bank_exchange_rate_obj.exchange_rate, parsed_date
            except BankExchangeRate.DoesNotExist:
                raise ValueError(
                    "No valid exchange rate found for the given currency code."
                )

    @staticmethod
    def get_user(request):
        user_id = request.headers.get("user-id")
        user = Users.objects.filter(id=user_id)
        return user[0]
        # token = request.META.get('HTTP_AUTHORIZATION', False)
        # if token:
        #     token = str(token).split()[1].encode("utf-8")
        #     knoxAuth = TokenAuthentication()
        #     user, auth_token = knoxAuth.authenticate_credentials(token)
        #     request.user = user
        #     return user

    def get_queryset(self):
        objs = CashAllocation.objects.filter(archived=False).order_by('-created_at', '-id')
        return objs

    @staticmethod
    def get_iso_week_number(date_obj):
        """Returns a string in the format YYYY MM W[1,2,3,4,5]."""
        year = date_obj.strftime("%Y")
        month = date_obj.strftime("%m")
        
        # Calculate which week of the month it is
        first_day_of_month = date_obj.replace(day=1)
        adjusted_day = (date_obj.day + first_day_of_month.weekday()) - 1
        week_of_month = (adjusted_day // 7) + 1
        
        return f"{year} {month} W{week_of_month}"

    @staticmethod
    def create_policy_info(request, cash_tracker_report_obj, bankTxn, bankExc, roe_date):

        policy_pk = request.data["policy_pk"]

        policy_info = CashAllocationViewSet.create_policy_fk(policy_pk)
        policy_line_ref = policy_info.Policy_Line_Ref

        cash_tracker_report_obj.Third_Party_Capacity = request.data.get('Third_Party_Capacity', '')

        cash_tracker_report_obj.Currency = policy_info.Original_Ccy
        # cash_tracker_report_obj.Mosaic_percent = policy_info.
        cash_tracker_report_obj.Coverholder_Fee_Amount = (
            policy_info.Coverholder_Commision_Pct
        )
        cash_tracker_report_obj.MGA_Commission = str(
            float(policy_info.Coverholder_Commision_Pct or 0)
            * float(policy_info.Net_Written_Premium_100_in_Orig or 0)
        )

        comment = request.data.get("comment")
        cash_tracker_report_obj.comment = comment

        cash_tracker_report_obj.save()

        logger.info(f'policy_line_ref {policy_line_ref}')
        try:
            cash_tracker_report_obj.Policy_Type = PolicyType.objects.filter(policy_start_letter=policy_line_ref[0]).first().policy_type
            logger.info(f'cash_tracker_report_obj.Policy_Type {cash_tracker_report_obj.Policy_Type}')

            cash_tracker_report_obj.LOB = LOB.objects.filter(lob_code=policy_line_ref[1:3]).first().lob

            logger.info(f'select lob {LOB.objects.filter(lob_code=policy_line_ref[1:3]).first().lob}')

            logger.info(f'cash_tracker_report_obj.LOB {cash_tracker_report_obj.LOB}')

            # Assuming the YOA is always the last 2 characters
            yoa_short = policy_line_ref[-4:-2]
            yoa_full = f"20{yoa_short}"
            cash_tracker_report_obj.YOA = yoa_full

        except Exception as e:
            logger.info(f'Policy type or LOB or YOA or Binding_Agreement not found for {policy_line_ref}')
            cash_tracker_report_obj.Policy_Type = 'Missing Policy Number'
            cash_tracker_report_obj.LOB = 'Missing Policy Number'
            cash_tracker_report_obj.YOA = 'Missing Policy Number'
            cash_tracker_report_obj.Binding_Agreement = 'Missing Policy Number'

        cash_tracker_report_obj.Remaining_Balance = Decimal(request.data["receivable_amt"] or 0) - Decimal(
            request.data["allocated_amt"] or 0)

        cash_tracker_report_obj.Total_Receivables_including_Bank_Charges = Decimal(
            request.data["receivable_amt"] or 0) + Decimal(bankTxn.Bank_Charges or 0)

        if bankExc:
            if bankExc:
                try:
                    cash_tracker_report_obj.Receivable_Amount_calculated = Decimal(
                        request.data["receivable_amt"] or 0) / Decimal(bankExc or 1)
                except:
                    pass
            try:
                cash_tracker_report_obj.Allocated_Amount_calculated = Decimal(
                    request.data["allocated_amt"] or 0) / Decimal(bankExc or 1)
            except:
                pass
            if bankExc and bankExc != 0:
                try:
                    bank_charges = Decimal(bankTxn.Bank_Charges or 0)
                    cash_tracker_report_obj.Bank_Charges_calculated = (bank_charges / Decimal(bankExc)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                except Exception as e:
                    cash_tracker_report_obj.Bank_Charges_calculated = Decimal('0.00')
            else:
                cash_tracker_report_obj.Bank_Charges_calculated = Decimal('0.00')

        cash_tracker_report_obj.Remaining_Balance_usd = Decimal(
            cash_tracker_report_obj.Receivable_Amount_calculated or 0) - Decimal(
            cash_tracker_report_obj.Allocated_Amount_calculated or 0)
        cash_tracker_report_obj.MGA_Commission_calculated = Decimal(request.data["bank_roe"] or 0) * Decimal(
            cash_tracker_report_obj.MGA_Commission or 0)

        if float(cash_tracker_report_obj.Remaining_Balance or 0) > 0:
            cash_tracker_report_obj.Payment_Status = "Under Payment"
        elif float(cash_tracker_report_obj.Remaining_Balance or 0) < 0:
            cash_tracker_report_obj.Payment_Status = "Over Payment"
        elif float(cash_tracker_report_obj.Remaining_Balance or 0) == 0:
            cash_tracker_report_obj.Payment_Status = "Allocated"
        if (
                request.data["allocation_systemid"] in ["XFI", "BOTH"]
                and request.data["allocation_status"] == "Allocated"
        ):
            cash_tracker_report_obj.XFI_Allocated_Date = request.data["allocation_date"]

        if (
            request.data["allocation_systemid"] in ["GXB", "BOTH"]
            and request.data["allocation_status"] == "Allocated"
        ):
            cash_tracker_report_obj.GXB_Status = "Allocated"

        if float(cash_tracker_report_obj.Receivable_Amount_calculated or 0) != 0.0:
            if not cash_tracker_report_obj.Allocated_Amount_calculated:
                cash_tracker_report_obj.Allocated_Amount_calculated = 0.00
            if not cash_tracker_report_obj.Receivable_Amount_calculated:
                cash_tracker_report_obj.Receivable_Amount_calculated = 0.00

            cash_tracker_report_obj.percent_Change = (
                                                             float(cash_tracker_report_obj.Receivable_Amount_calculated)
                                                             - float(
                                                         cash_tracker_report_obj.Allocated_Amount_calculated)
                                                     ) / float(cash_tracker_report_obj.Receivable_Amount_calculated)
        else:
            # Handle the case where Receivable_Amount_calculated is 0
            cash_tracker_report_obj.percent_Change = 0  # Or another default value

        cash_tracker_report_obj.Payment_Receive_Week = CashAllocationViewSet.get_iso_week_number(
            bankTxn.Payment_Receive_Date)

        if request.data["allocation_date"]:
            cash_tracker_report_obj.Allocation_Week = CashAllocationViewSet.get_iso_week_number(
                datetime.strptime(request.data["allocation_date"],
                                  "%Y-%m-%d")
            )
            cash_tracker_report_obj.Reporting_Week = CashAllocationViewSet.get_iso_week_number(
                datetime.strptime(request.data["allocation_date"],
                                  "%Y-%m-%d")
            )
            
        else:
            cash_tracker_report_obj.Allocation_Week = None
        if cash_tracker_report_obj.Payment_Receive_Date == None:
            cash_tracker_report_obj.no_of_Days_Overdue = None
        elif cash_tracker_report_obj.XFI_Allocated_Date == None:
            import numpy

            date1 = bankTxn.Payment_Receive_Date.date()
            date2 = date.today()
            cash_tracker_report_obj.no_of_Days_Overdue = numpy.busday_count(date1, date2)
        elif cash_tracker_report_obj.Payment_Receive_Date and cash_tracker_report_obj.XFI_Allocated_Date:
            import numpy

            date1 = bankTxn.Payment_Receive_Date.date()
            date2 = datetime.strptime(
                cash_tracker_report_obj.XFI_Allocated_Date.split(' ')[0], "%Y-%m-%d"
            ).date()
            cash_tracker_report_obj.no_of_Days_Overdue = numpy.busday_count(date1, date2)

        if cash_tracker_report_obj.no_of_Days_Overdue:
            cash_tracker_report_obj.SLA = "Status - Set Up"
        
        if bankExc:
            logger.info(f'bankExc: {bankExc} roe_date: {roe_date}')
            cash_tracker_report_obj.ROE = bankExc
            cash_tracker_report_obj.ROE_Date = roe_date
            logger.info(f'cash_tracker_report_obj.ROE: {cash_tracker_report_obj.ROE}')
            logger.info(f'cash_tracker_report_obj.ROE_Date: {cash_tracker_report_obj.ROE_Date}')

        cash_tracker_report_obj.save()

    @staticmethod
    def create_policy_fk(policy_pk):
        return PolicyInformation.objects.get(id=policy_pk)

    @staticmethod
    def get_cash_allocation_id_and_status(bank_transaction_obj):
        """ This method to get no of count of cash allocation for given bank transaction id """

        cash_allocations = CashAllocation.objects.filter(bank_txn=bank_transaction_obj, archived=False)

        cash_allocations_id_and_status = []

        for cash_allocation in cash_allocations:
            cash_allocations_id_and_status.append(cash_allocation.allocation_status)

        return cash_allocations_id_and_status

    @staticmethod
    def update_bank_transaction_status(bank_transaction_obj, cash_allocations_id_and_status):
        all_allocated = all(status_value.lower() == "allocated" for status_value in cash_allocations_id_and_status)

        bank_transaction_obj.transaction_status = "Completed" if all_allocated else "In Progress"
        bank_transaction_obj.save()

    @classmethod
    def create_cash_tracker_report(cls, request, obj):
        # Mapping saved object with cashtracker report object.

        bankTxn = BankTransaction.objects.get(id=request.data["bank_txn"])

        remaining_balance = obj.receivable_amt - obj.allocated_amt
        roe,roe_date = CashAllocationViewSet.get_bank_exchange_rate(bankTxn)
        logger.info(f'create cash tracker report roe: {roe} roe_date: {roe_date}')
        remaining_balance_usd = remaining_balance / roe
        logger.info(f'remaining_balance_usd: {remaining_balance_usd}')
        try:
            including_bank_charges = obj.receivable_amt + request.data['bank_charges']
        except:
            including_bank_charges = 0
        try:
            account_month = bankTxn.Payment_Receive_Date
            if account_month:
                # Get the last day of the month
                last_day = calendar.monthrange(account_month.year, account_month.month)[1]
                account_month = date(account_month.year, account_month.month, last_day)
        except Exception as e:
            logger.error("Error while getting account month",e)
            account_month = None
        payment_rec_date = bankTxn.Payment_Receive_Date.date() if bankTxn.Payment_Receive_Date else None
        user = cls.get_user(request)
        report_tracker = CashTrackerReport.objects.create(
            Accounting_Month=bankTxn.Accounting_Month,
            Invoice_Verification=request.data["allocation_invoice_verification"],
            Producing_Coverholder=request.data["allocation_entity"],
            PT_Receving_Bank_Account_Name=bankTxn.PT_Receving_Bank_Name,
            Receiving_Bank_Account=bankTxn.Receiving_Bank_Account,
            Payment_Receive_Date=payment_rec_date,
            Payment_Reference=bankTxn.Payment_Reference,
            Broker=bankTxn.Broker,
            Broker_Branch=bankTxn.Broker_Branch,
            Payment_Currency_Code=bankTxn.Bank_Currency_Code,
            Bank_Currency_Code=bankTxn.Bank_Currency_Code,
            Binding_Agreement=request.data["binding_agreement"],
            SCM_Partners=request.data["allocation_scm"],
            EEA_NonEEA=request.data.get("allocation_eea", None),
            ROE_Bank_Statement=request.data["bank_roe"],
            Master_Binder=request.data["allocation_binder"],
            Policy=request.data["policy_id"],
            Receivable_Amount=request.data["receivable_amt"],
            Allocated_Amount=request.data["allocated_amt"],
            Account_Handler=bankTxn.Assigned_User.user_name,
            Allocation_Status=request.data["allocation_status"],
            System_Correction=request.data["allocation_systemid"],
            bank_txn_id=request.data["bank_txn"],
            cash_allocation=obj,
            created_by=user,
            updated_by=user,
            policy_information= PolicyInformation.objects.get(id=request.data["policy_pk"]),
            # added on 13th july
            Remaining_Balance_usd=remaining_balance_usd,
            Total_Receivables_including_Bank_Charges=including_bank_charges,
            ROE=roe,
            ROE_Date=roe_date,
            Aging_Bucket='0 - 5'
        )

        return report_tracker

    # CMT-28
    @staticmethod
    def get_total_allocated_amount(bank_txn_id):
        return CashAllocation.objects.filter(bank_txn=bank_txn_id, allocation_status="Allocated", archived=False).aggregate(
            total_allocated=Sum('allocated_amt')
        )['total_allocated'] or 0

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        try:
            data=request.data
            old_ca_obj = None
            if data.get('deleted_ca'):
                data['deleted_time'] = datetime.now()
                old_ca_obj = CashAllocation.objects.get(id=data['deleted_ca'])
            serializer = CashAllocationCreateSerializer(data=data)

            if serializer.is_valid():
                serializer_saved = serializer.save()
                bankTxn = BankTransaction.objects.get(id=data["bank_txn"])

                # Handling Contra Allocation Creation
                if "is_contra_allocation" in data and data["is_contra_allocation"]:
                    serializer_saved.is_contra_allocation = True
                    serializer_saved.parent_contra_id = data["parent_contra_id"]

                    # Updating child contra allocation id to parent cash allocation
                    parent_cash_allocation = CashAllocation.objects.get(id=data["parent_contra_id"])
                    parent_cash_allocation.child_contra_id = serializer_saved.id
                    parent_cash_allocation.save()

                # Changed because Cash Allocation will never be allocated at the creation time.
                # if serializer_saved.allocation_status.lower() == "allocated":
                #     today = datetime.now()
                #     serializer_saved.accounting_monthyear = CashAllocationViewSet.get_accounting_date(today)
                #     serializer_saved.allocation_date = today.date()
                # else:
                #     serializer_saved.accounting_monthyear = CashAllocationViewSet.get_accounting_date(
                #         bankTxn.Payment_Receive_Date)
                #     try:
                #         serializer_saved.allocation_date = None
                #     except Exception as e:
                #         logger.error(f"Error saving allocation date: {e}")
                bank_recon_id = BankTransaction.objects.filter(archived=False, id=data["bank_txn"]).values(
                    'bank_reconciliation').first()
                serializer_saved.bank_reconciliation_id = bank_recon_id.get('bank_reconciliation')
                serializer_saved.policy_fk = CashAllocationViewSet.create_policy_fk(data['policy_pk'])
                user = self.get_user(request)
                if user:
                    serializer_saved.created_by = user
                    serializer_saved.updated_by = user

                # Explicitly set policy handler fields since serializer marks them read-only
                try:
                    serializer_saved.policy_handler = old_ca_obj.policy_handler if data.get('deleted_ca') else bankTxn.Assigned_User
                    serializer_saved.policy_assign_date = datetime.now().date()
                except Exception as e:
                    logger.error(f"Error setting policy handler on create: {e}")

                serializer_saved.save()
                CashAllocaionAudit.objects.create(
                    cash_allocation=serializer_saved,
                    audit_data={
                        "field_name": "Policy" if old_ca_obj else "-",
                        "old_value": old_ca_obj.policy_id if old_ca_obj else "-",
                        "new_value": serializer_saved.policy_id if old_ca_obj else "-",
                        "previous_edit_datetime": "-",
                        "current_edit_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "changed_by": user.pk,
                        "event_type": "add"
                    }
                )
                try:
                    bankTxn = BankTransaction.objects.get(id=data["bank_txn"])

                    cash_tracker_report_obj = CashAllocationViewSet.create_cash_tracker_report(request, serializer_saved)
                    if data.get('policy_id'):
                        bankExc, roe_date = CashAllocationViewSet.get_bank_exchange_rate(bank_transaction_obj=bankTxn)
                        logger.info(f'create cash allocation bankExc: {bankExc} roe_date: {roe_date}')
                        CashAllocationViewSet.create_policy_info(request, cash_tracker_report_obj, bankTxn, bankExc, roe_date)
                    else:
                        return Response({"msg": "Policy related data is not there!"}, status=400)
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    logger.error(f"Exception in creating report records!: {e}")
                
                # Get the serializer data
                response_data = serializer.data

                deleted_ca_id = serializer_saved.deleted_ca
                if deleted_ca_id:
                    policy_number = CashAllocation.objects.get(id=deleted_ca_id).policy_fk.Policy_Line_Ref
                    activities = cash_allocation_activities_helper(deleted_ca_id)
                    # Update the deleted_ca field in the response data
                    response_data['deleted_ca'] = {policy_number: activities}
            
                return Response(response_data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Transaction failed: {e}")
            raise Exception("Failed to create cash allocation")

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        try:
            # compare allocated_amount if change then update ct amount
            obj = self.get_object()

            user = self.get_user(request)

            bank_txn_obj = BankTransaction.objects.get(id=request.data["bank_txn"])

            if 'query' in request.data['allocation_status'].lower() and request.data['allocation_status'] != obj.allocation_status:
                current_time = datetime.now()
                if FollowUp.objects.filter(bank_transaction = bank_txn_obj, cash_allocation = obj, archived = False).exists():
                    old_followup_obj = FollowUp.objects.get(bank_transaction = bank_txn_obj, cash_allocation = obj, archived = False)
                    old_followup_obj.archived = True
                    old_followup_obj.save()

                FollowUp.objects.create(
                    bank_transaction = bank_txn_obj,
                    cash_allocation = obj,
                    query_status = request.data['allocation_status'],
                    date1_value = current_time.date(),
                    date1 = current_time,
                    created_by = user
                )

            # Save cashalllocation audit
            if request.data.get("updatedFields"):
                for key, value in dict(request.data["updatedFields"]).items():
                    CashAllocaionAudit.objects.create(
                        cash_allocation=obj,
                        audit_data={
                            "field_name": key,
                            "old_value": str(getattr(obj, key, "")),
                            "new_value": value,
                            "previous_edit_datetime": obj.updated_at.strftime("%Y-%m-%d %H:%M:%S") if obj.updated_at else "-",
                            "current_edit_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "changed_by": user.pk,
                            "event_type": "edit"
                        }
                    )

            if request.data["receivable_amt"]:
                receivable_amt = Decimal(request.data["receivable_amt"])
                cash_allocation_queryset = CashAllocation.objects.filter(bank_txn=bank_txn_obj, archived=False)
                total_receivable_amt = cash_allocation_queryset.aggregate(Sum('receivable_amt'))['receivable_amt__sum']
                if total_receivable_amt-obj.receivable_amt+receivable_amt > bank_txn_obj.Receivable_Amount:
                    return Response({"msg": "Receivable amount is greater than total receivable amount!"}, status=400)
                
            if request.data['allocation_status'] != "Allocated" and request.data['allocated_amt']!="0.00":
                return Response({"msg": "Allocated amount should 0(zero) when allocation_status is not Allocated!"}, status=400)

            current_allocated_amt = obj.allocated_amt

            allocated_amount = Decimal(request.data["allocated_amt"])
            receivable_amount = Decimal(request.data["receivable_amt"])

            is_allocated_amount_changed = False
            is_receivable_amount_changed = False

            if allocated_amount != current_allocated_amt:
                is_allocated_amount_changed = True

            if receivable_amount != obj.receivable_amt:
                is_receivable_amount_changed = True

            obj.bank_txn = bank_txn_obj
            obj.policy_id = request.data["policy_id"]

            obj.allocation_systemid = request.data["allocation_systemid"]
            obj.allocation_entity = request.data["allocation_entity"]
            obj.allocation_binder = request.data["allocation_binder"]
            obj.allocation_umr = request.data["allocation_umr"]
            obj.binding_agreement = request.data["binding_agreement"]
            obj.allocation_eea = request.data["allocation_eea"]
            obj.allocation_scm = request.data["allocation_scm"]
            obj.allocation_scmname = request.data["allocation_scmname"]
            obj.settlement_currency = request.data["settlement_currency"]
            obj.allocation_invoice_verification = request.data[
                "allocation_invoice_verification"
            ]
            obj.receivable_amt = request.data["receivable_amt"]
            obj.allocated_amt = request.data["allocated_amt"]
            obj.unallocated_amt = request.data["unallocated_amt"]
            obj.XFIbatchno = request.data["XFIbatchno"]
            obj.cashreference = request.data["cashreference"]
            obj.GXPbatchno = request.data["GXPbatchno"]
            obj.allocation_date = request.data["allocation_date"]

            today = datetime.today().date()
            last_day = calendar.monthrange(today.year, today.month)[1]
            accounting_date = date(today.year, today.month, last_day)
            if obj.allocation_status != request.data['allocation_status']:
                if request.data['allocation_status'].lower() == "allocated":
                    today = datetime.now()
                    new_ac = get_accounting_month(today)
                    if isinstance(new_ac, Response):
                        return new_ac
                    obj.accounting_monthyear = new_ac
                    obj.allocation_date = today.date()

                    accounting_date = new_ac
                #     if obj.allocation_date:
                #         last_day = calendar.monthrange(obj.allocation_date.year, obj.allocation_date.month)[1]
                #         accounting_date = date(obj.allocation_date.year, obj.allocation_date.month, last_day)
                #     else:
                #         accounting_date = None
                # else:
                #     obj.allocation_date = None
                if obj.allocation_status.lower() == "allocated":
                    obj.allocation_date = None

            obj.allocation_status = request.data["allocation_status"]
            obj.allocatedby = request.data["allocatedby"]
            obj.isActive = request.data["isActive"]
            obj.bank_roe = request.data["bank_roe"]
            obj.payment_currency = request.data["payment_currency"]
            obj.bank_curr = request.data["bank_curr"]
            obj.cash_reference = request.data["cash_reference"]
            obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_pk"])
            obj.policy_mf.set(request.data["policy_mf"])
            obj.comment = request.data["comment"]
            obj.installment_amount_org = request.data.get("installment_amount_org", obj.installment_amount_org)
            obj.policy_installment_number = request.data.get("policy_installment_number",
                                                            obj.policy_installment_number)
            obj.original_ccy = request.data["original_ccy"]
            
            if user:
                obj.updated_by = user

            obj.save()
            # CMT-28
            # if obj.allocation_status.lower() == "allocated":
            #     cash_allocations_id_and_status = CashAllocationViewSet.get_cash_allocation_id_and_status(obj.bank_txn)
            #     CashAllocationViewSet.update_bank_transaction_status(obj.bank_txn, cash_allocations_id_and_status)
            serializer = CashAllocationCreateSerializer(obj)

            if is_allocated_amount_changed:
                try:
                    total_allocated_amount = CashAllocation.objects.filter(
                        bank_reconciliation_id=obj.bank_reconciliation_id, archived=False
                    ).aggregate(total_allocated_amount=Sum('allocated_amt'))['total_allocated_amount']
                    BankReconciliationViewSet.update_ct_amount_and_ct_amount_var(total_allocated_amount,
                                                                                obj.bank_reconciliation_id)

                except Exception as e:
                    # Handle the error gracefully (e.g., log the error or return a user-friendly message)
                    logger.error(f"Error getting total allocated amount: {e}")

            if is_receivable_amount_changed:
                try:
                    total_receivable_amount = CashAllocation.objects.filter(
                        bank_reconciliation_id=obj.bank_reconciliation_id, archived=False
                    ).aggregate(total_receivable_amount=Sum('receivable_amt'))['total_receivable_amount']
                    BankReconciliationViewSet.update_ct_receivable_amount(total_receivable_amount, obj.bank_reconciliation_id)
                except Exception as e:
                    logger.error(f"Error updating ct receivable amount: {e}")

            report_tracker = CashTrackerReport.objects.filter(
                cash_allocation_id=obj.id
            ).update(
                Invoice_Verification=request.data["allocation_invoice_verification"],
                Producing_Coverholder=request.data["allocation_entity"],
                Binding_Agreement=request.data["binding_agreement"],
                SCM_Partners=request.data["allocation_scm"],
                EEA_NonEEA=request.data["allocation_eea"],
                ROE_Bank_Statement=request.data["bank_roe"],
                Master_Binder=request.data["allocation_binder"],
                Policy=request.data["policy_id"],
                Receivable_Amount=request.data["receivable_amt"],
                Allocated_Amount=request.data["allocated_amt"],
                Batch=request.data["XFIbatchno"],
                Cash_Reference=request.data["cash_reference"],
                GXB_Batch=request.data["GXPbatchno"],
                Allocation_Status=request.data["allocation_status"],
                System_Correction=request.data["allocation_systemid"],
                bank_txn_id=request.data["bank_txn"],
                updated_by=user if user else None,
                policy_information= PolicyInformation.objects.get(id=request.data["policy_pk"]),
                Accounting_Month=accounting_date if obj.allocation_status != request.data['allocation_status'] and request.data['allocation_status'].lower() == 'allocated' else obj.accounting_monthyear
            )
            cash_tracker_report_obj = CashTrackerReport.objects.filter(cash_allocation_id=obj.id)
            if obj.policy_id and cash_tracker_report_obj.exists():
                bankExc, roe_date = CashAllocationViewSet.get_bank_exchange_rate(
                    bank_transaction_obj=obj.bank_txn,
                )
                logger.info(f'update cash allocation bankExc: {bankExc} roe_date: {roe_date}')

                CashAllocationViewSet.create_policy_info(request, cash_tracker_report_obj.first(), obj.bank_txn, bankExc, roe_date)

            # Comment out because It should not update the accounting month in bank transaction in any condition
            # BankTransaction.objects.filter(archived=False, id=request.data["bank_txn"]).update(Accounting_Month=accounting_date)

            if obj.allocation_status != request.data['allocation_status']:
                if request.data['allocation_status'].lower() == "allocated":
                    update_cash_allocation(app_model_name="bankmanagement.CashAllocationIssues", cash_allocation_id_key='id',
                                        cash_allocation_id=obj.id,
                                        accounting_date=accounting_date)

                    update_cash_allocation(app_model_name="bankmanagement.CashAllocationCorrective",
                                        cash_allocation_id_key='cash_allocation_id',
                                        cash_allocation_id=obj.id,
                                        accounting_date=accounting_date)

                    update_cash_allocation(app_model_name="bankmanagement.CashAllocationWriteoff",
                                        cash_allocation_id_key='cash_allocation_id',
                                        cash_allocation_id=obj.id,
                                        accounting_date=accounting_date)

                    update_cash_allocation(app_model_name="bankmanagement.CashAllocationRefund",
                                        cash_allocation_id_key='cash_allocation_id',
                                        cash_allocation_id=obj.id,
                                        accounting_date=accounting_date)

                    update_cash_allocation(app_model_name="bankmanagement.CashAllocationCFI",
                                        cash_allocation_id_key='cash_allocation_id',
                                        cash_allocation_id=obj.id,
                                        accounting_date=accounting_date)

                    update_cash_allocation(app_model_name="bankmanagement.CashAllocationMSD",
                                        cash_allocation_id_key='cash_allocation_id',
                                        cash_allocation_id=obj.id,
                                        accounting_date=accounting_date)

                    update_cash_allocation(app_model_name="bankmanagement.PremiumPayment",
                                        cash_allocation_id_key='cash_allocation_id',
                                        cash_allocation_id=obj.id,
                                        accounting_date=accounting_date)

                    update_cash_allocation(app_model_name="bankmanagement.CorrectiveTRF",
                                        cash_allocation_id_key='cash_allocation_id',
                                        cash_allocation_id=obj.id,
                                        accounting_date=accounting_date)

            # CMT-28
            try:
                if CashAllocation.objects.filter(bank_txn=obj.bank_txn, allocation_status="Allocated", archived=False).exists():
                    total_allocated_amount = CashAllocationViewSet.get_total_allocated_amount(obj.bank_txn)
                    if total_allocated_amount == obj.bank_txn.Receivable_Amount:
                        BankTransaction.objects.filter(id=obj.bank_txn_id, archived=False).update(transaction_status="Completed")
            except Exception as e:
                print("Error in updating bank transaction status: ", e)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            logger.error(f"Transaction failed: {e}")
            raise Exception("Failed to update report tracker")
        
    @transaction.atomic
    def partial_update(self, request, *args, **kwargs):
        try:
            obj = self.get_object()
            data = request.data

            # Update fields selectively based on provided data
            for field in ['bank_txn', 'policy_id', 'allocation_systemid', 'allocation_entity', 'allocation_binder', 'allocation_umr', 'binding_agreement', 'allocation_eea', 'allocation_scm', 'allocation_scmname', 'settlement_currency', 'allocation_invoice_verification', 'receivable_amt', 'allocated_amt', 'unallocated_amt', 'XFIbatchno', 'cashreference', 'GXPbatchno', 'allocation_date', 'allocatedby', 'isActive', 'bank_roe', 'payment_currency', 'bank_curr', 'cash_reference', 'comment', 'installment_amount_org', 'policy_installment_number', 'original_ccy']:
                if field in data:
                    setattr(obj, field, data[field])

            # Handle special cases for dates and amounts
            if 'receivable_amt' in data:
                receivable_amt = Decimal(data['receivable_amt'])
                cash_allocation_queryset = CashAllocation.objects.filter(bank_txn=obj.bank_txn, archived=False)
                total_receivable_amt = cash_allocation_queryset.aggregate(Sum('receivable_amt'))['receivable_amt__sum']
                if total_receivable_amt - obj.receivable_amt + receivable_amt > obj.bank_txn.Receivable_Amount:
                    return Response({"msg": "Receivable amount is greater than total receivable amount!"}, status=400)

            if 'allocation_status' in data and data['allocation_status'] != "Allocated" and data.get('allocated_amt', "0.00") != "0.00":
                return Response({"msg": "Allocated amount should be 0 (zero) when allocation_status is not Allocated!"}, status=400)

            # Update the accounting date if the allocation date is provided or changed
            if 'allocation_date' in data:
                allocation_date = parse_date(data['allocation_date'])
                last_day = calendar.monthrange(allocation_date.year, allocation_date.month)[1]
                accounting_date = date(allocation_date.year, allocation_date.month, last_day)
                obj.accounting_monthyear = accounting_date
            obj.save()

            # Update related cash tracker report if necessary
            if 'policy_id' in data or 'receivable_amt' in data or 'allocated_amt' in data:
                cash_tracker_report_obj = CashTrackerReport.objects.get(cash_allocation_id=obj.id)
                if cash_tracker_report_obj:
                    update_fields = {
                        'Policy': data.get('policy_id', obj.policy_id),
                        'Receivable_Amount': data.get('receivable_amt', obj.receivable_amt),
                        'Allocated_Amount': data.get('allocated_amt', obj.allocated_amt),
                        'Accounting_Month': accounting_date
                    }
                    for field, value in update_fields.items():
                        setattr(cash_tracker_report_obj, field, value)
                    cash_tracker_report_obj.save()

            return Response({"msg": "Cash allocation updated successfully"}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Partial update failed: {e}")
            return Response({"msg": "Failed to update cash allocation"}, status=status.HTTP_400_BAD_REQUEST)

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.archived = True
        instance.save()
        logger.debug(f"Archived cash allocation: {instance.id}")

        CashAllocationWriteoff.objects.filter(cash_allocation=instance).update(archived=True)
        CashAllocationRefund.objects.filter(cash_allocation=instance).update(archived=True)
        CashAllocationIssues.objects.filter(cash_allocation=instance).update(archived=True)
        CashAllocationCorrective.objects.filter(cash_allocation=instance).update(archived=True)
        CrossAllocation.objects.filter(cash_allocation=instance).update(archived=True)
        CashAllocationCFI.objects.filter(cash_allocation=instance).update(archived=True)

        return Response({"message": "Cash allocation and their activities are archived successfully"}, status=status.HTTP_200_OK)


def get_user(request):
    user_id = request.headers.get("user-id")
    user = Users.objects.filter(id=user_id)
    return user[0]
    # token = request.META.get('HTTP_AUTHORIZATION', False)
    # if token:
    #     token = str(token).split()[1].encode("utf-8")
    #     knoxAuth = TokenAuthentication()
    #     user, auth_token = knoxAuth.authenticate_credentials(token)
    #     request.user = user
    #     return user

class CashAllocationUpdateAPIView(APIView):
    @transaction.atomic
    def patch(self, request, pk):
        try:
            instance = CashAllocation.objects.get(pk=pk)
            data = request.data
            logger.info(f"Received data for update: {data}")  # Log received data

            user = get_user(request)
            logger.info(f"user: {user}")
            
            # Save cashalllocation audit
            for key, value in dict(request.data).items():
                if key != 'policy_pk':
                    CashAllocaionAudit.objects.create(
                        cash_allocation=instance,
                        audit_data={
                            "field_name": key,
                            "old_value": str(getattr(instance, key, "")),
                            "new_value": value,
                            "previous_edit_datetime": instance.updated_at.strftime("%Y-%m-%d %H:%M:%S") if instance.updated_at else "-",
                            "current_edit_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "changed_by": user.pk,
                            "event_type": "edit"
                        }
                    )

            if not user:
                logger.warning("User authentication failed.")
                return Response({"message": "User is not authenticated"}, status=status.HTTP_401_UNAUTHORIZED)
            if user.role != "Manager":
                logger.warning("Unauthorized access attempt by user.")
                return Response({"message": "You are not authorized to update this cash allocation"}, status=status.HTTP_403_FORBIDDEN)

            if not instance.manual_entry and 'policy_pk' in data:
                policy_instance = PolicyInformation.objects.get(id=data['policy_pk'])
                umr_number = policy_instance.UMR_Number
                binding_agreement = policy_instance.Binding_Agreement
                producing_entity = policy_instance.Producing_Entity
                original_ccy = policy_instance.Original_Ccy

                # Check if the provided data matches the policy instance data
                mismatches = []

                if data.get('allocation_umr') and data['allocation_umr'] != umr_number:
                    mismatches.append(f"UMR_Number mismatch: expected {umr_number}, got {data['allocation_umr']}")
                if data.get('binding_agreement') and data['binding_agreement'] != binding_agreement:
                    mismatches.append(f"Binding_Agreement mismatch: expected {binding_agreement}, got {data['binding_agreement']}")
                if data.get('allocation_entity') and data['allocation_entity'] != producing_entity:
                    mismatches.append(f"Producing_Entity mismatch: expected {producing_entity}, got {data['allocation_entity']}")
                if data.get('original_ccy') and data['original_ccy'] != original_ccy:
                    mismatches.append(f"Original_CCY mismatch: expected {original_ccy}, got {data['original_ccy']}")

                logger.info(f"Mismatches found: {mismatches}")
                if mismatches:
                    error_message = "Data mismatch errors: " + ", ".join(mismatches)
                    print("error_message: ", error_message)
                    return Response({"message": error_message}, status=status.HTTP_400_BAD_REQUEST)

            with transaction.atomic():
                fields_to_update = [
                    'bank_txn', 'policy_id', 'allocation_systemid', 'allocation_entity', 'allocation_binder', 'bank_roe',
                    'allocation_umr', 'binding_agreement', 'allocation_eea', 'allocation_scm', 'allocation_scmname',
                    'settlement_currency', 'allocation_invoice_verification', 'receivable_amt', 'allocated_amt',
                    'unallocated_amt', 'XFIbatchno', 'cashreference','cash_reference', 'GXPbatchno', 'allocation_date', 'original_ccy', 'installment_amount_org'
                ]
                if 'settlement_currency' in data:
                    settlement_ccy_code = data.get('settlement_currency')
                    try:
                        settlement_ccy_instance = CurrencyDetails.objects.get(currency_code=settlement_ccy_code)
                        instance.settlement_ccy = settlement_ccy_instance
                    except CurrencyDetails.DoesNotExist:
                        return Response({"message": "CurrencyDetails not found for the provided code"}, status=status.HTTP_404_NOT_FOUND)
                # if 'installment_amount_org' in data:
                #     instance.xfi_amount = data['installment_amount_org']
                for field in fields_to_update:
                    if field in data:
                        print("field: ", field)
                        print("data[field]: ", data[field])
                        setattr(instance, field, data[field])

                # Check if manual_entry is True and update PolicyInformation
                if instance.manual_entry and 'policy_pk' in data:
                    policy_instance = PolicyInformation.objects.get(id=data['policy_pk'])
                    umr_number = policy_instance.UMR_Number
                    binding_agreement = policy_instance.Binding_Agreement
                    producing_entity = policy_instance.Producing_Entity
                    original_ccy = policy_instance.Original_Ccy

                    # Mapping of incoming data keys to PolicyInformation fields
                    policy_fields_mapping = {
                        'allocation_umr': 'UMR_Number',
                        'binding_agreement': 'Binding_Agreement',
                        'allocation_entity': 'Producing_Entity',
                        'original_ccy': 'Original_Ccy',
                        'Three_Party_Capacity_Deployed': 'Three_Party_Capacity_Deployed',
                        'allocation_scm': 'SCM_Partner',
                        'settlement_currency': 'Settlement_Ccy',
                        'allocation_binder': 'Syndicate_Binder',
                        'original_ccy': 'Installment_Ccy_in_Orig',
                        'installment_amount_org': 'Installment_Agency_Amount_in_Orig',
                    }
                    # Update fields if they exist in the request data
                    for incoming_key, model_field in policy_fields_mapping.items():
                        if incoming_key in data:
                            print("incoming_key: ", incoming_key)
                            print("model_field: ", model_field)
                            print("data[incoming_key]: ", data[incoming_key])
                            setattr(policy_instance, model_field, data[incoming_key])
                    policy_instance.save()

                instance.updated_by = user
                instance.save()
                logger.info("CashAllocation instance updated successfully.")

                # Update CashTrackerReport related to this CashAllocation
                cash_tracker_update_fields = {
                    'Invoice_Verification': 'allocation_invoice_verification',
                    'Producing_Coverholder': 'allocation_entity',
                    'Binding_Agreement': 'binding_agreement',
                    'SCM_Partners': 'allocation_scm',
                    'EEA_NonEEA': 'allocation_eea',
                    'ROE_Bank_Statement': 'bank_roe',
                    'Master_Binder': 'allocation_binder',
                    'Policy': 'policy_id',
                    'Receivable_Amount': 'receivable_amt',
                    'Allocated_Amount': 'allocated_amt',
                    'Batch': 'XFIbatchno',
                    'Cash_Reference': 'cash_reference',
                    'GXB_Batch': 'GXPbatchno',
                    'Allocation_Status': 'allocation_status',
                    'Payment_Currency_Code': 'original_ccy',
                    'Third_Party_Capacity': 'Three_Party_Capacity_Deployed',
                }
                update_data = {field: data[value] for field, value in cash_tracker_update_fields.items() if value in data and value in fields_to_update}
                if update_data:
                    print('update data--> ',update_data)
                    CashTrackerReport.objects.filter(cash_allocation_id=instance.id).update(**update_data, updated_by=user)
                    logger.info("CashTrackerReport updated successfully.")
                return Response({"message": "CashAllocation updated successfully"}, status=status.HTTP_200_OK)
        except CashAllocation.DoesNotExist:
            logger.error("CashAllocation not found.")
            return Response({"message": "CashAllocation not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Failed to update CashAllocation: {str(e)}")
            return Response({"message": f"Failed to update CashAllocation: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.archived = True
        instance.save()
        logger.debug(f"Archived cash allocation: {instance.id}")
        return Response({"message": "Cash allocation archived successfully"}, status=status.HTTP_200_OK)

class FileUploadToTransactionsViewSet(viewsets.ModelViewSet):
    model = BankTransaction
    serializer_class = BankTransactionSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def create(self, request, *args, **kwargs):
        data = request.data
        transaction_ids = json.loads(data["transaction_ids"])
        File_Name = data["File_Name"]
        transss = []
        for ids in transaction_ids:
            trans = BankTransaction.objects.get(id=ids["id"])
            if File_Name:
                trans.File_Name = data["File_Name"]
                trans.save()
                transss.append(trans)
        serializer = BankTransactionSerializer(transss, many=True)
        m = serializer.data
        return Response(m)


class MultiFileUploadToTransactionsViewSet(viewsets.ModelViewSet):
    model = BankTransaction
    serializer_class = BankTransactionSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def create(self, request, *args, **kwargs):
        data = request.data
        file = request.FILES["File_Name"]
        transaction_id = json.loads(data["transaction_id"])
        document_name = request.data["document_name"]
        document_details = request.data["document_details"]

        if file.content_type in [
            "application/pdf",
            "text/csv",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ]:
            if file.size > 15 * 1024 * 1024:
                return Response(
                    {"msg": "File size should be less than 15 MB!"}, status=400
                )

            new_doc = Documents.objects.create(
                document_name=document_name,
                document_date=date.today(),
                document_type=file.content_type,
                document_details=document_details,
            )
            new_doc.save()
            new_doc.document_file = file
            new_doc.save()

            reuired_data = {"module_name": "Cash Allocation", "bucket_name" : config("AWS_STORAGE_BUCKET_NAME")}
            filemanagment = reusable_file_upload(get_user(request), file, reuired_data, is_upload=False)

            if not isinstance(filemanagment, Response) or filemanagment.status_code >= 400:
                    # Extract error message from response
                    error_msg = filemanagment.data.get('error', 'Unable to upload the file to S3') if isinstance(filemanagment, Response) else 'Unable to upload the file to S3'
                    raise Exception(error_msg)

        else:
            return Response(
                {"msg": "Invalid file type! Upload only pdf, csv or xlsx"}, status=400
            )

        bankTxn = BankTransaction.objects.get(id=transaction_id)
        bankTxn.document_files.add(new_doc)
        bankTxn.save()

        bankSer = BankTransactionSerializer(bankTxn)
        attach_document_file_key(bankSer.data)
        return Response(bankSer.data)


@csrf_exempt
def getBankDetailsByTransactionId(request):
    if request.method == "GET":
        txn_id = request.GET.get("txn_id")
        bankTxnObject = (
            BankTransaction.objects.filter(archived=False, Bank_Transaction_Id=txn_id)
            .order_by("id")
            .first()
        )
        if bankTxnObject:
            serializer = BankTransactionSerializer(bankTxnObject)
            attach_document_file_key(serializer.data)
            return JsonResponse(serializer.data, status=200)
        else:
            return JsonResponse({"msg": "No data found for given txn_id!"}, status=404)

def attach_document_file_key(serializer_data):
    for data in serializer_data['document_files']:
        if data['document_file']:
            file_path = data['document_file']

            if '/media/documents/media/' in file_path:
                file_name = file_path.split('/media/documents/media/')[-1]
                data['document_file'] = file_name
            else:
                data['document_file'] = None


@csrf_exempt
def getTransactionsByTransactionId(request):
    if request.method == "GET":
        txn_id = request.GET.get("txn_id")
        bankTxnObject = CashAllocation.objects.filter(
            bank_txn__Bank_Transaction_Id=txn_id, archived=False
        )
        print(bankTxnObject)
        if bankTxnObject:
            serializer = CashAllocationSerializer(bankTxnObject, many=True)
            return JsonResponse(serializer.data, status=200)
        else:
            return JsonResponse(
                {"msg": "No data found for given txn_id!"}, status=404, safe=False
            )

@dataclass
class BucketConfig:
    bucket_name: str
    file_path_prefix: str = ""

class ModuleType(Enum):
    CASH_ALLOCATION = "cashAllocation"
    UPLOADED_STATEMENTS = "uploadedFiles"
    BANK_STATAEMENT_TRANSACTIONS = "bankStatementTransactions"
    FILE_MANAGEMENT = "fileManagement"
    WF = "workflow"

    def get_bucket_config(self, bucket_key: str, module_name: str = None) -> BucketConfig:
        configs = {
            self.CASH_ALLOCATION: BucketConfig(
                bucket_name=config("AWS_STORAGE_BUCKET_NAME"),
                file_path_prefix="media/documents/media/"
            ),
            self.UPLOADED_STATEMENTS: BucketConfig(
                bucket_name=config("AWS_S3_PROCESSED_BUCKET")
            ),
            self.BANK_STATAEMENT_TRANSACTIONS: BucketConfig(
                bucket_name=config("AWS_STORAGE_BUCKET_NAME"),
                file_path_prefix="media/bank/media/"
            ),
            self.FILE_MANAGEMENT: BucketConfig(
                bucket_name=config("AWS_STORAGE_BUCKET_NAME"),
                file_path_prefix=f"{module_name}/" if module_name else ""
            ),
            self.WF: BucketConfig(
                bucket_name=config("WF_BANK_AMOUNT_CHANGE_BUCKET_NAME"),
                file_path_prefix=f"{module_name}/" if module_name else ""
            )
        }
        bucket_config = configs[self]
        return BucketConfig(
            bucket_name=bucket_config.bucket_name,
            file_path_prefix=bucket_config.file_path_prefix
        )

class DownloadFileView(APIView):
    def get(self, request, *args, **kwargs):
        """
        Download a file from S3 based on the provided key and module

        query parameters:
            key (str): The key of the file to download
            module (str): The module that the file belongs to. Currently supported: ['cashAllocation', 'uploadedStatements']

        returns:
            HttpResponse: A response with the file contents and a Content-Disposition header set to attachment; filename=<filename>
        """
        bucket_key = request.query_params.get("key")
        module = request.query_params.get("module")
        module_name = request.query_params.get("moduleName")

        try:
            module_type = ModuleType(module)
            bucket_config = module_type.get_bucket_config(bucket_key, module_name)
            bucket_name = bucket_config.bucket_name
            file_path = f"{bucket_config.file_path_prefix}{bucket_key}" if bucket_config.file_path_prefix else bucket_key
        except ValueError:
            return Response(
                {"error": f"Invalid module type. Must be one of {[m.value for m in ModuleType]}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        presigned_url = create_presigned_url(bucket_name, file_path)
        if not presigned_url:
            return Response(
                {"error": "Failed to generate presigned URL"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            s3_response = requests.get(presigned_url, stream=True)
            if s3_response.status_code != 200:
                logging.error(
                    f"Failed to download file. Status code: {s3_response.status_code}"
                )
                return Response(
                    {"error": "Unable to download file from S3."},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            # Extract filename
            filename = bucket_key.split("/")[-1]

            # Guess the content type based on the file extension
            content_type = (
                s3_response.headers.get("Content-Type")
                or mimetypes.guess_type(filename)[0]
                or "application/octet-stream"
            )

            response = HttpResponse(
                content=s3_response.content, content_type=content_type
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            response["Content-Length"] = len(s3_response.content)

            return response
        except Exception as e:
            logging.error(f"Download error: {str(e)}")
            return Response(
                {"error": f"Error occurred while downloading file: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class CashAllocationIssuesViewSet(viewsets.ModelViewSet):
    queryset = CashAllocationIssues.objects.all()
    serializer_class = CashAllocationIssuesSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        issues = CashAllocationIssues.objects.all()
        return issues

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = CashAllocationIssuesCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            report_tracker = CashTrackerReport.objects.get(
                cash_allocation_id=request.data["cash_allocation"]
            )
            if report_tracker:
                report_tracker.Category = request.data["issue_category"]
                # report_tracker.Aging_Bucket = request.data["age_days"]
                report_tracker.Corretion_Type = CorrectionType.objects.get(
                    id=request.data["correction_type"]
                ).id
                report_tracker.Comment = request.data["comments"]
                report_tracker.Team = request.data["assignment"]
                report_tracker.Owner = request.data["issue_owner"]
                report_tracker.Initial_Query_Date = request.data["issue_date"]
                report_tracker.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.correction_type = CorrectionType.objects.get(
            id=request.data["correction_type"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.issue_category = request.data["issue_category"]
        obj.issue_owner = request.data["issue_owner"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.system_id = request.data["system_id"]
        obj.comments = request.data["comments"]
        obj.assignment = request.data["assignment"]
        obj.issue_date = request.data["issue_date"]
        obj.age_days = request.data["age_days"]
        obj.isActive = request.data["isActive"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = CashAllocationIssuesCreateSerializer(obj)
        report_tracker = CashTrackerReport.objects.filter(
            cash_allocation_id=obj.cash_allocation_id
        ).update(
            Category=request.data["issue_category"],
            # Aging_Bucket=request.data["age_days"],
            Corretion_Type_id=request.data["correction_type"],
            Comment=request.data["comments"],
            Team=request.data["assignment"],
            Owner=request.data["issue_owner"],
            Initial_Query_Date=request.data["issue_date"],
        )
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        data = request.data
        obj.bank_txn = (
            BankTransaction.objects.get(id=data.get("bank_txn"))
            if data.get("bank_txn")
            else obj.bank_txn
        )
        obj.cash_allocation = (
            CashAllocation.objects.get(id=data.get("cash_allocation"))
            if data.get("cash_allocation")
            else obj.cash_allocation
        )
        obj.correction_type = (
            CorrectionType.objects.get(id=data.get("correction_type"))
            if data.get("correction_type")
            else obj.correction_type
        )
        obj.policy_id = data.get("policy_id", obj.policy_id)
        obj.issue_category = data.get("issue_category", obj.issue_category)
        obj.issue_owner = data.get("issue_owner", obj.issue_owner)
        obj.accounting_monthyear = data.get(
            "accounting_monthyear", obj.accounting_monthyear
        )
        obj.system_id = data.get("system_id", obj.system_id)
        obj.comments = data.get("comments", obj.comments)
        obj.assignment = data.get("assignment", obj.assignment)
        obj.issue_date = data.get("issue_date", obj.issue_date)
        obj.age_days = data.get("age_days", obj.age_days)
        obj.isActive = data.get("isActive", obj.isActive)
        obj.policy_fk = (
            PolicyInformation.objects.get(id=data.get("policy_fk"))
            if data.get("policy_fk")
            else obj.policy_fk
        )
        if data.get("policy_mf"):
            obj.policy_mf.set(data.get("policy_mf"))
        obj.save()
        serializer = CashAllocationIssuesCreateSerializer(obj)

        report_tracker = CashTrackerReport.objects.filter(
            cash_allocation_id=obj.cash_allocation_id
        ).update(
            Category=data.get("issue_category", obj.issue_category),
            # Aging_Bucket=data.get("age_days", obj.age_days),
            Corretion_Type=(
                CorrectionType.objects.get(id=data.get("correction_type"))
                if data.get("correction_type")
                else obj.correction_type
            ),
            Comment=data.get("comments", obj.comments),
            Team=data.get("assignment", obj.assignment),
            Owner=data.get("issue_owner", obj.issue_owner),
            Initial_Query_Date=data.get("issue_date", obj.issue_date),
        )

        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CashAllocationCorrectiveViewSet(viewsets.ModelViewSet):
    queryset = CashAllocationCorrective.objects.all()
    serializer_class = CashAllocationCorrectiveSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = CashAllocationCorrective.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = CashAllocationCorrectiveCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            report_tracker = CashTrackerReport.objects.get(
                cash_allocation_id=request.data["cash_allocation"]
            )
            if report_tracker:
                report_tracker.Transfer_to_PT_Bank_Account_Name = request.data[
                    "transfer_pt_bank_account_name"
                ]
                report_tracker.Transfer_to_PT_Bank_Account = request.data[
                    "PT_bank_acct_Number"
                ]
                report_tracker.Treasury_Confirmed_Transfer_Date = request.data[
                    "treasury_confirmed_transfer_date"
                ]
                report_tracker.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.transfer_amt = request.data["transfer_amt"]
        obj.producing_entity_bankid = request.data["producing_entity_bankid"]
        obj.producing_bank_currid = request.data["producing_bank_currid"]
        obj.bank_roe = request.data["bank_roe"]
        obj.payment_currid = request.data["payment_currid"]
        obj.bank_curr = request.data["bank_curr"]
        obj.receivable_amt = request.data["receivable_amt"]
        obj.allocated_amt = request.data["allocated_amt"]
        obj.unallocated_amt = request.data["unallocated_amt"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.allocation_date = request.data["allocation_date"]
        obj.transfer_pt_bank_account_name = request.data[
            "transfer_pt_bank_account_name"
        ]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.treasury_confirmed_transfer_date = request.data[
            "treasury_confirmed_transfer_date"
        ]
        obj.PT_bank_acct_Number = request.data["PT_bank_acct_Number"]
        obj.save()
        serializer = CashAllocationCorrectiveCreateSerializer(obj)
        report_tracker = CashTrackerReport.objects.filter(
            cash_allocation_id=obj.cash_allocation_id
        ).update(
            Transfer_to_PT_Bank_Account_Name=request.data[
                "transfer_pt_bank_account_name"
            ],
            Transfer_to_PT_Bank_Account=request.data["PT_bank_acct_Number"],
            Treasury_Confirmed_Transfer_Date=request.data[
                "treasury_confirmed_transfer_date"
            ],
        )
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        data = request.data
        obj.bank_txn = (
            BankTransaction.objects.get(id=data.get("bank_txn"))
            if data.get("bank_txn")
            else obj.bank_txn
        )
        obj.cash_allocation = (
            CashAllocation.objects.get(id=data.get("cash_allocation"))
            if data.get("cash_allocation")
            else obj.cash_allocation
        )
        obj.policy_id = data.get("policy_id", obj.policy_id)
        obj.transfer_amt = data.get("transfer_amt", obj.transfer_amt)
        obj.producing_entity_bankid = data.get(
            "producing_entity_bankid", obj.producing_entity_bankid
        )
        obj.producing_bank_currid = data.get(
            "producing_bank_currid", obj.producing_bank_currid
        )
        obj.bank_roe = data.get("bank_roe", obj.bank_roe)
        obj.payment_currid = data.get("payment_currid", obj.payment_currid)
        obj.bank_curr = data.get("bank_curr", obj.bank_curr)
        obj.receivable_amt = data.get("receivable_amt", obj.receivable_amt)
        obj.allocated_amt = data.get("allocated_amt", obj.allocated_amt)
        obj.unallocated_amt = data.get("unallocated_amt", obj.unallocated_amt)
        obj.accounting_monthyear = data.get(
            "accounting_monthyear", obj.accounting_monthyear
        )
        obj.allocation_date = data.get("allocation_date", obj.allocation_date)
        obj.transfer_pt_bank_account_name = data.get(
            "transfer_pt_bank_account_name", obj.transfer_pt_bank_account_name
        )
        obj.policy_fk = (
            PolicyInformation.objects.get(id=data.get("policy_fk"))
            if data.get("policy_fk")
            else obj.policy_fk
        )
        if data.get("policy_mf"):
            obj.policy_mf.set(data.get("policy_mf"))
        obj.treasury_confirmed_transfer_date = data.get(
            "treasury_confirmed_transfer_date", obj.treasury_confirmed_transfer_date
        )
        obj.PT_bank_acct_Number = data.get(
            "PT_bank_acct_Number", obj.PT_bank_acct_Number
        )
        obj.save()
        serializer = CashAllocationCorrectiveCreateSerializer(obj)

        report_tracker = CashTrackerReport.objects.filter(
            cash_allocation_id=obj.cash_allocation_id
        ).update(
            Transfer_to_PT_Bank_Account_Name=data.get(
                "transfer_pt_bank_account_name", obj.transfer_pt_bank_account_name
            ),
            Transfer_to_PT_Bank_Account=data.get(
                "PT_bank_acct_Number", obj.PT_bank_acct_Number
            ),
            Treasury_Confirmed_Transfer_Date=data.get(
                "treasury_confirmed_transfer_date", obj.treasury_confirmed_transfer_date
            ),
        )

        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CashAllocationWriteoffViewSet(viewsets.ModelViewSet):
    queryset = CashAllocationWriteoff.objects.all()
    serializer_class = CashAllocationWriteoffSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = CashAllocationWriteoff.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = CashAllocationWriteoffCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.writeoff_reason = request.data["writeoff_reason"]
        obj.request_date = request.data["request_date"]
        obj.writeoff_amt = request.data["writeoff_amt"]
        obj.comments = request.data["comments"]
        obj.approved_date = request.data["approved_date"]
        obj.aprover_name = request.data["aprover_name"]
        obj.process_date = request.data["process_date"]
        obj.system_id = request.data["system_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.currency = request.data["currency"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.cash_reference = request.data["cash_reference"]
        obj.save()
        serializer = CashAllocationWriteoffCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.writeoff_reason = request.data["writeoff_reason"]
        obj.request_date = request.data["request_date"]
        obj.writeoff_amt = request.data["writeoff_amt"]
        obj.comments = request.data["comments"]
        obj.approved_date = request.data["approved_date"]
        obj.aprover_name = request.data["aprover_name"]
        obj.process_date = request.data["process_date"]
        obj.system_id = request.data["system_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.currency = request.data["currency"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.cash_reference = request.data["cash_reference"]
        obj.save()
        serializer = CashAllocationWriteoffCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CashAllocationRefundViewSet(viewsets.ModelViewSet):
    queryset = CashAllocationRefund.objects.all()
    serializer_class = CashAllocationRefundSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = CashAllocationRefund.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = CashAllocationRefundCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.refund_reason = request.data["refund_reason"]
        obj.request_date = request.data["request_date"]
        obj.refund_amt = request.data["refund_amt"]
        obj.comments = request.data["comments"]
        obj.approved_date = request.data["approved_date"]
        obj.aprover_name = request.data["aprover_name"]
        obj.process_date = request.data["process_date"]
        obj.system_id = request.data["system_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.currency = request.data["currency"]
        obj.bank_transaction_id = request.data["bank_transaction_id"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.broker_name = request.data["broker_name"]
        obj.save()
        serializer = CashAllocationRefundCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.refund_reason = request.data["refund_reason"]
        obj.request_date = request.data["request_date"]
        obj.refund_amt = request.data["refund_amt"]
        obj.comments = request.data["comments"]
        obj.approved_date = request.data["approved_date"]
        obj.aprover_name = request.data["aprover_name"]
        obj.process_date = request.data["process_date"]
        obj.system_id = request.data["system_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.currency = request.data["currency"]
        obj.bank_transaction_id = request.data["bank_transaction_id"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.broker_name = request.data["broker_name"]
        obj.save()
        serializer = CashAllocationRefundCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CashAllocationCFIViewSet(viewsets.ModelViewSet):
    queryset = CashAllocationCFI.objects.all()
    serializer_class = CashAllocationCFISerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = CashAllocationCFI.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = CashAllocationCFICreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.cfi_change = request.data["cfi_change"]
        obj.request_date = request.data["request_date"]
        obj.cfi_amt = request.data["cfi_amt"]
        obj.cfi_impact = request.data["cfi_impact"]
        obj.reallocation_date = request.data["reallocation_date"]
        obj.approved_date = request.data["approved_date"]
        obj.aprover_name = request.data["aprover_name"]
        obj.process_date = request.data["process_date"]
        obj.system_id = request.data["system_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = CashAllocationCFICreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.cfi_change = request.data["cfi_change"]
        obj.request_date = request.data["request_date"]
        obj.cfi_amt = request.data["cfi_amt"]
        obj.cfi_impact = request.data["cfi_impact"]
        obj.reallocation_date = request.data["reallocation_date"]
        obj.approved_date = request.data["approved_date"]
        obj.aprover_name = request.data["aprover_name"]
        obj.process_date = request.data["process_date"]
        obj.system_id = request.data["system_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = CashAllocationCFICreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CrossAllocationViewSet(viewsets.ModelViewSet):
    queryset = CrossAllocation.objects.all()
    serializer_class = CrossAllocationSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = CrossAllocation.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = CrossAllocationCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.ca_change = request.data["ca_change"]
        obj.request_date = request.data["request_date"]
        obj.ca_amt = request.data["ca_amt"]
        obj.ca_impact = request.data["ca_impact"]
        obj.approved_date = request.data["approved_date"]
        obj.aprover_name = request.data["aprover_name"]
        obj.process_date = request.data["process_date"]
        obj.system_id = request.data["system_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.new_allocation_date = request.data["new_allocation_date"]
        obj.save()
        serializer = CrossAllocationCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.ca_change = request.data["ca_change"]
        obj.request_date = request.data["request_date"]
        obj.ca_amt = request.data["ca_amt"]
        obj.ca_impact = request.data["ca_impact"]
        obj.approved_date = request.data["approved_date"]
        obj.aprover_name = request.data["aprover_name"]
        obj.process_date = request.data["process_date"]
        obj.system_id = request.data["system_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.new_allocation_date = request.data["new_allocation_date"]
        obj.save()
        serializer = CrossAllocationCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CashAllocationMSDViewSet(viewsets.ModelViewSet):
    queryset = CashAllocationMSD.objects.all()
    serializer_class = CashAllocationMSDSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = CashAllocationMSD.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = CashAllocationMSDCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.je_number = request.data["je_number"]
        obj.comments = request.data["comments"]
        obj.msd_amt = request.data["msd_amt"]
        obj.system_date = request.data["system_date"]
        obj.process_date = request.data["process_date"]
        obj.user_id = request.data["user_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = CashAllocationMSDCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.je_number = request.data["je_number"]
        obj.comments = request.data["comments"]
        obj.msd_amt = request.data["msd_amt"]
        obj.system_date = request.data["system_date"]
        obj.process_date = request.data["process_date"]
        obj.user_id = request.data["user_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = CashAllocationMSDCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CorrectiveTRFViewSet(viewsets.ModelViewSet):
    queryset = CorrectiveTRF.objects.all()
    serializer_class = CorrectiveTRFSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = CorrectiveTRF.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = CorrectiveTRFCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.receivable_amt = request.data["receivable_amt"]
        obj.payment_status = request.data["payment_status"]
        obj.payment_date = request.data["payment_date"]
        obj.wire_number = request.data["wire_number"]
        obj.comments = request.data["comments"]
        obj.system_date = request.data["system_date"]
        obj.user_id = request.data["user_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = CorrectiveTRFCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.receivable_amt = request.data["receivable_amt"]
        obj.payment_status = request.data["payment_status"]
        obj.payment_date = request.data["payment_date"]
        obj.wire_number = request.data["wire_number"]
        obj.comments = request.data["comments"]
        obj.system_date = request.data["system_date"]
        obj.user_id = request.data["user_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = CorrectiveTRFCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class PremiumPaymentViewSet(viewsets.ModelViewSet):
    queryset = PremiumPayment.objects.all()
    serializer_class = PremiumPaymentSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = PremiumPayment.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = PremiumPaymentCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.paid_amt = request.data["paid_amt"]
        obj.allocated_amt = request.data["allocated_amt"]
        obj.payment_status = request.data["payment_status"]
        obj.payment_date = request.data["payment_date"]
        obj.wire_number = request.data["wire_number"]
        obj.comments = request.data["comments"]
        obj.system_date = request.data["system_date"]
        obj.user_id = request.data["user_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = PremiumPaymentCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.paid_amt = request.data["paid_amt"]
        obj.allocated_amt = request.data["allocated_amt"]
        obj.payment_status = request.data["payment_status"]
        obj.payment_date = request.data["payment_date"]
        obj.wire_number = request.data["wire_number"]
        obj.comments = request.data["comments"]
        obj.system_date = request.data["system_date"]
        obj.user_id = request.data["user_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = PremiumPaymentCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class GetTransactionsByTransactionId(APIView):
    def get(self, request, format=None):
        txn_id = request.GET.get("txn_id")
        bankTxnObjects = CashAllocation.objects.filter(
            bank_txn__Bank_Transaction_Id=txn_id, archived=False
        )

        for bankTxnObject in bankTxnObjects:
            try:
                bankTxnObject.xfi_amount = PolicyInformation.objects.get(
                    id=bankTxnObject.policy_fk_id).Installment_Agency_Amount_in_Orig
                if bankTxnObject.xfi_amount:
                    bankTxnObject.xfi_variance_amount = float(bankTxnObject.allocated_amt) - float(
                        bankTxnObject.xfi_amount)
                if bankTxnObject.gxbamount:
                    bankTxnObject.gxb_variance_amount = float(bankTxnObject.allocated_amt) - float(
                        bankTxnObject.gxbamount)
                policy_info_list = PolicyCreateViewSet.get_policy_information_based_on_id(bankTxnObject.policy_id)
                bankTxnObject.policy_original_amount_list = [
                    {
                        'policy_pk': policy.id,
                        'Instalment_Nbr': policy.Instalment_Nbr,
                        'Installment_Agency_Amount_in_Orig': policy.Installment_Agency_Amount_in_Orig,
                        'Installment_Due_date': policy.Installment_Due_date,
                        'Original_Ccy': policy.Original_Ccy,
                        'Settlement_Ccy': policy.Settlement_Ccy,
                    } for policy in policy_info_list
                ]

            except:
                pass

        if bankTxnObjects:
            serializer = CashAllocationSerializer(bankTxnObjects, many=True)

            for data in serializer.data:
                deleted_ca_id = data['deleted_ca']
                if deleted_ca_id:
                    policy_number = CashAllocation.objects.get(id=deleted_ca_id).policy_fk.Policy_Line_Ref
                    activities = cash_allocation_activities_helper(deleted_ca_id)
                    
                    # Get workflow data
                    data_list = []
                    write_off_objects = CashAllocationWriteoff.objects.filter(
                        cash_allocation__id=deleted_ca_id
                    )
                    if write_off_objects:
                        writeoff_serializer = CashAllocationWriteoffCreateSerializer(write_off_objects, many=True)
                        data_list = writeoff_serializer.data
                        get_wf_id = WorkFlow.objects.get(workflow_name='WF_WRITEOFF')
                        for row in data_list:
                            workflow_ser = WorkflowBankTransactionsSerializer(row)
                            try:
                                row['txn_id'] = WorkflowBankTransactions.objects.filter(bank_txn_id=workflow_ser.data['id'],
                                                                                        workflow=get_wf_id.id).first().id
                                row['workflow'] = WorkflowBankTransactionsSerializer(
                                    WorkflowBankTransactions.objects.get(id=row['txn_id'])).data
                            except:
                                row['workflow'] = None
                    
                    # Update the deleted_ca field with both policy activities and workflow data
                    data['deleted_ca'] = {
                        'policy_number': policy_number,
                        'activities': activities,
                        'workflow_data': data_list
                    }
                ct_obj = CashTrackerReport.objects.filter(cash_allocation=data['id']).first()
                data['cash_tracker_data'] = CashTrackerReportSerializer(ct_obj).data
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response({"msg": "No data found for given txn_id!"}, status=status.HTTP_404_NOT_FOUND)


class GetTransactionsByTransactionIdValidation(APIView):
    def get(self, request, format=None):
        txn_id = request.GET.get("txn_id")
        bankTxnObjects = CashAllocation.objects.filter(
            bank_txn__Bank_Transaction_Id=txn_id, archived=False
        )

        if bankTxnObjects:
            return Response({"message": "Success"}, status=status.HTTP_200_OK)
        else:
            return Response({"message": "Not Success"}, status=status.HTTP_200_OK)


@csrf_exempt
def getCashAllocationIssuesFromCashAllocation(request):
    if request.method == "GET":
        cashallocation_id = request.GET.get("cashallocation_id")
        objects = CashAllocationIssues.objects.filter(
            cash_allocation__id=cashallocation_id
        )
        if objects:
            serializer = CashAllocationIssuesCreateSerializer(objects, many=True)
            data_list = serializer.data
            get_wf_id = WorkFlow.objects.get(workflow_name='WF_CORRECTION_TYPES')
            for data in data_list:
                workflow_ser = WorkflowBankTransactionsSerializer(data)
                data['txn_id'] = WorkflowBankTransactions.objects.filter(bank_txn_id=workflow_ser.data['id'],
                                                                         workflow=get_wf_id.id).first().id if WorkflowBankTransactions.objects.filter(
                    bank_txn_id=workflow_ser.data['id'], workflow=get_wf_id.id).first() else None
                data['workflow'] = WorkflowBankTransactionsSerializer(WorkflowBankTransactions.objects.filter(
                    id=data['txn_id']).first()).data if WorkflowBankTransactionsSerializer(
                    WorkflowBankTransactions.objects.filter(id=data['txn_id']).first()) else None
            return JsonResponse(data_list, status=200, safe=False)
        else:
            return JsonResponse({"msg": "No data found for given txn_id!"}, status=404)


@csrf_exempt
def getCashAllocationCorrectiveFromCashAllocation(request):
    if request.method == "GET":
        cashallocation_id = request.GET.get("cashallocation_id")
        objects = CashAllocationCorrective.objects.filter(
            cash_allocation__id=cashallocation_id
        )
        if objects:
            serializer = CashAllocationCorrectiveCreateSerializer(objects, many=True)
            data_list = serializer.data
            get_wf_id = WorkFlow.objects.get(workflow_name='WF_CORRECTIVE_TRANSFER')
            for data in data_list:
                workflow_ser = WorkflowBankTransactionsSerializer(data)
                data['txn_id'] = WorkflowBankTransactions.objects.filter(bank_txn_id=workflow_ser.data['id'],
                                                                         workflow=get_wf_id.id).first().id if WorkflowBankTransactions.objects.filter(
                    bank_txn_id=workflow_ser.data['id'], workflow=get_wf_id.id).first() else None
                data['workflow'] = WorkflowBankTransactionsSerializer(WorkflowBankTransactions.objects.filter(
                    id=data['txn_id']).first()).data if WorkflowBankTransactionsSerializer(
                    WorkflowBankTransactions.objects.filter(id=data['txn_id']).first()) else None
            return JsonResponse(data_list, status=200, safe=False)
        else:
            return JsonResponse({"msg": "No data found for given txn_id!"}, status=404)


@csrf_exempt
def getCashAllocationWriteoffFromCashAllocation(request):
    if request.method == "GET":
        cashallocation_id = request.GET.get("cashallocation_id")
        objects = CashAllocationWriteoff.objects.filter(
            cash_allocation__id=cashallocation_id
        )
        if objects:
            serializer = CashAllocationWriteoffCreateSerializer(objects, many=True)
            data_list = serializer.data
            get_wf_id = WorkFlow.objects.get(workflow_name='WF_WRITEOFF')
            for data in data_list:
                try:
                    workflow_ser = WorkflowBankTransactionsSerializer(data)
                    data['txn_id'] = WorkflowBankTransactions.objects.filter(bank_txn_id=workflow_ser.data['id'],
                                                                            workflow=get_wf_id.id).first().id
                    data['workflow'] = WorkflowBankTransactionsSerializer(
                        WorkflowBankTransactions.objects.get(id=data['txn_id'])).data
                except:
                    data['workflow'] = None
            return JsonResponse(data_list, status=200, safe=False)
        else:
            return JsonResponse({"msg": "No data found for given txn_id!"}, status=404)


@csrf_exempt
def getCashAllocationRefundFromCashAllocation(request):
    if request.method == "GET":
        cashallocation_id = request.GET.get("cashallocation_id")

        objects = CashAllocationRefund.objects.filter(
            cash_allocation__id=cashallocation_id
        )
        if objects:
            serializer = CashAllocationRefundCreateSerializer(objects, many=True)
            data_list = serializer.data
            get_wf_id = WorkFlow.objects.get(workflow_name='WF_REFUND')
            for data in data_list:
                workflow_ser = WorkflowBankTransactionsSerializer(data)
                data['txn_id'] = WorkflowBankTransactions.objects.filter(bank_txn_id=workflow_ser.data['id'],
                                                                         workflow=get_wf_id.id).first().id if WorkflowBankTransactions.objects.filter(
                    bank_txn_id=workflow_ser.data['id'], workflow=get_wf_id.id).first() else None
                data['workflow'] = WorkflowBankTransactionsSerializer(WorkflowBankTransactions.objects.filter(
                    id=data['txn_id']).first()).data if WorkflowBankTransactionsSerializer(
                    WorkflowBankTransactions.objects.filter(id=data['txn_id']).first()) else None

            return JsonResponse(data_list, status=200, safe=False)
        else:
            return JsonResponse({"msg": "No data found for given txn_id!"}, status=404)


@csrf_exempt
def getCashAllocationCFIFromCashAllocation(request):
    if request.method == "GET":
        cashallocation_id = request.GET.get("cashallocation_id")
        objects = CashAllocationCFI.objects.filter(
            cash_allocation__id=cashallocation_id
        )
        if objects:
            serializer = CashAllocationCFICreateSerializer(objects, many=True)
            data_list = serializer.data
            get_wf_id = WorkFlow.objects.get(workflow_name='WF_CFI')
            for data in data_list:
                workflow_ser = WorkflowBankTransactionsSerializer(data)
                data['txn_id'] = WorkflowBankTransactions.objects.filter(bank_txn_id=workflow_ser.data['id'],
                                                                         workflow=get_wf_id.id).first().id if WorkflowBankTransactions.objects.filter(
                    bank_txn_id=workflow_ser.data['id'], workflow=get_wf_id.id).first() else None
                data['workflow'] = WorkflowBankTransactionsSerializer(WorkflowBankTransactions.objects.filter(
                    id=data['txn_id']).first()).data if WorkflowBankTransactionsSerializer(
                    WorkflowBankTransactions.objects.filter(id=data['txn_id']).first()) else None

            return JsonResponse(data_list, status=200, safe=False)
        else:
            return JsonResponse({"msg": "No data found for given txn_id!"}, status=404)


@csrf_exempt
def getCrossAllocationFromCashAllocation(request):
    if request.method == "GET":
        cashallocation_id = request.GET.get("cashallocation_id")
        objects = CrossAllocation.objects.filter(cash_allocation__id=cashallocation_id)
        if objects:
            serializer = CrossAllocationCreateSerializer(objects, many=True)
            data_list = serializer.data
            get_wf_id = WorkFlow.objects.get(workflow_name='WF_CROSS_ALLOCATION')
            for data in data_list:
                workflow_ser = WorkflowBankTransactionsSerializer(data)
                data['txn_id'] = WorkflowBankTransactions.objects.filter(bank_txn_id=workflow_ser.data['id'],
                                                                         workflow=get_wf_id.id).first().id if WorkflowBankTransactions.objects.filter(
                    bank_txn_id=workflow_ser.data['id'], workflow=get_wf_id.id).first() else None
                data['workflow'] = WorkflowBankTransactionsSerializer(WorkflowBankTransactions.objects.filter(
                    id=data['txn_id']).first()).data if WorkflowBankTransactionsSerializer(
                    WorkflowBankTransactions.objects.filter(id=data['txn_id']).first()) else None
            return JsonResponse(data_list, status=200, safe=False)
        else:
            return JsonResponse({"msg": "No data found for given txn_id!"}, status=404)


@csrf_exempt
def getCashAllocationMSDFromCashAllocation(request):
    if request.method == "GET":
        cashallocation_id = request.GET.get("cashallocation_id")
        objects = CashAllocationMSD.objects.filter(
            cash_allocation__id=cashallocation_id
        )
        if objects:
            serializer = CashAllocationMSDCreateSerializer(objects, many=True)
            return JsonResponse(serializer.data, status=200, safe=False)
        else:
            return JsonResponse({"msg": "No data found for given txn_id!"}, status=404)


@csrf_exempt
def getPremiumPaymentFromCashAllocation(request):
    if request.method == "GET":
        cashallocation_id = request.GET.get("cashallocation_id")
        objects = PremiumPayment.objects.filter(cash_allocation__id=cashallocation_id)
        if objects:
            serializer = PremiumPaymentCreateSerializer(objects, many=True)
            return JsonResponse(serializer.data, status=200, safe=False)
        else:
            return JsonResponse({"msg": "No data found for given txn_id!"}, status=404)


@csrf_exempt
def getCorrectiveTRFFromCashAllocation(request):
    if request.method == "GET":
        cashallocation_id = request.GET.get("cashallocation_id")
        objects = CorrectiveTRF.objects.filter(cash_allocation__id=cashallocation_id)
        if objects:
            serializer = CorrectiveTRFCreateSerializer(objects, many=True)
            data_list = serializer.data
            get_wf_id = WorkFlow.objects.get(workflow_name='WF_CORRECTIVE_TRANSFER')
            for data in data_list:
                workflow_ser = WorkflowBankTransactionsSerializer(data)
                data['txn_id'] = WorkflowBankTransactions.objects.filter(bank_txn_id=workflow_ser.data['id'],
                                                                         workflow=get_wf_id.id).first().id if WorkflowBankTransactions.objects.filter(
                    bank_txn_id=workflow_ser.data['id'], workflow=get_wf_id.id).first() else None
                data['workflow'] = WorkflowBankTransactionsSerializer(WorkflowBankTransactions.objects.filter(
                    id=data['txn_id']).first()).data if WorkflowBankTransactionsSerializer(
                    WorkflowBankTransactions.objects.filter(id=data['txn_id']).first()) else None
            return JsonResponse(data_list, status=200, safe=False)
        else:
            return JsonResponse({"msg": "No data found for given txn_id!"}, status=404)


class WorkStepViewSet(viewsets.ModelViewSet):
    queryset = WorkStep.objects.all()
    serializer_class = WorkStepSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = WorkStep.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = WorkStepCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()

        obj.user.set(request.data["user"])
        obj.step_name = request.data["step_name"]
        obj.comments = request.data["comments"]
        obj.status = request.data["status"]
        obj.save()
        serializer = WorkStepCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.user.set(request.data["user"])
        obj.step_name = request.data["step_name"]
        obj.comments = request.data["comments"]
        obj.status = request.data["status"]
        obj.save()
        serializer = WorkStepCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class WorkFlowViewSet(viewsets.ModelViewSet):
    queryset = WorkFlow.objects.all()
    serializer_class = WorkFlowSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = WorkFlow.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = WorkFlowCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()

        obj.workflow_step.set(request.data["workflow_step"])
        obj.workflow_name = request.data["workflow_name"]
        obj.save()
        serializer = WorkFlowCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()

        obj.workflow_step.set(request.data["workflow_step"])
        obj.workflow_name = request.data["workflow_name"]
        obj.save()
        serializer = WorkFlowCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(["POST"])
def setWorkflowWithTransaction(request):
    if request.method == "POST":
        json_data = json.loads(request.body.decode("utf-8"))
        workflow_id = json_data["workflow_id"]
        bank_txn_id = json_data["bank_txn_id"]

        bankTxnObject = BankTransaction.objects.filter(archived=False, Bank_Transaction_Id=bank_txn_id)
        if bankTxnObject.exists():
            workflow = WorkFlow.objects.get(id=workflow_id)
            bankTxnObject.update(workflow=workflow)
            return Response(
                {"data": BankTransactionSerializer(bankTxnObject[0]).data}, status=200
            )
        else:
            return Response({"msg": "Bank txn id is not found"}, status=400)


import pandas as pd
from .forms import *


@csrf_exempt
def cash_tracker_excel_import(request):
    if request.method == "POST":
        form = CashTrackerImportForm(request.POST, request.FILES)
        if form.is_valid():
            dict_excel_data = []
            df = pd.read_excel(request.FILES["cash_tracker_excel_file"])
            dff = df.dropna(subset=["Accounting Month"])
            for index, row in dff.iterrows():
                ddd = {}
                Accounting_Monthy = row["Accounting Month"]
                if not Accounting_Monthy == "nan":
                    ddd["Accounting_Month"] = Accounting_Monthy

                    Allocation_Status = row["Allocation Status"]
                    ddd["Allocation_Status"] = Allocation_Status

                    Division = row["Division"]
                    ddd["Division"] = Division

                    Receiving_Bank_Account = row["Receiving Bank Account #"]
                    ddd["Receiving_Bank_Account"] = Receiving_Bank_Account

                    Cash_Transfers = row["Cash Transfers"]
                    ddd["Cash_Transfers"] = Cash_Transfers

                    Policy_Type = row["Policy Type"]
                    ddd["Policy_Type"] = Policy_Type

                    LOB = row["LOB"]
                    ddd["LOB"] = LOB

                    Broker = row["Broker"]
                    ddd["Broker"] = Broker

                    Broker_Branch = row["Broker Branch"]
                    ddd["Broker_Branch"] = Broker_Branch

                    Binding_Agreement = row["Binding Agreement"]
                    ddd["Binding_Agreement"] = Binding_Agreement

                    SCM_Partners = row["SCM Partners"]
                    ddd["SCM_Partners"] = SCM_Partners

                    EEA_Non_EEA = row["EEA / Non-EEA"]
                    ddd["EEA_Non_EEA"] = EEA_Non_EEA

                    Payment_Receive_Date = row["Payment Receive Date"]
                    ddd["Payment_Receive_Date"] = Payment_Receive_Date

                    Payment_Currency_Code = row["Payment Currency Code"]
                    ddd["Payment_Currency_Code"] = Payment_Currency_Code

                    Bank_Currency_Code = row["Bank Currency Code"]
                    ddd["Bank_Currency_Code"] = Bank_Currency_Code

                    ROE_Bank_Statement = row["ROE - Bank Statement"]
                    ddd["ROE_Bank_Statement"] = ROE_Bank_Statement

                    Master_Binder = row["Master Binder"]
                    ddd["Master_Binder"] = Master_Binder

                    Policy = row["Policy"]
                    ddd["Policy"] = Policy

                    Category = row["Category"]
                    ddd["Category"] = Category

                    Bank_Charges = row["Bank Charges"]
                    ddd["Bank_Charges"] = Bank_Charges

                    Receivable_Amount = row["Receivable Amount"]
                    ddd["Receivable_Amount"] = Receivable_Amount

                    Allocated_Amount = row["Allocated Amount"]
                    ddd["Allocated_Amount"] = Allocated_Amount

                    dict_excel_data.append(ddd)
                    #
                    CashTracker.objects.create(
                        Accounting_Monthy=Accounting_Monthy,
                        Allocation_Status=Allocation_Status,
                        Division=Division,
                        Receiving_Bank_Account=Receiving_Bank_Account,
                        Cash_Transfers=Cash_Transfers,
                        Policy_Type=Policy_Type,
                        LOB=LOB,
                        Broker=Broker,
                        Broker_Branch=Broker_Branch,
                        Binding_Agreement=Binding_Agreement,
                        SCM_Partners=SCM_Partners,
                        EEA_Non_EEA=EEA_Non_EEA,
                        Payment_Receive_Date=Payment_Receive_Date,
                        Payment_Currency_Code=Payment_Currency_Code,
                        ROE_Bank_Statement=ROE_Bank_Statement,
                        Master_Binder=Master_Binder,
                        Policy=Policy,
                        Category=Category,
                        Bank_Charges=Bank_Charges,
                        Receivable_Amount=Receivable_Amount,
                        Allocated_Amount=Allocated_Amount,
                    )
            return JsonResponse({"results": dict_excel_data})


class CashTrackerViewSet(viewsets.ModelViewSet):
    queryset = CashTracker.objects.all()
    serializer_class = CashTrackerSerializer


@csrf_exempt
def getCashTracker(request):
    ### Using the View SQL Query
    if request.method == "GET":
        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        skip = page_number * rows_per_page

        bank_name = request.GET.get("bankName", None)
        allocation_status = request.GET.get("status", None)
        transactionId = request.GET.get("transactionId", None)
        policyNumber = request.GET.get("policyNumber", None)
        searchKey = request.GET.get('searchKey', None)
        receivableAmount = request.GET.get("receivableAmount", None)

        allocated = request.GET.get("allocated", None)
        from_allocation_date = request.GET.get("fromAllocationDate", None)
        to_allocation_date = request.GET.get("toAllocationDate", None)

        # Constructing the base SQL query
        sql_query = "SELECT * FROM financial_data_onlinectr WHERE \"Cash Allocation ID\" != '' AND \"Cash Allocation ID\" ~ '^\d+$'"

        # Adding conditions based on the request parameters
        params = []

        if bank_name:
            sql_query += " AND \"Receiving Bank Name\" = %s"
            params.append(bank_name)

        if allocation_status:
            sql_query += " AND \"allocation status\" = %s"  # Adjust if the view doesn't have this field
            params.append(allocation_status)

        # Handle receivable amount with precision logic
        receivable_amount_lower = None
        receivable_amount_upper = None

        if receivableAmount:
            floatReceivableAmount = float(receivableAmount)
            intReceivableAmount = int(floatReceivableAmount)
            if intReceivableAmount != floatReceivableAmount:
                receivable_amount_lower = floatReceivableAmount
                receivable_amount_upper = floatReceivableAmount
            else:
                trailing_zero_count = len(str(intReceivableAmount)) - len(str(intReceivableAmount).rstrip('0'))
                if trailing_zero_count == 0:
                    receivable_amount_lower = receivableAmount
                    receivable_amount_upper = receivableAmount
                else:
                    if intReceivableAmount >= 0:
                        receivable_amount_lower = intReceivableAmount
                        receivable_amount_upper = intReceivableAmount + 10**trailing_zero_count
                    else:
                        receivable_amount_lower = intReceivableAmount - 10**trailing_zero_count
                        receivable_amount_upper = intReceivableAmount

        if receivable_amount_lower is not None and receivable_amount_upper is not None:
            sql_query += ' AND "Receivable Settlement Amount" >= %s AND "Receivable Settlement Amount" <= %s'
            params.extend([receivable_amount_lower, receivable_amount_upper])

        if transactionId:
            sql_query += " AND \"Bank Transaction Number\" ILIKE %s"
            params.append(f"%{transactionId}%")

        if policyNumber:
            sql_query += " AND \"Policy\" ILIKE %s"  # Adjust if the view doesn't have this field
            params.append(f"%{policyNumber}%")

        if allocated:
            if allocated == "Allocated":
                sql_query += " AND \"allocation status\" = %s"  # Adjust if the view doesn't have this field
                params.append("Allocated")
            elif allocated == "Unallocated":
                sql_query += " AND \"allocation status\" != %s"  # Adjust if the view doesn't have this field
                params.append("Allocated")

        if from_allocation_date and to_allocation_date:
            sql_query += " AND TO_CHAR(\"Allocation Date\"::date, 'YYYY-MM-DD') >= %s"  # Adjust if the view doesn't have this field
            params.append(from_allocation_date)
            sql_query += " AND TO_CHAR(\"Allocation Date\"::date, 'YYYY-MM-DD') <= %s"  # Adjust if the view doesn't have this field
            params.append(to_allocation_date)

        if searchKey:
            sql_query += " AND (\"Batch Reference\" ILIKE %s OR \"GXB Batch #\" ILIKE %s OR \"Cash Reference\" ILIKE %s)"
            params.extend([f"%{searchKey}%", f"%{searchKey}%", f"%{searchKey}%"])

        # 1) RUN FULL QUERY (NO LIMIT) → for summary calculations
        with connection.cursor() as cursor:
            cursor.execute(sql_query, params)
            summary_rows = cursor.fetchall()
            column_names = [desc[0] for desc in cursor.description]

        # 2) PAGINATED QUERY → ONLY rows_per_page rows for FE  
        paged_query = sql_query + ' ORDER BY "Cash Allocation ID"::int DESC LIMIT %s OFFSET %s'
        paged_params = params + [rows_per_page, skip]

        with connection.cursor() as cursor:
            cursor.execute(paged_query, paged_params)
            page_rows = cursor.fetchall()

        # COUNT QUERY (unchanged)
        count_query = "SELECT COUNT(*) FROM financial_data_onlinectr WHERE \"Cash Allocation ID\" != '' AND \"Cash Allocation ID\" ~ '^\d+$'"
        count_params = []

        if bank_name:
            count_query += " AND \"Receiving Bank Name\" = %s"
            count_params.append(bank_name)

        if allocation_status:
            count_query += " AND \"allocation status\" = %s"
            count_params.append(allocation_status)

        if receivable_amount_lower is not None:
            count_query += ' AND "Receivable Settlement Amount" >= %s AND "Receivable Settlement Amount" <= %s'
            count_params.extend([receivable_amount_lower, receivable_amount_upper])

        if transactionId:
            count_query += " AND \"Bank Transaction Number\" ILIKE %s"
            count_params.append(f"%{transactionId}%")

        if policyNumber:
            count_query += " AND \"Policy\" ILIKE %s"  # Adjust if the view doesn't have this field
            count_params.append(f"%{policyNumber}%")

        if allocated:
            if allocated == "Allocated":
                count_query += " AND \"allocation status\" = %s"  # Adjust if the view doesn't have this field
                count_params.append("Allocated")
            elif allocated == "Unallocated":
                count_query += " AND \"allocation status\" != %s"  # Adjust if the view doesn't have this field
                count_params.append("Allocated")

        if from_allocation_date and to_allocation_date:
            count_query += " AND TO_CHAR(\"Allocation Date\"::date, 'YYYY-MM-DD') >= %s"  # Adjust if the view doesn't have this field
            count_params.append(from_allocation_date)
            count_query += " AND TO_CHAR(\"Allocation Date\"::date, 'YYYY-MM-DD') <= %s"  # Adjust if the view doesn't have this field
            count_params.append(to_allocation_date)

        if searchKey:
            count_query += " AND (\"Batch Reference\" ILIKE %s OR \"GXB Batch #\" ILIKE %s OR \"Cash Reference\" ILIKE %s)"
            count_params.extend([f"%{searchKey}%", f"%{searchKey}%", f"%{searchKey}%"])

        # Execute the count query
        with connection.cursor() as cursor:
            cursor.execute(count_query, count_params)
            count = cursor.fetchone()[0]


        summary_data = {
            "total_premium_due_amt": 0,
            "receivable_amount": 0,
            "allocated_amount": 0,
            "outstanding_amt": 0,
            "total_receivable_amount": 0,
            "unallocated_amount": 0
        }

        transaction_id_set = set()

        bank_txn_ids = [row[column_names.index('Bank Transaction ID')] for row in summary_rows]
        ca_ids = [row[column_names.index('Cash Allocation ID')] for row in summary_rows]

        bank_txn_map = {obj.id: obj for obj in BankTransaction.objects.filter(id__in=bank_txn_ids)}
        ca_map = {obj.id: obj for obj in CashAllocation.objects.filter(id__in=ca_ids)}

        for row in summary_rows:
            bank_txn_id = row[column_names.index('Bank Transaction ID')]
            bank_txn_num = row[column_names.index('Bank Transaction Number')]
            ca_id = row[column_names.index('Cash Allocation ID')]

            bank_txn_obj = bank_txn_map.get(bank_txn_id)
            ca = ca_map.get(int(ca_id))

            receivable_amt = getattr(ca, 'receivable_amt', 0)
            allocated_amt = getattr(ca, 'allocated_amt', 0)
            receivable_amount = getattr(bank_txn_obj, 'Receivable_Amount', 0)

            summary_data["total_premium_due_amt"] += Decimal(row[column_names.index('Premium Due amt Original')] or 0)
            summary_data["receivable_amount"] += Decimal(row[column_names.index('Receivable Settlement Amount')] or 0)
            summary_data["allocated_amount"] += Decimal(row[column_names.index('Allocated Amount')] or 0)

            if bank_txn_num and bank_txn_num.strip() not in transaction_id_set:
                summary_data["total_receivable_amount"] += Decimal(receivable_amount or 0)
            transaction_id_set.add(bank_txn_num.strip())

            summary_data["unallocated_amount"] += Decimal(row[column_names.index('Unallocated Amount')] or 0)

        
        # Build PAGINATED DATA LIST
        data = []

        for row in page_rows:
            bank_txn_id = row[column_names.index('Bank Transaction ID')]
            bank_txn_obj = bank_txn_map.get(bank_txn_id)

            allocation_data = {
                "id": row[column_names.index('Cash Allocation ID')],
                "Accounting_Month": row[column_names.index('accounting_monthyear')],
                "Bank_Transaction_Id": row[column_names.index('Bank Transaction Number')],
                "Bank_Name": '',
                "Real_Bank_Name": row[column_names.index('Receiving Bank Name')],
                "Broker_Branch": row[column_names.index('Broker Branch')],
                "bank_txn": BankTransactionCreateSerializer(bank_txn_obj).data if bank_txn_obj else {},
                "Policy_Line_Ref": row[column_names.index('Policy')],
                "allocation_status": row[column_names.index('allocation status')],
                "bank_curr": row[column_names.index('Bank Currency')],
                "allocated_amt": row[column_names.index('Allocated Amount')],
                "cashreference": row[column_names.index('Cash Reference')],
                "GXPbatchno": row[column_names.index('GXB Batch #')],
                "XFIbatchno": row[column_names.index('Batch Reference')],
                "comment": row[column_names.index('Comment')],
                "historical": row[column_names.index('ca historical')],
                "original_curr": row[column_names.index('Original Currency')],
                "receivable_amt": row[column_names.index('Receivable Settlement Amount')],
                "unallocated_amt": row[column_names.index('Unallocated Amount')],
                "policy_allocation_status": row[column_names.index('allocation status')],
                "allocation_date": row[column_names.index('Allocation Date')],
                "Assigned_User": row[column_names.index('Account Handler')],
                "Assigned_Users": row[column_names.index('Account Handler')],
            }

            data.append(allocation_data)

        response_data = {
            "count": count,
            "data": data,
            "summary": summary_data
        }

        return JsonResponse(response_data)


def filter_qs_if_not_Nonee(qs, **kwargs):
    return qs.filter(
        **{
            key: value
            for key, value in kwargs.items()
            if value is not None and str(value).strip()
        }
    )


class BankTransactionSearchViewSet(ListAPIView):
    model = BankTransaction
    serializer_class = BankTransactionSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def post(self, request):
        data = request.data
        date_and_time_to = data["date_and_time_to"]
        date_and_time_frm = data["date_and_time_frm"]
        Bank_Account_Name_Entity = data["Bank_Account_Name_Entity"]
        Bank_Transaction_Id = data["Bank_Transaction_Id"]
        Receiving_Bank_Account = data["Receiving_Bank_Account"]
        skip = data["skip"]
        pageSize = data["pageSize"]

        from dateutil import parser

        if date_and_time_to:
            date_and_time_to = datetime.strptime(date_and_time_to, "%m-%d-%Y")
            print(date_and_time_to)
        if date_and_time_frm:
            date_and_time_frm = datetime.strptime(date_and_time_frm, "%m-%d-%Y")
            print(date_and_time_frm)
        if date_and_time_frm and date_and_time_to:
            bnk_txns_count = filter_qs_if_not_Nonee(
                BankTransaction.objects.filter(archived=False)
                .select_related("broker_information")
                .order_by("-id"),
                Bank_Account_Name_Entity=Bank_Account_Name_Entity,
                Bank_Transaction_Id=Bank_Transaction_Id,
                Receiving_Bank_Account=Receiving_Bank_Account,
                Date_And_Time__range=[date_and_time_frm, date_and_time_to],
            ).count()

            bnk_txns = filter_qs_if_not_Nonee(
                BankTransaction.objects.filter(archived=False)
                .select_related("broker_information")
                .order_by("-id"),
                Bank_Account_Name_Entity=Bank_Account_Name_Entity,
                Bank_Transaction_Id=Bank_Transaction_Id,
                Receiving_Bank_Account=Receiving_Bank_Account,
                Date_And_Time__range=[date_and_time_frm, date_and_time_to],
            )[skip: skip + pageSize]
        else:
            bnk_txns_count = filter_qs_if_not_Nonee(
                BankTransaction.objects.filter(archived=False)
                .select_related("broker_information")
                .order_by("-id"),
                Bank_Account_Name_Entity=Bank_Account_Name_Entity,
                Bank_Transaction_Id=Bank_Transaction_Id,
            ).count()

            bnk_txns = filter_qs_if_not_Nonee(
                BankTransaction.objects.filter(archived=False)
                .select_related("broker_information")
                .order_by("-id"),
                Bank_Account_Name_Entity=Bank_Account_Name_Entity,
                Bank_Transaction_Id=Bank_Transaction_Id,
                Receiving_Bank_Account=Receiving_Bank_Account,
            )[skip: skip + pageSize]

        serializer = BankTransactionSerializer(bnk_txns, many=True)
        return JsonResponse({"data": serializer.data, "totalCount": bnk_txns_count})


class CashTrackerReportViewSet(viewsets.ModelViewSet):
    queryset = CashTrackerReport.objects.all()
    serializer_class = CashTrackerReportSerializer


class CashAllocationMSDViewSet(viewsets.ModelViewSet):
    queryset = CashAllocationMSD.objects.all()
    serializer_class = CashAllocationMSDSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = CashAllocationMSD.objects.all()
        return objs

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = CashAllocationMSDCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer_saved = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.je_number = request.data["je_number"]
        obj.comments = request.data["comments"]
        obj.msd_amt = request.data["msd_amt"]
        obj.system_date = request.data["system_date"]
        obj.process_date = request.data["process_date"]
        obj.user_id = request.data["user_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = CashAllocationMSDCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.bank_txn = BankTransaction.objects.get(id=request.data["bank_txn"])
        obj.cash_allocation = CashAllocation.objects.get(
            id=request.data["cash_allocation"]
        )
        obj.policy_id = request.data["policy_id"]
        obj.je_number = request.data["je_number"]
        obj.comments = request.data["comments"]
        obj.msd_amt = request.data["msd_amt"]
        obj.system_date = request.data["system_date"]
        obj.process_date = request.data["process_date"]
        obj.user_id = request.data["user_id"]
        obj.accounting_monthyear = request.data["accounting_monthyear"]
        obj.policy_fk = PolicyInformation.objects.get(id=request.data["policy_fk"])
        obj.policy_mf.set(request.data["policy_mf"])
        obj.save()
        serializer = CashAllocationMSDCreateSerializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class WorkflowBankTransactionsViewSet(viewsets.ModelViewSet):
    queryset = WorkflowBankTransactions.objects.all()
    serializer_class = WorkflowBankTransactionsSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        objs = WorkflowBankTransactions.objects.all()
        return objs

    def workflow_specific_change_values(self, data, workflow_name):

        changefields = {}
        if workflow_name == 'CHANGE_BANK_TRANSACTION_AMOUNT':
            Receivable_Amount = data["Receivable_Amount"]
            changefields = {"Receivable_Amount": Receivable_Amount}
        return changefields

    def workflow_specific_change_implement(self, rec, bank_txn_id, workflow_name, stepp, user=None):
        bank_txn_object = None
        if workflow_name == 'CHANGE_BANK_TRANSACTION_AMOUNT' and stepp == 'changeamt':
            Receivable_Amount = safe_decimal_conversion(rec["Receivable_Amount"])
            bank_txn_object = BankTransaction.objects.filter(archived=False, 
            Bank_Transaction_Id=bank_txn_id).update(
                Receivable_Amount=Receivable_Amount)
        elif workflow_name == 'WF_WRITEOFF' and stepp == 'process_writeoff':
            try:
                obj = CashAllocationWriteoff.objects.get(id=bank_txn_id)
                obj.process_date = datetime.now()
                obj.approved_date = datetime.now()
                obj.approved_status = 'Approved'
                obj.approver_id = user
                obj.save()
            except:
                print("Failed updating write off with bank txn id: ", bank_txn_id)
                pass
        elif workflow_name == 'WF_CORRECTIVE_TRANSFER' and stepp == 'process_corrective_transfer':
            try:
                bank_txn_object = BankTransaction.objects.filter(archived=False, Bank_Transaction_Id=bank_txn_id).update(
                    Receivable_Amount=Receivable_Amount)
                obj = CashAllocationCorrective.objects.get(id=bank_txn_id)
                obj.approved_status = 'Approved'
                obj.approver_id = user
                obj.save()
            except:
                print("Failed updating corrective with bank txn id: ", bank_txn_id)
                pass
        elif workflow_name == 'WF_REFUND' and stepp == 'process_refund':
            try:
                obj = CashAllocationRefund.objects.get(id=bank_txn_id)
                obj.process_date = datetime.now()
                obj.approved_date = datetime.now()
                obj.approved_status = 'Approved'
                obj.approver_id = user
                obj.save()
            except:
                print("Failed updating refund with bank txn id: ", bank_txn_id)
                pass
        elif workflow_name == 'WF_CFI' and stepp == 'process_cfi':
            try:
                obj = CashAllocationCFI.objects.get(id=bank_txn_id)
                obj.process_date = datetime.now()
                obj.approved_date = datetime.now()
                obj.approved_status = 'Approved'
                obj.approver_id = user
                obj.save()
            except:
                print("Failed updating cfi with bank txn id: ", bank_txn_id)
                pass
        elif workflow_name == 'WF_CROSS_ALLOCATION' and stepp == 'process_cross_allocation':
            try:
                obj = CrossAllocation.objects.get(id=bank_txn_id)
                obj.process_date = datetime.now()
                obj.approved_date = datetime.now()
                obj.approved_status = 'Approved'
                obj.approver_id = user
                obj.save()
            except:
                print("Failed updating cross allocation with bank txn id: ", bank_txn_id)
                pass
        else:
            bank_txn_object = None

        return bank_txn_object

    def createnew_new(self, request, files):
        data = request.data
        workflow_name = data["workflow_name"]
        bank_txn_id = data["bank_txn_id"]
        print(" create new: ", bank_txn_id)
        comments = data["comments"]
        initiated_user_id = data["initiated_user_id"]
        workflow_name = data["workflow_name"]
        user_data = Users.objects.get(id=initiated_user_id)
        id = user_data.id
        initiated_by = initiated_user_id
        # email = user_data.get_decrypted_email()
        email = user_data.email
        initiate_step_email_recs = [user_data.get_decrypted_email(), ]
        user_name = user_data.user_name
        initiated_data = {}
        initiated_data["id"] = id
        initiated_data["email"] = email
        initiated_data["user_name"] = user_name
        gjjj = []
        gjjj.append(initiated_data)
        kl = {}
        kl["user"] = gjjj
        kl["comments"] = comments
        kl["status"] = "NEW"
        kl["ctime"] = str(datetime.now())
        kl["uptime"] = str()
        kl["step_name"] = "initiater"
        kl["step_process"] = None
        ff = WorkFlow.objects.get(workflow_name=workflow_name)
        gg = WorkFlowSerializer(ff)
        db_files = []
        for i in files:
            reuired_data = {"module_name": "WF_Bank_Amount_Change", "bucket_name" : config("WF_BANK_AMOUNT_CHANGE_BUCKET_NAME")}
            filemanagment = reusable_file_upload(get_user(request), i, reuired_data, is_upload=False)

            if not isinstance(filemanagment, Response) or filemanagment.status_code >= 400:
                # Extract error message from response
                error_msg = filemanagment.data.get('error', 'Unable to upload the file to S3') if isinstance(filemanagment, Response) else 'Unable to upload the file to S3'
                raise Exception(error_msg)
            
            db_files.append(filemanagment.data['file_name'])
        files_arr = ','.join(db_files)
        hhhh = WorkflowBankTransactions.objects.create(
            workflow=ff, bank_txn_id=bank_txn_id, file=files_arr
        )

        dict = {}
        g = gg.data
        id = g["id"]
        workflow_name = g["workflow_name"]
        kkk = g["workflow_step"]

        # print("kllllllllll", kkk)
        firststelement = 0
        f = []
        f.append(kl)
        # email_from = settings.EMAIL_HOST_USER
        # imagepath = config('MOSAIC_LOGO_IMAGE')
        current_step = None
        current_step_userids = []
        for i in kkk:
            # print("i........:", i)
            if firststelement == 0:
                current_step = i["step_name"]
                users = i["user"]
                initiator_user = users
                # create current step users id (used for displaying worklist)
                current_step_userids = []
                for user in users:
                    userid = user["id"]
                    current_step_userids.append(userid)
                current_step_userids = current_step_userids
                current_step_email_recs = current_step_userids
                past_step_email_recs = []

            firststelement = firststelement + 1
            l = {}
            l["id"] = i["id"]
            l["ctime"] = str()
            l["status"] = i["status"]
            l["uptime"] = str()
            l["comments"] = i["comments"]
            l["step_name"] = i["step_name"]
            l["step_process"] = i["step_process"]
            llll = i["user"]
            oo = []
            for m in llll:
                hh = {}
                hh["id"] = m["id"]
                hh["email"] = m["email"]
                hh["user_name"] = m["user_name"]

                oo.append(hh)
            l["user"] = oo
            f.append(l)
        workflow_step = f
        # current_step = "reviewer"
        hhhh.changefields = self.workflow_specific_change_values(data, workflow_name)
        dict["id"] = id
        dict["workflow_name"] = workflow_name
        dict["workflow_step"] = workflow_step
        hhhh.workflow_json_data = dict
        hhhh.current_step = current_step
        hhhh.current_step_userids = current_step_userids
        current_step_email_recs = current_step_userids

        hhhh.initiated_by = initiated_by
        hhhh.workflow_status = "In Process"
        hhhh.save()
        serializer = WorkflowBankTransactionsSerializer(hhhh)

        ############# SEND EMAIL TO REVIWER ##############
        email_ids = [Users.objects.get(id=user_id).get_decrypted_email() for user_id in current_step_email_recs]
        email_from = settings.EMAIL_HOST_USER
        try:
            subject = "New Worklist Item Assigned to You"
            imagepath = config('MOSAIC_LOGO_IMAGE')
            body = settings.WORK_FLOW_INITIATE_BODY.format(imagepath=imagepath, reviewer_approver='Reviewer',
                                                           initiator=user_name, transaction_id=bank_txn_id)
            send_email(sender_email=email_from, recipient_email=email_ids + initiate_step_email_recs, subject=subject,
                       body=body)
        except Exception as e:
            print("Error occured while sending mail to: {} and error is {}".format(current_step_email_recs, str(e)))
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def create(self, request, *args, **kwargs):
        data = request.data
        files = request.FILES.getlist('files')
        workflow_name = data["workflow_name"]

        # if workflow_name == 'CHANGE_BANK_TRANSACTION_AMOUNT':
        print(" executing create new: ", workflow_name)
        serializer = self.createnew_new(request, files)
        return Response(serializer.data)

        print(" executing existing: ", workflow_name)

        bank_txn_id = data["bank_txn_id"]
        Receivable_Amount = data["Receivable_Amount"]
        comments = data["comments"]
        initiated_user_id = data["initiated_user_id"]
        workflow_name = data["workflow_name"]
        user_data = Users.objects.get(id=initiated_user_id)
        id = user_data.id
        email = user_data.get_decrypted_email()
        user_name = user_data.user_name
        initiated_data = {}
        initiated_data["id"] = id
        initiated_data["email"] = email
        initiated_data["user_name"] = user_name
        gjjj = []
        gjjj.append(initiated_data)
        kl = {}
        kl["user"] = gjjj
        kl["comments"] = comments
        kl["status"] = "NEW"
        kl["ctime"] = str(datetime.now())
        kl["uptime"] = str()
        kl["step_name"] = "initiater"
        # workflow = 5
        work_flow_obj = WorkFlow.objects.get(workflow_name=workflow_name)
        work_flow_serializer = WorkFlowSerializer(work_flow_obj)
        hhhh = WorkflowBankTransactions.objects.create(
            workflow=work_flow_obj, bank_txn_id=bank_txn_id
        )

        dict = {}
        g = work_flow_serializer.data
        id = g["id"]
        workflow_name = g["workflow_name"]
        kkk = g["workflow_step"]

        # print("kllllllllll", kl)

        f = []
        f.append(kl)
        for i in kkk:
            l = {}
            l["id"] = i["id"]
            l["ctime"] = str()
            l["status"] = i["status"]
            l["uptime"] = str()
            l["comments"] = i["comments"]
            l["step_name"] = i["step_name"]
            llll = i["user"]
            oo = []
            for m in llll:
                hh = {}
                hh["id"] = m["id"]
                hh["email"] = m["email"]
                hh["user_name"] = m["user_name"]
                oo.append(hh)
            l["user"] = oo
            f.append(l)
        workflow_step = f
        current_step = "reviewer"
        hhhh.changefields = {"Receivable_Amount": Receivable_Amount}
        dict["id"] = id
        dict["workflow_name"] = workflow_name
        dict["workflow_step"] = workflow_step
        hhhh.workflow_json_data = dict
        hhhh.current_step = current_step
        hhhh.workflow_status = "In Process"
        hhhh.save()
        serializer = WorkflowBankTransactionsSerializer(hhhh)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update_new(self, request):

        obj = self.get_object()
        # print("object: ", obj)
        # print("object: ", obj._dict_)
        bank_txn_id = request.data["bank_txn_id"]
        obj.bank_txn_id = request.data["bank_txn_id"]
        step_name = request.data["step_name"]
        status = request.data["status"]
        comments = request.data["comments"]
        # building current user details
        user_id = request.data["user_id"]
        user_data = Users.objects.get(id=user_id)
        id = user_data.id

        email = user_data.get_decrypted_email()
        user_name = user_data.user_name
        user_data = {}
        user_data["id"] = id
        user_data["email"] = email
        user_data["user_name"] = user_name
        gjjj = []
        gjjj.append(user_data)
        f = obj.workflow_json_data
        id = f["id"]
        workflow_name = f["workflow_name"]
        obj.changefields = obj.changefields
        obj.bank_txn_id = bank_txn_id
        rec = obj.changefields
        if workflow_name == 'CHANGE_BANK_TRANSACTION_AMOUNT':
            Receivable_Amount = safe_decimal_conversion(rec["Receivable_Amount"])  # to be handled
            # print(Receivable_Amount, "Receivable_Amount")

        g = f["workflow_step"]
        # print("xxxxxxxx printing g : ", g)
        # print("xxxxxxxx printing f : ", f)

        lengthg = len(g) - 1

        step_name_found = 0
        jjj = []
        recipient_list = []
        current_step_email_recs = []
        for index, mm in enumerate(g):
            # print("Index 1: ", index)
            if step_name_found == 1 and status.lower() != 'reject':
                obj.current_step = mm["step_name"]
                obj.workflow_status = "in process"
                # print("not rejected: line no 2642 ", obj.current_step, obj.workflow_status)
                users = mm["user"]
                current_step_userids = []
                for user in users:
                    userid = user["id"]
                    current_step_userids.append(userid)
                obj.current_step_userids = current_step_userids
                current_step_email_recs = current_step_userids
                step_name_found = 0

            if mm["step_name"] == step_name:
                mm["ctime"] = str(datetime.now())
                mm["status"] = status
                mm["comments"] = comments
                mm["user"] = gjjj
                past_step_email_recs = gjjj
                step_name_found = 1
                recipient_list.append(mm['user'][0]['email'])
                if mm["step_process"] is not None and status.lower() == 'approve':
                    # print("executing step process: ", mm["step_process"])
                    stepp = mm["step_process"]
                    self.workflow_specific_change_implement(rec, bank_txn_id, workflow_name, stepp, user=mm["user"])
                    if index == lengthg:
                        obj.workflow_status = "completed"
                        obj.current_step = None
                elif status.lower() == 'reject':

                    obj.current_step = None
                    obj.workflow_status = "rejected"
                    # print("rejected: line no 2668 ", obj.current_step, obj.workflow_status)

            # print("Index and Lengthg and status: ", index, lengthg, status)

            # if status == 'Approve' and index == lengthg:
            # obj.current_step = None
            # obj.workflow_status = "completed"
            # self.workflow_specific_change_implement(rec, bank_txn_id, workflow_name)

            jjj.append(mm)

        workflow_step = jjj
        dictss = {}
        dictss["id"] = id
        dictss["workflow_name"] = workflow_name
        dictss["workflow_step"] = workflow_step

        obj.workflow_json_data = dictss
        obj.save()
        ##################### SEND EMAIL ####################
        email_ids = [Users.objects.get(id=user_id).get_decrypted_email() for user_id in current_step_email_recs]
        past_users_email_ids = [user_data['email'] for user_data in past_step_email_recs]

        email_from = settings.EMAIL_HOST_USER
        try:
            subject = "Status Update for {}".format(bank_txn_id)
            imagepath = config('MOSAIC_LOGO_IMAGE')
            body = settings.WORK_FLOW_CHANGE_STATUS_REVIEWER_BODY.format(imagepath=imagepath,
                                                                         reviewer_approver=user['user_name'],
                                                                         initiator=user_name,
                                                                         transaction_id=bank_txn_id)
            send_email(sender_email=email_from, recipient_email=email_ids + past_users_email_ids, subject=subject,
                       body=body)
        except Exception as e:
            print("Error occured while sending mail to: {} and error is {}".format(current_step_email_recs, str(e)))

        serializer = WorkflowBankTransactionsSerializer(obj)
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):

        obj = self.get_object()
        workflow_name = obj.workflow_json_data["workflow_name"]

        # if workflow_name == 'CHANGE_BANK_TRANSACTION_AMOUNT':
        print(" executing new patial update: ", workflow_name)
        serializer = self.partial_update_new(request)
        return Response(serializer.data)

        print(" executing existing patial update: ", workflow_name)

        bank_txn_id = request.data["bank_txn_id"]
        obj.bank_txn_id = request.data["bank_txn_id"]
        step_name = request.data["step_name"]
        status = request.data["status"]
        comments = request.data["comments"]
        user_id = request.data["user_id"]
        user_data = Users.objects.get(id=user_id)
        id = user_data.id
        email = user_data.get_decrypted_email()
        user_name = user_data.user_name
        user_data = {}
        user_data["id"] = id
        user_data["email"] = email
        user_data["user_name"] = user_name
        gjjj = []
        gjjj.append(user_data)
        f = obj.workflow_json_data
        id = f["id"]
        workflow_name = f["workflow_name"]
        obj.changefields = obj.changefields
        obj.bank_txn_id = bank_txn_id
        rec = obj.changefields
        Receivable_Amount = rec["Receivable_Amount"]
        print(Receivable_Amount, "Receivable_Amount")
        g = f["workflow_step"]
        jjj = []
        for mm in g:
            if mm["status"] == "NEW" and not mm["uptime"]:
                mm["uptime"] = str(datetime.now())
            if mm["status"] == "Approve" or mm["status"] == "Reject":
                mm["uptime"] = str(datetime.now())
            if mm["step_name"] == step_name:
                mm["ctime"] = str(datetime.now())
                mm["status"] = status
                mm["comments"] = comments
                mm["user"] = gjjj
            if step_name == "approver":
                mm["uptime"] = str(datetime.now())
            if step_name == "approver" and status == "Approve":
                BankTransaction.objects.filter(archived=False, Bank_Transaction_Id=bank_txn_id).update(
                    Receivable_Amount=Receivable_Amount
                )
            jjj.append(mm)
        workflow_step = jjj
        dictss = {}
        dictss["id"] = id
        dictss["workflow_name"] = workflow_name
        dictss["workflow_step"] = workflow_step

        cut = request.data["step_name"]

        if cut == "reviewer" and status == "Approve":
            print("sss")
            obj.current_step = "approver"
            obj.workflow_status = "In Process"

        if cut == "reviewer" and status == "Reject":
            obj.current_step = ""
            obj.workflow_status = "Rejected"

        if cut == "approver" and status == "Approve":
            obj.current_step = ""
            obj.workflow_status = "Approved"

        if cut == "approver" and status == "Reject":
            obj.current_step = ""
            obj.workflow_status = "Rejected"

        obj.workflow_json_data = dictss
        obj.save()
        serializer = WorkflowBankTransactionsSerializer(obj)
        return Response(serializer.data)


@csrf_exempt
def getWorkflowList(request):
    if request.method == "GET":
        user_id = request.GET.get("user_id", None)
        if not user_id:
            return Response(
                {"Error": "Please give a user id"}, status=status.HTTP_400_BAD_REQUEST
            )
        user_id = int(user_id)

        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        skip = page_number * rows_per_page

        try:
            user_data = Users.objects.get(id=user_id)
        except Users.DoesNotExist:
            return Response(
                {"Error": "User not found"}, status=status.HTTP_404_NOT_FOUND
            )

        initiated_data = {
            "user": [{"id": user_data.id, "email": user_data.email, "user_name": user_data.user_name}]
        }
        bank_tx_id = request.GET.get("transactionId", None)
        fromDateReceived = request.GET.get("fromDateReceived", None)
        toDateReceived = request.GET.get("toDateReceived", None)

        filter_conditions = Q()

        if bank_tx_id:
            filter_conditions &= Q(bank_txn_id=bank_tx_id)

        if fromDateReceived and toDateReceived:
            from_date = parse_datetime(fromDateReceived)
            to_date = parse_datetime(toDateReceived)
            if from_date and to_date:
                to_date = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)
                filter_conditions &= Q(created_at__range=(from_date, to_date))

        filter_conditions &= Q(workflow_json_data__contains={'workflow_step': [initiated_data]})
        queryset = WorkflowBankTransactions.objects.filter(filter_conditions).order_by('-id')

        g = []
        for i in queryset:
            workflow_name = i.workflow_json_data["workflow_name"]
            if workflow_name == "WF_WRITEOFF":
                obj = CashAllocationWriteoff.objects.get(id=i.bank_txn_id)
                if obj.cash_allocation.archived == True:
                    continue
            elif workflow_name == "WF_CORRECTIVE_TRANSFER":
                obj = CashAllocationCorrective.objects.get(id=i.bank_txn_id)
                if obj.cash_allocation.archived == True:
                    continue
            elif workflow_name == "WF_REFUND":
                obj = CashAllocationRefund.objects.get(id=i.bank_txn_id)
                if obj and obj.cash_allocation and obj.cash_allocation.archived == True:
                    continue
            elif workflow_name == "WF_CFI":
                obj = CashAllocationCFI.objects.get(id=i.bank_txn_id)
                if obj.cash_allocation.archived == True:
                    continue
            elif workflow_name == "WF_CROSS_ALLOCATION":
                obj = CrossAllocation.objects.get(id=i.bank_txn_id)
                if obj.cash_allocation.archived == True:
                    continue
            elif workflow_name == "WF_CORRECTION_TYPES":
                obj = CashAllocationIssues.objects.get(id=i.bank_txn_id)
                if obj.cash_allocation.archived == True:
                    continue
            
            curr = i.current_step
            c = i.workflow_json_data.get("workflow_step", [])
            for m in c:
                f = m.get("user", [])
                for ddd in f:
                    if ddd["id"] == user_id and m["step_name"] in (curr, "initiater"):
                        if i not in g:
                            g.append(i)
                            
        count = len(g)
        serializer = WorkflowBankTransactionsSerializer(g[skip: skip + rows_per_page], many=True)
        data = serializer.data
        response_data = {}
        response_data["count"] = count
        response_data["data"] = data
        return JsonResponse(response_data, safe=False)


class MyPageNumberPaginations(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page"
    max_page_size = 100


class BankTransactionList(generics.ListAPIView):
    queryset = BankTransaction.objects.filter(archived=False).order_by("-id")
    serializer_class = BankTransactionSerializer
    pagination_class = MyPageNumberPaginations

    def get_queryset(self):
        queryset = super().get_queryset()
        query_params = self.request.query_params
        ff = BankTransaction.objects.filter(archived=False)
        with_related = BankTransaction.objects.select_related(
            "Assigned_User", "broker_information", "bank_details", "workflow"
        ).filter(archived=False)
        queryset = with_related.prefetch_related(
            "Assigned_Users", "document_files"
        ).filter(archived=False)
        # print(queryset,"dddddddddddddddddddd")
        date_and_time_frmm = query_params.get("date_and_time_frm")
        date_and_time_too = query_params.get("date_and_time_to")
        bank_account_name_entity = query_params.get("Bank_Account_Name_Entity")
        bank_transaction_id = query_params.get("Bank_Transaction_Id")
        print(date_and_time_frmm, date_and_time_too, "dddddd")
        print(bank_account_name_entity, bank_transaction_id)
        if date_and_time_frmm and date_and_time_too:
            date_format = "%d-%m-%Y"
            date_and_time_frm = datetime.strptime(date_and_time_frmm, date_format)
            date_and_time_to = datetime.strptime(date_and_time_too, date_format)
            print(date_and_time_frm, date_and_time_to, "inside")
            queryset = queryset.filter(
                Date_And_Time__range=[date_and_time_frm, date_and_time_to]
            ).order_by("-id")
        if bank_account_name_entity:
            queryset = queryset.filter(
                Bank_Account_Name_Entity=bank_account_name_entity
            ).order_by("-id")
        if bank_transaction_id:
            queryset = queryset.filter(
                Bank_Transaction_Id=bank_transaction_id
            ).order_by("-id")

        return queryset


class CashAllocationList(generics.ListAPIView):
    queryset = CashAllocation.objects.filter(archived=False).order_by("-id")
    serializer_class = CashAllocationSerializer
    pagination_class = MyPageNumberPaginations

    def get_queryset(self):
        queryset = super().get_queryset()
        query_params = self.request.query_params
        ff = CashAllocation.objects.filter(archived=False)
        with_related = CashAllocation.objects.select_related(
            "bank_txn", "policy_fk"
        ).filter(archived=False)
        queryset = with_related.prefetch_related("policy_fk").filter(archived=False)
        # print(queryset,"dddddddddddddddddddd")
        date_and_time_frmm = query_params.get("date_and_time_frm")
        date_and_time_too = query_params.get("date_and_time_to")
        bank_account_name_entity = query_params.get("Bank_Account_Name_Entity")
        bank_transaction_id = query_params.get("Bank_Transaction_Id")
        print(date_and_time_frmm, date_and_time_too, "dddddd")
        print(bank_account_name_entity, bank_transaction_id)
        if date_and_time_frmm and date_and_time_too:
            date_format = "%d-%m-%Y"
            date_and_time_frm = datetime.strptime(date_and_time_frmm, date_format)
            date_and_time_to = datetime.strptime(date_and_time_too, date_format)
            print(date_and_time_frm, date_and_time_to, "inside")
            queryset = queryset.filter(
                allocation_date__range=[date_and_time_frm, date_and_time_to]
            ).order_by("-id")

        if bank_account_name_entity:
            queryset = queryset.filter(
                bank_txn__Bank_Account_Name_Entity=bank_account_name_entity
            ).order_by("-id")

        if bank_transaction_id:
            queryset = queryset.filter(
                bank_txn__Bank_Transaction_Id=bank_transaction_id
            ).order_by("-id")
        return queryset


class CashTrackerList(generics.ListAPIView):
    queryset = CashTrackerReport.objects.all().order_by("-id")
    serializer_class = CashTrackerReportSerializer
    pagination_class = MyPageNumberPaginations

    def get_queryset(self):
        queryset = super().get_queryset()
        query_params = self.request.query_params
        ff = CashTrackerReport.objects.all()
        queryset = CashTrackerReport.objects.select_related(
            "bank_txn",
            "cash_allocation",
            "ca_issues",
            "ca_corrective",
            "policy_information",
        ).all()
        # print(queryset,"dddddddddddddddddddd")
        Policy = query_params.get("Policy")
        if Policy:
            queryset = queryset.filter(Policy=Policy).order_by("-id")
        return queryset


class BankTransactionWorkflowStatusViewSet(generics.ListAPIView):
    queryset = WorkflowBankTransactions.objects.all().order_by("-id")
    serializer_class = WorkflowBankTransactionsSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        query_params = self.request.query_params
        ff = WorkflowBankTransactions.objects.all()
        queryset = WorkflowBankTransactions.objects.select_related("workflow").all()
        workflow_name = query_params.get("workflow_name")
        bank_txn_id = query_params.get("bank_txn_id")
        if bank_txn_id:
            queryset = queryset.filter(
                workflow_json_data__workflow_name=workflow_name,
                bank_txn_id=bank_txn_id,
                workflow_status="In Process",
            ).order_by("-id")
        return queryset

def safe_date_parse(date_str):
    if pd.isna(date_str):
        return None
    try:
        return pd.to_datetime(date_str, format='mixed')
    except Exception as e:
        logger.error(f"Error in safe_date_parse: {str(e)}")
        return None

def getCashTrackerExport(request):
    logger.info("Starting getCashTrackerExport function")

    # SQL query to select all data from the view
    sql_query = "SELECT * FROM financial_data_bankview02"

    # Execute the query
    with connection.cursor() as cursor:
        cursor.execute(sql_query)
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    # Convert the data to a pandas DataFrame
    df = pd.DataFrame(rows, columns=columns)

    # Convert the DataFrame to an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="my_view_data.xlsx"'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    return response

def get_original_amount(cash_allocation_id):
    try:
        cash_allocation = CashAllocation.objects.get(id=cash_allocation_id)
        policy = PolicyInformation.objects.get(id=cash_allocation.policy_fk_id)
        return policy.Installment_Amount_Syndicate_Share_in_Orig
    except Exception as e:
        logger.info(traceback.format_exc())
        logger.error(f"Error in get_original_amount: {str(e)}")
        return 0.0

class BankTransactionGetList(generics.ListAPIView):
    queryset = BankTransaction.objects.filter(archived=False).order_by("-id")
    serializer_class = BankTransactionSerializer
    pagination_class = MyPageNumberPaginations

    def get_queryset(self):
        queryset = super().get_queryset()
        query_params = self.request.query_params
        ff = BankTransaction.objects.filter(archived=False)
        with_related = BankTransaction.objects.select_related(
            "Assigned_User", "broker_information", "bank_details", "workflow"
        ).filter(archived=False)
        queryset = with_related.prefetch_related(
            "Assigned_Users", "document_files"
        ).filter(archived=False)
        # print(queryset,"dddddddddddddddddddd")
        pageSize = query_params.get("pageSize")
        skip = query_params.get("skip")
        if pageSize and skip:
            queryset = queryset.filter(Date_And_Time__range=[]).order_by("-id")
        return queryset


class BankReconciliationViewSet(viewsets.ModelViewSet):
    pagination_class = CustomPagination
    queryset = BankReconciliation.objects.all()
    serializer_class = BankReconciliationSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    @staticmethod
    def update_bank_reconciliation(receivable_amount, bank_charges, transaction_type, bank_details,
                                   receiving_bank_account,
                                   payment_receive_date, bank_reconciliation=None):
        """
        Calculates debit/credit amounts, total amount, and other reconciliation details based on provided arguments.

        Args:
            bank_reconciliation (obj): Primary Key of Bank Recon.
            bank_details (int): Primary Key of Bank Recon.
            receivable_amount (Decimal): The receivable amount (can be positive or negative).
            bank_charges (Decimal): The bank charges associated with the transaction.
            receiving_bank_account (str, optional): The receiving bank account number (if known). Defaults to None.
            transaction_type (str, optional): Create/Update.
            payment_receive_date (str, optional): The payment/receive date (if known). Defaults to None.

        Returns:
            dict: A dictionary containing the calculated values:
        """

        debit_amount = Decimal(0.0)
        credit_amount = Decimal(0.0)

        obj = {
        }

        try:
            bank_details_obj = BankDetails.objects.get(id=bank_details)

            if transaction_type == "create":
                if receivable_amount < 0:
                    debit_amount = -receivable_amount
                elif receivable_amount > 0:
                    credit_amount = receivable_amount

                total_amount = credit_amount - debit_amount
                ct_amount_car = total_amount

                obj.update({
                    'ct_amount': 0,
                    'ct_receivable_amount': 0,
                    'auto_created': 1,
                    'file_date': payment_receive_date,
                    'uploaded_date': payment_receive_date,
                    'uploaded_time': timezone.now(),
                    "debit_amount": debit_amount,
                    "credit_amount": credit_amount,
                    "total_amount": total_amount,
                    "ct_amount_car": ct_amount_car,
                    "bank_charges": bank_charges,
                    "bank_account_no": receiving_bank_account,
                    "bank_details": bank_details_obj,
                })

                bank_reconciliation_obj_details = BankReconciliation.objects.create(**obj)
                return bank_reconciliation_obj_details.id
            else:
                if receivable_amount < 0:
                    bank_reconciliation.debit_amount = bank_reconciliation.debit_amount - receivable_amount
                elif receivable_amount > 0:
                    bank_reconciliation.credit_amount = bank_reconciliation.credit_amount + receivable_amount
                bank_reconciliation.total_amount = bank_reconciliation.credit_amount - bank_reconciliation.debit_amount
                bank_reconciliation.ct_amount_car = bank_reconciliation.total_amount - bank_reconciliation.ct_amount

                bank_reconciliation.bank_charges = bank_reconciliation.bank_charges + bank_charges
                bank_reconciliation.bank_account_no = receiving_bank_account
                bank_reconciliation.bank_details = bank_details_obj
                bank_reconciliation.file_date = payment_receive_date
                bank_reconciliation.save()
                return bank_reconciliation.id
        except Exception as e:
            print("Exception e", str(e))

    @staticmethod
    def update_ct_amount_and_ct_amount_var(total_allocated_amount, bank_reconciliation_id):
        """
        Updates the CT amount and CT amount variance in the bank reconciliation table.

        Args:
            total_allocated_amount (Decimal): The new allocated amount (assumed to be numerical).
            bank_reconciliation_id (int): The ID of the bank reconciliation record.

        """

        try:
            bank_reconciliation = BankReconciliation.objects.get(pk=bank_reconciliation_id)
            bank_reconciliation.ct_amount = total_allocated_amount
            bank_reconciliation.ct_amount_car = bank_reconciliation.total_amount - bank_reconciliation.ct_amount
            bank_reconciliation.save()
        except BankReconciliation.DoesNotExist:
            raise ValueError("Bank reconciliation record with ID {} does not exist.".format(bank_reconciliation_id))
        except Exception as e:
            raise ValueError("Variable Data Type Mismatched Error: ".format(e))

    @staticmethod
    def update_ct_receivable_amount(total_allocated_amount, bank_reconciliation_id):
        bank_reconciliation = BankReconciliation.objects.get(id=bank_reconciliation_id)
        bank_reconciliation.ct_receivable_amount = total_allocated_amount
        bank_reconciliation.ct_receivable_amount_var = bank_reconciliation.total_amount - bank_reconciliation.ct_receivable_amount
        bank_reconciliation.save()

    def get_queryset(self):
        format = "%d/%m/%Y"
        objs = BankReconciliation.objects.all().order_by('-uploaded_time')

        # objs = objs.order_by('file_date')
        bank_name = self.request.GET.get("bankName", None)
        account_no = self.request.GET.get("accountNo", None)
        uploaded_status = self.request.GET.get("uploadedStatus", None)

        uploaded_statements = self.request.GET.get("uploaded_statements", None)

        to_date = self.request.GET.get("toDate", None)

        from_date = self.request.GET.get("fromDate", None)
        if from_date:
            from_date = datetime.strptime(from_date, format)
            formatted_from_date = from_date.strftime("%Y-%m-%d")
        else:
            formatted_from_date = None

        if to_date:
            to_date = datetime.strptime(to_date, format)
            formatted_to_date = to_date.strftime("%Y-%m-%d")
        else:
            formatted_to_date = None

        if bank_name:
            objs = objs.filter(bank_details__bank_name=bank_name)
        
        if account_no:
            objs = objs.filter(bank_account_no=account_no)

        if uploaded_status:
            objs = objs.filter(uploaded_status=uploaded_status)

        if formatted_from_date:
            objs = objs.filter(uploaded_date__gte=formatted_from_date)

        if formatted_to_date:
            objs = objs.filter(uploaded_date__lte=formatted_to_date)

        if uploaded_statements:
            if str(uploaded_statements) == '1':
                objs = objs.filter(~(Q(migrated_data=1) | Q(auto_created=1)))
        return objs

    def list(self, request):
        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        skip = page_number * rows_per_page

        queryset = self.filter_queryset(self.get_queryset())[
                   skip: skip + rows_per_page
                   ]

        serializer = self.get_serializer(queryset, many=True)
        all_data = serializer.data

        for data in all_data:
            if data['file_name_hyperlink']:
                bucket_key = data['file_name']
                data['file_name_hyperlink'] = bucket_key
        data = {"count": self.get_queryset().count(), "data": all_data}
        return Response(data)

    def retrieve(self, request, pk=None):
        if pk:
            doc = BankReconciliation.objects.get(id=pk)
            serializer = BankReconciliationSerializer(doc)
            dataa = serializer.data
            doc_files = dataa["file_name_hyperlink"]
            return Response(dataa)

    def create(self, request, *args, **kwargs):
        # If we're creating (POST) then we switch serializers to the one that doesn't include depth = 2
        serializer = BankReconciliation(data=request.data)

        if serializer.is_valid():
            serializer_saved = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.allocated_analyst_id = Users.objects.get(
            id=request.data["allocated_analyst_id"]
        )
        obj.bank_account_no = request.data["bank_account_no"]
        obj.file_name = request.data["file_name"]
        obj.uploaded_date = request.data["uploaded_date"]
        obj.uploaded_time = request.data["uploaded_time"]
        obj.file_date = request.data["file_date"]
        obj.credit_amount = request.data["credit_amount"]
        obj.debit_amount = request.data["debit_amount"]
        obj.total_amount = request.data["total_amount"]
        obj.ct_amount = request.data["ct_amount"]
        obj.ct_amount_car = request.data["ct_amount_car"]
        obj.bank_charges = request.data["bank_charges"]
        obj.ct_bank_charges = request.data["ct_bank_charges"]
        obj.ct_bank_charges_var = request.data["ct_bank_charges_var"]
        obj.category_total = request.data["category_total"]
        obj.error_message = request.data["error_message"]
        obj.locked = request.data["locked"]
        obj.allocated_date = request.data["allocated_date"]
        obj.final_status = request.data["final_status"]
        obj.analyst_comments = request.data["analyst_comments"]
        obj.ct_comments = request.data["ct_comments"]
        obj.resolution_date = request.data["resolution_date"]
        obj.file_name_hyperlink = request.data["file_name_hyperlink"]
        obj.save()
        serializer = BankReconciliationSerializer(obj)
        dataa = serializer.data
        doc_files = dataa["file_name_hyperlink"]
        return Response(dataa, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        doc_object = self.get_object()
        data = request.data
        doc_object.bank_account_no = data.get(
            "bank_account_no", doc_object.bank_account_no
        )
        doc_object.file_name = data.get("file_name", doc_object.file_name)
        doc_object.uploaded_date = data.get("upload_date", doc_object.uploaded_date)
        doc_object.uploaded_time = data.get("uploaded_time", doc_object.uploaded_time)
        doc_object.file_date = data.get("file_date", doc_object.file_date)
        doc_object.uploaded_status = data.get(
            "uploaded_status", doc_object.uploaded_status
        )
        doc_object.credit_amount = data.get("credit_amount", doc_object.credit_amount)
        doc_object.debit_amount = data.get("debit_amount", doc_object.debit_amount)
        doc_object.total_amount = data.get("total_amount", doc_object.total_amount)
        doc_object.ct_amount = data.get("ct_amount", doc_object.ct_amount)
        doc_object.ct_amount_car = data.get("ct_amount_car", doc_object.ct_amount_car)
        doc_object.bank_charges = data.get("bank_charges", doc_object.bank_charges)

        doc_object.ct_comments = data.get("ct_comments", doc_object.ct_comments)

        doc_object.ct_bank_charges = data.get(
            "ct_bank_charges", doc_object.ct_bank_charges
        )
        doc_object.ct_bank_charges_var = data.get(
            "ct_bank_charges_var", doc_object.ct_bank_charges_var
        )
        doc_object.category_total = data.get(
            "category_total", doc_object.category_total
        )
        doc_object.error_message = data.get("error_message", doc_object.error_message)
        doc_object.locked = data.get("locked", doc_object.locked)
        doc_object.allocated_date = data.get(
            "allocated_date", doc_object.allocated_date
        )
        doc_object.final_status = data.get("final_status", doc_object.final_status)
        doc_object.analyst_comments = data.get(
            "analyst_comments", doc_object.analyst_comments
        )
        doc_object.resolution_date = data.get(
            "resolution_date", doc_object.resolution_date
        )
        doc_object.allocated_analyst_id = (
            Users.objects.get(id=data.get("allocated_analyst_id"))
            if data.get("allocated_analyst_id")
            else doc_object.allocated_analyst_id
        )
        doc_object.file_name_hyperlink = data.get(
            "file_name_hyperlink", doc_object.file_name_hyperlink
        )
        doc_object.save()

        if data.get("allocated_analyst_id"):
            email_from = settings.EMAIL_HOST_USER
            email_to = Users.objects.get(id=data.get("allocated_analyst_id")).get_decrypted_email()
            recipient_list = [email_to]

            imagepath = config('MOSAIC_LOGO_IMAGE')

            subject = "New Rejected File Assigned to You"
            body = """
            <html>
                <head>
                </head>
                <body>
                    <div class="email-body">
                        <div class="email-header">
                            <img src="{imagepath}" alt="Mosaic Insurance Logo">
                        </div>
                        <p>Dear User,</p>
                        <p>A new rejected file {file_name} has been assigned to you. Please review it as soon as possible.</p>
                        <p>If you have any questions, email us at <a href="mailto:support@mosaicinsurance.com">support@mosaicinsurance.com</a>.</p>
                        <p>If you no longer wish to receive these email notifications, you can unsubscribe by replying to this email with "Unsubscribe" in the subject line.</p>

                        <div class="email-footer">
                            <p>Regards,<br>Mosaic Insurance</p>
                        </div>
                        <p>This message, including any attachments, may include proprietary or confidential material. Any distribution or use of this communication by anyone other than the intended recipient(s) is prohibited. If you are not the intended recipient, please notify the sender by replying to this message and then deleting it from your system.</p>
                    </div>
                </body>
            </html>
            """.format(imagepath=imagepath, file_name=doc_object.file_name)
            try:
                send_email(sender_email=email_from, recipient_email=recipient_list, subject=subject, body=body)
            except:
                pass

        serializer = BankReconciliationSerializer(doc_object)
        dataa = serializer.data
        return Response(dataa, status=status.HTTP_201_CREATED)


class BankReconciliationAccountNoViewSet(viewsets.ModelViewSet):
    pagination_class = CustomPagination
    queryset = BankReconciliation.objects.all().order_by('-file_date')  # changed
    serializer_class = BankReconciliationSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        variance = self.request.GET.get("variance", None)

        objs = self.queryset.filter(
            ~Q(ct_amount_car=0) & ~Q(uploaded_status='rejected')) if variance == "1" else self.queryset.filter(
            ~Q(uploaded_status='rejected'))

        format = "%d/%m/%Y"
        account_no = self.request.GET.get("accountNo", None)
        bank_name = self.request.GET.get("bankName", None)
        to_date = self.request.GET.get("toDate", None)
        from_date = self.request.GET.get("fromDate", None)

        entity_division = self.request.GET.get("entity_division", None)
        currency_code = self.request.GET.get("currency_code", None)

        if from_date:
            from_date = datetime.strptime(from_date, format)
            formatted_from_date = from_date.strftime("%Y-%m-%d")
        else:
            formatted_from_date = None

        if to_date:
            to_date = datetime.strptime(to_date, format)
            formatted_to_date = to_date.strftime("%Y-%m-%d")
        else:
            formatted_to_date = None

        if bank_name:
            objs = objs.filter(bank_details__bank_name=bank_name)

        if account_no:
            objs = objs.filter(bank_account_no=account_no)

        if formatted_from_date:
            objs = objs.filter(file_date__gte=formatted_from_date)

        if formatted_to_date:
            objs = objs.filter(file_date__lte=formatted_to_date)

        if entity_division:
            objs = objs.filter(bank_details__entity_number=entity_division)

        if currency_code:
            objs = objs.filter(bank_details__currency=currency_code)       
        return objs

    def list(self, request):
        format = "%d/%m/%Y"
        account_no = self.request.GET.get("accountNo", None)
        bank_name = self.request.GET.get("bankName", None)
        to_date = self.request.GET.get("toDate", None)
        from_date = self.request.GET.get("fromDate", None)
        entity_division = self.request.GET.get("entity_division", None)
        currency_code = self.request.GET.get("currency_code", None)
        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        skip = page_number * rows_per_page
        
        with connection.cursor() as cursor:
            query = """SELECT * FROM public.financial_bankrecon_app WHERE 1=1"""
            if account_no:
                query += f" AND \"Receiving_Bank_Account\" = '{account_no}'"
            if bank_name:
                query += f" AND \"bank_name\" = '{bank_name}'"
            if from_date:
                from_date = datetime.strptime(from_date, format)
                formatted_from_date = from_date.strftime("%Y-%m-%d")
                query += f" AND \"Payment_Receive_Date\" >= '{formatted_from_date}'"
            if to_date:
                to_date = datetime.strptime(to_date, format)
                formatted_to_date = to_date.strftime("%Y-%m-%d")
                query += f" AND \"Payment_Receive_Date\" <= '{formatted_to_date}'"
            if entity_division:
                query += f" AND \"entity_number\" = '{entity_division}'"
            if currency_code:
                query += f" AND \"currency\" = '{currency_code}'"
                       
            cursor.execute(query)
            results = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            dataa = [dict(zip(columns, row)) for row in results]

            summary_data = {
            "total_credit_amount" : 0,
            "total_debit_amount" : 0,
            "total_ca_rec_amount" : 0,
            "total_variance" : 0
            }

            for data in dataa: 
                summary_data["total_credit_amount"] += data["total_credit_amount"] or 0
                summary_data["total_debit_amount"] += data["total_debit_amount"] or 0
                summary_data["total_ca_rec_amount"] += data["ca rec amt"] or 0
                summary_data["total_variance"] += data["variance"] or 0
        # queryset = self.filter_queryset(self.get_queryset())[
        #            skip: skip + rows_per_page
        #            ]

        # # Apply variance condition here

        # serializer = self.get_serializer(queryset, many=True)
        # dataa = serializer.data
        data = {"count": len(dataa), "data": dataa[skip: skip + rows_per_page], "summary": summary_data}
        return Response(data)


class CashAllocationAllocatedTransaction(generics.GenericAPIView):
    """
    This class to do Locked/Unlocked for allocated transaction
    """

    queryset = CashAllocation.objects.filter(archived=False)
    serializer_class = CashAllocationAllocatedTransaction  # Update if needed for serialization
    lookup_field = ''

    def get_object(self):
        accounting_monthyear = self.request.data.get('allocation_datetime')
        if accounting_monthyear is not None:
            self.kwargs['accounting_monthyear'] = accounting_monthyear
        return super(CashAllocationAllocatedTransaction, self).get_object()

    def patch(self, request, *args, **kwargs):
        accounting_monthyear = request.data.get('allocation_datetime')
        locked = request.data.get('locked')

        start_date_str, end_date_str = accounting_monthyear.split('_')
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")

        if accounting_monthyear is None or locked is None:
            return Response({'error': 'Missing required fields: allocation_datetime, locked'},
                            status=status.HTTP_400_BAD_REQUEST)

        user = request.headers.get('user-id', None)

        if user and Users.objects.filter(id=user).exists():
            num_updated = lock_or_unlock_cash_allocation(locked, start_date, end_date, Users.objects.get(id=user))
        else:
            return Response({'error': 'Authentication required for locking/unlocking'},
                            status=status.HTTP_401_UNAUTHORIZED)

        locked_or_unlocked = 'locked' if locked else 'unlocked'

        if num_updated > 0:
            return Response({'message': f'{num_updated} CashAllocation records {locked_or_unlocked} successfully'})
        else:
            return Response({'message': 'No CashAllocation records found for the provided criteria'})



class BankmanagementDashboard(generics.GenericAPIView):
    """ This class for patient dashboard view """

    def get(self, request):
        result_dict = {}

        try:
            date_range = self.request.query_params.get('date_range', None)
            DATE_FORMAT = '%Y-%m-%d'

            start_date_str = ""
            end_date_str = ""

            if not date_range:
                end_date_str = str(date.today())
                start_date_str = str(date.today() - timedelta(days=7))
            else:
                start_date_str, end_date_str = date_range.split('_')[0], date_range.split('_')[1]

            request.data['start_date'] = datetime.strptime(start_date_str, DATE_FORMAT).date()
            request.data['end_date'] = datetime.strptime(end_date_str, DATE_FORMAT).date()

            DATE_F = '%d-%b-%Y'
            start_date = datetime.strptime(start_date_str, DATE_FORMAT).date().strftime(DATE_F)
            end_date = datetime.strptime(end_date_str, DATE_FORMAT).date().strftime(DATE_F)

            with connection.cursor() as cursor:
                query = """SELECT * FROM public.financial_data_bankview01 WHERE TO_DATE("Payment_Receive_Date", 'DD-Mon-YYYY') BETWEEN %s AND %s"""
                cursor.execute(query, [start_date, end_date])
                results = cursor.fetchall()
                columns = [col[0] for col in cursor.description]
                view_data = [dict(zip(columns, row)) for row in results]

                # Calculate open statement count
                unique_bank_transaction_ids = set(row["Bank Transaction ID"] for row in view_data if row["Broker Branch"] != "Non Premium")
                open_statement_count = len(unique_bank_transaction_ids)

                # Filter unique transactions based on "Bank Transaction ID"
                unique_transactions = {row["Bank Transaction ID"]: row for row in view_data if row["Broker Branch"] != "Non Premium"}.values()

                # Calculate open statement amount
                unique_bank_transaction_amounts = {}
                for row in view_data:
                    bank_transaction_id = row["Bank Transaction ID"]
                    amount = row["Bank Txn Amount"]
                    if bank_transaction_id and amount is not None:
                        if bank_transaction_id not in unique_bank_transaction_amounts:
                            unique_bank_transaction_amounts[bank_transaction_id] = 0
                        unique_bank_transaction_amounts[bank_transaction_id] += amount

                # open_statement_amount = sum(unique_bank_transaction_amounts.values())
                open_statement_amount = sum(row["Bank Txn Amount"] for row in view_data if row["Bank Txn Amount"] is not None and row["Broker Branch"] != "Non Premium" and row["serial_number"] == 1)

                # Calculate assigned transaction count
                assigned_transaction_count = len([row for row in unique_transactions if row["Account Handler"] != ""])

                # Calculate assigned transaction amount
                assigned_transaction_amount = sum(row["Bank Txn Amount"] for row in view_data if row["Account Handler"] != "" and row["Bank Txn Amount"] is not None and row["Broker Branch"] != "Non Premium" and row["serial_number"] == 1)

                # Calculate non assigned transaction count
                non_assigned_transaction_count = len([row for row in unique_transactions if row["Account Handler"] == ""])

                # Calculate non assigned transaction amount
                non_assigned_transaction_amount = sum(row["Bank Txn Amount"] for row in view_data if row["Account Handler"] == "" and row["Bank Txn Amount"] is not None and row["Broker Branch"] != "Non Premium" and row["serial_number"] == 1)

                # Calculate receivable settlement count
                receivable_settlement_count = len([row for row in view_data if row["Receivable Amount (USD)"] is not None and row["Broker Branch"] != "Non Premium"])

                # Calculate allocated transaction count
                allocated_transaction_count = len([row for row in view_data if row["allocation status"] == "Allocated" and row["Broker Branch"] != "Non Premium"])

                # Calculate allocated transaction amount
                allocated_transaction_amount = sum(row["Allocated Amount (USD)"] for row in view_data if row["Allocated Amount (USD)"] is not None and row["Broker Branch"] != "Non Premium")

                # Calculate unallocated transaction count
                unallocated_transaction_count = len([row for row in view_data if row["allocation status"] != "Allocated" and row["Broker Branch"] != "Non Premium"])

                # Calculate unallocated transaction amount
                unallocated_transaction_amount = sum(row["Remaining Balance (USD)"] for row in view_data if row["Remaining Balance (USD)"] is not None and row["Broker Branch"] != "Non Premium")

                # Receivable Trend
                receivable_settlement_totals = defaultdict(float)
                allocated_amount_totals = defaultdict(float)
                remaining_balance_totals = defaultdict(float)
                for row in view_data:
                    if row["Broker Branch"] != "Non Premium":
                        if row["Receivable Amount (USD)"]:
                            date_key = datetime.strptime(row["Payment_Receive_Date"], '%d-%b-%Y').strftime('%Y-%m-%d')
                            receivable_settlement_totals[date_key] += float(row["Receivable Amount (USD)"])
                        
                        if row["Allocated Amount (USD)"]:
                            date_key = datetime.strptime(row["Payment_Receive_Date"], '%d-%b-%Y').strftime('%Y-%m-%d')
                            allocated_amount_totals[date_key] += float(row["Allocated Amount (USD)"])
                        
                        if row["Remaining Balance (USD)"]:
                            date_key = datetime.strptime(row["Payment_Receive_Date"], '%d-%b-%Y').strftime('%Y-%m-%d')
                            remaining_balance_totals[date_key] += float(row["Remaining Balance (USD)"])

                receivable_settlement_amount = [{"Payment_Receive_Date": date, "total_receivable_amt": total} for date, total in receivable_settlement_totals.items()]
                allocated_amount = [{"allocation_date": date, "total_allocated_amt": total} for date, total in allocated_amount_totals.items()]
                remaining_balance = [{"date": date, "net_amount": total} for date, total in remaining_balance_totals.items()]

                # Progress Chart:
                progress_data = []
                for j in view_data:
                    if j["Broker Branch"] != "Non Premium":
                        handler = j.get("Account Handler", "Unknown")
                        status = j.get("allocation status", "Unknown")
                        found = False
                        for entry in progress_data:
                            if entry["user_name"] == handler and entry["allocation_status"] == status:
                                entry["status_count"] += 1
                                found = True
                                break
                        if not found:
                            progress_data.append({
                                "status_count": 1,
                                "user_name": handler,
                                "allocation_status": status
                            })

                # Target Goals:
                receivable_settlement_amounts = sum(row["Receivable Amount (USD)"] for row in view_data if row["Receivable Amount (USD)"] is not None and row["Broker Branch"] != "Non Premium")
                target_goals = {
                    'receivable_amt_sum': receivable_settlement_amounts,
                    'allocated_amt_sum': allocated_transaction_amount
                }

                # Unallocation Status:
                unallocation_statuswise_count = []
                unique_transactions = {row["Bank Transaction ID"]: row for row in view_data if row["Broker Branch"] != "Non Premium"}.values()
                status_groups = defaultdict(list)
                for row in unique_transactions:
                    if row["allocation status"] != "Allocated":
                        status_groups[row["allocation status"]].append(row)

                for status, transactions in status_groups.items():
                    status_count = len(transactions)
                    unallocated_amt = sum(row["Remaining Balance (USD)"] for row in transactions if row["Remaining Balance (USD)"] is not None)
                    unallocation_statuswise_count.append({
                        "allocation_status": status,
                        "unallocation_status_count": status_count,
                        "unallocated_amt": unallocated_amt
                    })

                result_dict.update({'unallocation_statuswise_count': unallocation_statuswise_count})
                allocation_status_count = allocated_transaction_count
                unallocation_statuswise_count = unallocation_statuswise_count
                total_unallocated_count = unallocated_transaction_count
                total_count = open_statement_count
                res = {
                    'allocation_status_count':allocation_status_count, 
                    'unallocation_statuswise_count':unallocation_statuswise_count, 
                    'total_unallocated_count':total_unallocated_count, 
                    'total_count':total_count,
                    'unallocated_transaction_amount': unallocated_transaction_amount,
                    'receivable_amt_sum': receivable_settlement_amounts
                }

                # Calculate SLA
                sla_met = len([
                    row for row in unique_transactions 
                    if row["Allocation Date"] and row["Payment_Receive_Date"] 
                    and (datetime.strptime(row["Allocation Date"], "%d-%b-%Y") - datetime.strptime(row["Payment_Receive_Date"], "%d-%b-%Y")).days <= 3
                ])

                sla_not_met = len([
                    row for row in unique_transactions 
                    if row["Allocation Date"] and row["Payment_Receive_Date"] 
                    and (datetime.strptime(row["Allocation Date"], "%d-%b-%Y") - datetime.strptime(row["Payment_Receive_Date"], "%d-%b-%Y")).days > 3
                ])

                sla = {
                    "sla_met": sla_met, 
                    "sla_not_met": sla_not_met
                }

                # Employee Performance Table
                employee_performance = {}
                for row in unique_transactions:
                    handler = row.get("Account Handler", "Unknown")
                    if handler not in employee_performance:
                        employee_performance[handler] = {
                            "user_name": handler,
                            "employee_bank_txn_count": 0,
                            "employee_bank_txn_amount": 0,
                            "employee_receivable_count": 0,
                            "employee_receivable_amount": 0,
                            "employee_allocated_count": 0,
                            "employee_allocated_amount": 0,
                            "employee_unallocated_count": 0,
                            "employee_unallocated_amount": 0
                        }


                # Sum amounts from view_data
                for row in view_data:
                    if row["Broker Branch"] != "Non Premium":
                        handler = row.get("Account Handler", "Unknown")
                        if handler in employee_performance:
                            if row["Receivable Amount (USD)"] is not None:
                                employee_performance[handler]["employee_receivable_count"] += 1
                                employee_performance[handler]["employee_receivable_amount"] += row["Receivable Amount (USD)"]

                            if row["allocation status"] == "Allocated":
                                employee_performance[handler]["employee_allocated_count"] += 1
                                if row["Allocated Amount (USD)"] is not None:
                                    employee_performance[handler]["employee_allocated_amount"] += row["Allocated Amount (USD)"]

                            if row["allocation status"] != "Allocated":
                                employee_performance[handler]["employee_unallocated_count"] += 1
                                if row["Remaining Balance (USD)"] is not None:
                                    employee_performance[handler]["employee_unallocated_amount"] += row["Remaining Balance (USD)"]

                            if row["Bank Txn Amount"] is not None and row["serial_number"] == 1:
                                employee_performance[handler]["employee_bank_txn_count"] += 1
                                employee_performance[handler]["employee_bank_txn_amount"] += row["Bank Txn Amount"]

                employee_performance_list = list(employee_performance.values())

            result_dict.update({'open_statement_count': open_statement_count})
            result_dict.update({'open_statement_amount': {'credit_amount__sum': open_statement_amount}})
            result_dict.update({'assigned_transaction_count': assigned_transaction_count})
            result_dict.update({'assigned_transaction_amount': {'receivable_amt': assigned_transaction_amount}})
            result_dict.update({'non_assigned_transaction_count': non_assigned_transaction_count})
            result_dict.update({'non_assigned_transaction_amount': {'receivable_amt': non_assigned_transaction_amount}})
            result_dict.update({'receivable_settlement_count': receivable_settlement_count})
            result_dict.update({'receivable_settlement_amounts': {'receivable_amt': receivable_settlement_amounts}})
            result_dict.update({'allocated_transaction_count': allocated_transaction_count})
            result_dict.update({'allocated_transaction_amount': {'receivable_amt': allocated_transaction_amount}})
            result_dict.update({'unallocated_transaction_count': unallocated_transaction_count})
            result_dict.update({'unallocated_transaction_amount': {'receivable_amt': unallocated_transaction_amount}})
            result_dict.update({'datewise_total_amount': {'datewise_receivable_amount': receivable_settlement_amount, 'datewise_allocated_amount': allocated_amount, 'datewise_unallocated_amount': remaining_balance}})
            result_dict.update({'target_goals': target_goals})
            result_dict.update({'employee_status_wise_count': progress_data})
            result_dict.update({'allocation_status_count': res})
            result_dict.update({'sla': sla})
            result_dict.update({'employee_allocation_data': employee_performance_list})

            result_dict.update({'user_count': DashboardOperations.get_user_count(start_date=request.data['start_date'],
                                                                                 end_date=request.data['end_date'])})
            result_dict.update({'total_categorywise_amt': DashboardOperations.get_total_amount(
                start_date=request.data['start_date'], end_date=request.data['end_date'])})
            result_dict.update({'credit_debit_amount': DashboardOperations.get_credit_debit_amount(
                start_date=request.data['start_date'], end_date=request.data['end_date'])})
            result_dict.update({'allocated_data_amt': DashboardOperations.get_allocated_data_amount(
                start_date=request.data['start_date'], end_date=request.data['end_date'])})

        except Exception as e:
            print("Exception e: ", str(e))

        return Response(data=result_dict)


class BankBalanceDashboard(generics.GenericAPIView):
    def get(self, request):
        date_range = request.query_params.get('date_range', None)
        start_date = date_range.split('_')[0]  # Extract start date
        end_date = date_range.split('_')[1]  # Extract end date
        bank_account = request.query_params.get('bank_account')
        bank_name = request.query_params.get('bank_name')

        filter_params = {'Payment_Receive_Date__range': [start_date, end_date]}
        if bank_account:
            filter_params['Receiving_Bank_Account'] = bank_account
        if bank_name:
            filter_params['bank_details__bank_name'] = bank_name

        # credit_debit_total_amount = BankTransaction.objects.filter(**filter_params).select_related(
        #     "bank_reconciliation", "bank_details").values(
        #     "Payment_Receive_Date", "Receiving_Bank_Account", bank_name=F("bank_details__bank_name")
        # ).annotate(
        #     credit_total_amount=Sum("bank_reconciliation__credit_amount"),
        #     debit_total_amount=Sum("bank_reconciliation__debit_amount")
        # )

        credit_debit_total_amount = BankTransaction.objects.filter(**filter_params, archived=False).select_related(
            "bank_reconciliation", "bank_details").values(
            "Payment_Receive_Date", "Receiving_Bank_Account", bank_name=F("bank_details__bank_name")
        ).annotate(
            credit_total_amount=Sum(
                Case(
                    When(Receivable_Amount__gt=0, then=F('Receivable_Amount')),
                    default=Value(0),
                    output_field=models.DecimalField()
                )
            ),
            debit_total_amount=Sum(
                Case(
                    When(Receivable_Amount__lt=0, then=-F('Receivable_Amount')),
                    default=Value(0),
                    output_field=models.DecimalField()
                )
            )
        )

        data = {
            "credit_debit_total_amount": credit_debit_total_amount
        }
        return Response(data=data)


class BankNames(generics.GenericAPIView):
    def get(self, request):
        date_range = request.query_params.get('date_range', None)
        start_date = date_range.split('_')[0]
        end_date = date_range.split('_')[1]
        bank_names = BankTransaction.objects.filter(archived=False, Payment_Receive_Date__range=[start_date, end_date]).values_list('bank_details__bank_name', flat=True).distinct()
        return Response(data={"bank_names": bank_names})


class BankAccountBasedOnBankName(generics.GenericAPIView):
    def get(self, request):
        bank_name = request.query_params.get('bank_name', None)
        bank_accounts = BankDetails.objects.filter(bank_name=bank_name).values_list('account_number', flat=True).distinct()
        return Response(data={"bank_accounts": bank_accounts})


class CashAllocationLockedUnlockedHistoryList(generics.ListAPIView, generics.RetrieveUpdateAPIView):
    pagination_class = CustomPagination
    serializer_class = CashAllocationLockedUnlockedHistorySerializer

    def get_queryset(self):
        """
        Limit queryset to objects the user has access to (e.g., based on permissions)
        """
        return CashAllocationLockedUnlockedHistory.objects.all().order_by('-id')

    def update(self, request, *args, **kwargs):
        cash_allocation_id = self.request.data.get('id')
        comment = self.request.data.get('comment')

        try:
            cash_allocation_history = CashAllocationLockedUnlockedHistory.objects.get(id=cash_allocation_id)
            start_date = cash_allocation_history.date_range.split('_')[0]  # Extract start date
            end_date = cash_allocation_history.date_range.split('_')[1]  # Extract end date

            cash_allocations_to_unlock = CashAllocation.objects.filter(
                locked=True,
                accounting_monthyear__gte=start_date,  # Use allocation_date for filtering
                accounting_monthyear__lte=end_date, archived=False
            )

            for cash_allocation in cash_allocations_to_unlock:
                # Update CashAllocation data
                cash_allocation.locked = False
                cash_allocation.save(update_fields=['locked'])

            num_unlocked_records = cash_allocations_to_unlock.count()

            commented_by = request.headers.get('user-id', None)  # Assuming user is authenticated

            cash_allocation_history.comment = comment
            cash_allocation_history.commented_by = Users.objects.get(id=commented_by) if commented_by else None
            cash_allocation_history.locked = False
            cash_allocation_history.unlocked = True
            cash_allocation_history.save(update_fields=['locked', 'comment', 'locked_unlocked_by', 'unlocked'])

            message = f'{num_unlocked_records} Record Unlocked Successfully.'

            return Response({'message': message}, status=status.HTTP_200_OK)

        except CashAllocationLockedUnlockedHistory.DoesNotExist:
            return Response({'message': 'CashAllocationLockedUnlockedHistory not found.'},
                            status=status.HTTP_404_NOT_FOUND)

        except serializers.ValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)


class CashTrackerReportForCashAllocation(generics.ListAPIView, generics.RetrieveUpdateAPIView):
    queryset = CashTrackerReport.objects.all()
    serializer_class = CashTrackerReportSerializer

    def list(self, request):
        cash_allocation_id = request.query_params.get('cash_allocation_id')

        try:
            if not cash_allocation_id:
                return Response(status=status.HTTP_400_BAD_REQUEST,
                                data={'error': 'Missing cash_allocation_id in query parameters'})

            cash_tracker_report = CashTrackerReport.objects.get(cash_allocation=cash_allocation_id)
            return Response(self.serializer_class(cash_tracker_report).data)
        except CashTrackerReport.DoesNotExist:
            raise NotFound('Cash Tracker Report not found for cash allocation ID {}'.format(cash_allocation_id))

    def partial_update(self, request):
        cash_allocation_id = request.query_params.get('cash_allocation_id')

        try:
            if not cash_allocation_id:
                return Response(status=status.HTTP_400_BAD_REQUEST,
                                data={'error': 'Missing cash_allocation_id in query parameters'})

            cash_tracker_report = CashTrackerReport.objects.get(cash_allocation=cash_allocation_id)

            # Parse the request data
            patched_data = JSONParser().parse(request)

            # Update the cash tracker report with the provided data
            serializer = CashTrackerReportSerializer(cash_tracker_report, data=patched_data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            return Response(serializer.data)
        except CashTrackerReport.DoesNotExist:
            raise NotFound('Cash Tracker Report not found for cash allocation ID {}'.format(cash_allocation_id))
        except Exception as e:
            return Response(status=status.HTTP_400_BAD_REQUEST, data={'error': str(e)})


def cash_allocation_activities_helper(id):
    data = dict()

    # Allocation Id from Request
    cash_allocation_id = CashAllocation.objects.get(id=int(id))
    
    # Check If CFI Activities Exists
    cash_allocation_cfi = CashAllocationCFI.objects.filter(cash_allocation=cash_allocation_id)
    if cash_allocation_cfi.exists():
        data["CFI"] = CashAllocationCFISerializer(cash_allocation_cfi, many=True).data
        wf_id = WorkFlow.objects.get(workflow_name='WF_CFI')
        for record in data["CFI"]:
            try:
                txn_id = WorkflowBankTransactions.objects.filter(bank_txn_id=record['id'],
                                                                            workflow=wf_id.id).first().id
                workflow_status = WorkflowBankTransactions.objects.get(id=txn_id).workflow_status
                record['workflow_status'] = workflow_status
            except Exception as e:
                record['workflow_status'] = None

    # Check If Write off Activities Exists
    cash_allocation_writeoff = CashAllocationWriteoff.objects.filter(cash_allocation=cash_allocation_id)
    if cash_allocation_writeoff.exists():
        data["Write off"] = CashAllocationWriteoffSerializer(cash_allocation_writeoff, many=True).data
        wf_id = WorkFlow.objects.get(workflow_name='WF_WRITEOFF')
        for record in data["Write off"]:
            try:
                txn_id = WorkflowBankTransactions.objects.filter(bank_txn_id=record['id'],
                                                                            workflow=wf_id.id).first().id
                workflow_status = WorkflowBankTransactions.objects.get(id=txn_id).workflow_status
                record['workflow_status'] = workflow_status
            except Exception as e:
                record['workflow_status'] = None

    # Check If Refund Activities Exists
    cash_allocation_refund = CashAllocationRefund.objects.filter(cash_allocation=cash_allocation_id)
    if cash_allocation_refund.exists():
        data["Refund"] = CashAllocationRefundSerializer(cash_allocation_refund, many=True).data
        wf_id = WorkFlow.objects.get(workflow_name='WF_REFUND')
        for record in data["Refund"]:
            try:
                txn_id = WorkflowBankTransactions.objects.filter(bank_txn_id=record['id'],
                                                                            workflow=wf_id.id).first().id
                workflow_status = WorkflowBankTransactions.objects.get(id=txn_id).workflow_status
                record['workflow_status'] = workflow_status
            except Exception as e:
                record['workflow_status'] = None

    # Check If Corrective Transfer Activities Exists
    cash_allocation_corrective = CashAllocationCorrective.objects.filter(cash_allocation=cash_allocation_id)
    if cash_allocation_corrective.exists():
        data["Corrective Transfer"] = CashAllocationCorrectiveSerializer(cash_allocation_corrective, many=True).data
        wf_id = WorkFlow.objects.get(workflow_name='WF_CORRECTIVE_TRANSFER')
        for record in data["Corrective Transfer"]:
            try:
                txn_id = WorkflowBankTransactions.objects.filter(bank_txn_id=record['id'],
                                                                            workflow=wf_id.id).first().id
                workflow_status = WorkflowBankTransactions.objects.get(id=txn_id).workflow_status
                record['workflow_status'] = workflow_status
            except Exception as e:
                record['workflow_status'] = None

    # Check If Correction type Activities Exists
    cash_allocation_correction = CashAllocationIssues.objects.filter(cash_allocation=cash_allocation_id)
    if cash_allocation_correction.exists():
        data["Correction type"] = CashAllocationIssuesSerializer(cash_allocation_correction, many=True).data
        wf_id = WorkFlow.objects.get(workflow_name='WF_CORRECTION_TYPES')
        for record in data["Correction type"]:
            try:
                txn_id = WorkflowBankTransactions.objects.filter(bank_txn_id=record['id'],
                                                                            workflow=wf_id.id).first().id
                workflow_status = WorkflowBankTransactions.objects.get(id=txn_id).workflow_status
                record['workflow_status'] = workflow_status
            except Exception as e:
                record['workflow_status'] = None
    
    # Check If Cross Allocation Activities Exists
    cash_allocation_issues = CrossAllocation.objects.filter(cash_allocation=cash_allocation_id)
    if cash_allocation_issues.exists():
        data["Cross Allocation"] = CrossAllocationSerializer(cash_allocation_issues, many=True).data
        wf_id = WorkFlow.objects.get(workflow_name='WF_CROSS_ALLOCATION')
        for record in data["Cross Allocation"]:
            try:
                txn_id = WorkflowBankTransactions.objects.filter(bank_txn_id=record['id'],
                                                                            workflow=wf_id.id).first().id
                workflow_status = WorkflowBankTransactions.objects.get(id=txn_id).workflow_status
                record['workflow_status'] = workflow_status
            except Exception as e:
                record['workflow_status'] = None

    return data


class CheckCashAllocationActivities(APIView):
    def get(self, request, id):
        try:
            data = cash_allocation_activities_helper(id)
            if data:
                return Response({"message": "Not Success", "data": data}, status=status.HTTP_200_OK)
            else:
                return Response({"message": "Success"}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class WorkstepUserUpdateAPIView(APIView):
    def patch(self, request, pk):
        try:
            user_id = request.data.get("user_id", None)
            if pk and user_id:
                workstep = WorkStep.objects.get(id=pk)
                workstep.user.add(user_id)
                workstep.save()
                return Response({"message":"WorkStep user updated successfully"})
            else:
                return Response({"Error": "Insufficient data"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"message": f"Failed to add workStep user: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk, *args, **kwargs):
        try:
            user_id = request.query_params.get("user_id", None)
            if pk and user_id:
                workstep = WorkStep.objects.get(id=pk)
                user = Users.objects.get(id=user_id)
                workstep.user.remove(user)
                return Response({"message":"WorkStep user removed successfully"})
            else:
                return Response({"Error": "Insufficient data"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"message": f"Failed to remove workStep user: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

class AccountingMonthEndPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100

class AccountingMonthEndViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Accounting Month End dates.
    Handles CRUD operations with validations for:
    - Date format (YYYY-MM-DD)
    - UK time-based changes
    - No date overlaps allowed
    - No gaps between dates allowed
    - No past date modifications
    """
    pagination_class = AccountingMonthEndPagination
    queryset = AccountingMonthEnd.objects.all().order_by('-accounting_month_date')
    serializer_class = AccountingMonthEndSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        """
        Filters queryset based on from_date and to_date parameters.
        """
        try:
            queryset = AccountingMonthEnd.objects.all().order_by('-id')
    
            from_date = self.request.query_params.get('from_date')
            to_date = self.request.query_params.get('to_date')
    
            if from_date and to_date:
                queryset = queryset.filter(
                    accounting_month_date__range=[from_date, to_date]
                )
            return queryset
    
        except Exception as e:
            logger.error(f"Error in get_queryset: {str(e)}")
            return AccountingMonthEnd.objects.none()
    
    def list(self, request):
        """
        List accounting month end records with pagination.
        """
        try:
            queryset = self.filter_queryset(self.get_queryset())
            result = []
            
            for item in queryset:
                data = {
                    'id': item.id,
                    'accounting_month_date': item.accounting_month_date,
                    'accounting_month_start_date': item.accounting_month_start_date,
                    'accounting_month_end_date': item.accounting_month_end_date,
                    'created_at': item.created_at,
                    'updated_at': item.updated_at,
                    'created_by': {
                        'id': item.created_by.id if item.created_by else None,
                        'user_name': item.created_by.user_name if item.created_by else None,
                    },
                    'updated_by': {
                        'id': item.updated_by.id if item.updated_by else None,
                        'user_name': item.updated_by.user_name if item.updated_by else None,
                    }
                }
                result.append(data)
                
            page = self.paginate_queryset(result)
            return self.get_paginated_response(page)
            
        except Exception as e:
            logger.error(f"Failed to list accounting month end: {str(e)}")
            return Response(
                {"message": [str(e)]}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    def create(self, request, *args, **kwargs):
        try:
            user = get_user(request)
            if not user:
                return Response(
                    {"message": "User authentication required"}, 
                    status=status.HTTP_401_UNAUTHORIZED
                )

            # Prepare the user details to be returned
            user_details = {
                "id": user.id,
                "user_name": user.user_name,  
                "email": user.email
            }

            request.data['created_by'] = user.id
            request.data['updated_by'] = user.id
    
            # Check if accounting_month_date exists
            accounting_month_date = request.data.get('accounting_month_date')
            dt = datetime.strptime(accounting_month_date, '%Y-%m-%d').date()
            month = dt.month
            year = dt.year
            if accounting_month_date:
                existing = AccountingMonthEnd.objects.filter(
                    accounting_month_date__month=month,
                    accounting_month_date__year=year
                ).exists()
                if existing:
                    return Response(
                        {"message": [f"Record already exists for {dt.strftime('%b-%Y')}"]}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
    
            serializer = self.get_serializer(data=request.data)
            if not serializer.is_valid():
                error_message = serializer.errors.get('message', 
                    next(iter(serializer.errors.values()))[0]
                )
                return Response(
                    {"message": error_message}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
    
            self.perform_create(serializer)

            # Add the user details to the response data
            response_data = serializer.data
            response_data["created_by"] = user_details
            response_data["updated_by"] = user_details

            return Response(response_data, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"Failed to create accounting month end: {str(e)}")
            return Response(
                {"message": [str(e)]}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def update(self, request, *args, **kwargs):
        try:
            user = get_user(request)
            if not user:
                return Response(
                    {"message": ["User authentication required"]}, 
                    status=status.HTTP_401_UNAUTHORIZED
                )

            # Prepare the user details to be returned
            user_details = {
                "id": user.id,
                "user_name": user.user_name,  
                "email": user.email
            }

            request.data['updated_by'] = user.id
            instance = self.get_object()
    
            # Check if accounting_month_date exists
            accounting_month_date = request.data.get('accounting_month_date')
            if accounting_month_date:
                existing = AccountingMonthEnd.objects.filter(
                    accounting_month_date=accounting_month_date
                ).exclude(id=instance.id).exists()
                if existing:
                    return Response(
                        {"message": [f"Record already exists for {accounting_month_date}"]}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
    
            serializer = self.get_serializer(
                instance,
                data=request.data,
                partial=True
            )
            
            if not serializer.is_valid():
                error_message = serializer.errors.get('message',
                    next(iter(serializer.errors.values()))[0]
                )
                return Response(
                    {"message": error_message}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
    
            self.perform_update(serializer)

            # Add the user details to the response data
            response_data = serializer.data
            response_data["created_by"] = user_details  # Include created_by details as well
            response_data["updated_by"] = user_details  # Include updated_by details

            return Response(response_data)

        except Exception as e:
            logger.error(f"Failed to update accounting month end: {str(e)}")
            return Response(
                {"message": [str(e)]}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    def destroy(self, request, *args, **kwargs):
        """
        Deletes accounting month end record.
        
        Args:
            request: HTTP request object
            
        Returns:
            Response: Success message on deletion
            
        Raises:
            400: If deletion fails
        """
        try:
            instance = self.get_object()
            instance.delete()
            return Response({"message": "Accounting month deleted successfully"},
                            status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Failed to delete accounting month: {str(e)}")
            return Response({"message": [str(e)]}, status=status.HTTP_400_BAD_REQUEST)


class ChaserViewset(viewsets.ModelViewSet):
    pagination_class = AccountingMonthEndPagination
    queryset = FollowUp.objects.filter(archived=False).order_by('-id')
    serializer_class = ChaserSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def list(self, request):
        try:
            user = get_user(request)
            filter_params = {}

            # Role-based filter
            if user.role == "CC Processor":
                filter_params['bank_transaction__Assigned_User'] = user

            # Toggle filter
            toggle_value = request.GET.get("toggle_value")
            if toggle_value == "Allocated":
                filter_params['cash_allocation__allocation_status__icontains'] = "Allocated"
            else:
                filter_params['cash_allocation__allocation_status__icontains'] = "query"

            # Other filters
            if policy_no := request.GET.get("policyNumber"):
                filter_params['cash_allocation__policy_id__icontains'] = policy_no

            if txn_id := request.GET.get("transactionId"):
                filter_params['bank_transaction__Bank_Transaction_Id__icontains'] = txn_id

            if broker_branch := request.GET.get("brokerBranch"):
                filter_params['bank_transaction__Broker_Branch'] = broker_branch

            if handler_name := request.GET.get("handlerName"):
                filter_params['bank_transaction__Assigned_User__user_name'] = handler_name

            # Date range filter
            q_object = None
            from_date = request.GET.get("fromAllocationDate")
            to_date = request.GET.get("toAllocationDate")
            if from_date and to_date:
                from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
                to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
                q_object = (
                    Q(date1_value__range=(from_date, to_date)) |
                    Q(date2_value__range=(from_date, to_date)) |
                    Q(date3_value__range=(from_date, to_date)) |
                    Q(escalation_date_value__range=(from_date, to_date))
                )

            # Pagination
            page_number = int(request.GET.get("skip", 0))
            rows_per_page = int(request.GET.get("pageSize", 20))
            offset = page_number * rows_per_page

            # Base queryset
            queryset = (
                self.queryset
                .filter(**filter_params)
                .select_related(
                    'bank_transaction',
                    'cash_allocation',
                    'bank_transaction__Assigned_User'
                )
            )
            if q_object is not None:
                queryset = queryset.filter(q_object)

            # Count and aggregate in DB (faster than Python loop)
            total = queryset.count()
            premium_due_amount = (
                queryset.aggregate(total=Sum('cash_allocation__receivable_amt'))['total'] or 0
            )
            premium_due_amount = round(float(premium_due_amount), 2)

            # Use annotate to avoid Python-side counting
            summary_counts = queryset.aggregate(
                in_date2=Count('id', filter=Q(date2_value__isnull=False)),
                in_date3=Count('id', filter=Q(date3_value__isnull=False)),
                in_escalation=Count('id', filter=Q(escalation_date_value__isnull=False)),
            )

            # Paginate at DB level
            paginated_queryset = queryset[offset: offset + rows_per_page]

            serializer_data = self.serializer_class(paginated_queryset, many=True).data

            # Post-filter "Over Due" in Python
            if toggle_value == "Over Due":
                serializer_data = [j for j in serializer_data if j['indicator'] == 'Red']

            summary = {
                "query_status": total,
                "premium_due_amount": premium_due_amount,
                **summary_counts
            }

            return Response({"count": total, "data": serializer_data, "summary": summary})

        except Exception as e:
            logger.error(f"Failed to list chaser data: {str(e)}")
            return Response(
                {"message": [str(e)]},
                status=status.HTTP_400_BAD_REQUEST
            )

    def partial_update(self, request, *args, **kwargs):
        data = request.data
        followup_id = data.get('id')
        
        if not followup_id:
            return Response({"message": "FollowUp ID is required"}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            followup = FollowUp.objects.get(id=followup_id)
        except FollowUp.DoesNotExist:
            return Response({"message": "FollowUp not found"}, status=status.HTTP_404_NOT_FOUND)
        
        try:
            audit_obj = None
            # Save follow up audit
            for key, value in dict(data).items():
                if key in ['date2_value', 'date3_value', 'escalation_date_value', 'comments']:
                    audit_obj = FollowUpAudit.objects.create(
                        follow_up=followup,
                        audit_data={
                            "field_name": key,
                            "old_value": str(getattr(followup, key, "")),
                            "new_value": value,
                            "previous_edit_datetime": followup.updated_at.strftime("%Y-%m-%d %H:%M:%S") if followup.updated_at else "-",
                            "current_edit_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "changed_by": get_user(request).pk,
                            "event_type": "edit"
                        }
                    )

            # Convert date strings to date objects
            if data.get('date2_value'):
                data['date2_value'] = datetime.strptime(data['date2_value'], "%d-%m-%Y").date()
            if data.get('date3_value'):
                data['date3_value'] = datetime.strptime(data['date3_value'], "%d-%m-%Y").date()
            if data.get('escalation_date_value'):
                data['escalation_date_value'] = datetime.strptime(data['escalation_date_value'], "%d-%m-%Y").date()
                
            # Update the FollowUp instance
            serializer = FollowUpSerializer(followup, data=data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            
            return Response(serializer.data, status=status.HTTP_200_OK)
        except:
            if audit_obj:
                audit_obj.delete()
            return Response({'error': 'Unexpected error occurs!'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def safe_decimal_conversion(amount_str):
    """
    Safely convert string amount with currency symbols and commas to Decimal.
    
    Args:
        amount_str: String containing amount (e.g., "£172,598.28", "172,598.28", "172598.28")
    
    Returns:
        Decimal: Converted decimal value
        
    Raises:
        ValidationError: If the string cannot be converted to a valid decimal
    """
    if amount_str is None:
        return Decimal('0')
    
    # Convert to string if it's not already
    amount_str = str(amount_str).strip()
    
    if not amount_str or amount_str.lower() in ['nan', 'nat', 'none', '']:
        return Decimal('0')
    
    try:
        # Remove currency symbols (£, $, €, etc.) and other non-numeric characters except . and -
        import re
        cleaned_amount = re.sub(r'[^\d.,-]', '', amount_str)
        
        # Handle negative amounts
        is_negative = cleaned_amount.startswith('-')
        if is_negative:
            cleaned_amount = cleaned_amount[1:]
        
        # Remove commas
        cleaned_amount = cleaned_amount.replace(',', '')
        
        # Convert to Decimal
        decimal_value = Decimal(cleaned_amount)
        
        # Apply negative sign if needed
        if is_negative:
            decimal_value = -decimal_value
            
        return decimal_value
        
    except (ValueError, TypeError, Exception) as e:
        raise ValidationError(f"'{amount_str}' value must be a decimal number.")

@api_view(["GET"])
def downloadAccountingMonthEnd(request):
    if request.method == "GET":
        """
        Download accounting month end as xlsx file
        """
        try:
            queryset = AccountingMonthEnd.objects.all().order_by("-created_at")

            # Convert queryset to list of dictionaries
            accounting_month_end_data = []
            for record in queryset:
                accounting_month_end_data.append(
                    {
                        "id": record.id,
                        "Accounting Month Date": record.accounting_month_date,
                        "Accounting Month Start Date": record.accounting_month_start_date,
                        "Accounting Month End Date": record.accounting_month_end_date,
                        "Created By": record.created_by,
                        "Created At": record.created_at.strftime("%d-%m-%Y %H:%M:%S")
                        if record.created_at else "",
                        "Updated By": record.updated_by,
                        "Updated At": record.updated_at.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updated_at else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(accounting_month_end_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Accounting Month End")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Accounting Month End"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=accounting_month_end_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AssignPolicyHandler(generics.ListAPIView, generics.UpdateAPIView):
    pagination_class = AccountingMonthEndPagination
    queryset = CashAllocation.objects.filter(~Q(allocation_status="Allocated"), archived=False).order_by('-id')
    serializer_class = AssignPolicyHandlerSerializer

    def list(self, request):
        try:
            format = "%d/%m/%Y"
            filter_params = {}
            user = get_user(request)
            if user.role == "CC Processor":
                filter_params['policy_handler'] = user
            page_number = int(request.GET.get("skip", 0))
            rows_per_page = int(request.GET.get("pageSize", 20))
            skip = page_number * rows_per_page

            txn_id = request.GET.get("transactionId", None)
            if txn_id:
                filter_params['bank_txn__Bank_Transaction_Id__icontains'] = txn_id

            bank_handler_name = request.GET.get("bank_handler_name", None)
            if bank_handler_name:
                filter_params['bank_txn__Assigned_User__id'] = bank_handler_name

            policy_handler_name = request.GET.get("policy_handler_name", None)
            if policy_handler_name:
                filter_params['policy_handler__id'] = policy_handler_name

            to_date = self.request.GET.get("toDate", None)
            from_date = self.request.GET.get("fromDate", None)
            if from_date:
                from_date = datetime.strptime(from_date, format)
                formatted_from_date = from_date.strftime("%Y-%m-%d")
            else:
                formatted_from_date = None

            if to_date:
                to_date = datetime.strptime(to_date, format)
                formatted_to_date = to_date.strftime("%Y-%m-%d")
            else:
                formatted_to_date = None

            if formatted_from_date:
                filter_params['policy_assign_date__gte']=formatted_from_date

            if formatted_to_date:
                filter_params['policy_assign_date__lte']=formatted_to_date

            queryset = self.queryset.filter(**filter_params)
            total = queryset.count()
            page_queryset = queryset[skip: skip + rows_per_page]
            serializer = self.serializer_class(page_queryset, many=True)

            bank_assign_users = (
                BankTransaction.objects
                .filter(~Q(Assigned_User=None))
                .distinct()
                .annotate(user_id=F('Assigned_User__id'), username=F('Assigned_User__user_name'))
                .values('user_id', 'username')
            )
            ca_assign_users = (
                CashAllocation.objects
                .filter(~Q(policy_handler=None))
                .distinct()
                .annotate(user_id=F('policy_handler__id'), username=F('policy_handler__user_name'))
                .values('user_id', 'username')
            )

            data = {
                "count": total, 
                "data": serializer.data,
                "bank_assign_users": bank_assign_users,
                "ca_assign_users": ca_assign_users
            }
            return Response(data)

        except Exception as e:
            logger.error(f"Failed to list Assign Policy Handler data: {str(e)}")
            return Response(
                {"message": [str(e)]}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    def partial_update(self, request):
        try:
            data = request.data
            Ids = data.get('Ids')
            user_id = data.get('user_login_id')
            user_id = Users.objects.get(id=user_id)
            for i in Ids:
                obj = CashAllocation.objects.get(id=i['id'])

                try:
                    # Update audit data for policy handler
                    CashAllocaionAudit.objects.create(
                        cash_allocation=obj,
                        audit_data={
                            "field_name": "Policy Handler",
                            "old_value": obj.policy_handler.user_name if obj.policy_handler else "-",
                            "new_value": user_id.user_name if user_id else "-",
                            "previous_edit_datetime": obj.updated_at.strftime("%Y-%m-%d %H:%M:%S") if obj.updated_at else "-",
                            "current_edit_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "changed_by": get_user(request).pk,
                            "event_type": "edit"
                        }
                    )
                except Exception as e:
                    logger.error(f"Failed to update CashAllocation Audit for Assign Policy Handler: {str(e)}")

                obj.policy_handler=user_id
                obj.policy_assign_date=datetime.now().date()
                obj.save()
            return Response(status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Failed to update Assign Policy Handler: {str(e)}")
            return Response(
                {"message": [str(e)]}, 
                status=status.HTTP_400_BAD_REQUEST
            )
