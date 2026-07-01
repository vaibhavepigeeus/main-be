from django.shortcuts import render

# Create your views here.

from django.db.models import Q
from django.shortcuts import render
from rest_framework.decorators import api_view, permission_classes, action
from .serializers import *
from rest_framework.response import Response
from rest_framework import generics, viewsets
from rest_framework.views import APIView

from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.generics import ListCreateAPIView, ListAPIView
from .models import *
from django.conf import settings
from django.core.mail import send_mail
from django.utils.crypto import get_random_string
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse, HttpResponse

from rest_framework_swagger import renderers
from rest_framework.decorators import api_view, renderer_classes
from django.views.decorators.csrf import csrf_exempt
from .serializers import *
from cmtbackend.storage_backends import create_presigned_url
from datetime import datetime as dt
import json
from rest_framework import status
from decouple import config
ENVIRONMENT = config('ENVIRONMENT')
LEGACY_DOCUMENTS_BUCKET = settings.LEGACY_DOCUMENTS_BUCKET
LEGACY_DOCUMENTS_S3_PREFIX = f"https://{LEGACY_DOCUMENTS_BUCKET}.s3.amazonaws.com/"
from django.db import transaction
from django.db.models import Max
from .tasks import update_policy_calculations
import re
from django.utils import timezone
import logging
logger = logging.getLogger('bankmanagement')
import os
from filemanagement.views import reusable_file_upload
from knox.auth import TokenAuthentication
import pandas as pd
import math, re
from rest_framework.pagination import PageNumberPagination
from openpyxl.styles import PatternFill
from datetime import date
import time
from users.models import Users
from bankmanagement.models import CashAllocationIssues
from bankmanagement.serializers import CashAllocationIssuesSerializer
from documents.utils.encryption_util import decrypt_text
from io import BytesIO
from openpyxl import load_workbook
import ast
from django.db import connection
from django.db.models import Max
from django.db.models import Sum

class CountModelMixin(object):
    @action(detail=False)
    def count(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        content = {"count": queryset.count()}
        return Response(content)


class DocumentsViewSet(viewsets.ModelViewSet, CountModelMixin):
    model = Documents
    serializer_class = DocumentsSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        doc = Documents.objects.all()
        return doc

    def list(self, request):
        doc = Documents.objects.all()
        serializer = DocumentsSerializer(doc, many=True)
        dataa = serializer.data
        for i in dataa:
            print(i, "ggggggg")
            doc_files = i["document_file"]
            if doc_files:
                bucket_key = doc_files.replace(
                    LEGACY_DOCUMENTS_S3_PREFIX, ""
                )
                print("buket", bucket_key)
                i["document_file"] = create_presigned_url(LEGACY_DOCUMENTS_BUCKET, bucket_key)
        return Response(dataa)

    def retrieve(self, request, pk=None):
        if pk:
            doc = Documents.objects.get(id=pk)
            serializer = DocumentsSerializer(doc)
            dataa = serializer.data
            doc_files = dataa["document_file"]
            print(dataa, "fffffffffffffffffffffff")
            bucket_key = doc_files.replace(LEGACY_DOCUMENTS_S3_PREFIX, "")
            print("buket", bucket_key)
            dataa["document_file"] = create_presigned_url(LEGACY_DOCUMENTS_BUCKET, bucket_key)
            return Response(dataa)

    def create(self, request, *args, **kwargs):

        upload_date = datetime.now()
        data = request.data
        new_doc = Documents.objects.create(
            document_name=data["document_name"],
            document_date=data["document_date"],
            upload_date=upload_date,
            document_type=data["document_type"],
        )
        new_doc.save()
        doc_f = data["document_file"]
        print("dc", doc_f)
        if doc_f:
            print("hereee")
            new_doc.document_file = data["document_file"]
            new_doc.save()
        serializer = DocumentsSerializer(new_doc)
        dataa = serializer.data
        doc_files = dataa["document_file"]
        print(dataa, "fffffffffffffffffffffff")
        bucket_key = doc_files.replace(LEGACY_DOCUMENTS_S3_PREFIX, "")
        print("buket", bucket_key)
        dataa["document_file"] = create_presigned_url(LEGACY_DOCUMENTS_BUCKET, bucket_key)
        return Response(dataa)

    def update(self, request, *args, **kwargs):

        doc_object = self.get_object()
        print("doc_object", doc_object)
        data = request.data
        doc_object.document_name = data["document_name"]
        doc_object.document_date = data["document_date"]
        doc_object.upload_date = data["upload_date"]
        doc_object.document_file = data["document_file"]
        doc_object.document_type = data["document_type"]
        doc_object.save()
        serializer = DocumentsSerializer(doc_object)
        dataa = serializer.data
        doc_files = dataa["document_file"]
        print(dataa, "fffffffffffffffffffffff")
        bucket_key = doc_files.replace(LEGACY_DOCUMENTS_S3_PREFIX, "")
        print("buket", bucket_key)
        dataa["document_file"] = create_presigned_url(LEGACY_DOCUMENTS_BUCKET, bucket_key)
        return Response(dataa)

    def partial_update(self, request, *args, **kwargs):

        doc_object = self.get_object()
        data = request.data
        doc_object.document_name = data.get("document_name", doc_object.document_name)
        doc_object.document_date = data.get("document_date", doc_object.document_date)
        doc_object.upload_date = data.get("upload_date", doc_object.upload_date)
        doc_object.document_file = data.get("document_file", doc_object.document_file)
        doc_object.document_type = data.get("document_type", doc_object.document_type)
        doc_object.save()
        serializer = DocumentsSerializer(doc_object)
        dataa = serializer.data
        doc_files = dataa["document_file"]
        print(dataa, "fffffffffffffffffffffff")
        bucket_key = doc_files.replace(LEGACY_DOCUMENTS_S3_PREFIX, "")
        print("buket", bucket_key)
        dataa["document_file"] = create_presigned_url(LEGACY_DOCUMENTS_BUCKET, bucket_key)
        return Response(dataa)

    def destroy(self, request, *args, **kwargs):
        doc_object = self.get_object()
        doc_object.delete()
        return Response({"message": "documents deleted successfully"})


import pandas as pd
from .forms import *
from .models import *
from .serializers import BrokerInformationSerializer


@csrf_exempt
def broker_excel_import(request):
    if request.method == "POST":
        form = BrokerInformationImportForm(request.POST, request.FILES)
        if form.is_valid():
            dict_excel_data = []
            df = pd.read_excel(request.FILES["broker_excel_file"], skiprows=[0])
            excel_data = df.dropna(subset=["Broker Name"])
            for index, row in excel_data.iterrows():
                ddd = {}
                broker_name = row["Broker Name"]
                if not broker_name == "nan":
                    ddd["broker_name"] = broker_name
                    broker = row["Broker #"]
                    ddd["broker"] = broker
                    branch = row["Branch"]
                    ddd["branch"] = branch
                    duplicate_count = row["Duplicate Count"]
                    ddd["duplicate_count"] = duplicate_count
                    soa_received_from_broker = row["SOA received from Broker"]
                    ddd["soa_received_from_broker"] = soa_received_from_broker
                    name = row["Name"]
                    ddd["name"] = name
                    email = row["Email"]
                    ddd["email"] = email
                    secondary_email = row.get("Secondary Email", None)
                    ddd["secondary_email"] = secondary_email
                    phone_number = row["Phone"]
                    ddd["phone_number"] = phone_number
                    broker_branch_location = row["Broker Branch Location"]
                    ddd["broker_branch_location"] = broker_branch_location
                    dict_excel_data.append(ddd)
                    if not BrokerInformation.objects.filter(email=email).exists():
                        BrokerInformation.objects.create(
                            broker_name=broker_name,
                            broker=broker,
                            branch=branch,
                            duplicate_count=duplicate_count,
                            soa_received_from_broker=soa_received_from_broker,
                            name=name,
                            email=email,
                            secondary_email=secondary_email,
                            phone_number=phone_number,
                            broker_branch_location=broker_branch_location,
                        )
            return JsonResponse({"results": dict_excel_data})


class BankDetailsAPIView(APIView):
    def post(self, request):
        file = request.FILES["filedata"]
        df = pd.read_excel(file)
        rows = []
        for row in df.to_dict("records"):
            record = {
                "region": row.get("Region"),
                "entity_number": row.get("Entity N"),
                "msd_entity_number": row.get("MSD  Entity N"),
                "entity_name": row.get("Entity Name"),
                "bank_name": row.get("Bank Name"),
                "account_number": row.get("Bank Account"),
                "account_type": row.get("Bank Account Type"),
                "currency": row.get("Currency "),
                "msd_acct_number": row.get("MSD Acct N"),
                "msd_acct_name": row.get("MSD Acct Name"),
            }
            rows.append(record)

        serializer = BankDetailsSerializer(data=rows, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Data imported successfully!"})
        else:
            return Response(serializer.errors, status=400)

    def get(self, request, format=None):
        ac_no = self.request.query_params.get("ac_no")
        bank_detail_obj = BankDetails.objects.get(account_number=ac_no)
        serializer = BankDetailsSerializer(bank_detail_obj)
        return Response(serializer.data)


@csrf_exempt
def get_all_unique_bank_names(request):
    if request.method == "GET":
        bank_names = BankDetails.objects.values_list('bank_name', flat=True).distinct()
        bank_namess = {"bank_names": sorted(bank_names)}
        return JsonResponse(bank_namess)


@csrf_exempt
def get_all_unique_account_numbers_by_bank_name(request):
    if request.method == "GET":
        bank_name = request.GET.get("bank_name")
        account_numbers = BankDetails.objects.all().values_list('account_number', flat=True).distinct()
        if bank_name:
            account_numbers = BankDetails.objects.filter(bank_name=bank_name).values_list('account_number', flat=True).distinct()
        account_numberss = {"account_numbers": sorted(account_numbers)}
        return JsonResponse(account_numberss)


@csrf_exempt
def getAllBankNames(request):
    if request.method == "GET":
        dict = {}
        bank_name = []
        account_number = []
        bankDetailsObjects = BankDetails.objects.all().distinct("bank_name")
        for i in bankDetailsObjects:
            b_n = i.bank_name
            bank_name.append(b_n)
            ac_no = i.account_number
            account_number.append(ac_no)
        bank_namess = {"bank_names": sorted(bank_name)}
        dict.update(bank_namess)
        account_numberss = {"account_numbers": sorted(account_number)}
        dict.update(account_numberss)
        return JsonResponse(dict)


@csrf_exempt
def getDetailsByBankName(request):
    if request.method == "GET":
        bank_name = request.GET.get("bank_name")
        bankDetailsObjects = (
            BankDetails.objects.filter(bank_name=bank_name).order_by("id").first()
        )
        if bankDetailsObjects:
            serializer = BankDetailsSerializer(bankDetailsObjects)
            data = serializer.data
        else:
            data = {"message": "Data not Found"}
        return JsonResponse(data)

@csrf_exempt
def getBankDetailsForFilter(request):
    if request.method == "POST":
        try:
            # Load JSON data from request body
            json_data = json.loads(request.body.decode("utf-8"))

            # Ensure that some filters are provided in the request
            if not json_data:
                return JsonResponse({"message": "At least one filter is required."}, status=400)

            # Query the bank details
            bankDetailsObjects = BankDetails.objects.filter(**json_data).order_by("id")
            if bankDetailsObjects.exists():
                # Serialize data if objects are found
                bankDetails = bankDetailsObjects.all()
            else:
                # If no records are found, return all bank details
                bankDetails = BankDetails.objects.all()

            serializer = BankDetailsSerializer(bankDetails, many=True)
            return JsonResponse(serializer.data, status=200, safe=False)

        except json.JSONDecodeError:
            return JsonResponse({"message": "Invalid JSON format."}, status=400)
        except Exception as e:
            return JsonResponse({"message": f"An error occurred: {str(e)}"}, status=500)


@csrf_exempt
def getAllBrokerList(request):
    if request.method == "GET":
        dict = {}
        Broker_Name = []
        Broker_Branch_Name = []
        Brokerdata = BrokerInformation.objects.all()
        for i in Brokerdata:
            bro_n = i.broker_name
            if not bro_n in Broker_Name:
                Broker_Name.append(bro_n)
            bro_branch = i.branch
            if not bro_branch in Broker_Branch_Name:
                Broker_Branch_Name.append(bro_branch)
        broker_namess = {"Broker_Name": sorted(Broker_Name)}
        dict.update(broker_namess)
        broker_branchs = {"Broker_Branch_Name": sorted(Broker_Branch_Name)}
        dict.update(broker_branchs)
        return JsonResponse(dict)


@csrf_exempt
def currency_excel_import(request):
    if request.method == "POST":
        form = CurrencyInformationImportForm(request.POST, request.FILES)
        if form.is_valid():
            dict_excel_data = []
            df = pd.read_excel(request.FILES["currency_excel_file"])
            for index, row in df.iterrows():
                ddd = {}
                currency_and_country = row["Country and currency"]
                if not currency_and_country == "nan":
                    ddd["currency_and_country"] = currency_and_country
                    currency_code = row["Currency code"]
                    ddd["currency_code"] = currency_code
                    symbol = row["Symbol"]
                    ddd["symbol"] = symbol
                    dict_excel_data.append(ddd)
                    if not CurrencyDetails.objects.filter(
                        country_and_currency=currency_and_country
                    ).exists():
                        CurrencyDetails.objects.create(
                            country_and_currency=currency_and_country,
                            currency_code=currency_code,
                            symbol=symbol,
                        )
            return JsonResponse({"results": dict_excel_data})


class CurrencyViewSet(APIView):

    def get(self, request):
        data = CurrencyDetails.objects.all().order_by('-addedDateAndTime')
        currency_code = request.GET.get("currency_code", None)
        country_and_currency = request.GET.get("country", None)

        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        skip = page_number * rows_per_page
        filter_conditions = Q()
        if currency_code:
            filter_conditions &= Q(currency_code__icontains=currency_code)

        if country_and_currency:
            filter_conditions &= Q(country_and_currency__icontains=country_and_currency)

        filtered_data = data.filter(filter_conditions)
        count = data.filter(filter_conditions).count()
        serializer = CurrencyInformationSerializer(filtered_data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        data = request.data
        all_data = []
        failures = {}
        try:
            for j in range(len(data)):
                data[j]["updatedDateAndTime"] = datetime.datetime.now().isoformat()
                data[j]["addedDateAndTime"] = datetime.datetime.now().isoformat()
                serializer = CurrencyInformationSerializer(data=data[j])
                if serializer.is_valid():
                    new_doc=serializer.save()
                    data_item = [data[j]]
                    fields = json.dumps(data_item)
                    if len(fields) > 65535:
                        return Response(
                            {
                                "msg": "Length of the value for update_fields property got exceeded!"
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    new_doc.updated_fields = fields
                    new_doc.save()
                    all_data.append(serializer.data)
                else:
                    failures[str(j + 1)] = serializer.errors
            if failures:
                return Response(
                    {
                        "Message": "Data is not inserted for row number {}".format(
                            failures
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            return Response(all_data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        try:
            instance = CurrencyDetails.objects.get(id=pk)
            request.data["updatedDateAndTime"]= datetime.datetime.now().isoformat()
            serializer = CurrencyInformationSerializer(
                instance, data=request.data, partial=True
            )
            if serializer.is_valid():
                updated_doc=serializer.save()
                data_items=[]
                if instance.updated_fields:
                    old_list=instance.updated_fields
                    old_list=json.loads(old_list)
                    data_items.extend(old_list)
                data_items.append(request.data)
                changedFields = json.dumps(data_items)
                if len(changedFields)>65535:
                    return Response(
                        {
                            "msg": "Length of the value for update_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                else:
                    updated_doc.updated_fields = changedFields
                    updated_doc.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        curr_object = CurrencyDetails.objects.get(id=pk)
        curr_object.delete()
        return Response(
            {"message": "currecny deleted successfully"}, status=status.HTTP_200_OK
        )


@csrf_exempt
def get_all_currencies(request):
    if request.method == "GET":
        currencies = CurrencyDetails.objects.all().values_list('currency_code', flat=True).distinct()
        currencies_dict = {"currencies": sorted(currencies)}
        return JsonResponse(currencies_dict)

class BrokerNameInfoViewSet(ListAPIView):
    model = BrokerInformation
    serializer_class = BrokerInformationSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def post(self, request):
        data = request.data
        Broker_Name = data["Broker_Name"]
        brokerDetailsObjects = (
            BrokerInformation.objects.filter(broker_name=Broker_Name)
            .order_by("id")
            .first()
        )
        if brokerDetailsObjects:
            serializer = BrokerInformationSerializer(brokerDetailsObjects)
            data = serializer.data
        else:
            data = {"message": "Data not Found"}
        return Response(data)


class BrokerBranchNameInfoViewSet(ListAPIView):
    model = BrokerInformation
    serializer_class = BrokerInformationSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def post(self, request):
        data = request.data
        Broker_Branch_Name = data["Broker_Branch_Name"]
        print(Broker_Branch_Name, ".............br")
        brokerDetailsObjects = (
            BrokerInformation.objects.filter(branch=Broker_Branch_Name)
            .order_by("id")
            .first()
        )
        if brokerDetailsObjects:
            serializer = BrokerInformationSerializer(brokerDetailsObjects)
            data = serializer.data
        else:
            data = {"message": "Data not Found"}
        return Response(data)


class AllocationStatusViewSet(APIView):

    def get(self, request):
        data = AllocationStatus.objects.all().order_by('-addedDateAndTime')
        allocation_status = request.GET.get("allocation_status", None)
        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        skip = page_number * rows_per_page
        filter_conditions = Q()
        if allocation_status:
            search_filter = Q(allocation_status__icontains=allocation_status)
            filter_conditions &= search_filter
        filtered_data = data.filter(filter_conditions)
        count = data.filter(filter_conditions).count()
        serializer = AllocationStatusSerializer(filtered_data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        data = request.data
        all_data = []
        failures = {}
        try:
            for j in range(len(data)):
                data[j]["updatedDateAndTime"] = datetime.datetime.now().isoformat()
                data[j]["addedDateAndTime"] = datetime.datetime.now().isoformat()
                serializer = AllocationStatusSerializer(data=data[j])
                if serializer.is_valid():
                    new_doc=serializer.save()
                    data_item = [data[j]]
                    fields = json.dumps(data_item)
                    if len(fields) > 65535:
                        return Response(
                            {
                                "msg": "Length of the value for update_fields property got exceeded!"
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    new_doc.updated_fields = fields
                    new_doc.save()
                    all_data.append(serializer.data)
                else:
                    failures[str(j + 1)] = serializer.errors
            if failures:
                return Response(
                    {
                        "Message": "Data is not inserted for row number {}".format(
                            failures
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            return Response(all_data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        try:
            instance = AllocationStatus.objects.get(id=pk)
            request.data["updatedDateAndTime"] = datetime.datetime.now().isoformat()
            serializer = AllocationStatusSerializer(
                instance, data=request.data, partial=True
            )
            if serializer.is_valid():
                updated_doc=serializer.save()
                data_items=[]
                if instance.updated_fields:
                    old_list=instance.updated_fields
                    old_list=json.loads(old_list)
                    data_items.extend(old_list)
                data_items.append(request.data)
                changedFields = json.dumps(data_items)
                if len(changedFields)>65535:
                    return Response(
                        {"msg": "Length of the value for update_fields property got exceeded!"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                else:
                    updated_doc.updated_fields = changedFields
                    updated_doc.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class PolicyTypeViewSet(APIView):

    def get(self, request):
        data = PolicyType.objects.all().order_by('-addedDateAndTime')
        search_query = request.GET.get("search_query", None)
        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        skip = page_number * rows_per_page
        filter_conditions = Q()
        if search_query:
            search_filter = Q(policy_type__icontains=search_query) | Q(
                id__icontains=search_query
            )
            filter_conditions &= search_filter
        filtered_data = data.filter(filter_conditions)
        count = data.filter(filter_conditions).count()
        serializer = PolicyTypeSerializer(filtered_data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        data = request.data
        all_data = []
        failures = {}
        try:
            for j in range(len(data)):
                data[j]["updatedDateAndTime"] = datetime.datetime.now().isoformat()
                data[j]["addedDateAndTime"] = datetime.datetime.now().isoformat()
                serializer = PolicyTypeSerializer(data=data[j])
                if serializer.is_valid():
                    new_doc=serializer.save()
                    data_item = [data[j]]
                    fields = json.dumps(data_item)
                    if len(fields) > 65535:
                        return Response(
                            {
                                "msg": "Length of the value for update_fields property got exceeded!"
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    new_doc.updated_fields = fields
                    new_doc.save()
                    all_data.append(serializer.data)
                else:
                    failures[str(j + 1)] = serializer.errors
            if failures:
                return Response(
                    {
                        "Message": "Data is not inserted for row number {}".format(
                            failures
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            return Response(all_data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        try:
            instance = PolicyType.objects.get(id=pk)
            request.data["updatedDateAndTime"]= datetime.datetime.now().isoformat()
            serializer = PolicyTypeSerializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                updated_doc=serializer.save()
                data_items=[]
                if instance.updated_fields:
                    old_list=instance.updated_fields
                    old_list=json.loads(old_list)
                    data_items.extend(old_list)
                data_items.append(request.data)
                changedFields = json.dumps(data_items)
                if len(changedFields)>65535:
                    return Response(
                        {
                            "msg": "Length of the value for update_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                else:
                    updated_doc.updated_fields = changedFields
                    updated_doc.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class BrokerInfoViewSet(APIView):
    model = BrokerInformation
    serializer_class = BrokerInformationSerializer

    def get(self, request):
        data = BrokerInformation.objects.all().order_by('-addedDateAndTime')
        broker_name = request.GET.get("brokerName", None)
        broker_branch = request.GET.get("brokerBranch", None)
        broker_email = request.GET.get("brokerEmail", None)
        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        skip = page_number * rows_per_page
        filter_conditions = Q()
        if broker_name:
            search_filter = Q(broker_name__icontains=broker_name)
            filter_conditions &= search_filter

        if broker_branch:
            search_filter = Q(branch__icontains=broker_branch)
            filter_conditions &= search_filter

        initial_filtered_data = data.filter(filter_conditions)
        
        filtered_data=[]
        if broker_email:
            for record in initial_filtered_data:
                email = record.email
                secondary_email = record.secondary_email
                decrypted_email = decrypt_text(email)
                if broker_email.lower() in decrypted_email.lower() or (secondary_email and broker_email.lower() in secondary_email.lower()):
                    filtered_data.append(record)
        else:
            filtered_data = initial_filtered_data
        
        count = len(filtered_data)
        serializer = BrokerInformationSerializer(filtered_data[skip: skip + rows_per_page], many=True)
        for item in serializer.data:
            item.pop('email', None)
            item.pop('phone_number', None)
        return Response({ "data": serializer.data, "count": count }, status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        data = request.data
        all_data = []
        failures = {}
        try:
            for j in range(len(data)):
                dataa=[data[j].copy()]
                data[j]["updatedDateAndTime"] = datetime.datetime.now()
                data[j]["addedDateAndTime"] = datetime.datetime.now()
                serializer = BrokerInformationSerializer(data=data[j])
                if serializer.is_valid():
                    print("==========")
                    new_obj=serializer.save()
                    print("saved===>")
                    dataa[0]["addedDateAndTime"]= datetime.datetime.now().isoformat()
                    dataa[0]["updatedDateAndTime"]= datetime.datetime.now().isoformat()
                    changedFields=json.dumps(dataa)
                    new_obj.updated_fields = changedFields
                    new_obj.save()
                    print("new_obj--", new_obj)
                    updated_serializer = BrokerInformationSerializer(new_obj)
                    all_data.append(updated_serializer.data)
                else:
                    print("----------", serializer.errors)
                    failures[str(j + 1)] = serializer.errors
            if failures:
                return Response(
                    {
                        "Message": "Data is not inserted for row number {}".format(
                            failures
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            return Response(all_data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        try:
            instance = BrokerInformation.objects.get(id=pk)
            serializer = BrokerInformationSerializer(
                instance, data=request.data, partial=True
            )
            request.data["updatedDateAndTime"]= datetime.datetime.now().isoformat()
            if serializer.is_valid():
                updated_doc=serializer.save()
                data_items=[]
                if instance.updated_fields:
                    old_list=instance.updated_fields
                    old_list=json.loads(old_list)
                    data_items.extend(old_list)
                data_items.append(request.data)
                changedFields = json.dumps(data_items)
                if len(changedFields)>65535:
                    return Response(
                        {"msg": "Length of the value for update_fields property got exceeded!"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                else:
                    updated_doc.updated_fields = changedFields
                    updated_doc.save()
                updated_serializer=BrokerInformationSerializer(updated_doc)
                return Response(updated_serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class LOBViewSet(viewsets.ModelViewSet):
    model = LOB
    serializer_class = LOBSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        doc = LOB.objects.all()
        return doc

    def list(self, request):
        lob = LOB.objects.all().order_by('-addedDateAndTime')
        serializer = LOBSerializer(lob, many=True)
        dataa = serializer.data
        return Response(dataa)

    def retrieve(self, request, pk=None):
        if pk:
            doc = LOB.objects.get(id=pk)
            serializer = LOBSerializer(doc)
            dataa = serializer.data
            return Response(dataa)

    # def create(self, request, *args, **kwargs):
    #     try:
    #         data_list = request.data  # Expecting a list of dictionaries

    #         # Add addedDateAndTime to each data item
    #         for data in data_list:
    #             data["addedDateAndTime"] = dt.now().isoformat()
    #             data["updatedDateAndTime"] = dt.now().isoformat()

    #         serializer = LOBSerializer(data=data_list, many=True)

    #         if serializer.is_valid():
    #             new_docs = serializer.save()

    #             # Update updated_fields for each new document
    #             for new_doc, data in zip(new_docs, data_list):
    #                 data_items = [data]
    #                 fields = json.dumps(data_items)
    #                 if len(fields) > 65535:
    #                     return Response(
    #                         {
    #                             "msg": "Length of the value for update_fields property got exceeded!"
    #                         },
    #                         status=status.HTTP_400_BAD_REQUEST,
    #                     )
    #                 new_doc.updated_fields = fields
    #                 new_doc.save()

    #             return Response(serializer.data, status=status.HTTP_201_CREATED)
    #         else:
    #             return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    #     except Exception as e:
    #         return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def create(self, request, *args, **kwargs):
        try:
            dataa = request.data
            res_data = []
            failures = {}
            for j, data in enumerate(dataa):
                if not LOB.objects.filter(
                    lob_code = data["lob_code"]
                ).exists():
                    new_doc = LOB.objects.create(
                        lob_code=data["lob_code"],
                        lob=data["lob"],
                        created_by=data["created_by"],
                        addedDateAndTime=dt.now(),
                    )
                    addedDateAndTime = dt.now()
                    data["addedDateAndTime"] = str(addedDateAndTime)
                    data["updatedDateAndTime"] = str(addedDateAndTime)
                    data_items = []
                    data_items.append(data)
                    fields = json.dumps(data_items)
                    if len(fields) > 65535:
                        return Response(
                            {
                                "msg": "Length of the value for update_fields property got exceeded!"
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    else:
                        new_doc.updated_fields = fields
                    new_doc.save()
                    serializer = LOBSerializer(new_doc)
                    res_data.append(serializer.data)
                else:
                    failures[str(j + 1)] = "This lob code already exists"

            if failures:
                return Response(
                    {
                        "Message": "Data is not inserted for row number {}".format(failures),
                        "Duplicates": "There are duplicates in rows: {}".format(",".join(failures.keys()))
                    },
                        status=status.HTTP_400_BAD_REQUEST,
                )

            return Response(res_data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        doc_object = self.get_object()
        doc_object.delete()
        return Response({"message": "lob deleted successfully"})

    def partial_update(self, request, *args, **kwargs):
        try:
            instance = (
                self.get_object()
            )  # Assume get_object method is defined elsewhere to get the instance
            data = request.data
            lob_code = data.get("lob_code", None)
            if lob_code:
                if LOB.objects.filter(
                        ~Q(id=instance.id),
                        lob_code = data["lob_code"]
                    ).exists():
                    return Response(
                        {
                            "Message": "This lob code already exists",
                            "Duplicates": "This lob already exists"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            # Set updated_by and updatedDateAndTime fields
            serializer = LOBSerializer(instance, data=data, partial=True)

            if serializer.is_valid():
                updated_instance = serializer.save()

                # Update updated_fields
                updated_data_items = serializer.validated_data
                updated_data_items["updatedDateAndTime"] = datetime.datetime.now().isoformat()
                data_list=[]
                if instance.updated_fields:
                    old_list=instance.updated_fields
                    old_list=json.loads(old_list)
                    data_list.extend(old_list)
                data_list.append(updated_data_items)
                updated_fields=json.dumps(data_list)
                if len(updated_fields) > 65535:
                    return Response(
                        {
                            "msg": "Length of the value for updated_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                updated_instance.updated_fields = updated_fields
                updated_instance.save()

                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class SCMPartnersViewSet(viewsets.ModelViewSet):
    queryset = SCMPartners.objects.all()
    serializer_class = SCMPartnersSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def create(self, request, *args, **kwargs):
        try:
            dataa = request.data
            added_date_and_time = dt.now()
            res_data=[]
            for data in dataa:
                new_doc = SCMPartners.objects.create(
                    partner_name=data.get("partner_name"),
                    created_by=data.get("created_by"),
                    addedDateAndTime=added_date_and_time,
                )

                # Prepare data for updated_fields
                data["addedDateAndTime"] = added_date_and_time.isoformat()
                data["updatedDateAndTime"] = added_date_and_time.isoformat()
                data_items = [data]
                fields = json.dumps(data_items)

                # Check the length of the serialized data
                if len(fields) > 65535:
                    return Response(
                        {
                            "msg": "Length of the value for update_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                else:
                    new_doc.updated_fields = fields

                # Save the new document
                new_doc.save()
                serializer = SCMPartnersSerializer(new_doc)
                res_data.append(serializer.data)

            return Response(res_data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, *args, **kwargs):
        try:
            doc_object = self.get_object()
            data = request.data
            doc_object.partner_name = data.get("partner_name", doc_object.partner_name)
            doc_object.updated_by = data.get("updated_by", doc_object.updated_by)
            doc_object.updatedDateAndTime = dt.now()
            doc_object.save()

            updatedatetime = dt.now()
            data["updatedDateAndTime"] = str(updatedatetime)
            data_items = []
            if doc_object.updated_fields:
                old_list = doc_object.updated_fields
                old_list = json.loads(old_list)
                data_items.extend(old_list)
            data_items.append(data)
            changedFields = json.dumps(data_items)
            if len(changedFields) > 65535:
                return Response(
                    {"msg": "Length of the value for update_fields property got exceeded!"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            else:
                doc_object.updated_fields = changedFields

            doc_object.save()
            serializer = SCMPartnersSerializer(doc_object)
            dataa = serializer.data
            return Response(dataa, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class BindingAgreementViewSet(viewsets.ModelViewSet):
    queryset = BindingAgreement.objects.all()
    serializer_class = BindingAgreementSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def create(self, request, *args, **kwargs):
        try:
            data_list = request.data  # Expecting a list of dictionaries
            new_docs = []
            failures = {}

            for j, data in enumerate(data_list):
                added_date_and_time = dt.now()

                if not BindingAgreement.objects.filter(binding_aggrement_type__iexact=data.get("binding_aggrement_type")).exists():
                    new_doc = BindingAgreement.objects.create(
                        binding_aggrement_type=data.get("binding_aggrement_type"),
                        created_by=data.get("created_by"),
                        addedDateAndTime=added_date_and_time,
                    )

                    # Prepare data for updated_fields
                    data["addedDateAndTime"] = added_date_and_time.isoformat()
                    data["updatedDateAndTime"] = added_date_and_time.isoformat()

                    data_items = [data]
                    fields = json.dumps(data_items)

                    # Check the length of the serialized data
                    if len(fields) > 65535:
                        return Response(
                            {
                                "msg": "Length of the value for update_fields property got exceeded!"
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    else:
                        new_doc.updated_fields = fields

                    # Save the new document
                    new_doc.save()
                    serializer = BindingAgreementSerializer(new_doc)
                    new_docs.append(serializer.data)

                else:
                    failures[str(j + 1)] = "This entity division already exists"

            if failures:
                return Response(
                    {
                        "Message": "Data is not inserted for row number {}".format(failures),
                        "Duplicates": "There are duplicates in rows: {}".format(",".join(failures.keys()))
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            return Response(new_docs, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, *args, **kwargs):
        try:
            doc_object = self.get_object()
            data = request.data

            binding_aggrement_type= data.get("binding_aggrement_type", None)
            if binding_aggrement_type:
                if BindingAgreement.objects.filter( ~Q(id=doc_object.id), binding_aggrement_type__iexact = data["binding_aggrement_type"]).exists():
                    return Response(
                        {
                            "Message": "This entity division already exists",
                            "Duplicates": "This entity division already exists"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            doc_object.binding_aggrement_type = data.get(
                "binding_aggrement_type", doc_object.binding_aggrement_type
            )
            doc_object.updated_by = data.get("updated_by", doc_object.updated_by)
            doc_object.updatedDateAndTime = dt.now().isoformat()
            doc_object.save()

            updatedatetime = dt.now().isoformat()
            data["updatedDateAndTime"] = str(updatedatetime)
            data_items = []
            if doc_object.updated_fields:
                old_list = doc_object.updated_fields
                old_list = json.loads(old_list)
                data_items.extend(old_list)
            data_items.append(data)
            changedFields = json.dumps(data_items)
            if len(changedFields) > 65535:
                return Response(
                    {"msg": "Length of the value for update_fields property got exceeded!"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            else:
                doc_object.updated_fields = changedFields

            doc_object.save()
            serializer = BindingAgreementSerializer(doc_object)
            dataa = serializer.data
            return Response(dataa, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CorrectionTypeViewSet(APIView):

    def get(self, request, pk=None):
        if pk is not None:
            try:
                obj = CashAllocationIssues.objects.get(pk=pk)
                serializer = CashAllocationIssuesSerializer(obj)
                return Response(serializer.data, status=status.HTTP_200_OK)
            except CashAllocationIssues.DoesNotExist:
                return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        else:
            archived = request.GET.get("archived", None)
            print('archived',archived)
            if archived == "true":
                # Include archived (for history)
                data = CorrectionType.objects.all().order_by('-addedDateAndTime')
            else:
                # archived=false OR not provided → only active (dropdown default)
                data = CorrectionType.objects.filter(
                    archived=False
                ).order_by('-addedDateAndTime')
                
            correction_type = request.GET.get("correction_type", None)
            allocation_status = request.GET.get("allocation_status", None)
            page_number = int(request.GET.get("skip", 0))
            rows_per_page = int(request.GET.get("pageSize", 20))
            skip = page_number * rows_per_page
            filter_conditions = Q()
            if correction_type:
                filter_conditions &= Q(correction_type__icontains=correction_type)
            if allocation_status:
                filter_conditions &= Q(allocation_status__icontains=allocation_status)
            filtered_data = data.filter(filter_conditions)
            count = data.filter(filter_conditions).count()
            serializer = CorrectionTypeSerializer(filtered_data, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        data = request.data
        all_data = []
        failures = {}
        try:
            for j in range(len(data)):
                data[j]["updatedDateAndTime"] = datetime.datetime.now().isoformat()
                data[j]["addedDateAndTime"] = datetime.datetime.now().isoformat()
                serializer = CorrectionTypeSerializer(data=data[j])
                if serializer.is_valid():
                    new_doc=serializer.save()
                    data_item = [data[j]]
                    fields = json.dumps(data_item)
                    if len(fields) > 65535:
                        return Response(
                            {
                                "msg": "Length of the value for update_fields property got exceeded!"
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    new_doc.updated_fields = fields
                    new_doc.save()
                    all_data.append(serializer.data)
                else:
                    failures[str(j + 1)] = serializer.errors
            if failures:
                return Response(
                    {
                        "Message": "Data is not inserted for row number {}".format(
                            failures
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            return Response(all_data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        try:
            instance = CorrectionType.objects.get(id=pk, archived=False)
            request.data["updatedDateAndTime"]= datetime.datetime.now().isoformat()
            serializer = CorrectionTypeSerializer(instance, data=request.data, partial=True)

            if serializer.is_valid():
                updated_doc=serializer.save()
                data_items=[]
                if instance.updated_fields:
                    old_list=instance.updated_fields
                    old_list=json.loads(old_list)
                    data_items.extend(old_list)
                data_items.append(request.data)
                changedFields = json.dumps(data_items)
                if len(changedFields)>65535:
                    return Response(
                        {"msg": "Length of the value for update_fields property got exceeded!"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                else:
                    updated_doc.updated_fields = changedFields
                    updated_doc.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class TransactionCategoryViewSet(APIView):

    def get(self, request):
        data = TransactionCategory.objects.all().order_by('-id')
        serializer = TransactionCategorySerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        data = request.data
        max_id = TransactionCategory.objects.aggregate(Max('id'))['id__max'] or 0
        data['id'] = max_id + 1
        serializer = TransactionCategorySerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def patch(self, request, pk):
        instance = TransactionCategory.objects.get(id=pk)
        serializer = TransactionCategorySerializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)


class CashTransferViewSet(viewsets.ModelViewSet):
    queryset = CashTransfer.objects.all()
    serializer_class = CashTransferSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def create(self, request, *args, **kwargs):
        try:
            dataa = request.data
            res_data=[]
            for data in dataa:
                new_doc = CashTransfer.objects.create(
                    cash_transfer_value=data["cash_transfer_value"],
                    created_by=data["created_by"],
                    addedDateAndTime=dt.now(),
                )
                addedDateAndTime = dt.now()
                data["addedDateAndTime"] = str(addedDateAndTime)
                data["updatedDateAndTime"] = str(addedDateAndTime)
                data_items = []
                data_items.append(data)
                fields = json.dumps(data_items)
                if len(fields) > 65535:
                    return Response(
                        {
                            "msg": "Length of the value for update_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                else:
                    new_doc.updated_fields = fields
                new_doc.save()
                serializer = CashTransferSerializer(new_doc)
                res_data.append(serializer.data)
            return Response(res_data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, *args, **kwargs):
        try:
            doc_object = self.get_object()
            data = request.data
            doc_object.cash_transfer_value = data.get(
                "cash_transfer_value", doc_object.cash_transfer_value
            )
            doc_object.updated_by = data.get("updated_by", doc_object.updated_by)
            doc_object.updatedDateAndTime = dt.now()
            doc_object.save()

            updatedatetime = dt.now()
            data["updatedDateAndTime"] = str(updatedatetime)
            data_items = []
            if doc_object.updated_fields:
                old_list = doc_object.updated_fields
                old_list = json.loads(old_list)
                data_items.extend(old_list)
            data_items.append(data)
            changedFields = json.dumps(data_items)
            if len(changedFields) > 65535:
                return Response(
                    {"msg": "Length of the value for update_fields property got exceeded!"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            else:
                doc_object.updated_fields = changedFields

            doc_object.save()
            serializer = CashTransferSerializer(doc_object)
            dataa = serializer.data
            return Response(dataa, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class EntityViewSet(viewsets.ModelViewSet):
    model = Entity
    serializer_class = EntitySerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        en = Entity.objects.all()
        return en

    def list(self, request):
        en = Entity.objects.all().order_by('-addedDateAndTime')
        serializer = EntitySerializer(en, many=True)
        dataa = serializer.data
        return Response(dataa)

    def retrieve(self, request, pk=None):
        if pk:
            en = Entity.objects.get(id=pk)
            serializer = EntitySerializer(en)
            dataa = serializer.data
            return Response(dataa)

    # def create(self, request, *args, **kwargs):
    #     try:
    #         data = request.data
    #         new_doc = Entity.objects.create(entity_divisions=data["entity_divisions"], created_by=data["created_by"], addedDateAndTime=datetime.now())

    #         addedDateAndTime = datetime.now()
    #         data["addedDateAndTime"] = str(addedDateAndTime)
    #         data_items = []
    #         data_items.append(data)
    #         fields = json.dumps(data_items)
    #         if len(fields) > 65535:
    #             return Response({"msg": "Length of the value for update_fields property got exceeded!"})
    #         else:
    #             new_doc.updated_fields = fields
    #         new_doc.save()
    #         serializer = EntitySerializer(new_doc)
    #         dataa = serializer.data
    #         return Response(dataa, status=status.HTTP_201_CREATED)
    #     except Exception as e:
    #         return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    def create(self, request, *args, **kwargs):
        try:
            dataa = request.data
            res_data = []
            failures = {}
            for j, data in enumerate(dataa):
                if not Entity.objects.filter(
                    entity_divisions = data["entity_divisions"]
                ).exists():
                    new_doc = Entity.objects.create(
                        entity_name=data["entity_name"],
                        entity_divisions=data["entity_divisions"],
                        created_by=data["created_by"],
                        addedDateAndTime=dt.now(),
                    )
                    addedDateAndTime = dt.now()
                    data["addedDateAndTime"] = str(addedDateAndTime)
                    data["updatedDateAndTime"] = str(addedDateAndTime)
                    data_items = []
                    data_items.append(data)
                    fields = json.dumps(data_items)
                    if len(fields) > 65535:
                        return Response(
                            {
                                "msg": "Length of the value for update_fields property got exceeded!",

                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    else:
                        new_doc.updated_fields = fields
                    new_doc.save()
                    serializer = EntitySerializer(new_doc)
                    res_data.append(serializer.data)
                else:
                    failures[str(j + 1)] = "This entity division already exists"

            if failures:
                return Response(
                    {
                        "Message": "Data is not inserted for row number {}".format(failures),
                        "Duplicates": "There are duplicates in rows: {}".format(",".join(failures.keys()))},
                        status=status.HTTP_400_BAD_REQUEST,
                )

            return Response(res_data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        doc_object = self.get_object()
        data = request.data
        doc_object.entity_divisions = data.get(
            "entity_divisions", doc_object.entity_divisions
        )
        doc_object.updated_by = data.get("updated_by", doc_object.updated_by)
        doc_object.entity_name = data.get("entity_name", doc_object.entity_name)
        doc_object.save()
        serializer = EntitySerializer(doc_object)
        dataa = serializer.data
        return Response(dataa)

    def partial_update(self, request, *args, **kwargs):
        try:
            doc_object = self.get_object()
            data = request.data

            entity_divisions = data.get("entity_divisions", None)

            if entity_divisions:
                if Entity.objects.filter(
                    ~Q(id=doc_object.id),
                        entity_divisions = data["entity_divisions"]
                    ).exists():
                    return Response(
                        {
                            "Message": "This entity division already exists",
                            "Duplicates": "This entity division already exists"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            doc_object.entity_divisions = data.get(
                "entity_divisions", doc_object.entity_divisions
            )
            doc_object.updated_by = data.get("updated_by", doc_object.updated_by)
            doc_object.entity_name = data.get("entity_name", doc_object.entity_name)
            doc_object.updatedDateAndTime = dt.now()
            doc_object.save()

            updatedatetime = dt.now()
            data["updatedDateAndTime"] = str(updatedatetime)
            data_items = []
            if doc_object.updated_fields:
                old_list = doc_object.updated_fields
                old_list = json.loads(old_list)
                data_items.extend(old_list)
            data_items.append(data)
            changedFields = json.dumps(data_items)
            if len(changedFields) > 65535:
                return Response(
                    {"msg": "Length of the value for update_fields property got exceeded!"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            else:
                doc_object.updated_fields = changedFields
            doc_object.save()
            serializer = EntitySerializer(doc_object)
            dataa = serializer.data
            return Response(dataa, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        doc_object = self.get_object()
        doc_object.delete()
        return Response({"message": "entity deleted successfully"})


@csrf_exempt
def get_all_entity_divisions(request):
    if request.method == "GET":
        entity_divisions = Entity.objects.all().values_list('entity_divisions', flat=True).distinct()
        entity_divisions_dict = {"entity_divisions": sorted(entity_divisions)}
        return JsonResponse(entity_divisions_dict)

class IssueCatergoryViewSet(viewsets.ModelViewSet):
    queryset = IssueCatergory.objects.all()
    serializer_class = IssueCatergorySerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def create(self, request, *args, **kwargs):
        dataa = request.data
        objs = []
        for data in dataa:
            new_doc = IssueCatergory.objects.create(
                issue_catergory=data["issue_catergory"],
                created_by=data["created_by"],
                addedDateAndTime=datetime.now(),
            )
            new_doc.save()

            addedDateAndTime = datetime.now()
            data["addedDateAndTime"] = str(addedDateAndTime)
            data_items = []
            data_items.append(data)
            fields = json.dumps(data_items)
            if len(fields) > 65535:
                return Response(
                    {
                        "msg": "Length of the value for update_fields property got exceeded!"
                    }
                )
            else:
                new_doc.updated_fields = fields
                new_doc.save()
            objs.append(new_doc)
        serializer = IssueCatergorySerializer(objs, many=True)
        dataa = serializer.data
        return Response(dataa)

    def partial_update(self, request, *args, **kwargs):
        doc_object = self.get_object()
        data = request.data
        doc_object.issue_catergory = data.get(
            "issue_catergory", doc_object.issue_catergory
        )
        doc_object.updated_by = data.get("updated_by", doc_object.updated_by)
        doc_object.updatedDateAndTime = datetime.now()
        doc_object.save()

        updatedatetime = datetime.now()
        data["updatedDateAndTime"] = str(updatedatetime)
        data_items = []
        if doc_object.updated_fields:
            old_list = doc_object.updated_fields
            old_list = json.loads(old_list)
            data_items.extend(old_list)
        data_items.append(data)
        changedFields = json.dumps(data_items)
        if len(changedFields) > 65535:
            return Response(
                {"msg": "Length of the value for update_fields property got exceeded!"}
            )
        else:
            doc_object.updated_fields = changedFields

        doc_object.save()
        serializer = IssueCatergorySerializer(doc_object)
        dataa = serializer.data
        return Response(dataa)

def parse_date(date_str):
    """
    Parse date string in various formats and return YYYY-MM-DD format.
    Returns None if parsing fails.
    """
    if pd.isna(date_str) or str(date_str).lower() in ['nat', 'nan', '']:
        return None

    date_formats = [
        "%Y-%m-%d %H:%M:%S.%f",  # 2022-06-30 00:00:00.000
        "%Y-%m-%d %H:%M:%S",     # 2022-06-30 00:00:00
        "%Y-%m-%d",              # 2022-06-30
        "%d-%m-%Y %H:%M",        # 30-06-2022 00:00
        "%d-%m-%Y",              # 30-06-2022
        "%d/%m/%Y %H:%M",        # 30/06/2022 00:00
        "%d/%m/%Y",              # 30/06/2022
        "%m/%d/%Y %H:%M",        # 06/30/2022 00:00
        "%m/%d/%Y",              # 06/30/2022
    ]

    for date_format in date_formats:
        try:
            return dt.strptime(str(date_str), date_format).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return None


@csrf_exempt
def policy_data_excel_import(request):
    if request.method == "POST":
        uploaded_by = request.POST["uploaded_by"]
        form = PolicyInformationImportForm(request.POST, request.FILES)
        dict_excel_data = []
        if form.is_valid():
            df = pd.read_excel(request.FILES["policy_excel_file"])
            chunk_size = 3000
            archived_count = 0
            
            file_name = request.FILES["policy_excel_file"].name
            if PolicyInformation.objects.filter(file_name=file_name).exists():
                return JsonResponse({"message": "This file is already uploaded!"}, status=400)

            file = request.FILES["policy_excel_file"]
            match = re.search(r'(\d{2})-(\d{2})-(\d{4})', file_name)
            if match:
                day = match.group(1)  
                month = match.group(2)  
                year = match.group(3)  
            else:
                logger.info(f"File name wrong: The file name should contain date in the format DD-MM-YYYY")
                return JsonResponse({"message": "The file name should contain date in the format DD-MM-YYYY"}, status=400)

        try:
            # Fetch the current maximum ID to handle new IDs correctly
            max_id = PolicyInformation.objects.aggregate(Max('id'))['id__max'] or 0
            new_ids = []
            policy_objects = []
            with transaction.atomic():
                for start in range(0, len(df), chunk_size):
                    # Policy chunk will be refreshed.
                    policy_chunks = []
                    end = min(start + chunk_size, len(df))
                    df_chunk = df.iloc[start:end]
                    for index, row in df_chunk.iterrows():
                        max_id += 1
                        new_ids.append(max_id)
                        ddd = {}
                        Producing_Entity = row["Producing Entity"]
                        if not Producing_Entity == "nan":
                            ddd["Producing_Entity"] = Producing_Entity

                            Class_of_Business = (
                                None
                                if pd.isna(row["Class of Business"])
                                else row["Class of Business"]
                            )
                            ddd["Class_of_Business"] = Class_of_Business

                            Year_of_Account = (
                                None
                                if pd.isna(row["Year of Account"])
                                else row["Year of Account"]
                            )
                            ddd["Year_of_Account"] = Year_of_Account

                            Syndicate_Binder = (
                                None
                                if pd.isna(row["Syndicate Binder"])
                                else row["Syndicate Binder"]
                            )
                            ddd["Syndicate_Binder"] = Syndicate_Binder

                            Policy_Line_Ref = (
                                None
                                if pd.isna(row["Policy Line Ref"])
                                else row["Policy Line Ref"]
                            )
                            ddd["Policy_Line_Ref"] = Policy_Line_Ref

                            Policy_Status = (
                                None
                                if pd.isna(row["Policy Status"])
                                else row["Policy Status"]
                            )
                            ddd["Policy_Status"] = Policy_Status

                            Policy_Activity_Status = (
                                None
                                if row.get('Policy Activity Status', None) == None
                                or pd.isna(row["Policy Activity Status"])
                                else row["Policy Activity Status"]
                            )
                            ddd["Policy_Activity_Status"] = Policy_Activity_Status

                            Inception_Date = (
                                None
                                if row.get('Inception Date', None) == None
                                or pd.isna(row["Inception Date"])
                                or str(row["Inception Date"]) in ["NaT", "nan"]
                                else parse_date(str(row["Inception Date"]))
                            )
                            ddd["Inception_Date"] = Inception_Date

                            # Expired_Date = row['Expired Date']
                            # ddd["Expired_Date"] = Expired_Date

                            Expired_Date = (
                                None
                                if row.get('Expiry Date', None) == None
                                or pd.isna(row["Expiry Date"])
                                or str(row["Expiry Date"]) in ["NaT", "nan"]
                                else parse_date(str(row["Expiry Date"]))
                            )
                            ddd["Expired_Date"] = Expired_Date

                            Date_Cancelled = (
                                None
                                if row.get('Date Cancelled', None) == None
                                or pd.isna(row["Date Cancelled"])
                                or str(row["Date Cancelled"]) in ["NaT", "nan"]
                                else parse_date(str(row["Date Cancelled"]))
                            )
                            ddd["Date_Cancelled"] = Date_Cancelled

                            Cancellation_Reason = (
                                None
                                if row.get('Cancellation Reason', None) == None
                                or pd.isna(row["Cancellation Reason"])
                                else row["Cancellation Reason"]
                            )
                            ddd["Cancellation_Reason"] = Cancellation_Reason

                            Transaction_Status = (
                                None
                                if row.get('Transaction Status', None) == None
                                or pd.isna(row["Transaction Status"])
                                else row["Transaction Status"]
                            )
                            ddd["Transaction_Status"] = Transaction_Status

                            UMR_Number = (
                                None
                                if pd.isna(row["UMR Number"])
                                else row["UMR Number"]
                            )
                            ddd["UMR_Number"] = UMR_Number

                            Three_Party_Capacity_Deployed = (
                                None
                                if pd.isna(row["3rd Party Capacity Deployed"])
                                else row["3rd Party Capacity Deployed"]
                            )
                            ddd["Three_Party_Capacity_Deployed"] = (
                                Three_Party_Capacity_Deployed
                            )

                            SCM_Partner = (
                                None
                                if pd.isna(row["SCM Partner"])
                                else row["SCM Partner"]
                            )
                            ddd["SCM_Partner"] = SCM_Partner

                            Signed_Line_Pct = (
                                None
                                if row.get('Signed Line Pct', None) == None
                                or pd.isna(row["Signed Line Pct"])
                                else row["Signed Line Pct"]
                            )
                            ddd["Signed_Line_Pct"] = Signed_Line_Pct

                            Broker_Order_Pct = (
                                None
                                if row.get('Broker Order Pct', None) == None
                                or pd.isna(row["Broker Order Pct"])
                                else row["Broker Order Pct"]
                            )
                            ddd["Broker_Order_Pct"] = Broker_Order_Pct

                            Signed_Order_Pct = (
                                None
                                if row.get('Signed Order Pct', None) == None
                                or pd.isna(row["Signed Order Pct"])
                                else row["Signed Order Pct"]
                            )
                            ddd["Signed_Order_Pct"] = Signed_Order_Pct

                            Broker_Commision_Pct = (
                                None
                                if row.get('Broker Commision Pct', None) == None
                                or pd.isna(row["Broker Commision Pct"])
                                else row["Broker Commision Pct"]
                            )
                            ddd["Broker_Commision_Pct"] = Broker_Commision_Pct

                            Coverholder_Commision_Pct = (
                                None
                                if row.get('Coverholder Commision Pct', None) == None
                                or pd.isna(row["Coverholder Commision Pct"])
                                else row["Coverholder Commision Pct"]
                            )
                            ddd["Coverholder_Commision_Pct"] = Coverholder_Commision_Pct

                            Broker_Reference = (
                                None
                                if row.get('Broker Reference', None) == None
                                or pd.isna(row["Broker Reference"])
                                else row["Broker Reference"]
                            )
                            ddd["Broker_Reference"] = Broker_Reference

                            Underwriter = (
                                None 
                                if row.get('Underwriter', None) == None
                                or pd.isna(row["Underwriter"]) 
                                else row["Underwriter"]
                            )
                            ddd["Underwriter"] = Underwriter

                            MOP = None if row.get('MOP', None) == None or pd.isna(row["MOP"]) else row["MOP"]
                            ddd["MOP"] = MOP

                            Broker = None if pd.isna(row["Broker"]) else row["Broker"]
                            ddd["Broker"] = Broker

                            Master_Broker = (
                                None if pd.isna(row["Master Broker"]) else row["Master Broker"]
                            )
                            ddd["Master_Broker"] = Master_Broker

                            Insured = None if pd.isna(row["Insured"]) else row["Insured"]
                            ddd["Insured"] = Insured

                            Summary_Currency = (
                                None
                                if row.get('Summary Currency', None) == None
                                or pd.isna(row["Summary Currency"])
                                else row["Summary Currency"]
                            )
                            ddd["Summary_Currency"] = Summary_Currency

                            Summary_ROE = (
                                None if row.get('Summary ROE', None) == None or pd.isna(row["Summary ROE"]) else row["Summary ROE"]
                            )
                            ddd["Summary_ROE"] = Summary_ROE

                            Settlement_Ccy = (
                                None
                                if row.get('Settlement Ccy', None) == None
                                or pd.isna(row["Settlement Ccy"])
                                else row["Settlement Ccy"]
                            )
                            ddd["Settlement_Ccy"] = Settlement_Ccy

                            Settlement_ROE = (
                                None
                                if row.get('Settlement ROE', None) == None
                                or pd.isna(row["Settlement ROE"])
                                else row["Settlement ROE"]
                            )
                            ddd["Settlement_ROE"] = Settlement_ROE

                            Gross_Written_Premium_100_in_Sett = (
                                None
                                if row.get('Gross Written Premium 100 (Sett)', None) == None
                                or pd.isna(row["Gross Written Premium 100 (Sett)"])
                                else row["Gross Written Premium 100 (Sett)"]
                            )
                            ddd["Gross_Written_Premium_100_in_Sett"] = (
                                Gross_Written_Premium_100_in_Sett
                            )

                            Net_Written_Premium_100_in_Sett = (
                                None
                                if row.get('Net Written Premium 100 (Sett)', None) == None
                                or pd.isna(row["Net Written Premium 100 (Sett)"])
                                else row["Net Written Premium 100 (Sett)"]
                            )
                            ddd["Net_Written_Premium_100_in_Sett"] = (
                                Net_Written_Premium_100_in_Sett
                            )

                            True_Net_Written_Premium_100_in_Sett = (
                                None
                                if row.get('True Net Written Premium 100 (Sett)', None) == None
                                or pd.isna(row["True Net Written Premium 100 (Sett)"]) 
                                else row["True Net Written Premium 100 (Sett)"]
                            )
                            ddd["True_Net_Written_Premium_100_in_Sett"] = (
                                True_Net_Written_Premium_100_in_Sett
                            )

                            Original_Ccy = (
                                None if pd.isna(row["Original Ccy"]) else row["Original Ccy"]
                            )
                            ddd["Original_Ccy"] = Original_Ccy

                            Original_ROE = (
                                None if row.get('Original ROE', None) == None or pd.isna(row["Original ROE"]) else row["Original ROE"]
                            )
                            ddd["Original_ROE"] = Original_ROE

                            Gross_Written_Premium_100_in_Orig = (
                                None
                                if row.get('Gross Written Premium 100 (Orig)', None) == None
                                or pd.isna(row["Gross Written Premium 100 (Orig)"])
                                else row["Gross Written Premium 100 (Orig)"]
                            )
                            ddd["Gross_Written_Premium_100_in_Orig"] = (
                                Gross_Written_Premium_100_in_Orig
                            )

                            Gross_Written_Premium_Agency_Share_in_Orig = (
                                None
                                if row.get('Gross Written Premium Agency Share (Orig)', None) == None
                                or pd.isna(row["Gross Written Premium Agency Share (Orig)"])
                                else row["Gross Written Premium Agency Share (Orig)"]
                            )
                            ddd["Gross_Written_Premium_Agency_Share_in_Orig"] = (
                                Gross_Written_Premium_Agency_Share_in_Orig
                            )

                            Gross_Written_Premium_Syndicate_Share_in_Orig = (
                                None
                                if row.get('Gross Written Premium Syndicate Share (Orig)', None) == None
                                or pd.isna(row["Gross Written Premium Syndicate Share (Orig)"])
                                else row["Gross Written Premium Syndicate Share (Orig)"]
                            )
                            ddd["Gross_Written_Premium_Syndicate_Share_in_Orig"] = (
                                Gross_Written_Premium_Syndicate_Share_in_Orig
                            )

                            Gross_Written_Premium_Non_Syndicate_Share_in_Orig = (
                                None
                                if row.get('Gross Written Premium Non-Syndicate Share (Orig)', None) == None
                                or pd.isna(row["Gross Written Premium Non-Syndicate Share (Orig)"])
                                else row["Gross Written Premium Non-Syndicate Share (Orig)"]
                            )
                            ddd["Gross_Written_Premium_Non_Syndicate_Share_in_Orig"] = (
                                Gross_Written_Premium_Non_Syndicate_Share_in_Orig
                            )

                            Net_Written_Premium_100_in_Orig = (
                                None
                                if row.get('Net Written Premium 100 (Orig)', None) == None
                                or pd.isna(row["Net Written Premium 100 (Orig)"])
                                else row["Net Written Premium 100 (Orig)"]
                            )
                            ddd["Net_Written_Premium_100_in_Orig"] = (
                                Net_Written_Premium_100_in_Orig
                            )

                            Net_Written_Premium_Agency_Share_in_Orig = (
                                None
                                if row.get('Net Written Premium Agency Share (Orig)', None) == None
                                or pd.isna(row["Net Written Premium Agency Share (Orig)"])
                                else row["Net Written Premium Agency Share (Orig)"]
                            )
                            ddd["Net_Written_Premium_Agency_Share_in_Orig"] = (
                                Net_Written_Premium_Agency_Share_in_Orig
                            )

                            Net_Written_Premium_Syndicate_Share_in_Orig = (
                                None
                                if row.get('Net Written Premium Syndicate Share (Orig)', None) == None
                                or pd.isna(row["Net Written Premium Syndicate Share (Orig)"])
                                else row["Net Written Premium Syndicate Share (Orig)"]
                            )
                            ddd["Net_Written_Premium_Syndicate_Share_in_Orig"] = (
                                Net_Written_Premium_Syndicate_Share_in_Orig
                            )

                            Net_Written_Premium_Non_Syndicate_Share_in_Orig = (
                                None
                                if row.get('Net Written Premium Non-Syndicate Share (Orig)', None) == None
                                or pd.isna(row["Net Written Premium Non-Syndicate Share (Orig)"])
                                else row["Net Written Premium Non-Syndicate Share (Orig)"]
                            )
                            ddd["Net_Written_Premium_Non_Syndicate_Share_in_Orig"] = (
                                Net_Written_Premium_Non_Syndicate_Share_in_Orig
                            )

                            True_Net_Written_Premium_100_in_Orig = (
                                None
                                if row.get('True Net Written Premium 100 (Orig)', None) == None
                                or pd.isna(row["True Net Written Premium 100 (Orig)"])
                                else row["True Net Written Premium 100 (Orig)"]
                            )
                            ddd["True_Net_Written_Premium_100_in_Orig"] = (
                                True_Net_Written_Premium_100_in_Orig
                            )

                            True_Net_Written_Premium_Syndicate_Share_in_Orig = (
                                None
                                if row.get('True Net Written Premium Syndicate Share (Orig)', None) == None
                                or pd.isna(row["True Net Written Premium Syndicate Share (Orig)"])
                                else row["True Net Written Premium Syndicate Share (Orig)"]
                            )
                            ddd["True_Net_Written_Premium_Syndicate_Share_in_Orig"] = (
                                True_Net_Written_Premium_Syndicate_Share_in_Orig
                            )

                            True_Net_Written_Premium_Agency_Share_in_Orig = (
                                None
                                if row.get('True Net Written Premium Agency Share (Orig)', None) == None
                                or pd.isna(row["True Net Written Premium Agency Share (Orig)"])
                                else row["True Net Written Premium Agency Share (Orig)"]
                            )
                            ddd["True_Net_Written_Premium_Agency_Share_in_Orig"] = (
                                True_Net_Written_Premium_Agency_Share_in_Orig
                            )

                            True_Net_Written_Premium_Non_Syndicate_Share_in_Orig = (
                                None
                                if row.get('True Net Written Premium Non-Syndicate Share (Orig)', None) == None
                                or pd.isna(row["True Net Written Premium Non-Syndicate Share (Orig)"])
                                else row["True Net Written Premium Non-Syndicate Share (Orig)"]
                            )
                            ddd["True_Net_Written_Premium_Non_Syndicate_Share_in_Orig"] = (
                                True_Net_Written_Premium_Non_Syndicate_Share_in_Orig
                            )

                            Gross_Written_Premium_Syndicate_Share_in_USD = (
                                None
                                if row.get('Gross Written Premium Syndicate Share (USD)', None) == None
                                or pd.isna(row["Gross Written Premium Syndicate Share (USD)"])
                                else row["Gross Written Premium Syndicate Share (USD)"]
                            )
                            ddd["Gross_Written_Premium_Syndicate_Share_in_USD"] = (
                                Gross_Written_Premium_Syndicate_Share_in_USD
                            )

                            Gross_Written_Premium_Agency_Share_in_USD = (
                                None
                                if row.get('Gross Written Premium Agency Share (USD)', None) == None
                                or pd.isna(row["Gross Written Premium Agency Share (USD)"])
                                else row["Gross Written Premium Agency Share (USD)"]
                            )
                            ddd["Gross_Written_Premium_Agency_Share_in_USD"] = (
                                Gross_Written_Premium_Agency_Share_in_USD
                            )

                            Gross_Written_Premium_100_in_USD = (
                                None
                                if row.get('Gross Written Premium 100 (USD)', None) == None
                                or pd.isna(row["Gross Written Premium 100 (USD)"])
                                else row["Gross Written Premium 100 (USD)"]
                            )
                            ddd["Gross_Written_Premium_100_in_USD"] = (
                                Gross_Written_Premium_100_in_USD
                            )

                            Net_Written_Premium_100_in_USD = (
                                None
                                if row.get('Net Written Premium 100 (USD)', None) == None
                                or pd.isna(row["Net Written Premium 100 (USD)"])
                                else row["Net Written Premium 100 (USD)"]
                            )
                            ddd["Net_Written_Premium_100_in_USD"] = (
                                Net_Written_Premium_100_in_USD
                            )

                            True_Net_Written_Premium_100_in_USD = (
                                None
                                if row.get('True Net Written Premium 100 (USD)', None) == None
                                or pd.isna(row["True Net Written Premium 100 (USD)"])
                                else row["True Net Written Premium 100 (USD)"]
                            )
                            ddd["True_Net_Written_Premium_100_in_USD"] = (
                                True_Net_Written_Premium_100_in_USD
                            )

                            PremiumBasis = (
                                None 
                                if row.get('PremiumBasis', None) == None
                                or pd.isna(row["PremiumBasis"]) 
                                else row["PremiumBasis"]
                            )
                            ddd["PremiumBasis"] = PremiumBasis

                            Instalment_Nbr = (
                                None
                                if row.get('Instalment Nbr', None) == None
                                or pd.isna(row["Instalment Nbr"])
                                else row["Instalment Nbr"]
                            )
                            ddd["Instalment_Nbr"] = Instalment_Nbr

                            Installment_Category = (
                                None
                                if row.get('Installment Category', None) == None
                                or pd.isna(row["Installment Category"])
                                else row["Installment Category"]
                            )
                            ddd["Installment_Category"] = Installment_Category

                            Installment_Due_date = (
                                None
                                if row.get('Installment Due date', None) == None
                                or pd.isna(row["Installment Due date"]) 
                                or str(row["Installment Due date"]) in ["NaT", "nan"]
                                else parse_date(str(row["Installment Due date"]))
                            )
                            ddd["Installment_Due_date"] = Installment_Due_date

                            Installment_Ccy_in_Orig = (
                                None
                                if row.get('Installment Ccy (Orig)', None) == None
                                or pd.isna(row["Installment Ccy (Orig)"])
                                else row["Installment Ccy (Orig)"]
                            )
                            ddd["Installment_Ccy_in_Orig"] = Installment_Ccy_in_Orig

                            Installment_Agency_Amount_in_Orig = row['Installment Agency Amount (Orig)']
                            ddd["Installment_Agency_Amount_in_Orig"] = Installment_Agency_Amount_in_Orig

                            Installment_Agency_Amount_in_Sett = (
                                None
                                if row.get('Installment Agency Amount (Sett)', None) == None
                                or pd.isna(row["Installment Agency Amount (Sett)"])
                                else row["Installment Agency Amount (Sett)"]
                            )
                            ddd["Installment_Agency_Amount_in_Sett"] = (
                                Installment_Agency_Amount_in_Sett
                            )

                            Installment_Agency_Amount_in_USD = (
                                None
                                if row.get('Installment Agency Amount (USD)', None) == None
                                or pd.isna(row["Installment Agency Amount (USD)"])
                                else row["Installment Agency Amount (USD)"]
                            )
                            ddd["Installment_Agency_Amount_in_USD"] = (
                                Installment_Agency_Amount_in_USD
                            )

                            Installment_Amount_Syndicate_Share_in_Orig = (
                                None
                                if row.get('Installment Amount Syndicate Share (Sett)', None) == None
                                or pd.isna(row["Installment Amount Syndicate Share (Sett)"])
                                else row["Installment Amount Syndicate Share (Sett)"]
                            )
                            ddd["Installment_Amount_Syndicate_Share_in_Orig"] = (
                                Installment_Amount_Syndicate_Share_in_Orig
                            )

                            Installment_Amount_Syndicate_Share_in_Sett = (
                                None
                                if row.get('Installment Amount Syndicate Share (Orig)', None) == None
                                or pd.isna(row["Installment Amount Syndicate Share (Orig)"])
                                else row["Installment Amount Syndicate Share (Orig)"]
                            )
                            ddd["Installment_Amount_Syndicate_Share_in_Sett"] = (
                                Installment_Amount_Syndicate_Share_in_Sett
                            )

                            Installment_Amount_Syndicate_Share_in_USD = (
                                None
                                if row.get('Installment Amount Syndicate Share (USD)', None) == None
                                or pd.isna(row["Installment Amount Syndicate Share (USD)"])
                                else row["Installment Amount Syndicate Share (USD)"]
                            )
                            ddd["Installment_Amount_Syndicate_Share_in_USD"] = (
                                Installment_Amount_Syndicate_Share_in_USD
                            )

                            Paid_Amount_in_USD = (
                                None
                                if row.get('Paid Amount (USD)', None) == None
                                or pd.isna(row["Paid Amount (USD)"])
                                else row["Paid Amount (USD)"]
                            )
                            ddd["Paid_Amount_in_USD"] = Paid_Amount_in_USD

                            Last_Allocation_Date = (
                                None
                                if row.get('Last Allocation Date', None) == None
                                or pd.isna(row["Last Allocation Date"]) 
                                or str(row["Last Allocation Date"]) in ["NaT", "nan"]
                                else parse_date(str(row["Last Allocation Date"]))
                            )
                            ddd["Last_Allocation_Date"] = Last_Allocation_Date

                            Diff_in_USD = (
                                None 
                                if row.get('Diff (USD)', None) == None
                                or pd.isna(row["Diff (USD)"]) 
                                else row["Diff (USD)"]
                            )
                            ddd["Diff_in_USD"] = Diff_in_USD

                            Overdue_Days = (
                                None 
                                if row.get('Overdue Days', None) == None
                                or pd.isna(row["Overdue Days"]) 
                                else row["Overdue Days"]
                            )
                            ddd["Overdue_Days"] = Overdue_Days

                            Overdue_Category = (
                                None
                                if row.get('Overdue Category', None) == None
                                or pd.isna(row["Overdue Category"])
                                else row["Overdue Category"]
                            )
                            ddd["Overdue_Category"] = Overdue_Category
                            dict_excel_data.append(ddd)

                            market_source = (
                                None if pd.isna(row["Market Source"]) else row["Market Source"]
                            )
                            ddd["market_source"] = market_source

                            i_policy_key = (
                                None if pd.isna(row["lPolicyKey"]) else row["lPolicyKey"]
                            )
                            ddd["i_policy_key"] = i_policy_key

                            cancellation_type = (
                                None if pd.isna(row["Cancellation Type"]) else row["Cancellation Type"]
                            )
                            ddd["cancellation_type"] = cancellation_type

                            line_ref_1609_5399 = row.get("1609/5399 Line Ref")
                            line_ref_2610_5431 = row.get("2610/5431 Line Ref")

                            if pd.isna(line_ref_1609_5399):
                                line_ref_1609_5399 = None

                            if pd.isna(line_ref_2610_5431):
                                line_ref_2610_5431 = None

                            # last_record = PolicyInformation.objects.all().order_by('id').last()
                            # print("last_record", last_record.id)

                            # # Fetch the existing record that needs to be archived
                            # existing_records = PolicyInformation.objects.filter(
                            #     Policy_Line_Ref=Policy_Line_Ref, Instalment_Nbr=Instalment_Nbr
                            # )

                            # if existing_records.exists():
                            #     for record in existing_records:
                            #         record.archived = True
                            #         record.save()
                            #         print(f"Archived record with ID: {record.id}")
                            # else:
                            #     print("No existing record found to archive.")
                            # policy_objects.append(PolicyInformation(

                            # Appended the record to policy Chunk.
                            policy_objects.append(PolicyInformation(
                                id=max_id,
                                Producing_Entity=Producing_Entity,
                                Class_of_Business=Class_of_Business,
                                Year_of_Account=Year_of_Account,
                                Syndicate_Binder=Syndicate_Binder,
                                Policy_Line_Ref=Policy_Line_Ref,
                                Policy_Status=Policy_Status,
                                Policy_Activity_Status=Policy_Activity_Status,
                                Inception_Date=Inception_Date,
                                Expired_Date=Expired_Date,
                                Date_Cancelled=Date_Cancelled,
                                Cancellation_Reason=Cancellation_Reason,
                                Transaction_Status=Transaction_Status,
                                UMR_Number=UMR_Number,
                                Three_Party_Capacity_Deployed=Three_Party_Capacity_Deployed,
                                SCM_Partner=SCM_Partner,
                                Signed_Line_Pct=Signed_Line_Pct,
                                Broker_Order_Pct=Broker_Order_Pct,
                                Signed_Order_Pct=Signed_Order_Pct,
                                Broker_Commision_Pct=Broker_Commision_Pct,
                                Coverholder_Commision_Pct=Coverholder_Commision_Pct,
                                Broker_Reference=Broker_Reference,
                                Underwriter=Underwriter,
                                MOP=MOP,
                                Broker=Broker,
                                Master_Broker=Master_Broker,
                                Insured=Insured,
                                Summary_Currency=Summary_Currency,
                                Summary_ROE=Summary_ROE,
                                Settlement_Ccy=Settlement_Ccy,
                                Settlement_ROE=Settlement_ROE,
                                Gross_Written_Premium_100_in_Sett=Gross_Written_Premium_100_in_Sett,
                                Net_Written_Premium_100_in_Sett=Net_Written_Premium_100_in_Sett,
                                True_Net_Written_Premium_100_in_Sett=True_Net_Written_Premium_100_in_Sett,
                                Original_Ccy=Original_Ccy,
                                Original_ROE=Original_ROE,
                                Gross_Written_Premium_100_in_Orig=Gross_Written_Premium_100_in_Orig,
                                Gross_Written_Premium_Agency_Share_in_Orig=Gross_Written_Premium_Agency_Share_in_Orig,
                                Gross_Written_Premium_Syndicate_Share_in_Orig=Gross_Written_Premium_Syndicate_Share_in_Orig,
                                Gross_Written_Premium_Non_Syndicate_Share_in_Orig=Gross_Written_Premium_Non_Syndicate_Share_in_Orig,
                                Net_Written_Premium_100_in_Orig=Net_Written_Premium_100_in_Orig,
                                Net_Written_Premium_Agency_Share_in_Orig=Net_Written_Premium_Agency_Share_in_Orig,
                                Net_Written_Premium_Syndicate_Share_in_Orig=Net_Written_Premium_Syndicate_Share_in_Orig,
                                Net_Written_Premium_Non_Syndicate_Share_in_Orig=Net_Written_Premium_Non_Syndicate_Share_in_Orig,
                                True_Net_Written_Premium_100_in_Orig=True_Net_Written_Premium_100_in_Orig,
                                True_Net_Written_Premium_Syndicate_Share_in_Orig=True_Net_Written_Premium_Syndicate_Share_in_Orig,
                                True_Net_Written_Premium_Agency_Share_in_Orig=True_Net_Written_Premium_Agency_Share_in_Orig,
                                True_Net_Written_Premium_Non_Syndicate_Share_in_Orig=True_Net_Written_Premium_Non_Syndicate_Share_in_Orig,
                                Gross_Written_Premium_Syndicate_Share_in_USD=Gross_Written_Premium_Syndicate_Share_in_USD,
                                Gross_Written_Premium_Agency_Share_in_USD=Gross_Written_Premium_Agency_Share_in_USD,
                                Gross_Written_Premium_100_in_USD=Gross_Written_Premium_100_in_USD,
                                Net_Written_Premium_100_in_USD=Net_Written_Premium_100_in_USD,
                                True_Net_Written_Premium_100_in_USD=True_Net_Written_Premium_100_in_USD,
                                PremiumBasis=PremiumBasis,
                                Instalment_Nbr=Instalment_Nbr,
                                Installment_Category=Installment_Category,
                                Installment_Due_date=Installment_Due_date,
                                Installment_Ccy_in_Orig=Installment_Ccy_in_Orig,
                                Installment_Agency_Amount_in_Orig=Installment_Agency_Amount_in_Orig,
                                Installment_Agency_Amount_in_Sett=Installment_Agency_Amount_in_Sett,
                                Installment_Agency_Amount_in_USD=Installment_Agency_Amount_in_USD,
                                Installment_Amount_Syndicate_Share_in_Orig=Installment_Amount_Syndicate_Share_in_Orig,
                                Installment_Amount_Syndicate_Share_in_Sett=Installment_Amount_Syndicate_Share_in_Sett,
                                Installment_Amount_Syndicate_Share_in_USD=Installment_Amount_Syndicate_Share_in_USD,
                                Paid_Amount_in_USD=Paid_Amount_in_USD,
                                Last_Allocation_Date=Last_Allocation_Date,
                                Diff_in_USD=Diff_in_USD,
                                Overdue_Days=Overdue_Days,
                                Overdue_Category=Overdue_Category,
                                uploaded_by=uploaded_by,
                                file_month=month,
                                file_year=year,
                                file_name=file_name,
                                market_source=market_source,
                                i_policy_key=i_policy_key,
                                cancellation_type=cancellation_type,
                                line_ref_1609_5399=line_ref_1609_5399,
                                line_ref_2610_5431=line_ref_2610_5431
                            ))

                        logger.info(f"Outer FOr loop :  {start}  {end}")
                    
                    
            uploaded_date_time = timezone.now() 
            logger.info(f"uploaded_date_time : {uploaded_date_time} ")

            upload_success = False
            try:
                # Step 1: Save records to the DB first within a transaction
                with transaction.atomic():
                    logger.info("Applying Bulk Create 1 ")

                    # Define batch size
                    batch_size = 2000

                    logger.info(f"Policy Object Length : {len(policy_objects)}")
                    # Split the policy_objects into batches of 2000 records
                    for i in range(0, len(policy_objects), batch_size):
                        batch = policy_objects[i:i + batch_size]
                        try:
                            # Bulk create each batch
                            PolicyInformation.objects.bulk_create(batch)
                            logger.info(f"Batch {i // batch_size + 1} inserted successfully. Total records saved: {len(batch)}")
                            
                            # Simulate a delay of 1 second before processing the next batch
                            time.sleep(1)
                        except Exception as e:
                            # If an error occurs during bulk create, we log the error and stop further processing
                            logger.error(f"Error occurred during bulk create in batch {i // batch_size + 1}: {str(e)}")
                            raise Exception(f"Error during database insertion, aborting process. No file upload will be done.")
                    
                    logger.info(f"All records successfully inserted. Total records saved: {len(policy_objects)}")
                
                # Step 2: If bulk create is successful, attempt to upload the file to S3
                reuired_data = {
                    "module_name": "Remittance Upload",
                    "no_of_records": len(policy_objects),
                    "bucket_name": config("AWS_STORAGE_BUCKET_NAME")
                }

                user = get_user(request)
                filemanagment = reusable_file_upload(user, file, reuired_data, is_upload=False)

                # Step 3: Check if the file upload was successful
                if isinstance(filemanagment, Response) and filemanagment.status_code >= 400:
                    error_msg = filemanagment.data.get('error', 'Unable to upload the file to S3')
                    
                    # If file upload fails, rollback the DB transaction by raising an exception
                    logger.error(f"File upload to S3 failed: {error_msg}")
                    # raise Exception(f"File upload to S3 failed: {error_msg}")

                logger.info(f"File uploaded to S3 successfully.")

                AgedDeptFileRecord.objects.create(
                    file_name=file_name,
                    no_of_records=len(policy_objects),
                    month=month,  # Assuming month is passed in the request
                    year=year,    # Assuming year is passed in the request
                    uploaded_date_time=uploaded_date_time,
                    uploaded_by=uploaded_by,
                    archived=False,  # Set to False since these records are newly created
                    status="Pending",
                )
                logger.info("AgedDeptFileRecord INserted")
                upload_success = True

            except Exception as e:
                # In case of any error (either during file upload or DB insert)
                logger.error(f"Error: {str(e)}")
                return JsonResponse({"error": f"Upload failed: {str(e)}"}, status=400)

            archived_count = 0

            if upload_success:
                try:
                    archived_count = PolicyInformation.objects.exclude(id__in=new_ids)\
                        .filter(archived=False).update(archived=True)
                except Exception as e:
                    logger.error(f"Archiving failed: {str(e)}")

            if upload_success:
                return JsonResponse({
                    "results": f"Age data inserted successfully. Total records saved: {len(policy_objects)}. Total records archived: {archived_count}"
                })
            else:
                return JsonResponse({"error": "Upload failed"}, status=400)

        except TimeoutError:
            return JsonResponse({"error": "Upload timed out. Please try again."}, status=408)
            
        except Exception as e:
            print(e)
            logger.info(f"Error while saving record: {str(e)}")
            return JsonResponse({"error": str(e)}, status=400)
        


class PolicyViewSet(viewsets.ModelViewSet):
    queryset = PolicyInformation.objects.filter(archived=False)
    serializer_class = PolicySerializer


@csrf_exempt
def getDetailsByPolicyRef(request):
    if request.method == "GET":
        Policy_Line_Ref = request.GET.get("Policy_Line_Ref")
        policyDetailsObjects = (
            PolicyInformation.objects.filter(
                Policy_Line_Ref=Policy_Line_Ref, archived=False
            )
            .order_by("id")
            .first()
        )
        if policyDetailsObjects:
            serializer = PolicyRefNoSerializer(policyDetailsObjects)
            data = serializer.data
        else:
            data = {"message": "Data not Found"}
        return JsonResponse(data)


class PolicyInformationViewSet(viewsets.ModelViewSet):
    model = PolicyInformation
    serializer_class = PolicyRefNoSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        pol = PolicyInformation.objects.filter(archived=False)
        return pol

    def list(self, request):
        search_key = request.GET.get("policy_line_ref", None)
        policies = PolicyInformation.objects.filter(
            Policy_Line_Ref__icontains=search_key, archived=False
        )

        response_data = []
        for policy in policies:
            response_data.append(
                {"policy_pk": policy.id, "policy_id": policy.Policy_Line_Ref, "installment_nuber": policy.Instalment_Nbr, "orignal_amount": policy.Installment_Agency_Amount_in_Orig, "installment_due_date": policy.Installment_Due_date, "policy_status": policy.Policy_Status}
            )

        return Response({"all_policies": response_data})


class PolicyCreateViewSet(viewsets.ModelViewSet, CountModelMixin):
    model = PolicyInformation
    serializer_class = PolicyRefNoSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    @staticmethod
    def get_policy_information_based_on_id(policy_number):
        return PolicyInformation.objects.filter(
            Policy_Line_Ref=policy_number, archived=False
        )

    def get_queryset(self):
        pol = PolicyInformation.objects.filter(archived=False)
        return pol

    def list(self, request):
        policy = PolicyInformation.objects.filter(archived=False)
        serializer = PolicyRefNoSerializer(policy, many=True)
        dataa = serializer.data
        return Response(dataa)

    def create(self, request, *args, **kwargs):
        data = request.data
        id = PolicyInformation.objects.last().id
        try:
            obj = PolicyInformation.objects.filter(
                Policy_Line_Ref=data["Policy_Line_Ref"], archived=False
            ).first()
            if obj:
                return Response(
                    {"Message": "Policy with that id already exists!"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            pass

        new_policy = PolicyInformation.objects.create(
            id=int(id) + 1,
            Producing_Entity=data["Producing_Entity"],
            Syndicate_Binder=data["Syndicate_Binder"],
            Policy_Line_Ref=data["Policy_Line_Ref"],
            UMR_Number=data["UMR_Number"],
            Three_Party_Capacity_Deployed=data["Three_Party_Capacity_Deployed"],
            SCM_Partner=data["SCM_Partner"],
            SCM_Insurer_partner_name=data["SCM_Insurer_partner_name"],
            Binding_Agreement=data["Binding_Agreement"],
            Settlement_Ccy=data["settlement_currency"],
            Original_Ccy=data["original_ccy"],
            Installment_Agency_Amount_in_Orig=data["installment_amount_org"]
        )
        new_policy.save()
        serializer = PolicyRefNoSerializer(new_policy)
        data = serializer.data
        return Response(data)

    def retrieve(self, request, pk=None):
        if pk:
            doc = PolicyInformation.objects.get(id=pk)
            serializer = PolicyRefNoSerializer(doc)
            dataa = serializer.data
            return Response(dataa)

    def update(self, request, *args, **kwargs):
        policy_object = self.get_object()
        print("policy_object", policy_object)
        data = request.data
        policy_object.Producing_Entity = data["Producing_Entity"]
        policy_object.Syndicate_Binder = data["Syndicate_Binder"]
        policy_object.Policy_Line_Ref = data["Policy_Line_Ref"]
        policy_object.UMR_Number = data["UMR_Number"]
        policy_object.Three_Party_Capacity_Deployed = data[
            "Three_Party_Capacity_Deployed"
        ]
        policy_object.SCM_Partner = data["SCM_Partner"]
        policy_object.SCM_Insurer_partner_name = data["SCM_Insurer_partner_name"]
        policy_object.Binding_Agreement = data["Binding_Agreement"]
        policy_object.save()
        serializer = PolicyRefNoSerializer(policy_object)
        dataa = serializer.data
        return Response(dataa)

    def partial_update(self, request, *args, **kwargs):
        policy_object = self.get_object()
        data = request.data
        policy_object.Producing_Entity = data.get(
            "Producing_Entity", policy_object.Producing_Entity
        )
        policy_object.Syndicate_Binder = data.get(
            "Syndicate_Binder", policy_object.Syndicate_Binder
        )
        policy_object.Policy_Line_Ref = data.get(
            "Policy_Line_Ref", policy_object.Policy_Line_Ref
        )
        policy_object.UMR_Number = data.get("UMR_Number", policy_object.UMR_Number)
        policy_object.Three_Party_Capacity_Deployed = data.get(
            "Three_Party_Capacity_Deployed", policy_object.Three_Party_Capacity_Deployed
        )
        policy_object.SCM_Partner = data.get("SCM_Partner", policy_object.SCM_Partner)
        policy_object.SCM_Insurer_partner_name = data.get(
            "SCM_Insurer_partner_name", policy_object.SCM_Insurer_partner_name
        )
        policy_object.Binding_Agreement = data.get(
            "Binding_Agreement", policy_object.Binding_Agreement
        )
        policy_object.save()
        serializer = DocumentsSerializer(policy_object)
        dataa = serializer.data
        return Response(dataa)

    def destroy(self, request, *args, **kwargs):
        policy_object = self.get_object()
        policy_object.delete()
        return Response({"message": "policy deleted successfully"})


@csrf_exempt
def exchange_rate_excel_import(request):
    if request.method == "POST":
        form = ExchangeRateImportForm(request.POST, request.FILES)
        if form.is_valid():
            dict_excel_data = []
            monthh = request.POST.get("month")
            yearr = request.POST.get("year")
            year = int(yearr)
            month = int(monthh)
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - pd.Timedelta(seconds=1)
            else:
                end_date = datetime(year, month + 1, 1) - pd.Timedelta(seconds=1)
            df = pd.read_excel(request.FILES["exchange_excel_file"])
            for index, row in df.iterrows():
                ddd = {}
                Month = row["Month"]
                print(start_date, end_date, Month)
                if (Month >= start_date) & (Month <= end_date):
                    print("mmmmm", Month)
                    currency_code = row["Currency Code"]
                    if not currency_code == "nan":
                        Month = row["Month"]
                        # print("ssssss excel",Month)
                        ddd["Month"] = Month
                        currency_code = row["Currency Code"]
                        ddd["currency_code"] = currency_code
                        exchange_rate = row["Month-End"]
                        ddd["exchange_rate"] = exchange_rate
                        if not BankExchangeRate.objects.filter(
                            month=Month, currency_code=currency_code
                        ).exists():
                            dict_excel_data.append(ddd)
                            BankExchangeRate.objects.create(
                                month=Month,
                                currency_code=currency_code,
                                exchange_rate=exchange_rate,
                            )
                else:
                    pass
            if not dict_excel_data:
                return JsonResponse({"results": "No Data Found By Given Month"})
            else:
                return JsonResponse({"results": dict_excel_data})


class BankExchangeRateViewSet(viewsets.ModelViewSet):
    model = BankExchangeRate
    serializer_class = BankExchangeRateSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        en = BankExchangeRate.objects.all()
        return en

    def list(self, request):
        data = BankExchangeRate.objects.all().order_by('-month', 'currency_code')
        currency_code = request.query_params.get('currencyCode', None)
        transaction_year = request.query_params.get('transactionYear', None)
        transaction_month = request.query_params.get('transactionMonth', None)
        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        skip = page_number * rows_per_page
        filter_conditions = Q()
        if currency_code:
            filter_conditions &= Q(currency_code=currency_code)

        if transaction_year:
            filter_conditions &= Q(month__icontains = f"{transaction_year}-")

        if transaction_month:
            filter_conditions &= Q(month__icontains = f"-{transaction_month}-" )

        filtered_data = data.filter(filter_conditions)
        count = data.filter(filter_conditions).count()

        serializer = BankExchangeRateSerializer(filtered_data[skip: skip + rows_per_page], many=True)
        dataa = serializer.data
        return Response({ "data": dataa, "count": count }, status=status.HTTP_200_OK)

    def retrieve(self, request, pk=None):
        if pk:
            en = BankExchangeRate.objects.get(id=pk)
            serializer = BankExchangeRateSerializer(en)
            dataa = serializer.data
            return Response(dataa)

    def create(self, request, *args, **kwargs):
        try:
            dataa = request.data
            results = []
            for data in dataa:
                if not BankExchangeRate.objects.filter(
                    month=data["month"], currency_code=data["currency_code"]
                ).exists():
                    id=1
                    max_id=BankExchangeRate.objects.aggregate(Max('id'))
                    if max_id:
                        id=max_id['id__max']+1
                    new_doc = BankExchangeRate.objects.create(
                        id=id,
                        month=data["month"],
                        currency_code=data["currency_code"],
                        exchange_rate=data["exchange_rate"],
                        created_by=data["created_by"],
                        addedDateAndTime=dt.now(),
                    )
                    new_doc.save()

                    addedDateAndTime = dt.now()
                    data["addedDateAndTime"] = str(addedDateAndTime)
                    data["updatedDateAndTime"] = str(addedDateAndTime)
                    data_items = []
                    data_items.append(data)
                    fields = json.dumps(data_items)
                    if len(fields) > 65535:
                        return Response(
                            {
                                "msg": "Length of the value for update_fields property got exceeded!"
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    else:
                        new_doc.updated_fields = fields
                        new_doc.save()
                    results.append(new_doc)
            serializer = BankExchangeRateSerializer(results, many=True)
            dataa = serializer.data
            return Response(dataa, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        doc_object = self.get_object()
        data = request.data
        doc_object.month = data["month"]
        doc_object.currency_code = data["currency_code"]
        doc_object.exchange_rate = data["exchange_rate"]
        doc_object.save()
        serializer = BankExchangeRateSerializer(doc_object)
        dataa = serializer.data
        return Response(dataa)

    def partial_update(self, request, *args, **kwargs):
        try:
            doc_object = self.get_object()
            data = request.data
            doc_object.month = data.get("month", doc_object.month)
            doc_object.currency_code = data.get("currency_code", doc_object.currency_code)
            doc_object.exchange_rate = data.get("exchange_rate", doc_object.exchange_rate)
            doc_object.updated_by = data.get("updated_by", doc_object.updated_by)
            doc_object.updatedDateAndTime = dt.now()
            doc_object.save()

            updatedatetime = dt.now()
            data["updatedDateAndTime"] = str(updatedatetime)
            data_items = []
            if doc_object.updated_fields:
                old_list = doc_object.updated_fields
                old_list = json.loads(old_list)
                data_items.extend(old_list)
            data_items.append(data)
            changedFields = json.dumps(data_items)
            if len(changedFields) > 65535:
                return Response(
                    {"msg": "Length of the value for update_fields property got exceeded!"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            else:
                doc_object.updated_fields = changedFields

            doc_object.save()

            serializer = BankExchangeRateSerializer(doc_object)
            dataa = serializer.data
            return Response(dataa, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        doc_object = self.get_object()
        doc_object.delete()
        return Response({"message": "bank exchange rate deleted successfully"})


class BankInfoViewSet(APIView):

    def get(self, request):
        data = BankDetails.objects.all().order_by('-created_at')
        msd_account_no = request.GET.get("msdAccountNo", None)
        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 20))
        skip = page_number * rows_per_page
        filter_conditions = Q()
        account_number_or_bank_name = request.GET.get("accountNo", None)
        entity_number = request.GET.get("entityNumber", None)
        if msd_account_no:
            filter_conditions &= Q(msd_acct_number__icontains=msd_account_no)

        if entity_number:
            filter_conditions &= Q(entity_number__icontains=entity_number)

        if account_number_or_bank_name:
            search_filter = Q(
                account_number__icontains=account_number_or_bank_name
            ) | Q(bank_name__icontains=account_number_or_bank_name)
            filter_conditions &= search_filter

        # filtered_data = data.filter(filter_conditions)[skip: skip + rows_per_page]  // To be added
        filtered_data = data.filter(filter_conditions)

        count = data.filter(filter_conditions).count()
        serializer = BankDetailsSerializer(filtered_data[skip: skip + rows_per_page], many=True)
        return Response({ "data": serializer.data, "count": count }, status=status.HTTP_200_OK)

    def post(self, request):
        data = request.data
        all_data = []
        failures = {}
        try:
            for j in range(len(data)):
                data[j]["updated_at"] = datetime.datetime.now().isoformat()
                data[j]["created_at"] = datetime.datetime.now().isoformat()
                if(data[j]["account_opening_date"]==""):
                    data[j]["account_opening_date"] = None
                serializer = BankDetailsSerializer(data=data[j])
                if serializer.is_valid():
                    new_doc=serializer.save()
                    data_item = [data[j]]
                    fields = json.dumps(data_item)
                    if len(fields) > 65535:
                        return Response(
                            {
                                "msg": "Length of the value for update_fields property got exceeded!"
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    new_doc.updated_fields = fields
                    new_doc.save()
                    all_data.append(serializer.data)
                else:
                    failures[str(j + 1)] = serializer.errors
            if failures:
                return Response(
                    {"Message": "Data is not inserted for row number {}".format(failures)},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            return Response(all_data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        try:
            instance = BankDetails.objects.get(id=pk)
            request.data["updated_at"]= datetime.datetime.now().isoformat()
            if(("account_opening_date" in request.data) and request.data["account_opening_date"]==""):
                    request.data["account_opening_date"] = None
            serializer = BankDetailsSerializer(instance, data=request.data, partial=True)

            if serializer.is_valid():
                updated_doc=serializer.save()
                data_items=[]
                if instance.updated_fields:
                    old_list=instance.updated_fields
                    old_list=json.loads(old_list)
                    data_items.extend(old_list)
                data_items.append(request.data)
                changedFields = json.dumps(data_items)
                if len(changedFields)>65535:
                    return Response(
                        {
                            "msg": "Length of the value for update_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                else:
                    updated_doc.updated_fields = changedFields
                    updated_doc.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


@csrf_exempt
def getDetailsByBrokerBranchName(request):
    pass
    if request.method == "GET":
        Broker_Branch_Name = request.GET.get("Broker_Branch_Name")
        print(Broker_Branch_Name)
        brokerDetailsObjects = (
            BrokerInformation.objects.filter(branch=Broker_Branch_Name)
            .order_by("id")
            .first()
        )
        serializer = BrokerInformationSerializer(brokerDetailsObjects)
        return JsonResponse(serializer.data)


class PowerBIReportViewSet(viewsets.ModelViewSet):
    serializer_class = PowerBIReportSerializer

    def get_queryset(self):
        return PowerBIReport.objects.filter(active=True)

    def list(self, request):
        try:
            queryset = self.get_queryset()
            if ENVIRONMENT:
                queryset = queryset.filter(environment=ENVIRONMENT).order_by("report_sequence")
            else:
                queryset = queryset.filter(environment="PRODUCTION").order_by("report_sequence")
            serializer = PowerBIReportSerializer(queryset, many=True)
            return Response(serializer.data)
        except Exception as e:
            print("Error while retrieving Power BI reports", e)
            return Response({"error": "Error while retrieving Power BI reports"}, status=500)

# class ROEFileUploadViewSet(viewsets.ModelViewSet):
#     parser_classes = (MultiPartParser, FormParser, JSONParser)

#     def create(self, request):
#         if request.method == "POST":
#             file = request.FILES["file"]
#             if not file:
#                 return Response({"message": "No file provided"}, status=400)

#             file_name = file.name
#             extension = os.path.splitext(file_name)[1].lower()

#             try:
#                 if extension in ['.csv']:
#                     df = pd.read_csv(file)
#                 elif extension in ['.xls', '.xlsx']:
#                     df = pd.read_excel(file)
#                 else:
#                     return Response({"message": "Unsupported file format"}, status=400)
#             except Exception as e:
#                 print(e)
#                 return Response({"message": "Invalid file format or corrupted file"}, status=400)

#             try:
#                 with transaction.atomic():
#                     df.columns = [col.strip().lower().replace(' ', '').replace("_", "") for col in df.columns]
#                     for index, row in df.iterrows():
#                         if pd.isna(row['month']) or type(row['month']) == str or row['month'] == "":
#                             raise Exception("File not uploaded! Please check the month data")
#                         if pd.isna(row['currencycode']) or row['currencycode'] == "" or type(row['currencycode']) != str:
#                             raise Exception("File not uploaded! Please check the currency code data")
#                         if pd.isna(row['exchangerate']) or row['exchangerate'] == "" or row['exchangerate'] == 0 or row['exchangerate'] == "0":
#                             raise Exception("File not uploaded! Please check the exchange rate data")

#                         month = row['month'].strftime('%Y-%m-%d')
#                         currency_code = row['currencycode']

#                         if BankExchangeRate.objects.filter(month=month, currency_code=currency_code).exists():
#                             print("Object already exists-------")
#                             raise Exception(f"File not uploaded! Record already exists with month: {month} and currency code: {currency_code}")
#                         else:
#                             print("Creating new object")
#                             BankExchangeRate.objects.create(
#                                 month=month,
#                                 currency_code=currency_code,
#                                 exchange_rate=row['exchangerate'],
#                                 addedDateAndTime=row['addeddateandtime'],
#                                 created_by=row['createdby'],
#                                 updatedDateAndTime=row['updateddateandtime'],
#                                 updated_by=row['updatedby'],
#                                 updated_fields=row['updatedfields']
#                             )
#             except ValueError as e:
#                 return Response({"message": "File not uploaded! Please check the file data"}, status=400)
#             except Exception as e:
#                 return Response({"message": str(e)}, status=400)
#             return Response({"message": "File uploaded and data stored successfully"})
#         return Response({"message": "File not uploaded"})

class ROEFileUploadViewSet(viewsets.ModelViewSet):
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def create(self, request):
        if request.method == "POST":
            file = request.FILES["file"]
            if not file:
                return Response({"message": "No file provided"}, status=400)

            file_name = file.name
            created_by=request.POST.get('created_by')
            extension = os.path.splitext(file_name)[1].lower()

            try:
                if extension in ['.csv']:
                    df = pd.read_csv(file)
                elif extension in ['.xls', '.xlsx']:
                    df = pd.read_excel(file, sheet_name=0)
                else:
                    return Response({"message": "Unsupported file format"}, status=400)
            except Exception as e:
                print(e)
                return Response({"message": "Invalid file format or corrupted file"}, status=400)

            try:
                day=int(file_name[4:6])
                month=int(file_name[7:9])
                year=int(file_name[10:14])

                formatted_month = date(year, month, day).strftime('%Y-%m-%d')

                current_datetime=dt.now().isoformat()

                df_columns=[col for col in df.columns]

                if df_columns[0].strip()!="month" or df_columns[1].strip()!="Premium Currency" or df_columns[2].strip()!="ROE":
                    return Response({"message": "Invalid file columns"}, status=400)

                with transaction.atomic() :
                    df.columns = [col.strip().lower().replace(' ', '').replace("_", "") for col in df.columns]
                    for index, row in df.iterrows():
                        if pd.isna(row['premiumcurrency']) or type(row['premiumcurrency']) != str or row['premiumcurrency'].strip() == "":
                            raise Exception("File not uploaded! Please check the currency code data")

                        try:
                            if pd.isna(row['roe']) or (type(row['roe'])==str and row["roe"].strip()=="") or float(row["roe"])<=0:
                                raise Exception("File not uploaded! Please check the exchange rate data")
                        except Exception as e:
                            raise Exception("File not uploaded! Please check the exchange rate data")

                        currency_code = row['premiumcurrency'].strip()

                        if BankExchangeRate.objects.filter(month=formatted_month, currency_code=currency_code).exists():
                            print("Object already exists-------")
                            raise Exception(f"File not uploaded! Record already exists with month: {formatted_month} and currency code: {currency_code}")
                        else:
                            id=1
                            max_id=BankExchangeRate.objects.aggregate(Max('id'))
                            if max_id:
                                id=max_id['id__max']+1
                            print("Creating new object")

                            #set updated_fields
                            data_items=[]
                            updated_fields={}

                            updated_fields["month"] = formatted_month
                            updated_fields["currency_code"] = currency_code
                            updated_fields["exchange_rate"] = row["roe"]
                            updated_fields["created_by"] = created_by
                            updated_fields["addedDateAndTime"] = current_datetime
                            updated_fields["updatedDateAndTime"] = current_datetime
                            data_items.append(updated_fields)
                            changedFields=json.dumps(data_items)


                            BankExchangeRate.objects.create(
                                id=id,
                                month=formatted_month,
                                currency_code=currency_code,
                                exchange_rate=row["roe"],
                                addedDateAndTime=current_datetime,
                                created_by=created_by,
                                updatedDateAndTime=current_datetime,
                                updated_fields=changedFields
                            )
                    reuired_data = {"module_name": "ROE", "bucket_name" : config("AWS_STORAGE_BUCKET_NAME")}
                    filemanagment = reusable_file_upload(get_user(request), file, reuired_data, is_upload=False)

                    if not isinstance(filemanagment, Response) or filemanagment.status_code >= 400:
                        # Extract error message from response
                        error_msg = filemanagment.data.get('error', 'Unable to upload the file to S3') if isinstance(filemanagment, Response) else 'Unable to upload the file to S3'
                        raise Exception(error_msg)
            except ValueError as e:
                return Response({"message": "File not uploaded! Please check the file data"}, status=400)
            except Exception as e:
                return Response({"message": str(e)}, status=400)
            return Response({"message": "File uploaded and data stored successfully"})
        return Response({"message": "File not uploaded"})

class SLAViewSet(viewsets.ModelViewSet):
    model = SLA
    serializer_class = SLASerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        doc = SLA.objects.all()
        return doc

    def list(self, request):
        sla = SLA.objects.all().order_by('-addedDateAndTime')
        serializer = SLASerializer(sla, many=True)
        dataa = serializer.data
        return Response(dataa)

    def retrieve(self, request, pk=None):
        if pk:
            doc = SLA.objects.get(id=pk)
            serializer = SLASerializer(doc)
            dataa = serializer.data
            return Response(dataa)

    def create(self, request, *args, **kwargs):
        try:
            data = request.data

            # Add addedDateAndTime to data item
            date=dt.now().isoformat()
            data["addedDateAndTime"] = date
            data["updatedDateAndTime"] = date

            # Add is_active to data item
            data["is_active"] = True

            serializer = SLASerializer(data=data)

            if serializer.is_valid():
                # Retrieve past SLA
                past_sla=SLA.objects.order_by("-id").first()

                # Update past SLA
                if past_sla:
                    past_sla.updatedDateAndTime = date
                    past_sla.updated_by=data.get("created_by", None)
                    past_sla.is_active=False

                    updated_past_sla={}
                    updated_past_sla["is_active"] = False
                    updated_past_sla["updatedDateAndTime"] = date
                    updated_past_sla["updated_by"] = data.get("created_by", None)
                    data_items=[]
                    if past_sla.updated_fields:
                        old_list=past_sla.updated_fields
                        old_list=json.loads(old_list)
                        data_items.extend(old_list)
                    data_items.append(updated_past_sla)
                    changedFields=json.dumps(data_items)
                    if len(changedFields)>65535:
                        return Response(
                        {
                            "msg": "Length of the value for update_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                    past_sla.updated_fields=changedFields
                    past_sla.save()

                new_doc = serializer.save()

                data_item = [data]
                fields = json.dumps(data_item)
                if len(fields) > 65535:
                    return Response(
                        {
                            "msg": "Length of the value for update_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                new_doc.updated_fields = fields
                new_doc.save()

                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        doc_object = self.get_object()
        doc_object.delete()
        return Response({"message": "sla deleted successfully"})

    def partial_update(self, request, *args, **kwargs):
        try:
            doc_object = self.get_object()

            data = request.data

            # set updated values
            doc_object.sla = data.get("sla", doc_object.sla)
            doc_object.updated_by = data.get("updated_by", doc_object.updated_by)
            doc_object.updatedDateAndTime = dt.now()
            doc_object.save()

            # update updated_fields
            updatedatetime = dt.now()
            data["updatedDateAndTime"] = str(updatedatetime)
            data_items = []

            if doc_object.updated_fields:
                old_list = doc_object.updated_fields
                old_list = json.loads(old_list)
                data_items.extend(old_list)
            data_items.append(data)
            changedFields = json.dumps(data_items)

            if len(changedFields) > 65535:
                return Response(
                    {"msg": "Length of the value for update_fields property got exceeded!"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            else:
                doc_object.updated_fields = changedFields

            doc_object.save()
            serializer = SLASerializer(doc_object)
            response_data = serializer.data
            return Response(response_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ParticipatingInsurerViewSet(viewsets.ModelViewSet):
    model = ParticipatingInsurer
    serializer_class = ParticipatingInsurerSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        doc = ParticipatingInsurer.objects.all()
        return doc

    def list(self, request):
        participating_insurers = ParticipatingInsurer.objects.all().order_by('-addedDateAndTime')
        serializer = ParticipatingInsurerSerializer(participating_insurers, many=True)
        dataa = serializer.data
        return Response(dataa)

    def retrieve(self, request, pk=None):
        if pk:
            doc = ParticipatingInsurer.objects.get(id=pk)
            serializer = ParticipatingInsurerSerializer(doc)
            dataa = serializer.data
            return Response(dataa)

    def create(self, request, *args, **kwargs):
        try:
            data = request.data

            # Add addedDateAndTime to data item
            date=dt.now().isoformat()
            data["addedDateAndTime"] = date
            data["updatedDateAndTime"] = date

            serializer = ParticipatingInsurerSerializer(data=data)

            if serializer.is_valid():
                new_doc = serializer.save()
                data_item = [data]
                fields = json.dumps(data_item)
                if len(fields) > 65535:
                    return Response(
                        {
                            "msg": "Length of the value for update_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                new_doc.updated_fields = fields
                new_doc.save()

                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        doc_object = self.get_object()
        doc_object.delete()
        return Response({"message": "Participating insurer deleted successfully"})

    def partial_update(self, request, *args, **kwargs):
        try:
            doc_object = self.get_object()
            data = request.data
            doc_object.participating_insurer = data.get("participating_insurer", doc_object.participating_insurer)
            doc_object.updated_by = data.get("updated_by", doc_object.updated_by)
            doc_object.updatedDateAndTime = dt.now()
            doc_object.save()

            updatedatetime = dt.now()
            data["updatedDateAndTime"] = str(updatedatetime)
            data_items = []
            if doc_object.updated_fields:
                old_list = doc_object.updated_fields
                old_list = json.loads(old_list)
                data_items.extend(old_list)
            data_items.append(data)
            changedFields = json.dumps(data_items)
            if len(changedFields) > 65535:
                return Response(
                    {"msg": "Length of the value for update_fields property got exceeded!"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            else:
                doc_object.updated_fields = changedFields

            doc_object.save()
            serializer = ParticipatingInsurerSerializer(doc_object)
            response_data = serializer.data
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class EscalationViewSet(viewsets.ModelViewSet):
    model = Escalation
    serializer_class = EscalationSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        doc = Escalation.objects.all()
        return doc

    def list(self, request):
        escalation = Escalation.objects.select_related('escalation_level_one', 'escalation_level_two', 'escalation_level_three').all().order_by('-addedDateAndTime')
        serializer = EscalationSerializer(escalation, many=True)
        dataa = serializer.data
        return Response(dataa)

    def retrieve(self, request, pk=None):
        if pk:
            doc = Escalation.objects.get(id=pk)
            serializer = EscalationSerializer(doc)
            dataa = serializer.data
            return Response(dataa)

    def create(self, request, *args, **kwargs):
        try:
            data = request.data
            # Add addedDateAndTime to data item
            date=dt.now().isoformat()
            data["addedDateAndTime"] = date
            data["updatedDateAndTime"] = date

            serializer = EscalationSerializer(data=data)

            if serializer.is_valid():
                new_doc = serializer.save()
                data_item = [data]
                fields = json.dumps(data_item)
                if len(fields) > 65535:
                    return Response(
                        {
                            "msg": "Length of the value for update_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                new_doc.updated_fields = fields
                new_doc.save()

                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        doc_object = self.get_object()
        doc_object.delete()
        return Response({"message": "Escalation deleted successfully"})

    def partial_update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            request_data = request.data
            request_data["updatedDateAndTime"]= datetime.datetime.now().isoformat()
            serializer = EscalationSerializer(instance, data=request_data, partial=True)
            if serializer.is_valid():
                updated_doc=serializer.save()
                data_items=[]
                if instance.updated_fields:
                    old_list=instance.updated_fields
                    old_list=json.loads(old_list)
                    data_items.extend(old_list)
                data_items.append(request.data)
                changedFields = json.dumps(data_items)
                if len(changedFields)>65535:
                    return Response(
                        {"msg": "Length of the value for update_fields property got exceeded!"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                else:
                    updated_doc.updated_fields = changedFields
                    updated_doc.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class DecryptDataAPIView(APIView):
    def post(self, request):
        data = request.data
        broker_id = data.get('id')
        type = data.get('type')  # 'email' or 'phone_number'

        try:
            broker_info = BrokerInformation.objects.get(id=broker_id)
        except BrokerInformation.DoesNotExist:
            return Response({"error": "Broker information with this ID does not exist"}, status=status.HTTP_404_NOT_FOUND)

        if type == 'decrypted_email':
            decrypted_value = broker_info.get_decrypted_email()
        elif type == 'decrypted_phone_number':
            decrypted_value = broker_info.get_decrypted_phone_number()
        else:
            return Response({"error": "Invalid text type provided"}, status=status.HTTP_400_BAD_REQUEST)

        response_data = {type: decrypted_value}
        return Response(response_data, status=status.HTTP_200_OK)

        # updatedatetime = dt.now()
        # data["updatedDateAndTime"] = str(updatedatetime)
        # data_items = []
        # if doc_object.updated_fields:
        #     old_list = doc_object.updated_fields
        #     old_list = json.loads(old_list)
        #     data_items.extend(old_list)
        # data_items.append(data)
        # changedFields = json.dumps(data_items)
        # if len(changedFields) > 65535:
        #     return Response(
        #         {"msg": "Length of the value for update_fields property got exceeded!"}
        #     )
        # else:
        #     doc_object.updated_fields = changedFields

        # doc_object.save()
        # serializer = ParticipatingInsurerSerializer(doc_object)
        # response_data = serializer.data
        # return Response(response_data)

def clean_value(v):
    """
    Helper function to clean and validate individual values
    """
    # Handle various types of null or empty values
    if pd.isna(v) or v is None:
        return None
    if isinstance(v, str):
        # Remove commas from string representations of numbers
        v = v.replace(',', '')
        cleaned = v.lower().strip()
        if cleaned in ['nan', 'nat', '', 'none']:
            return None
        try:
            # Attempt to convert to float if possible
            float_val = float(cleaned)
            return int(float_val) if float_val.is_integer() else f"{float_val:.2f}"
        except ValueError:
            pass
    if isinstance(v, int) and v == 0:
        return 0
    if isinstance(v, float) and math.isnan(v):
        return 0
    # Format float values to two decimal places
    if isinstance(v, float):
        return int(v) if v.is_integer() else f"{v:.2f}"
    return v

def get_user(request):
    """
    Authenticate and retrieve the user from the request token
    """
    user_id = request.headers.get("user-id")
    user = Users.objects.filter(id=user_id)
    return user[0]
    # try:
    #     token = request.META.get('HTTP_AUTHORIZATION', '').split()[1].encode("utf-8")
    #     knoxAuth = TokenAuthentication()
    #     user, auth_token = knoxAuth.authenticate_credentials(token)
    #     return user
    # except (IndexError, AttributeError):
    #     return None

class SiriusDataPagination(PageNumberPagination):
    """
    Custom pagination class for SiriusData.
    """
    page_size_query_param = 'page_size'
    page_size = 10
    max_page_size = 100

class SiriusDataViewSet(viewsets.ModelViewSet):
    queryset = SiriusData.objects.all()
    serializer_class = SiriusDataSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    pagination_class = SiriusDataPagination

    # route for getting data from database based on pagination
    def get(self, request):
        """
        Retrieve paginated SiriusData records.
        """
        # Get all SiriusData records
        data = SiriusData.objects.all()

        # Use custom pagination class
        paginator = SiriusDataPagination()
        paginated_data = paginator.paginate_queryset(data, request)

        # Serialize the paginated data
        serializer = SiriusDataSerializer(paginated_data, many=True)
        print(serializer.data)
        # Return paginated response
        return paginator.get_paginated_response(serializer.data)

    def create(self, request, *args, **kwargs):
        """
        Handle file upload with comprehensive validation
        1. Validate file presence
        2. Validate file extension
        3. Validate required columns
        4. Process data
        """
        try:
            # Validate file presence
            if 'file' not in request.FILES:
                return Response(
                    {'message': 'No file provided in the request'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            #validate file name
            file_name = request.FILES['file'].name
            if not file_name.startswith('sirius_point_'):
                return Response(
                    {'message': 'Invalid file name. The file name must start with sirius_point_'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not re.search(r'\d{2}-\d{2}-\d{4}', file_name):
                return Response(
                    {'message': 'Invalid file name. The file name must contain a date in the format DD-MM-YYYY'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            file = request.FILES['file']
            if SiriusData.objects.filter(file_name=file.name).exists():
                return Response(
                    {'message': 'This file is already uploaded!'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate file extension
            if not file.name.lower().endswith(('.xlsx', '.csv')):
                return Response(
                    {'message': 'Invalid file format. Only .xlsx and .csv files are allowed'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Read file
            try:
                if file.name.lower().endswith('.xlsx'):
                    df = pd.read_excel(file)
                else:
                    df = pd.read_csv(file)
            except Exception as e:
                return Response(
                    {'message': f'Error reading file: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate file contents
            if df.empty:
                return Response(
                    {'message': 'The uploaded file is empty'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate required columns
            required_columns = ['Policy Line Reference', 'UMR', 'Agreement Name', 'XFI Policy Status',
                'Expiry Date', 'Producing Mosaic Entity', 'Service Company',
                'Class of Business', 'Industry', 'Insured Name', 'Insured Domicile',
                'Insured State', 'Master Broker', 'Producing Broker', 'Program',
                'Placing Broker', '100%  Limit ($)', '100% Premium ($)', 'Excess ($)',
                'Deductible ($)', 'Service Company Line %', 'Service Company Limit ($)',
                'Service Company Premium ($)', 'Revenue/ Turnover ($)', 'Partner Name',
                'Brokerage (%)', 'Brokerage ($)', 'Partner %', 'Partner Limit ($)',
                'Partner Premium ($)', 'Sirius Point %', 'UMR.1',
                'ARCH has no participant', 'UMR YOA', 'UMR ', '23_24 UMR Excluded',
                'Policy Line Reference.1', 'Partner %.1'
            ]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response(
                    {'message': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Process data
            with transaction.atomic():
                records = []
                user = get_user(request)

                # Iterate through rows with error handling
                for index, row in df.iterrows():
                    data = {
                        'policy_line_reference': clean_value(row.get('Policy Line Reference')),
                        'umr': clean_value(row.get('UMR')),
                        'agreement_name': clean_value(row.get('Agreement Name')),
                        'xfi_policy_status': clean_value(row.get('XFI Policy Status')),
                        'expiry_date': parse_date(clean_value(row.get('Expiry Date'))),
                        'producing_mosaic_entity': clean_value(row.get('Producing Mosaic Entity')),
                        'service_company': clean_value(row.get('Service Company')),
                        'class_of_business': clean_value(row.get('Class of Business')),
                        'industry': clean_value(row.get('Industry')),
                        'insured_name': clean_value(row.get('Insured Name')),
                        'insured_domicile': clean_value(row.get('Insured Domicile')),
                        'insured_state': clean_value(row.get('Insured State')),
                        'master_broker': clean_value(row.get('Master Broker')),
                        'producing_broker': clean_value(row.get('Producing Broker')),
                        'program': clean_value(row.get('Program')),
                        'placing_broker': clean_value(row.get('Placing Broker')),
                        'limit_100_percent': clean_value(row.get('100%  Limit ($)')),
                        'premium_100_percent': clean_value(row.get('100% Premium ($)')),
                        'excess': clean_value(row.get('Excess ($)')),
                        'deductible': clean_value(row.get('Deductible ($)')),
                        'service_company_line_percent': clean_value(row.get('Service Company Line %')),
                        'service_company_limit': clean_value(row.get('Service Company Limit ($)')),
                        'service_company_premium': clean_value(row.get('Service Company Premium ($)')),
                        'revenue_turnover': clean_value(row.get('Revenue/ Turnover ($)')),
                        'partner_name': clean_value(row.get('Partner Name')),
                        'brokerage_percent': clean_value(row.get('Brokerage (%)')),
                        'brokerage': clean_value(row.get('Brokerage ($)')),
                        'partner_percent': clean_value(row.get('Partner %')),
                        'partner_limit': clean_value(row.get('Partner Limit ($)')),
                        'partner_premium': clean_value(row.get('Partner Premium ($)')),
                        'sirius_point_percent': clean_value(row.get('Sirius Point %')),
                        'umr_1': clean_value(row.get('UMR.1')),
                        'arch_no_participant': clean_value(row.get('ARCH has no participant')),
                        'umr_yoa': clean_value(row.get('UMR YOA')),
                        'umr_2': clean_value(row.get('UMR ')),
                        'umr_23_24': clean_value(row.get('23_24 UMR Excluded')),
                        'policy_line_reference_1': clean_value(row.get('Policy Line Reference.1')),
                        'partner_percent_1': clean_value(row.get('Partner %.1')),
                        'file_name': file.name,
                        'uploaded_by': user.user_name if user else None
                    }
                    record = SiriusData(**{k: v for k, v in data.items() if v is not None})
                    records.append(record)

                SiriusData.objects.bulk_create(records)
                reuired_data = {"module_name": "Sirius Point", "no_of_records": len(df), "bucket_name" : config("AWS_STORAGE_BUCKET_NAME")}
                filemanagment = reusable_file_upload(user, file, reuired_data, is_upload=False)

                if not isinstance(filemanagment, Response) or filemanagment.status_code >= 400:
                    # Extract error message from response
                    error_msg = filemanagment.data.get('error', 'Unable to upload the file to S3') if isinstance(filemanagment, Response) else 'Unable to upload the file to S3'
                    raise Exception(error_msg)

            return Response({
                'message': f'Successfully uploaded {len(records)} records',
                'total_records': len(records)
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            print(e)
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class RBSDetailsPagination(PageNumberPagination):
    """
    Custom pagination class for RBSDetails.
    """
    page_size_query_param = 'page_size'
    page_size = 10
    max_page_size = 100

class RBSDetailsViewSet(viewsets.ModelViewSet):
    queryset = RBSDetails.objects.all()
    serializer_class = RBSDetailsSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    pagination_class = RBSDetailsPagination

    def get(self, request):
        """
        Retrieve paginated RBSDetails records.
        """
        # Get all RBSDetails records
        data = RBSDetails.objects.all()

        # Use custom pagination class
        paginator = RBSDetailsPagination()
        paginated_data = paginator.paginate_queryset(data, request)

        # Serialize the paginated data
        serializer = RBSDetailsSerializer(paginated_data, many=True)

        # Return paginated response
        return paginator.get_paginated_response(serializer.data)

    def create(self, request, *args, **kwargs):
        """
        Handle file upload with comprehensive validation
        1. Validate file presence
        2. Validate file extension
        3. Validate required columns
        4. Process data
        """
        try:
            # Validate file presence
            if 'file' not in request.FILES:
                return Response(
                    {'message': 'No file provided in the request'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            #validate file name
            file_name = request.FILES['file'].name
            if not file_name.startswith('rbs_details_'):
                return Response(
                    {'message': 'Invalid file name. The file name must start with rbs_details_'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not re.search(r'\d{2}-\d{2}-\d{4}', file_name):
                return Response(
                    {'message': 'Invalid file name. The file name must contain a date in the format DD-MM-YYYY'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            file = request.FILES['file']
            if RBSDetails.objects.filter(file_name=file.name).exists():
                return Response(
                    {'message': 'This file is already uploaded!'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # Validate file extension
            if not file.name.lower().endswith(('.xlsx', '.csv')):
                return Response(
                    {'message': 'Invalid file format. Only .xlsx and .csv files are allowed'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Read file
            try:
                if file.name.lower().endswith('.xlsx'):
                    df = pd.read_excel(file)
                else:
                    df = pd.read_csv(file)
            except Exception as e:
                return Response(
                    {'message': f'Error reading file: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            print(df.columns)
            # Validate file contents
            if df.empty:
                return Response(
                    {'message': 'The uploaded file is empty'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate required columns
            required_columns = [
                "Class of Business", "Service Company", "Producing Mosaic Entity", "Inception Year", "Agreement YOA",
                "Inception Date", "FON Date", "Accumulates On Agreement", "Accumulates On Dec", "Carrier Reference",
                "Master Policy Number", "Policy Line Reference", "Unique Market Reference", "Asta UMR (Agreement)",
                "Is Third Party Capacity Deployed?", "Underwriter Name", "Producing Underwriter Name", "Master Broker",
                "Broker Name", "Broker Key", "Insured", "Obligor", "Commodity", "Fund Code as per XFI", "Reinsured",
                "Expiry Date", "Policy Period Days", "Tenor in Month", "Method of Placement", "Mapped MOP", "Territory",
                "Insured Domicile", "Insured State", "Reinsured Domicile", "Industry", "Bank/Non Bank", "NAIC Company Code",
                "NAIC Description", "Slip Lead", "Bureau Lead", "Agency Line/Share %", "Avg Agency Line/Share %",
                "Mosaic 1609 Line %", "Avg Mosaic 1609 Line %", "Mosaic 1609 Order %", "Limit Currency",
                "Limit (Original Currency)", "Type of Layer", "Enterprise Value (Original)", "Enterprise Value (USD)",
                "Excess (Original Currency)", "Excess (USD)", "Deductible (Original Currency)", "Deductible (USD)",
                "100% Gross or Estimated Written Premium (Original Currency)", "Premium Currency",
                "100% Gross or Estimated Written Premium (USD)", "100% Gross or Estimated Written Premium by YOA (USD)",
                "100% Minimum or Deposit Premium (Original Currency)", "Premium Type", "Lloyd’s Risk Code", "Agency GELR",
                "Mosaic GELR", "Brokerage (%)", "Ceding Commission (%)", "Third Party Coverholder Commission (%)",
                "Original Commission (%)", "Mosaic 1609 Agency Commission (%)", "Mosaic 1609 Total Commission (%)",
                "Third Party Agency Commission (%)", "Third Party Total Commission (%)",
                "Agency Share Gross Written Premium (Original Currency)", "Mosaic 1609 Share Gross Written Premium (Original Currency)",
                "Business Plan Loss Ratio (%)", "Mosaic 1609 Share Benchmark Premium (Original Currency)",
                "Mosaic 1609 Share Benchmark Premium (USD)", "Achieved Price (%)",
                "Mosaic 1609 Exposure (Original Currency)", "Mosaic 1609 Share Gross Written Premium (GBP)",
                "Mosaic 1609 Original Commission Amount (GBP)", "Mosaic 1609 Agency Commission Amount (GBP)",
                "Mosaic Expected claims (Original Currency)", "Mosaic Expected claims (USD)", "Mosaic Expected claims (GBP)",
                "Mosaic Expected claims by YOA (GBP)", "Agency Expected claims (Original Currency)",
                "Agency Expected claims (USD)", "Agency Expected claims (GBP)", "Agency Expected claims by YOA (GBP)",
                "Mosaic 1609 Exposure (USD)", "Mosaic 1609 Share Gross Written Premium (USD)",
                "Mosaic 1609 Original Commission Amount (USD)", "Mosaic 1609 Agency Commission Amount (USD)",
                "Agency Share Gross Written Premium (USD)", "Agency Exposure (Original Currency)",
                "Agency Exposure (USD)", "Third Party Agency Commission  (USD)", "Mosaic Brokerage (USD)",
                "Mosaic 1609 Exposure by YOA (USD)", "Mosaic 1609 Share Gross Written Premium by YOA (USD)",
                "Mosaic 1609 Original Commission Amount by YOA (USD)",
                "Mosaic 1609 Agency Commission Amount by YOA (USD)",
                "Agency Share Gross Written Premium by YOA (USD)", "Agency Exposure by YOA (USD)",
                "Third Party Agency Commission by YOA (USD)", "Mosaic Brokerage by YOA (USD)",
                "Mosaic 1609 Share Gross Written Premium by YOA (GBP)",
                "Mosaic 1609 Original Commission Amount by YOA (GBP)",
                "Mosaic 1609 Agency Commission Amount by YOA (GBP)", "Booking Completed Date", "Policy Created By",
                "Date Written", "Peer Reviewer", "Peer Reviewed Date", "Assets Under Management", "Asset Size",
                "Jurisdiction Country", "Jurisdiction State", "Cyber Clause Status", "Clause Code", "Clause Title",
                "Rev/Turnover", "No. of Employees", "Policy Status", "XFI-Policy Level Status", "XFI-Policy Activity Status",
                "XFI-Policy Line Status", "Renewal Status", "Defense Costs Covered", "Onshore/Offshore", "Limit Basis",
                "Renewal Found in XFI", "Renewed Policy Line Reference", "RARC %", "Expired Gross Premium (USD) 100% of Previous Policy",
                "Expired Gross Premium Agency (USD) of previous policy",
                "Mosaic 1609 Expired Gross Premium (USD) of previous policy", "Amount Paid By Insured (USD)", "Net Premium (USD)",
                "Amount Paid By Insured by YOA (USD)", "Net Premium by YOA (USD)"
            ]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response(
                    {'message': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Process data
            with transaction.atomic():
                records = []
                user = get_user(request)
                for _, row in df.iterrows():
                    data = {
                        'class_of_business': clean_value(row.get('Class of Business')),
                        'service_company': clean_value(row.get('Service Company')),
                        'producing_mosaic_entity': clean_value(row.get('Producing Mosaic Entity')),
                        'inception_year': clean_value(row.get('Inception Year')),
                        'agreement_yoa': clean_value(row.get('Agreement YOA')),
                        'inception_date': parse_date(clean_value(row.get('Inception Date'))),
                        'fon_date': parse_date(clean_value(row.get('FON Date'))),
                        'expiry_date': parse_date(clean_value(row.get('Expiry Date'))),
                        'accumulates_on_agreement': clean_value(row.get('Accumulates On Agreement')),
                        'accumulates_on_dec': clean_value(row.get('Accumulates On Dec')),
                        'carrier_reference': clean_value(row.get('Carrier Reference')),
                        'master_policy_number': clean_value(row.get('Master Policy Number')),
                        'policy_line_reference': clean_value(row.get('Policy Line Reference')),
                        'unique_market_reference': clean_value(row.get('Unique Market Reference')),
                        'asta_umr_agreement': clean_value(row.get('Asta UMR (Agreement)')),
                        'is_third_party_capacity_deployed': clean_value(row.get('Is Third Party Capacity Deployed?')),
                        'underwriter_name': clean_value(row.get('Underwriter Name')),
                        'producing_underwriter_name': clean_value(row.get('Producing Underwriter Name')),
                        'master_broker': clean_value(row.get('Master Broker')),
                        'broker_name': clean_value(row.get('Broker Name')),
                        'broker_key': clean_value(row.get('Broker Key')),
                        'insured': clean_value(row.get('Insured')),
                        'obligor': clean_value(row.get('Obligor')),
                        'reinsured': clean_value(row.get('Reinsured')),
                        'commodity': clean_value(row.get('Commodity')),
                        'fund_code_as_per_xfi': clean_value(row.get('Fund Code as per XFI')),
                        'policy_period_days': clean_value(row.get('Policy Period Days')),
                        'tenor_in_month': clean_value(row.get('Tenor in Month')),
                        'method_of_placement': clean_value(row.get('Method of Placement')),
                        'mapped_mop': clean_value(row.get('Mapped MOP')),
                        'territory': clean_value(row.get('Territory')),
                        'insured_domicile': clean_value(row.get('Insured Domicile')),
                        'insured_state': clean_value(row.get('Insured State')),
                        'reinsured_domicile': clean_value(row.get('Reinsured Domicile')),
                        'industry': clean_value(row.get('Industry')),
                        'bank_non_bank': clean_value(row.get('Bank/Non Bank')),
                        'naic_company_code': clean_value(row.get('NAIC Company Code')),
                        'naic_description': clean_value(row.get('NAIC Description')),
                        'slip_lead': clean_value(row.get('Slip Lead')),
                        'bureau_lead': clean_value(row.get('Bureau Lead')),
                        'agency_line_share_pct': clean_value(row.get('Agency Line/Share %')),
                        'avg_agency_line_share_pct': clean_value(row.get('Avg Agency Line/Share %')),
                        'mosaic_1609_line_pct': clean_value(row.get('Mosaic 1609 Line %')),
                        'avg_mosaic_1609_line_pct': clean_value(row.get('Avg Mosaic 1609 Line %')),
                        'mosaic_1609_order_pct': clean_value(row.get('Mosaic 1609 Order %')),
                        'limit_currency': clean_value(row.get('Limit Currency')),
                        'limit_original_currency': clean_value(row.get('Limit (Original Currency)')),
                        'type_of_layer': clean_value(row.get('Type of Layer')),
                        'enterprise_value_original': clean_value(row.get('Enterprise Value (Original)')),
                        'enterprise_value_usd': clean_value(row.get('Enterprise Value (USD)')),
                        'excess_original_currency': clean_value(row.get('Excess (Original Currency)')),
                        'excess_usd': clean_value(row.get('Excess (USD)')),
                        'deductible_original_currency': clean_value(row.get('Deductible (Original Currency)')),
                        'deductible_usd': clean_value(row.get('Deductible (USD)')),
                        'gross_written_premium_original_currency': clean_value(row.get('100% Gross or Estimated Written Premium (Original Currency)')),
                        'premium_currency': clean_value(row.get('Premium Currency')),
                        'gross_written_premium_usd': clean_value(row.get('100% Gross or Estimated Written Premium (USD)')),
                        'gross_written_premium_by_yoa_usd': clean_value(row.get('100% Gross or Estimated Written Premium by YOA (USD)')),
                        'minimum_deposit_premium_original_currency': clean_value(row.get('100% Minimum or Deposit Premium (Original Currency)')),
                        'premium_type': clean_value(row.get('Premium Type')),
                        'lloyds_risk_code': clean_value(row.get('Lloyd\'s Risk Code')),
                        'agency_gelr': clean_value(row.get('Agency GELR')),
                        'mosaic_gelr': clean_value(row.get('Mosaic GELR')),
                        'brokerage_pct': clean_value(row.get('Brokerage (%)')),
                        'ceding_commission_pct': clean_value(row.get('Ceding Commission (%)')),
                        'third_party_coverholder_commission_pct': clean_value(row.get('Third Party Coverholder Commission (%)')),
                        'original_commission_pct': clean_value(row.get('Original Commission (%)')),
                        'mosaic_1609_agency_commission_pct': clean_value(row.get('Mosaic 1609 Agency Commission (%)')),
                        'mosaic_1609_total_commission_pct': clean_value(row.get('Mosaic 1609 Total Commission (%)')),
                        'third_party_agency_commission_pct': clean_value(row.get('Third Party Agency Commission (%)')),
                        'third_party_total_commission_pct': clean_value(row.get('Third Party Total Commission (%)')),
                        'agency_share_gwp_original_currency': clean_value(row.get('Agency Share Gross Written Premium (Original Currency)')),
                        'mosaic_1609_share_gwp_original_currency': clean_value(row.get('Mosaic 1609 Share Gross Written Premium (Original Currency)')),
                        'business_plan_loss_ratio_pct': clean_value(row.get('Business Plan Loss Ratio (%)')),
                        'mosaic_1609_share_benchmark_premium_original_currency': clean_value(row.get('Mosaic 1609 Share Benchmark Premium (Original Currency)')),
                        'mosaic_1609_share_benchmark_premium_usd': clean_value(row.get('Mosaic 1609 Share Benchmark Premium (USD)')),
                        'achieved_price_pct': clean_value(row.get('Achieved Price (%)')),
                        'mosaic_1609_exposure_original_currency': clean_value(row.get('Mosaic 1609 Exposure (Original Currency)')),
                        'mosaic_1609_share_gwp_gbp': clean_value(row.get('Mosaic 1609 Share Gross Written Premium (GBP)')),
                        'mosaic_1609_original_commission_amount_gbp': clean_value(row.get('Mosaic 1609 Original Commission Amount (GBP)')),
                        'mosaic_expected_claims_original_currency': clean_value(row.get('Mosaic Expected claims (Original Currency)')),
                        'mosaic_expected_claims_usd': clean_value(row.get('Mosaic Expected claims (USD)')),
                        'mosaic_expected_claims_gbp': clean_value(row.get('Mosaic Expected claims (GBP)')),
                        'mosaic_expected_claims_by_yoa_gbp': clean_value(row.get('Mosaic Expected claims by YOA (GBP)')),
                        'agency_expected_claims_original_currency': clean_value(row.get('Agency Expected claims (Original Currency)')),
                        'agency_expected_claims_usd': clean_value(row.get('Agency Expected claims (USD)')),
                        'agency_expected_claims_gbp': clean_value(row.get('Agency Expected claims (GBP)')),
                        'agency_expected_claims_by_yoa_gbp': clean_value(row.get('Agency Expected claims by YOA (GBP)')),
                        'mosaic_1609_exposure_usd': clean_value(row.get('Mosaic 1609 Exposure (USD)')),
                        'mosaic_1609_share_gwp_usd': clean_value(row.get('Mosaic 1609 Share Gross Written Premium (USD)')),
                        'mosaic_1609_original_commission_amount_usd': clean_value(row.get('Mosaic 1609 Original Commission Amount (USD)')),
                        'mosaic_1609_agency_commission_amount_usd': clean_value(row.get('Mosaic 1609 Agency Commission Amount (USD)')),
                        'agency_share_gwp_usd': clean_value(row.get('Agency Share Gross Written Premium (USD)')),
                        'agency_exposure_original_currency': clean_value(row.get('Agency Exposure (Original Currency)')),
                        'agency_exposure_usd': clean_value(row.get('Agency Exposure (USD)')),
                        'third_party_agency_commission_usd': clean_value(row.get('Third Party Agency Commission  (USD)')),
                        'mosaic_brokerage_usd': clean_value(row.get('Mosaic Brokerage (USD)')),
                        'mosaic_1609_exposure_by_yoa_usd': clean_value(row.get('Mosaic 1609 Exposure by YOA (USD)')),
                        'mosaic_1609_share_gwp_by_yoa_usd': clean_value(row.get('Mosaic 1609 Share Gross Written Premium by YOA (USD)')),
                        'mosaic_1609_original_commission_amount_by_yoa_usd': clean_value(row.get('Mosaic 1609 Original Commission Amount by YOA (USD)')),
                        'mosaic_1609_agency_commission_amount_by_yoa_usd': clean_value(row.get('Mosaic 1609 Agency Commission Amount by YOA (USD)')),
                        'agency_share_gwp_by_yoa_usd': clean_value(row.get('Agency Share Gross Written Premium by YOA (USD)')),
                        'agency_exposure_by_yoa_usd': clean_value(row.get('Agency Exposure by YOA (USD)')),
                        'third_party_agency_commission_by_yoa_usd': clean_value(row.get('Third Party Agency Commission by YOA (USD)')),
                        'mosaic_brokerage_by_yoa_usd': clean_value(row.get('Mosaic Brokerage by YOA (USD)')),
                        'mosaic_1609_share_gwp_by_yoa_gbp': clean_value(row.get('Mosaic 1609 Share Gross Written Premium by YOA (GBP)')),
                        'mosaic_1609_original_commission_amount_by_yoa_gbp': clean_value(row.get('Mosaic 1609 Original Commission Amount by YOA (GBP)')),
                        'mosaic_1609_agency_commission_amount_by_yoa_gbp': clean_value(row.get('Mosaic 1609 Agency Commission Amount by YOA (GBP)')),
                        'booking_completed_date': parse_date(clean_value(row.get('Booking Completed Date'))),
                        'policy_created_by': clean_value(row.get('Policy Created By')),
                        'date_written': parse_date(clean_value(row.get('Date Written'))),
                        'peer_reviewer': clean_value(row.get('Peer Reviewer')),
                        'peer_reviewed_date': parse_date(clean_value(row.get('Peer Reviewed Date'))),
                        'assets_under_management': clean_value(row.get('Assets Under Management')),
                        'asset_size': clean_value(row.get('Asset Size')),
                        'jurisdiction_country': clean_value(row.get('Jurisdiction Country')),
                        'jurisdiction_state': clean_value(row.get('Jurisdiction State')),
                        'cyber_clause_status': clean_value(row.get('Cyber Clause Status')),
                        'clause_code': clean_value(row.get('Clause Code')),
                        'clause_title': clean_value(row.get('Clause Title')),
                        'rev_turnover': clean_value(row.get('Rev/Turnover')),
                        'no_of_employees': clean_value(row.get('No. of Employees')),
                        'policy_status': clean_value(row.get('Policy Status')),
                        'xfi_policy_level_status': clean_value(row.get('XFI-Policy Level Status')),
                        'xfi_policy_activity_status': clean_value(row.get('XFI-Policy Activity Status')),
                        'xfi_policy_line_status': clean_value(row.get('XFI-Policy Line Status')),
                        'renewal_status': clean_value(row.get('Renewal Status')),
                        'defense_costs_covered': clean_value(row.get('Defense Costs Covered')),
                        'onshore_offshore': clean_value(row.get('Onshore/Offshore')),
                        'limit_basis': clean_value(row.get('Limit Basis')),
                        'renewal_found_in_xfi': clean_value(row.get('Renewal Found in XFI')),
                        'renewed_policy_line_reference': clean_value(row.get('Renewed Policy Line Reference')),
                        'rarc_pct': clean_value(row.get('RARC %')),
                        'expired_gross_premium_usd_100pct': clean_value(row.get('Expired Gross Premium (USD) 100% of Previous Policy')),
                        'expired_gross_premium_agency_usd': clean_value(row.get('Expired Gross Premium Agency (USD) of previous policy')),
                        'mosaic_1609_expired_gross_premium_usd': clean_value(row.get('Mosaic 1609 Expired Gross Premium (USD) of previous policy')),
                        'amount_paid_by_insured_usd': clean_value(row.get('Amount Paid By Insured (USD)')),
                        'net_premium_usd': clean_value(row.get('Net Premium (USD)')),
                        'amount_paid_by_insured_by_yoa_usd': clean_value(row.get('Amount Paid By Insured by YOA (USD)')),
                        'net_premium_by_yoa_usd': clean_value(row.get('Net Premium by YOA (USD)')),
                        'file_name': file.name,
                        'uploaded_by': user.user_name if user else None
                    }
                    record = RBSDetails(**{k: v for k, v in data.items() if v is not None})
                    records.append(record)

                RBSDetails.objects.bulk_create(records)
                reuired_data = {"module_name": "RBS Details", "no_of_records": len(df), "bucket_name" : config("AWS_STORAGE_BUCKET_NAME")}
                filemanagment = reusable_file_upload(user, file, reuired_data, is_upload=False)

                if not isinstance(filemanagment, Response) or filemanagment.status_code >= 400:
                    # Extract error message from response
                    error_msg = filemanagment.data.get('error', 'Unable to upload the file to S3') if isinstance(filemanagment, Response) else 'Unable to upload the file to S3'
                    raise Exception(error_msg)

            return Response({
                'message': f'Successfully uploaded {len(records)} records',
                'total_records': len(records)
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class MOPMappingViewSet(viewsets.ModelViewSet):
    queryset = MOP_mapping.objects.all()
    serializer_class = MOPMappingSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def create(self, request, *args, **kwargs):
        """
        Handle file upload with validation and data processing.
        This allows bulk creation by uploading a file.
        """
        try:
            # Validate file presence
            if 'file' not in request.FILES:
                return Response(
                    {'message': 'No file provided in the request'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            #validate file name
            file_name = request.FILES['file'].name
            if not file_name.startswith('mop_mapping_'):
                return Response(
                    {'message': 'Invalid file name. The file name must start with mop_mapping_'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not re.search(r'\d{2}-\d{2}-\d{4}', file_name):
                return Response(
                    {'message': 'Invalid file name. The file name must contain a date in the format DD-MM-YYYY'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            file = request.FILES['file']
            if MOP_mapping.objects.filter(file_name=file.name).exists():
                return Response(
                    {'message': 'This file is already uploaded!'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # Validate file extension
            if not file.name.lower().endswith(('.xlsx', '.csv')):
                return Response(
                    {'message': 'Invalid file format. Only .xlsx and .csv files are allowed'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Read file
            try:
                if file.name.lower().endswith('.xlsx'):
                    df = pd.read_excel(file)
                else:
                    df = pd.read_csv(file)
            except Exception as e:
                return Response(
                    {'message': f'Error reading file: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # Validate file contents
            if df.empty:
                return Response(
                    {'message': 'The uploaded file is empty'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate required columns
            required_columns = ['Method of Placement','Mapped MOP']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response(
                    {'message': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Process data
            with transaction.atomic():
                records = []
                user = get_user(request)
                for _, row in df.iterrows():
                    data = {
                        'method_of_placement': row.get('Method of Placement'),
                        'mapped_mop': row.get('Mapped MOP'),
                        'file_name': file.name,
                        'uploaded_by': user.user_name if user else None
                    }
                    record = MOP_mapping(**{k: v for k, v in data.items() if v is not None})
                    records.append(record)

                # Bulk create records
                MOP_mapping.objects.bulk_create(records)
                reuired_data = {"module_name": "MOP Mapping", "no_of_records": len(df), "bucket_name" : config("AWS_STORAGE_BUCKET_NAME")}
                user = get_user(request)
                filemanagment = reusable_file_upload(user, file, reuired_data, is_upload=False)

                if not isinstance(filemanagment, Response) or filemanagment.status_code >= 400:
                    # Extract error message from response
                    error_msg = filemanagment.data.get('error', 'Unable to upload the file to S3') if isinstance(filemanagment, Response) else 'Unable to upload the file to S3'
                    raise Exception(error_msg)

            return Response({
                'message': f'Successfully uploaded {len(records)} records',
                'total_records': len(records)
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['POST'])
    def create_record(self, request):
        """
        Create a single record without needing a file upload.
        Expects JSON payload with `method_of_placement` and `mapped_mop`.
        """
        try:
            # Validate incoming data
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                # Save the single record
                serializer.save(
                    uploaded_by=request.user.username if hasattr(request, 'user') else None
                )
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None):
        """
        Partial update for an individual MOPMapping record.
        """
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=True)

            # Validate and save the serializer
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        """
        Handle DELETE request with a custom response instead of the default 204 No Content.
        """
        try:
            # Get the object to delete
            instance = self.get_object()

            # Perform the deletion
            instance.delete()

            # Return a 200 OK with a success message
            return Response(
                {"message": "Record deleted successfully."},
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class AONLedgerPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class AONLedgerViewSet(viewsets.ModelViewSet):
    queryset = AON_Ledger.objects.all()
    serializer_class = AONLedgerSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    pagination_class = AONLedgerPagination

    def get(self, request):
        """
        Retrieve paginated AONLedger records.
        """
        # Get all AONLedger records
        data = AON_Ledger.objects.all()

        # Use custom pagination class
        paginator = AONLedgerPagination()
        paginated_data = paginator.paginate_queryset(data, request)

        # Serialize the paginated data
        serializer = AONLedgerSerializer(paginated_data, many=True)

        # Return paginated response
        return paginator.get_paginated_response(serializer.data)

    def create(self, request, *args, **kwargs):
        """
        Handle file upload via the create method instead of a custom action.
        """
        try:
            # Validate file presence
            if 'file' not in request.FILES:
                return Response(
                    {'message': 'No file provided in the request'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            #validate file name
            file_name = request.FILES['file'].name
            if not file_name.startswith('aon_ledger_'):
                return Response(
                    {'message': 'Invalid file name. The file name must start with aon_ledger_'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not re.search(r'\d{2}-\d{2}-\d{4}', file_name):
                return Response(
                    {'message': 'Invalid file name. The file name must contain a date in the format DD-MM-YYYY'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            file = request.FILES['file']
            if AON_Ledger.objects.filter(file_name=file.name).exists():
                return Response(
                    {'message': 'This file is already uploaded!'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # Validate file extension
            if not file.name.lower().endswith(('.xlsx', '.csv')):
                return Response(
                    {'message': 'Invalid file format. Only .xlsx and .csv files are allowed'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Read file
            try:
                if file.name.lower().endswith('.xlsx'):
                    df = pd.read_excel(file)
                else:
                    df = pd.read_csv(file)
            except Exception as e:
                return Response(
                    {'message': f'Error reading file: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate file contents
            if df.empty:
                return Response(
                    {'message': 'The uploaded file is empty'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate required columns
            required_columns = ['UW BA id', 'Underwriter Reference Number', 'Entry Number',
                'Burea/Non-Bureau', 'Undewriter Name', 'L2 Name', 'Policy N0.',
                'Your Reference Number', 'Assured', 'Client Name', 'Transaction No.',
                'Tran Version', 'Instalment No.', 'Txn Type', 'Trans Desc 1',
                'Trans Desc 2', 'Narrative', 'Entry Date', 'Inception Date',
                'Expiry Date', 'UW Due Date', 'UW Due Age Band', 'PPW Date', 'EBOT',
                'EBOT Status Date', 'PAP', 'Status', 'O. Ccy', 'S. Ccy',
                'Gross Prem Hard Amount', 'Amount Outstanding (Settlement Currency)',
                'IPT, Hard', 'Gross Commission, Hard', 'Amount O/S (Original Currency)',
                'GBP O/S', 'Division', 'Department', 'Team', 'Broker',
                'Processing Technician Name', 'Underwriter A/C Handler',
                'Exposure AgeBand', 'Insured'
            ]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response(
                    {'message': f"Missing required fields: {', '.join(missing_columns)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Process data
            with transaction.atomic():
                user = get_user(request)
                records = []
                for _, row in df.iterrows():
                    data = {
                        'uw_ba_id': clean_value(row.get('UW BA id')),
                        'underwriter_reference_number': clean_value(row.get('Underwriter Reference Number')),
                        'entry_number': clean_value(row.get('Entry Number')),
                        'bureau_non_bureau': clean_value(row.get('Burea/Non-Bureau')),
                        'underwriter_name': clean_value(row.get('Undewriter Name')),
                        'l2_name': clean_value(row.get('L2 Name')),
                        'policy_no': clean_value(row.get('Policy N0.')),
                        'your_reference_number': clean_value(row.get('Your Reference Number')),
                        'assured': clean_value(row.get('Assured')),
                        'client_name': clean_value(row.get('Client Name')),
                        'transaction_no': clean_value(row.get('Transaction No.')),
                        'tran_version': clean_value(row.get('Tran Version')),
                        'instalment_no': clean_value(row.get('Instalment No.')),
                        'txn_type': clean_value(row.get('Txn Type')),
                        'trans_desc_1': clean_value(row.get('Trans Desc 1')),
                        'trans_desc_2': clean_value(row.get('Trans Desc 2')),
                        'narrative': clean_value(row.get('Narrative')),
                        'entry_date': parse_date(clean_value(row.get('Entry Date'))),
                        'inception_date': parse_date(clean_value(row.get('Inception Date'))),
                        'expiry_date': parse_date(clean_value(row.get('Expiry Date'))),
                        'uw_due_date': parse_date(clean_value(row.get('UW Due Date'))),
                        'uw_due_age_band': clean_value(row.get('UW Due Age Band')),
                        'ppw_date': parse_date(clean_value(row.get('PPW Date'))),
                        'ebot': clean_value(row.get('EBOT')),
                        'ebot_status_date': parse_date(clean_value(row.get('EBOT Status Date'))),
                        'pap': clean_value(row.get('PAP')),
                        'status': clean_value(row.get('Status')),
                        'original_currency': clean_value(row.get('Original Currency')),
                        'settlement_currency': clean_value(row.get('Settlement Currency')),
                        'gross_prem_hard_amount': clean_value(row.get('Gross Prem Hard Amount')),
                        'amount_outstanding_settlement_currency': clean_value(row.get('Amount Outstanding (Settlement Currency)')),
                        'ipt_hard': clean_value(row.get('IPT, Hard')),
                        'gross_commission_hard': clean_value(row.get('Gross Commission, Hard')),
                        'amount_os_original_currency': clean_value(row.get('Amount OS Original Currency')),
                        'gbp_os': clean_value(row.get('GBP OS')),
                        'division': clean_value(row.get('Division')),
                        'department': clean_value(row.get('Department')),
                        'team': clean_value(row.get('Team')),
                        'broker': clean_value(row.get('Broker')),
                        'processing_technician_name': clean_value(row.get('Processing Technician Name')),
                        'underwriter_ac_handler': clean_value(row.get('Underwriter A/C Handler')),
                        'exposure_ageband': clean_value(row.get('Exposure AgeBand')),
                        'insured': clean_value(row.get('Insured')),
                        'file_name': file.name,
                        'uploaded_by': user.user_name if user else None
                    }
                    record = AON_Ledger(**{k: v for k, v in data.items() if v is not None})
                    records.append(record)

                AON_Ledger.objects.bulk_create(records)
                reuired_data = {"module_name": "AON Ledger", "no_of_records": len(df), "bucket_name" : config("AWS_STORAGE_BUCKET_NAME")}
                filemanagment = reusable_file_upload(user, file, reuired_data, is_upload=False)

                if not isinstance(filemanagment, Response) or filemanagment.status_code >= 400:
                    # Extract error message from response
                    error_msg = filemanagment.data.get('error', 'Unable to upload the file to S3') if isinstance(filemanagment, Response) else 'Unable to upload the file to S3'
                    raise Exception(error_msg)

            return Response({
                'message': f'Successfully uploaded {len(records)} records',
                'total_records': len(records)
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class TxnStatusViewset(APIView):

    def get(self, request):
        data = TransactionStatus.objects.all().order_by('-id')
        serializer = TransactionStatusSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        data = request.data
        print(data)
        max_id = TransactionStatus.objects.aggregate(Max('id'))['id__max'] or 0
        data['id'] = max_id + 1
        serializer = TransactionStatusSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def patch(self, request, pk):
        instance = TransactionStatus.objects.get(id=pk)
        serializer = TransactionStatusSerializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)


class FileRecordViewSet(viewsets.ModelViewSet):
    queryset = AgedDeptFileRecord.objects.all()
    serializer_class = AgedDeptFileRecordSerializer

    def get_queryset(self):
        queryset = AgedDeptFileRecord.objects.all()
        queryset = queryset.filter(archived=False)

        # Get filter parameters from the request
        id = self.request.query_params.get('id', None)
        month = self.request.query_params.get('month', None)
        year = self.request.query_params.get('year', None)

        if id:
            queryset = queryset.filter(id=id)
        if month:
            queryset = queryset.filter(month=month)
        if year:
            queryset = queryset.filter(year=year)

        return queryset.order_by('-year', 'month', 'id')

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        page_number = int(request.query_params.get("skip", 0))
        page_size = int(request.query_params.get("pageSize", 20))
        skip = page_number * page_size

        # Apply pagination
        paginated_queryset = queryset.order_by('-id')[skip: skip + page_size]

        serializer = self.get_serializer(paginated_queryset, many=True)
        return Response({
            'data': serializer.data,
            'count': queryset.count()
        }, status=status.HTTP_200_OK)

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.archived = True
        instance.save()
        logger.debug(f"Archived aged file: {instance.id}")
        return Response({"message": "aged file archived successfully"}, status=status.HTTP_200_OK)

def get_calulation_files_list(request):
    # Get all file names from each model
    AON_Ledger_file_names = set()  # Use a set to avoid duplicates
    MOP_mapping_file_names = set()  # Use a set to avoid duplicates
    RBSDetails_file_names = set()  # Use a set to avoid duplicates
    SiriusData_file_names = set()  # Use a set to avoid duplicates

    # Query each model and add file names to the set
    AON_Ledger_file_names.update(AON_Ledger.objects.values_list('file_name', flat=True))
    MOP_mapping_file_names.update(MOP_mapping.objects.values_list('file_name', flat=True))
    RBSDetails_file_names.update(RBSDetails.objects.values_list('file_name', flat=True))
    SiriusData_file_names.update(SiriusData.objects.values_list('file_name', flat=True))

    # Return the response as JSON
    return JsonResponse({
        'AON_Ledger_files': list(AON_Ledger_file_names),
        'MOP_mapping_files': list(MOP_mapping_file_names),
        'RBS_Details_files': list(RBSDetails_file_names),
        'Sirius_Data_files': list(SiriusData_file_names)
        }, status=status.HTTP_200_OK)
def get_agedept_details(request):
    try:
        # Get the file_name, skip, and pageSize from query parameters
        file_name = request.GET.get('file_name', None)
        skip = int(request.GET.get('skip', 0))  # Default to 0 if not provided
        page_size = int(request.GET.get('pageSize', 25))  # Default to 25 if not provided

        # Filter the policies by file_name
        policies_queryset = PolicyInformation.objects.filter(file_name=file_name)

        # Pagination logic: slicing the queryset based on skip and pageSize
        policies_paginated = policies_queryset[skip:skip + page_size]

        # Get the total count of policies matching the file_name (for pagination purposes)
        total_count = policies_queryset.count()

        # Convert the queryset to a list (so it can be returned in the response)
        policy_details = list(policies_paginated.values())

        return JsonResponse({
            "count": total_count,
            "data": policy_details
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error fetching policy details: {str(e)}")
        return JsonResponse({"error": "Error fetching policy details"}, status=status.HTTP_500_INTERNAL_SERVER)
def get_aged_dept_years(request):
    try:
        # Query for distinct years and their corresponding months
        year_month_data = AgedDeptFileRecord.objects.values('year', 'month').distinct().order_by('year', 'month')

        # Create a dictionary to hold year and months
        result = {}

        # Iterate over the data and group months by year
        for entry in year_month_data:
            year = entry['year']
            month = entry['month']

            # If the year is not in the dictionary, add it
            if year not in result:
                result[year] = []

            # Add the month to the list of months for that year
            result[year].append(month)

        # Convert the dictionary to a list of objects as per your desired format
        response_data = [{"year": year, "months": sorted(months)} for year, months in result.items()]

        return JsonResponse(response_data, safe=False)

    except Exception as e:
        logger.error(f"Error retrieving year and month data: {e}")
        return JsonResponse({"error": "Failed to retrieve year and month data"}, status=500)





# API to return data as JSON
def policy_data_json(request):
    try:
        file_name = request.GET.get('file_name', None)
        file_year = request.GET.get('file_year', None)
        file_month = request.GET.get('file_month', None)

        # Filter the data based on the query parameters
        policies = PolicyInformation.objects.all()

        if file_name:
            policies = policies.filter(file_name=file_name)
        if file_year:
            policies = policies.filter(file_year=file_year)
        if file_month:
            policies = policies.filter(file_month=file_month)

        # Convert the queryset to a list of dictionaries
        policy_list = list(policies.values())

        # Log the successful retrieval of data
        logger.info(f"Retrieved {len(policy_list)} policies for file_name={file_name}, file_year={file_year}, file_month={file_month}")

        # Return data as JSON
        return JsonResponse(policy_list, safe=False)

    except Exception as e:
        logger.error(f"Error retrieving policy data: {e}")
        return JsonResponse({"error": "Failed to retrieve policy data"}, status=500)

# Export the data for Policy Information.
def policy_data_excel(request):
    try:
        file_name = request.GET.get('file_name', None)
        file_year = request.GET.get('file_year', None)
        file_month = request.GET.get('file_month', None)

        # Filter the data based on the query parameters
        policies = PolicyInformation.objects.all()

        if file_name:
            policies = policies.filter(file_name=file_name)
        if file_year:
            policies = policies.filter(file_year=file_year)
        if file_month:
            policies = policies.filter(file_month=file_month)

        # Convert the queryset to a pandas DataFrame
        policy_df = pd.DataFrame(list(policies.values()))
        print("plocy record--> ",len(policy_df))

        # If no data is found, return a message
        if policy_df.empty:
            logger.warning("The policy data DataFrame is empty.")
            return JsonResponse({"error": "No data available for the given filters."}, status=404)

        # Reorder columns according to specified sequence
        columns = [
            "Producing_Entity", "Class_of_Business", "file_month", "file_year", "Year_of_Account", "Syndicate_Binder",
            "Policy_Line_Ref", "Policy_Status", "Policy_Activity_Status", "Inception_Date",
            "Expired_Date", "Date_Cancelled", "Cancellation_Reason", "Transaction_Status",
            "UMR_Number", "Three_Party_Capacity_Deployed", "SCM_Partner", "Signed_Line_Pct",
            "Broker_Order_Pct", "Signed_Order_Pct", "Broker_Commision_Pct", "Coverholder_Commision_Pct",
            "Broker_Reference", "Underwriter", "MOP", "Broker", "Master_Broker", "Insured",
            "Summary_Currency", "Summary_ROE", "Settlement_Ccy", "Settlement_ROE",
            "Gross_Written_Premium_100_in_Sett", "Net_Written_Premium_100_in_Sett",
            "True_Net_Written_Premium_100_in_Sett", "Original_Ccy", "Original_ROE",
            "Gross_Written_Premium_100_in_Orig", "Gross_Written_Premium_Agency_Share_in_Orig",
            "Gross_Written_Premium_Syndicate_Share_in_Orig", "Gross_Written_Premium_Non_Syndicate_Share_in_Orig",
            "Net_Written_Premium_100_in_Orig", "Net_Written_Premium_Agency_Share_in_Orig",
            "Net_Written_Premium_Syndicate_Share_in_Orig", "Net_Written_Premium_Non_Syndicate_Share_in_Orig",
            "True_Net_Written_Premium_100_in_Orig", "True_Net_Written_Premium_Syndicate_Share_in_Orig",
            "True_Net_Written_Premium_Agency_Share_in_Orig", "True_Net_Written_Premium_Non_Syndicate_Share_in_Orig",
            "Gross_Written_Premium_Syndicate_Share_in_USD", "Gross_Written_Premium_Agency_Share_in_USD",
            "Gross_Written_Premium_100_in_USD", "Net_Written_Premium_100_in_USD",
            "True_Net_Written_Premium_100_in_USD", "PremiumBasis", "Instalment_Nbr",
            "Installment_Category", "Installment_Due_date", "Installment_Ccy_in_Orig",
            "Installment_Agency_Amount_in_Orig", "Installment_Agency_Amount_in_Sett",
            "Installment_Agency_Amount_in_USD", "Installment_Amount_Syndicate_Share_in_Orig",
            "Installment_Amount_Syndicate_Share_in_Sett", "Installment_Amount_Syndicate_Share_in_USD",
            "Paid_Amount_in_USD", "Last_Allocation_Date", "Diff_in_USD", "Overdue_Days",
            "Overdue_Category", "Class_of_Business_Remapped", "Facility", "SP_PER", "MOP_Mapped",
            "Agency_Commission", "Brokerage_Installment_Sett", "Agency_Commission_USD",
            "Sirius_Point_Amount_GWP_USD", "Archre_Amount_GWP_USD", "ArchRe_Amount_Received",
            "ArchRe_Outstanding", "Commission", "Gross_Written_Premium_100_USD_Agency_DUA_Earned",
            "Gross_Written_Premium_100_USD_Syndicate", "Gross_Written_Premium_100_USD_SCM",
            "Net_Written_Premium_100_USD_Agency", "Net_Written_Premium_100_USD_Syndicate",
            "Net_Written_Premium_100_USD_SCM", "CT_Receivable_Total_Agency_Sett",
            "CT_Receivable_Total_Syndicate", "CT_Receivable_Total_SCM", "CT_Allcoated_Total_Agency",
            "CT_Rcvd_VS_Inst_Amt_Agcy_Unalloc_To_Clt", "CT_Rcvd_VS_Inst_Amt_Agcy_Unalloc_To_Clt_USD",
            "Money_To_Collect_Syndicate", "Money_To_Collect_USD_SCM",
            "Future_Due_45_Days_From_Reporting_Period", "CT_Rcvd_vs_Instalment_Syndicate",
            "CT_Rcvd_vs_Instalment_SCM", "USM", "CT_Unallocated", "CT_Unallocated_USD",
            "CT_Unallocated_USD_Syndicate", "CT_Unallocated_USD_SCM", "Brokerage_USD",
            "Agency_Commission_USD2", "Aged_Bucket_By_Period_Receivable", "AON_Collection_Status",
            "Aged_Bucket_By_Period_All", "Policy_Version", "Last_Allocation_Date_CT",
            "Progress_Status", "CC", "Status_update_date", "Comments",
            "CT_Receivable_Total_Agency_USD_Gross", "ARCH_Status", "Sum_of_Inst_NWP_USD_values_25",
            "No_Funds_Receive", "SP_Percent_0_Not_on_SP_File", "Payment_Received",
            "Pending_Client_Payment", "Comments_to_follow_up", "Claims",
            "Original_Cur_vs_Settlement", "Currency_Test", "exception"
        ]

        # Reorder columns and handle missing columns gracefully
        existing_columns = [col for col in columns if col in policy_df.columns]
        policy_df = policy_df[existing_columns]
        # Identify datetime columns
        datetime_columns = policy_df.select_dtypes(include=['datetime64[ns, UTC]']).columns

        # Remove timezone information from datetime columns
        for column in datetime_columns:
            policy_df[column] = policy_df[column].dt.tz_localize(None)

        # Convert the DataFrame to an Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="policy_data.xlsx"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            policy_df.to_excel(writer, index=False, sheet_name="Policy Data")

            workbook = writer.book
            worksheet = workbook['Policy Data']

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(start_color='ffc619', end_color='ffc619', fill_type='solid')

            # Apply the fill to the header row (assumes the header is in the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

        # Log the successful generation of the Excel file
        logger.info(f"Excel file generated successfully for {len(policy_df)} policies.")

        return response

    except Exception as e:
        logger.error(f"Error generating Excel file: {e}")
        return JsonResponse({"error": "Failed to generate Excel file"}, status=500)

# def get_user(request):
#     token = request.META.get('HTTP_AUTHORIZATION', False)
#     if token:
#         token = str(token).split()[1].encode("utf-8")
#         knoxAuth = TokenAuthentication()
#         user, auth_token = knoxAuth.authenticate_credentials(token)
#         request.user = user
#         return user

class TriggerPolicyCalculationView(APIView):
    """
    API endpoint to trigger policy calculations as a background task
    """
    def post(self, request):
        # Get required parameters from request
        file_name = request.data.get('file_name')
        file_month = request.data.get('file_month')
        file_year = request.data.get('file_year')
        record_id = request.data.get('id')
        is_rerun = request.data.get('is_rerun')

        AONLedger = request.data.get('AONLedger')
        RBSDetails = request.data.get('RBSDetails')
        SiriusPoints = request.data.get('SiriusPoints')
        MOPMapping = request.data.get('MOPMapping')


        print(file_name, file_month, file_year, record_id)
        print("File : ", AONLedger, RBSDetails, SiriusPoints, MOPMapping)
        calculation_data = {
            'AONLedger': AONLedger,
            'RBSDetails': RBSDetails,
            'SiriusPoints': SiriusPoints,
            'MOPMapping': MOPMapping,
        }

        try:
            try:
                # Retrieve and update the AgedDeptFileRecord with processing status
                record = AgedDeptFileRecord.objects.get(pk=record_id)
                record.status = 'Processing'  # Update status to 'PROCESSING'
                record.last_run_time = timezone.now()
                record.last_run_by = get_user(request).user_name
                record.calculation_files = calculation_data
                record.save()
            except Exception as e:
                return Response({
                    'status': 'error',
                    'message': f'Record with id {record_id} not found.'
                }, status=status.HTTP_404_NOT_FOUND)

            # Schedule the task with parameters using Celery
            # file_month and file_year are kept for backward compatibility but not used in the function
            task = update_policy_calculations.apply_async(
                args=[file_name, file_month, file_year, record_id, AONLedger, RBSDetails, SiriusPoints, MOPMapping, is_rerun]
            )

            return Response({
                'status': 'success',
                'message': 'Policy calculation task has been scheduled',
                'task_id': task.id,
                'parameters': {
                    'file_name': file_name,
                    'file_month': file_month,
                    'file_year': file_year,
                    'record_id': record_id
                }
            }, status=status.HTTP_202_ACCEPTED)

        except Exception as e:
            record = AgedDeptFileRecord.objects.get(pk=record_id)
            if record:
                record.status = 'Failed'
                record.error_message = str(e)
                record.save()
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AgedDebtDueViewSet(viewsets.ModelViewSet):
    queryset = AgeDebtAllocations.objects.all().order_by('-id')
    serializer_class = AgedebtSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_date(self, date_str):
        format = "%d/%m/%Y"
        # Convert from dd-mm-yyyy to yyyy-mm-dd
        converted_date = dt.strptime(date_str, format).strftime("%Y-%m-%d")
        return converted_date

    def list(self, request):
        """
        Retrieve paginated AgedDebtDue records.
        """
        page_number = int(request.GET.get("skip", 0))
        rows_per_page = int(request.GET.get("pageSize", 25))
        skip = page_number * rows_per_page
        status_list = ['0-3 Months', '3-6 Months', '6-12 Months', '12-24 Months', '24+ Months', 'Not Yet Due']

        due = request.GET.get("due", None)
        if due=="1":
            self.queryset = self.queryset.filter(status_usd__in=status_list)

        broker = request.GET.get("Broker", None)
        if broker:
            self.queryset = self.queryset.filter(broker_branch__in=broker.split(","))

        lob = request.GET.get("lob", None)
        if lob:
            self.queryset = self.queryset.filter(policy__regex=f'^.[{lob}]')

        yoa = request.GET.get("yoa", None)
        if yoa:
            self.queryset = self.queryset.filter(yoa__in=yoa.split(","))

        policy_ref = request.GET.get("policy_ref", None)
        if policy_ref:
            self.queryset = self.queryset.filter(policy=policy_ref)

        mop_mapped = request.GET.get("mop_mapped", None)
        if mop_mapped:
            self.queryset = self.queryset.filter(mop__in=mop_mapped.split(","))

        market_source = request.GET.get("market_source", None)
        if market_source:
            self.queryset = self.queryset.filter(market_source__in=market_source.split(","))

        broker_reference = request.GET.get("broker_reference", None)
        if broker_reference:
            self.queryset = self.queryset.filter(broker_reference__in=broker_reference.split(","))

        insured = request.GET.get("insured", None)
        if insured:
            self.queryset = self.queryset.filter(insured__in=insured.split(","))

        umr = request.GET.get("umr", None)
        if umr:
            self.queryset = self.queryset.filter(umr__in=umr.split(","))

        file_name = request.GET.get("file_name", None)
        if file_name:
            self.queryset = self.queryset.filter(file_name=file_name)
        else:
            self.queryset = self.queryset.filter(file_name=PolicyInformation.objects.filter(archived=False, file_name__isnull=False).exclude(file_name__exact='').last().file_name)

        entity = request.GET.get("entity", None)
        if entity:
            self.queryset = self.queryset.filter(producing_entity__in=entity.split(","))
        
        underwriter = request.GET.get("underwriter", None)
        if underwriter:
            self.queryset = self.queryset.filter(underwriter__in=underwriter.split(";"))

        status_sett = request.GET.get("status_sett", None)
        if status_sett:
            self.queryset = self.queryset.filter(status_sett__in=status_sett.split(","))

        status_usd = request.GET.get("status_usd", None)
        if status_usd:
            self.queryset = self.queryset.filter(status_usd__in=status_usd.split(","))

        from_inception_date = request.GET.get("fromInceptionDate", None)
        to_inception_date = request.GET.get("toInceptionDate", None)
        if from_inception_date:
            from_inception_date_value = self.get_date(from_inception_date)
        else:
            from_inception_date_value = None
        if to_inception_date:
            to_inception_date_value = self.get_date(to_inception_date)
        else:
            to_inception_date_value = None

        if from_inception_date_value and to_inception_date_value:
            self.queryset = self.queryset.filter(inception_date__range=[from_inception_date_value, to_inception_date_value])

        from_due_date = request.GET.get("fromDueDate", None)
        to_due_date = request.GET.get("toDueDate", None)
        if from_due_date:
            from_due_date_value = self.get_date(from_due_date)
        else:
            from_due_date_value = None
        if to_due_date:
            to_due_date_value = self.get_date(to_due_date)
        else:
            to_due_date_value = None

        if from_due_date_value and to_due_date_value:
            self.queryset = self.queryset.filter(installment_due_date__range=[from_due_date_value, to_due_date_value])

        queryset_data = self.queryset.order_by('-id')

        if request.GET.get("download"):
            serializer = AgedebtDownloadSerializer(queryset_data, many=True)
            df = pd.DataFrame(serializer.data)

            columns = [
                "yoa",
                "umr",
                "cob",
                "policy",
                "policy_status",
                "market_source",
                "original_ccy",
                "settlement_ccy",
                "broker_reference",
                "mop",
                "broker_branch",
                "binding_agreement",
                "installment_number",
                "installment_due_date",
                "insured",
                "installment_amount_sett",
                "total_receivable_sett",
                "ct_unallocated",
                "balance_after_subtraction_sett",
                "producing_entity",
                "underwriter",
                "status_sett",
                "status_usd",
                "expired_policy_status",
                "partial_payment_status",

                # Not present on powerBI report 
                "master_broker",
                "transaction_status",
                "inception_date",
                "expired_date",
                "total_allocated_sett",
                "installment_amount_usd",
                "total_receivable_usd",
                "total_allocated_usd",
                "ct_unallocated_usd",
                "balance_after_subtraction_usd",
                "bank_name",
                "bank_account_number",
                
                "cc_comments",
                "action",
                "category",
                "underwriter_comments"
            ]

            df = df[[c for c in columns if c in df.columns]]

            # Map old field names to new header names
            header_rename = {
                "yoa": "YOA",
                "umr": "UMR", 
                "cob": "LOB",
                "policy": "Policy",
                "policy_status": "Policy Status",
                "market_source": "Master Source",
                "original_ccy": "Original Currency",
                "settlement_ccy": "Settlement Currency",
                "broker_reference": "Broker Reference",
                "mop": "MOP",
                "broker_branch": "Broker Branch",
                "binding_agreement": "Binding Agreement",
                "installment_number": "Installment Number",
                "installment_due_date": "Installment Due Date",
                "insured": "Insured",
                "installment_amount_sett": "Installment Amount (Sett)",
                "total_receivable_sett": "Total Receiveable Amt (sett)",
                "ct_unallocated": "Unallocated Amt Sett", 
                "balance_after_subtraction_sett": "Uncollected Amt Sett", 
                "producing_entity": "Producing Entity",
                "underwriter": "Underwriter",
                "status_sett": "Status Sett", 
                "status_usd": "Sub Status Sett",
                "expired_policy_status" : "Expired policy status",
                "partial_payment_status" : "Partial payment status",

                # Not present on powerBI report 
                "master_broker": "Master Broker",
                "transaction_status": "Transaction Status",
                "inception_date": "Inception Date",
                "expired_date": "Expiry Date",
                "total_allocated_sett": "Allocated Amount Sett",
                "installment_amount_usd": "Total Premium Due USD",
                "total_receivable_usd": "Total Cash Received USD",
                "total_allocated_usd": "Allocated Amount USD",
                "ct_unallocated_usd": "Cash Unallocated USD",
                "balance_after_subtraction_usd": "Uncollected Premium USD",
                "bank_name": "Bank Name",
                "bank_account_number": "Bank Account Number",

                "cc_comments": "CC Comments",
                "action": "Action Required",
                "category": "Category",
                "underwriter_comments": "Underwriter Comments"
            }

            df = df.rename(columns=header_rename)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="aged_report.xlsx"'

            with pd.ExcelWriter(response, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Aged Debt Report')
                
                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Aged Debt Report']
                
                # Define decimal fields that need number formatting
                decimal_fields = [
                    'ct_unallocated_usd', 'ct_unallocated', 'installment_amount_usd', 
                    'total_receivable_usd', 'balance_after_subtraction_usd',
                    'installment_amount_sett', 'total_receivable_sett', 
                    'balance_after_subtraction_sett', 'brokerage_usd', 'brokerage_sett',
                    'commission_usd', 'commission_sett', 'gross_written_premium_100_sett',
                    'gross_written_premium_100_usd', 'net_written_premium_100_sett',
                    'net_written_premium_100_usd', 'total_allocated_usd',
                    'ct_unallocated_usd', 'total_allocated_sett',
                    'ct_unallocated'
                ]
                
                # Apply number formatting to decimal columns
                for col_name in decimal_fields:
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1  # Excel is 1-indexed
                        # Format as number with 2 decimal places
                        for row in range(2, len(df) + 2):  # Start from row 2 (skip header)
                            cell = worksheet.cell(row=row, column=col_idx)
                            if cell.value is not None:
                                try:
                                    cell.number_format = '#,##0.00'
                                except:
                                    pass

            return response

        serializer = self.serializer_class(queryset_data[skip: skip + rows_per_page], many=True)
        dataa = serializer.data
        
        # Optimize: Single query to get both fields, then process them
        distinct_values = PolicyInformation.objects.filter(archived=False).values_list("market_source", "Broker_Reference", "Insured", "UMR_Number", "Producing_Entity", "Underwriter", flat=False).distinct()
        aged_allocations_distinct_values = AgeDebtAllocations.objects.all().values_list("status_sett", "status_usd", "mop", flat=False).distinct()
        
        # Separate and filter out null values
        market_source = list(set(filter(None, [item[0] for item in distinct_values])))
        broker_reference = list(set(filter(None, [item[1] for item in distinct_values])))
        insured = list(set(filter(None, [item[2] for item in distinct_values])))
        umr = list(set(filter(None, [item[3] for item in distinct_values])))
        entity = list(set(filter(None, [item[4] for item in distinct_values])))
        underwriter = list(set(filter(None, [item[5] for item in distinct_values])))
        status_sett = list(set(filter(None, [item[0] for item in aged_allocations_distinct_values])))
        status_usd = list(set(filter(None, [item[1] for item in aged_allocations_distinct_values])))
        mop = list(set(filter(None, [item[2] for item in aged_allocations_distinct_values])))
        file_name = PolicyInformation.objects.filter(file_name__isnull=False) \
                    .values("file_name") \
                    .annotate(latest_id=Max("id")) \
                    .order_by("-latest_id") \
                    .values_list("file_name", flat=True)

        summary = queryset_data.aggregate(
            total_prem_due_sett=Sum('installment_amount_sett'),
            total_prem_due_usd=Sum('installment_amount_usd'),
            total_cash_rec_sett=Sum('total_receivable_sett'),
            total_cash_rec_usd=Sum('total_receivable_usd'),
            total_alloc_amt_sett=Sum('total_allocated_sett'),
            total_alloc_amt_usd=Sum('total_allocated_usd'),
            total_cash_unalloc_sett=Sum('ct_unallocated'),
            total_cash_unalloc_usd=Sum('ct_unallocated_usd'),
            total_uncollected_prem_sett=Sum('balance_after_subtraction_sett'),
            total_uncollected_prem_usd=Sum('balance_after_subtraction_usd'),

            # 🔹 SETT – Not yet due
            total_uncollected_notyet_due_sett=Sum(
                'balance_after_subtraction_sett',
                filter=Q(status_sett='Not Yet Due')
            ),
            # 🔹 USD – Not yet due
            total_uncollected_notyet_due_usd=Sum(
                'balance_after_subtraction_usd',
                filter=Q(status_usd='Not Yet Due')
            ),
            # 🔹 SETT – Overdue (everything except "Not yet due")
            total_uncollected_overdue_sett=Sum(
                'balance_after_subtraction_sett',
                filter=~Q(status_sett='Not Yet Due')
            ),
            # 🔹 USD – Overdue
            total_uncollected_overdue_usd=Sum(
                'balance_after_subtraction_usd',
                filter=~Q(status_usd='Not Yet Due')
            ),
        )


        data = {
            "count": queryset_data.count(),
            "data": dataa,
            "market_source": market_source,
            "broker_reference": broker_reference,
            "insured": insured,
            "umr": umr,
            "entity": entity,
            "underwriter": underwriter,
            "status_sett": status_sett,
            "status_usd": status_usd,
            "mop_mapped": mop,
            "file_name": file_name,
            "summary": {k: round(v or 0.00, 2) for k, v in summary.items()}
        }
        return Response(data)


@csrf_exempt
def LOBFilterViewSet(request):
    if request.method == "GET":
        data = list(LOB.objects.all().order_by('-id').values_list('lob_code', flat=True).distinct())
        return JsonResponse({"data": data})


@csrf_exempt
def YOAPolicyFilterViewSet(request):
    if request.method == "GET":
        data = list(PolicyInformation.objects.exclude(Year_of_Account__isnull=True).exclude(Year_of_Account='').values_list('Year_of_Account', flat=True).distinct().order_by('Year_of_Account'))
        return JsonResponse({"data": data})


@csrf_exempt
def agedDebtActionFilter(request):
    if request.method == "GET":
        data = list(set(filter(None, list(AgedDebtAction.objects.all().order_by('-id').values_list('aged_debt_action', flat=True)))))
        return JsonResponse({"data": data})


@csrf_exempt
def agedDebtCategoryFilter(request):
    if request.method == "GET":
        data = list(set(filter(None, list(AgedDebtCategory.objects.all().order_by('-id').values_list('aged_debt_category', flat=True)))))
        return JsonResponse({"data": data})


class AgedDebtDueManagementViewSet(viewsets.ModelViewSet):
    queryset = AgedDebtDueManagement.objects.all()
    serializer_class = AgedDebtDueManagementSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    # def list(self, request):
    #     queryset = self.get_queryset()

    #     page_number = int(request.query_params.get("skip", 0))
    #     page_size = int(request.query_params.get("pageSize", 20))
    #     skip = page_number * page_size

    #     # Apply pagination
    #     paginated_queryset = queryset[skip: skip + page_size]

    #     serializer = self.get_serializer(paginated_queryset, many=True)
    #     return Response({
    #         'data': serializer.data,
    #         'count': queryset.count()
    #     }, status=status.HTTP_200_OK)

    # def retrieve(self, request, pk=None):
    #     instance = self.get_object()
    #     serializer = self.get_serializer(instance)
    #     return Response(serializer.data, status=status.HTTP_200_OK)

    def create(self, request, *args, **kwargs):
        payload = request.data.copy()

        def _normalize_comment(value):
            if isinstance(value, list):
                value = " ".join([str(v) for v in value if v is not None])
            elif value is None:
                value = ""
            else:
                value = str(value)
            return value

        updated_by_value = payload.get("updated_by")
        file_name = payload.get("file_name")
        agedebt_id = payload.get("agedebt_id")
        status_usd = payload.get("status_usd")

        if "cc_comments" in payload:
            cc_val = _normalize_comment(payload.get("cc_comments"))
            cc_obj = {
                "no": 1,
                "value": cc_val,
                "date": dt.now().strftime("%d-%m-%Y %H:%M:%S"),
                "by": updated_by_value,
                "file_name": file_name if file_name else "-",
                "agedebt_id": agedebt_id if agedebt_id else "-",
                "status": status_usd if status_usd else "-",
            }
            payload["cc_comments"] = json.dumps([cc_obj])

        if "underwriter_comments" in payload:
            uw_val = _normalize_comment(payload.get("underwriter_comments"))
            uw_obj = {
                "no": 1,
                "value": uw_val,
                "date": dt.now().strftime("%d-%m-%Y %H:%M:%S"),
                "by": updated_by_value,
                "file_name": file_name if file_name else "-",
                "agedebt_id": agedebt_id if agedebt_id else "-",
                "status": status_usd if status_usd else "-",
            }
            payload["underwriter_comments"] = json.dumps([uw_obj])

        serializer = self.get_serializer(data=payload)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        payload = request.data.copy()
        def _normalize_comment(value):
            if isinstance(value, list):
                value = " ".join([str(v) for v in value if v is not None])
            elif value is None:
                value = ""
            else:
                value = str(value)
            return value
        updated_by_value = payload.get("updated_by")
        file_name = payload.get("file_name")
        agedebt_id = payload.get("agedebt_id")
        status_usd = payload.get("status_usd")
        instance = self.get_object()  # Get the object to update

        def _load_comments(raw):
            if not raw:
                return []
            # Already a list/dict
            if isinstance(raw, list):
                return raw
            if isinstance(raw, dict):
                return [raw]
            # Try JSON then Python literal
            for loader in (json.loads, ast.literal_eval):
                try:
                    parsed = loader(raw)
                    if isinstance(parsed, list):
                        return parsed
                    if isinstance(parsed, dict):
                        return [parsed]
                    if isinstance(parsed, str) and parsed.strip():
                        return [{
                            "no": 1,
                            "value": parsed,
                            "date": dt.now().strftime("%d-%m-%Y %H:%M:%S"),
                            "by": updated_by_value,
                            "file_name": file_name if file_name else "-",
                            "agedebt_id": agedebt_id if agedebt_id else "-",
                            "status": status_usd if status_usd else "-",
                        }]
                except Exception:
                    continue
            # Fallback: wrap raw string
            return [{
                "no": 1,
                "value": str(raw),
                "date": dt.now().strftime("%d-%m-%Y %H:%M:%S"),
                "by": updated_by_value,
                "file_name": file_name if file_name else "-",
                "agedebt_id": agedebt_id if agedebt_id else "-",
                "status": status_usd if status_usd else "-",
            }]

        old_cc_comments_value = _load_comments(instance.cc_comments)
        old_underwriter_comments_value = _load_comments(instance.underwriter_comments)
        if "cc_comments" in payload:
            cc_val = _normalize_comment(payload.get("cc_comments"))
            last_no = old_cc_comments_value[-1]['no'] if (old_cc_comments_value and isinstance(old_cc_comments_value[-1], dict) and 'no' in old_cc_comments_value[-1]) else 0
            cc_obj = {
                "no": last_no + 1,
                "value": cc_val,
                "date": dt.now().strftime("%d-%m-%Y %H:%M:%S"),
                "by": updated_by_value,
                "file_name": file_name if file_name else "-",
                "agedebt_id": agedebt_id if agedebt_id else "-",
                "status": status_usd if status_usd else "-",
            }
            payload["cc_comments"] = json.dumps(old_cc_comments_value + [cc_obj])

        if "underwriter_comments" in payload:
            uw_val = _normalize_comment(payload.get("underwriter_comments"))
            last_no_uw = old_underwriter_comments_value[-1]['no'] if (old_underwriter_comments_value and isinstance(old_underwriter_comments_value[-1], dict) and 'no' in old_underwriter_comments_value[-1]) else 0
            uw_obj = {
                "no": last_no_uw + 1,
                "value": uw_val,
                "date": dt.now().strftime("%d-%m-%Y %H:%M:%S"),
                "by": updated_by_value,
                "file_name": file_name if file_name else "-",
                "agedebt_id": agedebt_id if agedebt_id else "-",
                "status": status_usd if status_usd else "-",
            }
            payload["underwriter_comments"] = json.dumps(old_underwriter_comments_value + [uw_obj])
        serializer = self.get_serializer(instance, data=payload, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(["GET"])
def getBankDetailsList(request):
    if request.method == "GET":
        data = BankDetails.objects.all().order_by('-created_at')
        serializer = BankDetailsSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ChaserIndicatorViewSet(viewsets.ModelViewSet):
    model = ChaserIndicator
    serializer_class = ChaserIndicatorSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        doc = ChaserIndicator.objects.all()
        return doc

    def list(self, request):
        chaser_indicator = ChaserIndicator.objects.all()
        serializer = ChaserIndicatorSerializer(chaser_indicator, many=True)
        dataa = serializer.data
        return Response(dataa)

    def retrieve(self, request, pk=None):
        if pk:
            doc = ChaserIndicator.objects.get(id=pk)
            serializer = ChaserIndicatorSerializer(doc)
            dataa = serializer.data
            return Response(dataa)

    def create(self, request, *args, **kwargs):
        try:
            existing_chaser_indicator=ChaserIndicator.objects.all()
            if len(existing_chaser_indicator)>0:
                return Response({"msg": "There can be only one Chaser Indicator record"}, status=status.HTTP_400_BAD_REQUEST)

            data = request.data

            # Add addedDateAndTime to data item
            date=dt.now().isoformat()
            data["addedDateAndTime"] = date
            data["updatedDateAndTime"] = date

            serializer = ChaserIndicatorSerializer(data=data)

            if serializer.is_valid():
                new_doc = serializer.save()
                data_item = [data]
                fields = json.dumps(data_item)
                if len(fields) > 65535:
                    return Response(
                        {
                            "msg": "Length of the value for update_fields property got exceeded!"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                new_doc.updated_fields = fields
                new_doc.save()

                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        doc_object = self.get_object()
        doc_object.delete()
        return Response({"message": "Chaser Indicator deleted successfully"})

    def partial_update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            request_data = request.data
            request_data["updatedDateAndTime"]= datetime.datetime.now().isoformat()
            serializer = ChaserIndicatorSerializer(instance, data=request_data, partial=True)
            if serializer.is_valid():
                updated_doc=serializer.save()
                data_items=[]
                if instance.updated_fields:
                    old_list=instance.updated_fields
                    old_list=json.loads(old_list)
                    data_items.extend(old_list)
                data_items.append(request.data)
                changedFields = json.dumps(data_items)
                if len(changedFields)>65535:
                    return Response(
                        {"msg": "Length of the value for update_fields property got exceeded!"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                else:
                    updated_doc.updated_fields = changedFields
                    updated_doc.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
@api_view(["GET"])
def downloadAllocationStatus(request):
    if request.method == "GET":
        """
        Download allocation status data as xlsx file
        """
        try:
            queryset = AllocationStatus.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            allocation_status_data = []
            for record in queryset:
                allocation_status_data.append(
                    {
                        "id": record.id,
                        "Allocation Status": record.allocation_status,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated Date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(allocation_status_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Allocation Status")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Allocation Status"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=allocation_status_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
@api_view(["GET"])
def downloadExchangeRate(request):
    if request.method == "GET":
        """
        Download exchange rate data as xlsx file
        """
        try:
            queryset = BankExchangeRate.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            exchange_rate_data = []
            for record in queryset:
                exchange_rate_data.append(
                    {
                        "id": record.id,
                        "Month": record.month,
                        "Currency Code": record.currency_code,
                        "Exchange Rate": record.exchange_rate,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated Date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(exchange_rate_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Exchange Rate")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Exchange Rate"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=exchange_rate_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
@api_view(["GET"])
def downloadBankInfo(request):
    if request.method == "GET":
        """
        Download bank info data as xlsx file
        """
        try:
            queryset = BankDetails.objects.all().order_by("-created_at")

            # Convert queryset to list of dictionaries
            bank_info_data = []
            for record in queryset:
                bank_info_data.append(
                    {
                        "id": record.id,
                        "Region": record.region,
                        "Entity Number": record.entity_number,
                        "MSD Entity Number": record.msd_entity_number,
                        "Entity Name": record.entity_name,
                        "Bank Name": record.bank_name,
                        "Account Number": record.account_number,
                        "Account Opening Date": record.account_opening_date,
                        "Account Type": record.account_type,
                        "Currency": record.currency,
                        "MSD Account Number": record.msd_acct_number,
                        "MSD Account Name": record.msd_acct_name,
                        "Prime Bank Account": record.prime_bank_account,
                        "Created By": record.created_by,
                        "Created At": record.created_at.strftime("%d-%m-%Y %H:%M:%S")
                        if record.created_at else "",
                        "Updated By": record.updated_by,
                        "Updated At": record.updated_at.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updated_at else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(bank_info_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Bank Information")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Bank Information"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=bank_information_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
@api_view(["GET"])
def downloadBindingAgreement(request):
    if request.method == "GET":
        """
        Download binding agreement data as xlsx file
        """
        try:
            queryset = BindingAgreement.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            binding_agreement_data = []
            for record in queryset:
                binding_agreement_data.append(
                    {
                        "id": record.id,
                        "Binding Agreement Type": record.binding_aggrement_type,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(binding_agreement_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Binding Agreement")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Binding Agreement"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=binding_agreement_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
@api_view(["GET"])
def downloadBrokerInfo(request):
    if request.method == "GET":
        """
        Download broker info data as xlsx file
        """
        try:
            user = get_user(request)
            permissions = (
                user.user_permissions.permissions_list
                if user and user.user_permissions
                else {}
            )
            can_decrypt_broker = (
                permissions
                    .get("Broker Table", {})
                    .get("Upload decrypted file") == "Y"
            )
            queryset = BrokerInformation.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            broker_info_data = []
            for record in queryset:
                broker_info_data.append(
                    {
                        "id": record.id,
                        "Broker Name": record.broker_name,
                        "Broker": record.broker,
                        "Branch": record.branch,
                        "Duplicate Count": record.duplicate_count,
                        "SOA Received from Broker": record.soa_received_from_broker,
                        "Name": record.name,
                        "Email": (
                            user.get_decrypted_value(record.email)
                            if can_decrypt_broker
                            else record.email
                        ),
                        "Secondary Email": record.secondary_email,
                        "Phone Number": (
                            user.get_decrypted_value(record.phone_number)
                            if can_decrypt_broker
                            else record.phone_number
                        ),
                        "Broker Branch Location": record.broker_branch_location,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(broker_info_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Broker Information")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Broker Information"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=broker_information_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
@api_view(["GET"])
def downloadCorrectionType(request):
    if request.method == "GET":
        """
        Download correction type data as xlsx file
        """
        try:
            queryset = CorrectionType.objects.filter(archived=False).order_by('-addedDateAndTime')

            # Convert queryset to list of dictionaries
            correction_type_data = []
            for record in queryset:
                correction_type_data.append(
                    {
                        "id": record.id,
                        "Correction Reason": record.correction_type,
                        "Correction Description": record.correction_description,
                        "Allocation Status": record.allocation_status,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(correction_type_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Correction Reason")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Correction Reason"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=correction_reason_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(["GET"])
def downloadCorrectiveTransfer(request):
    if request.method == "GET":
        """
        Download corrective transfer data as xlsx file
        """
        try:
            queryset = CashTransfer.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            corrective_transfer_data = []
            for record in queryset:
                corrective_transfer_data.append(
                    {
                        "id": record.id,
                        "Cash Transfer Value": record.cash_transfer_value,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(corrective_transfer_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Corrective Transfer")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Corrective Transfer"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=corrective_transfer_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(["GET"])
def downloadCurrency(request):
    if request.method == "GET":
        """
        Download currency data as xlsx file
        """
        try:
            queryset = CurrencyDetails.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            currency_data = []
            for record in queryset:
                currency_data.append(
                    {
                        "id": record.id,
                        "Currency Code": record.currency_code,
                        "Country and Currency": record.country_and_currency,
                        "Symbol": record.symbol,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(currency_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Currency Details")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Currency Details"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=currency_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
@api_view(["GET"])
def downloadEntity(request):
    if request.method == "GET":
        """
        Download entity data as xlsx file
        """
        try:
            queryset = Entity.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            entity_data = []
            for record in queryset:
                entity_data.append(
                    {
                        "id": record.id,
                        "Entity Division": record.entity_divisions,
                        "Entity Name": record.entity_name,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(entity_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Entity Details")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Entity Details"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=entity_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        
@api_view(["GET"])
def downloadLOB(request):
    if request.method == "GET":
        """
        Download lob data as xlsx file
        """
        try:
            queryset = LOB.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            lob_data = []
            for record in queryset:
                lob_data.append(
                    {
                        "id": record.id,
                        "LOB Code": record.lob_code,
                        "LOB": record.lob,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(lob_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="LOB Details")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["LOB Details"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=lob_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(["GET"])
def downloadPolicyType(request):
    if request.method == "GET":
        """
        Download policy type data as xlsx file
        """
        try:
            queryset = PolicyType.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            policy_type_data = []
            for record in queryset:
                policy_type_data.append(
                    {
                        "id": record.id,
                        "Policy Start Letter": record.policy_start_letter,
                        "Policy Type": record.policy_type,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(policy_type_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Policy Type")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Policy Type"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=policy_type_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(["GET"])
def downloadSCMPartners(request):
    if request.method == "GET":
        """
        Download scm partners data as xlsx file
        """
        try:
            queryset = SCMPartners.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            scm_partners_data = []
            for record in queryset:
                scm_partners_data.append(
                    {
                        "id": record.id,
                        "Partner Name": record.partner_name,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(scm_partners_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="SCM Partners")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["SCM Partners"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=scm_partners_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(["GET"])
def downloadParticipatingInsurer(request):
    if request.method == "GET":
        """
        Download participating insurer data as xlsx file
        """
        try:
            queryset = ParticipatingInsurer.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            participating_insurer_data = []
            for record in queryset:
                participating_insurer_data.append(
                    {
                        "id": record.id,
                        "Participating Insurer": record.participating_insurer,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(participating_insurer_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Participating Insurer")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Participating Insurer"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=participating_insurer_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
@api_view(["GET"])
def downloadEscalation(request):
    if request.method == "GET":
        """
        Download escalation data as xlsx file
        """
        try:
            queryset = Escalation.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            escalation_data = []
            for record in queryset:
                escalation_data.append(
                    {
                        "id": record.id,
                        "Organization": record.organization,
                        "Transaction Type": record.transaction_type,
                        "Status": record.status,
                        "Escalation Level One": record.escalation_level_one,
                        "Escalation Level Two": record.escalation_level_two,
                        "Escalation Level Three": record.escalation_level_three,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(escalation_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Escalation")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Escalation"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=escalation_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
@api_view(["GET"])
def downloadTransactionCategory(request):
    if request.method == "GET":
        """
        Download transaction category as xlsx file
        """
        try:
            queryset = TransactionCategory.objects.all().order_by("-addedDateAndTime")

            # Convert queryset to list of dictionaries
            transaction_category_data = []
            for record in queryset:
                transaction_category_data.append(
                    {
                        "id": record.id,
                        "Transaction Category": record.txn_category,
                        "Category Description": record.category_description,
                        "Created By": record.created_by,
                        "Added Date and Time": record.addedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.addedDateAndTime else "",
                        "Updated By": record.updated_by,
                        "Updated date and Time": record.updatedDateAndTime.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updatedDateAndTime else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(transaction_category_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Transaction Category")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Transaction Category"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=transaction_category_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(["GET"])
def downloadTransactionStatus(request):
    if request.method == "GET":
        """
        Download transaction status as xlsx file
        """
        try:
            queryset = TransactionStatus.objects.all().order_by("-created_at")

            # Convert queryset to list of dictionaries
            transaction_status_data = []
            for record in queryset:
                transaction_status_data.append(
                    {
                        "id": record.id,
                        "Name": record.name,
                        "Created By": record.created_by,
                        "Created At": record.created_at.strftime("%d-%m-%Y %H:%M:%S")
                        if record.created_at else "",
                        "Updated By": record.updated_by,
                        "Updated At": record.updated_at.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updated_at else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(transaction_status_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Transaction Status")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Transaction Status"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=transaction_status_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AgedDebtActionViewSet(APIView):

    def get(self, request):
        data = AgedDebtAction.objects.all().order_by('-id')
        serializer = AgedDebtActionSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        data = request.data
        for i in data:
            max_id = AgedDebtAction.objects.aggregate(Max('id'))['id__max'] or 0
            i['id'] = max_id + 1
        serializer = AgedDebtActionSerializer(data=data, many=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def patch(self, request, pk):
        instance = AgedDebtAction.objects.get(id=pk)
        serializer = AgedDebtActionSerializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)


class AgedDebtCategoryViewSet(APIView):

    def get(self, request):
        data = AgedDebtCategory.objects.all().order_by('-id')
        serializer = AgedDebtCategorySerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        data = request.data
        for i in data:
            max_id = AgedDebtCategory.objects.aggregate(Max('id'))['id__max'] or 0
            i['id'] = max_id + 1
        serializer = AgedDebtCategorySerializer(data=data, many=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def patch(self, request, pk):
        instance = AgedDebtCategory.objects.get(id=pk)
        serializer = AgedDebtCategorySerializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
